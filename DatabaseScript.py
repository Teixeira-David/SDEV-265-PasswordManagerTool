"""
.______    _______ .___________.    ___         .______   .______       _______     ___       __  ___ 
|   _  \  |   ____||           |   /   \        |   _  \  |   _  \     |   ____|   /   \     |  |/  / 
|  |_)  | |  |__   `---|  |----`  /  ^  \       |  |_)  | |  |_)  |    |  |__     /  ^  \    |  '  /  
|   _  <  |   __|      |  |      /  /_\  \      |   _  <  |      /     |   __|   /  /_\  \   |    <   
|  |_)  | |  |____     |  |     /  _____  \     |  |_)  | |  |\  \----.|  |____ /  _____  \  |  .  \  
|______/  |_______|    |__|    /__/     \__\    |______/  | _| `._____||_______/__/     \__\ |__|\__\                                                                                                       
"""
"""
Project Name: Auto Belay Inspection Database
Developer: David Teixeira
Date: 11/02/2023
Abstract: This project is vol 0.0.2 to store and retrieve data for PPE Inspections
"""

# Import Python Libraries
import os
import getpass
import shutil
import sqlite3 
import configparser
import glob
import pandas as pd
import openpyxl
import subprocess
import sys
import re
import platform
from pyisemail import is_email
from datetime import datetime, timedelta, date
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tkinter import *
from tkinter import messagebox, filedialog
import tkinter as tk
from datetime import datetime 
import pyzipper
import tempfile


#######################################################################################################
# Database Class
#######################################################################################################
class Database():
    """
    Class Name: Database
    Class Description: This class is the main NOSQL database for the entire program
    """    
    def __init__(self, db_name=None, db_path=None, db_password=None):
        """ 
        Function Name: __init__
        Function Purpose: Instantiate the class objects and attributes 
        """           
        self.db_name = db_name
        self.db_path = db_path
        self.db_password = db_password
        self.conn = None
        self.currentUser = None    

    # Create a string of the database name
    def __str__(self):
        return str(self.db_name)
        
    # Property decorator object get function to access private db_name
    @property
    def db_name(self):
        return self._db_name

    # Property decorator object get function to access private db_path
    @property
    def db_path(self):
        return self._db_path
    
    # Property decorator object get function to access private db_password
    @property
    def db_password(self):
        return self._db_password

    # setter method 
    @db_name.setter 
    def db_name(self, db_name): 
        # Return true if specified object is of str type
        if not isinstance(db_name, str): 
            raise TypeError('Database Name must be a string') 
        # Check if the value is empty, otherwise check if the value is alpha
        if db_name.isspace(): 
            raise ValueError('Database Name cannot be empty') 
        # Set the attribute to the value if true
        elif db_name.isascii():
            self._db_name = db_name

    # setter method 
    @db_path.setter 
    def db_path(self, db_path): 
        # Return true if specified object is of str type
        if not isinstance(db_path, str): 
            raise TypeError('Database Path must be a string') 
        # Check if the value is empty, otherwise check if the value is alpha
        if db_path.isspace(): 
            raise ValueError('Database Path cannot be empty') 
        # Set the attribute to the value if true
        elif db_path.isascii():
            self._db_path = db_path
            
    # setter method 
    @db_password.setter 
    def db_password(self, db_password): 
        # Return true if specified object is of str type
        if not isinstance(db_password, str): 
            raise TypeError('Database Password must be a string') 
        # Check if the value is empty, otherwise check if the value is alpha
        if db_password.isspace(): 
            raise ValueError('Database Password cannot be empty') 
        # Set the attribute to the value if true
        elif db_password.isascii():
            self._db_password = db_password

    # setter method 
    def dbConnect(self):
        if not self.db_path:  # We're checking db_path instead of conn here.
            raise ValueError("Database path has not been set!")
        try:
            self.conn = sqlite3.connect(self.db_path)
            # Optionally handle password, etc. But SQLite doesn't have password by default.
        except sqlite3.Error as e:
            print(f"Error connecting to the database: {e}")
            self._conn = None

    def dbDisconnect(self):
        """ 
        Function Name: dbDisconnect
        Function Purpose: Disconnect the SQLite database 
        """          
        if self.conn:
            self.conn.close()
            self.conn = None

    def Get_Os_Desktop():
        """
        Function Name: Get_Os_Desktop
        Function Abstract: This function determines the os desktop dir and returns the desktop path
        """              
        operating_system = platform.system()
    
        if operating_system == 'Windows':
            # Code for Windows
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        elif operating_system == 'Linux' or operating_system == 'Darwin':
            # Code for Linux and macOS (Darwin)
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        else:
            # For unknown or unsupported operating systems
            print(f"Unsupported operating system: {operating_system}")
            return None
        
        return desktop_path        
    
    def resource_path(relative_path):
        """
        Function Name: resource_path
        Function Abstract: This function Get absolute path to resource, works for dev and for PyInstaller 
        """                
        try:
            # PyInstaller creates a temp folder and stores path in _MEIPASS
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")

        return os.path.join(base_path, relative_path)

    def Replace_Bundled_DB_With_Backup(bundled_db_path, backup_path):
        """
        Function Name: Replace_Bundled_DB_With_Backup
        Function Abstract: This function replaces the default database with the backup database snapshot.
        """            
        # Check the backup file's existence and its extension
        if not os.path.exists(backup_path) or not backup_path.endswith('.bak'):
            print(f"Invalid backup file: {backup_path}")
            return False

        # Create a temporary copy using the tempfile module
        temp_file_handle, temp_db_path = tempfile.mkstemp(suffix='.db')
        os.close(temp_file_handle)  # Close the file handle returned by mkstemp

        # Copy the original bundled DB to the temporary path
        shutil.copy2(bundled_db_path, temp_db_path)

        # Rename the .bak file to .db, then overwrite the temporary copy
        temp_backup_db_path = os.path.join(os.path.dirname(backup_path), 'temp_backup_database.db')
        os.rename(backup_path, temp_backup_db_path)

        try:
            # Copy the backup database over the temporary database
            shutil.copy2(temp_backup_db_path, temp_db_path)

            # Merge the backup database with the new tables into the temp database
            Database.Merge_Databases(Database, bundled_db_path, temp_db_path, temp_backup_db_path)
            
            # Replace the original bundled DB with the modified temp DB
            shutil.move(temp_db_path, bundled_db_path)
            return True
        
        except Exception as e:
            print(f"Error replacing the database: {e}")
            return False
        finally:
            # Clean up any temporary files
            if os.path.exists(temp_backup_db_path):
                os.remove(temp_backup_db_path)
            if os.path.exists(temp_db_path):
                os.remove(temp_db_path)
        
    def dbSet_Database_Attr(self):
        """ 
        Function Name: dbSet_Database_Attr
        Function Abstract: Set the database attributes from the configuration file.
        """      
        # Ensure RecordsDir exists on the desktop
        records_dir = os.path.join(self.Get_Os_Desktop(), "RecordsDir")
        if not os.path.exists(records_dir):
            os.makedirs(records_dir)
            
        # Set the records directory path
        Records.strRecords_path = records_dir
        BackupDb.dbBackup_path = records_dir
        
        # First get the database and config file from the primary package
        fConfigPath = Database.resource_path('config.ini')
        fDatabasePath = Database.resource_path('Database.db')
        
        # Check if the config path exists, if not None, proceed to extract the data from the config file
        if not fConfigPath:
            return

        # Make sure to comment this out when pushed to production    
        fConfigPath = Database.Config_File_Path(fConfigPath)
        fDatabasePath = Database.Config_File_Path(fDatabasePath)
                        
        # Create a ConfigParser instance
        config = configparser.ConfigParser()
        try:
            config.read(fConfigPath)
            dbName = config.get('Database', 'database')
            dbPswrd = config.get('Database', 'password')
        except configparser.NoSectionError as e:
            print(f"Section error: {e}")
            return
        except configparser.NoOptionError as e:
            print(f"Option error: {e}")
            return

        # Set the database password and find the oldest .bak file. If no .bak file, create the default database.
        # Instantiate the Database class
        db = Database(db_name=dbName, db_path=fDatabasePath, db_password=dbPswrd)
        Database.db_name = db.db_name
        Database.db_path = db.db_path
        Database.db_password = db.db_password        
        oldestBackupDB = BackupDb.Find_Oldest_BackupFile(BackupDb)            

        # Check if the database path exists, if not None, proceed to make the database
        if not oldestBackupDB:
            # Check if the DB file exists, if not, create and execute the database script
            if not os.path.exists(fDatabasePath): 
                # Write the database script
                db.dbCreateScript()
                    
                # Close the connection after creating it
                db.dbDisconnect()
        
        # If the .bak file exists, proceed to connect to it
        else:
            # Ensure you're disconnected from the DB before modifying the file
            db.dbDisconnect()
                        
            # Set the default database to the latest backup database snapshot.
            Database.Replace_Bundled_DB_With_Backup(fDatabasePath, oldestBackupDB[1])            
                
        # Connect to the database
        db.dbConnect()
            
    def Check_Table_Exists(cursor, table_name, db_alias="main"):
        """
        Function Name: Check_Table_Exists
        Function Purpose: Check if a table exists in the currently connected database.
        """        
        try:
            query = f"SELECT name FROM {db_alias}.sqlite_master WHERE type='table';"
            cursor.execute(query)
            tables = cursor.fetchall()
            # Print the tables for debugging purposes
            # print(f"Tables in {db_alias} database: {[table[0] for table in tables]}")
            
            # Check if the table name exists in the list of tables
            if any(table[0] == table_name for table in tables):
                return True
            else:
                return False

        except sqlite3.Error as e:
            print(f"Error checking table existence: {e}")
            return False

    def Check_View_Exists(cursor, view_name, db_alias="main"):
        """
        Function Name: Check_View_Exists
        Function Purpose: Check if a view exists in the currently connected database.
        """
        try:
            # Modify the query to check for views instead of tables
            query = f"SELECT name FROM {db_alias}.sqlite_master WHERE type='view';"
            cursor.execute(query)
            views = cursor.fetchall()

            # Check if the view name exists in the list of views
            if any(view[0] == view_name for view in views):
                return True
            else:
                return False

        except sqlite3.Error as e:
            print(f"Error checking view existence: {e}")
            return False
        
    def Check_ForeignKeys_Exists(cursor, table_foreign_key_pairs, db_alias="main"):
        """
        Function Name: Check_ForeignKeys_Exists
        Function Purpose: Check if specific foreign keys exist in the currently connected database.
        Params:
            cursor: The database cursor.
            table_foreign_key_pairs: A dictionary mapping table names to a list of foreign key names to check.
            db_alias: The database alias, default is "main".
        """
        try:
            for table, foreign_keys_to_check in table_foreign_key_pairs.items():
                # Retrieve the list of foreign keys for the table
                query_fk = f"PRAGMA {db_alias}.foreign_key_list('{table}');"
                cursor.execute(query_fk)
                foreign_keys = cursor.fetchall()

                # Check if the specified foreign keys exist in the table
                existing_foreign_keys = {fk_info[3] for fk_info in foreign_keys}  # Extract the column names of foreign keys
                for fk in foreign_keys_to_check:
                    if fk not in existing_foreign_keys:
                        print(f"Foreign key '{fk}' missing in table: {table}")
                        return False

            # If all specified foreign keys are found
            return True

        except sqlite3.Error as e:
            print(f"Error checking foreign key existence: {e}")
            return False
        
    def Merge_Databases(self, bundled_db_path, temp_db_path, backup_db_path):
        """
        Function Name: Merge_Databases
        Function Purpose: Merge new tables to the rolling database.
        """
        try:
            # Get the current user 
            Database.Get_Current_User(Database)
            
            # Connect to the temporary database
            self.conn = sqlite3.connect(temp_db_path)
            cursor = self.conn.cursor()
            
            # Drop all the new views in the backup database
            Database.dropAll_New_ViewsInDatabase(Database, "main", cursor)

            # Attach the backup database
            cursor.execute(f"ATTACH DATABASE '{backup_db_path}' AS backup_db")
            cursor.execute(f"ATTACH DATABASE '{bundled_db_path}' AS bundled_db")
            
            # Drop all the new views in the backup and bundled databases
            Database.dropAll_New_ViewsInDatabase(Database, "backup_db", cursor)
            Database.dropAll_New_ViewsInDatabase(Database, "bundled_db", cursor)

            # Drop all views in both backup_db and bundled_db
            # Database.dbDropViews(Database, "backup_db")
            
            # Create views in both databases
            # Database.Create_Base_Views(Database, "backup_db")
            
            # Dictionary mapping table names to functions to be executed if those tables are missing
            loadList = {
                "TLocations": Database.createLocationTables,
                "TAutoBelayReserviceReports": Database.createAutoBelay_ReserviceReports_Tables,
                "TRopes": Database.createRopeTables,
                "TRopeVisTextSelects": Database.createRopeVisTextSelTables,
                "TRopeVisualInspections": Database.createRopeVisInspectTables,
                "TRopePhysTextSelects": Database.createRopePhysTextSelTables,
                "TRopePhysicalInspections": Database.createRopePhysInspectTables,
                "TStandardRopeInspections": Database.createStandRopeInspectTables,                
                "TRopeInspections": Database.createRope_InspectionTables,
                "TRopeWallLocations": Database.createRope_WallLocationTables,
                "TRopeInspectors": Database.createRope_InspectorTables,
                "TConnectors": Database.createConnectorTables,
                "TConnectorVisMetalSelects": Database.createConnectorVisMetalSelTables,
                "TConnectorVisualInspections": Database.createConnectorVisInspectTables,
                "TConnectorPhysMetalSelects": Database.createConnectorPhysMetalSelTables,
                "TConnectorPhysicalInspections": Database.createConnectorPhysInspectTables,
                "TConnectorFunctions": Database.createConnectorFunctionsTables,
                "TConnectorFunctSelects": Database.createConnectorFunctSelTables,
                "TConnectorFunctionInspections": Database.createConnectorFunctInspectTables,
                "TStandardConnectorInspections": Database.createStandConnectorInspectTables,
                "TConnectorInspections": Database.createConnector_InspectionTables,
                "TConnectorWallLocations": Database.createConnector_WallLocationTables,
                "TConnectorInspectors": Database.createConnector_InspectorTables,
                "TBelayDevices": Database.createBelayDevicesTables,
                "TBelayDeviceVisMetalSelects": Database.createBelayDevicesVisMetalSelTables,
                "TBelayDeviceVisualInspections": Database.createBelayDevicesVisInspectTables,
                "TBelayDevicePhysMetalSelects": Database.createBelayDevicesPhysMetalSelTables,
                "TBelayDevicePhysicalInspections": Database.createBelayDevicesPhysInspectTables,
                "TBelayDeviceFunctions": Database.createBelayDevicesFunctionsTables,
                "TBelayDeviceFunctSelects": Database.createBelayDevicesFunctSelTables,
                "TBelayDeviceFunctionInspections": Database.createBelayDevicesFunctInspectTables,
                "TStandardBelayDeviceInspections": Database.createStandBelayDevicesInspectTables,
                "TBelayDeviceInspections": Database.createBelayDevices_InspectionTables,
                "TBelayDeviceWallLocations": Database.createBelayDevices_WallLocationTables,
                "TBelayDeviceInspectors": Database.createBelayDevices_InspectorTables,
                "TCustomRopeSystems": Database.createCustomRopeSystemsTables,
                "TRopeRetiredReports": Database.createRope_RetiredReports_Tables,
                "TConnectorRetiredReports": Database.createConnector_RetiredReports_Tables,
                "TBelayDeviceRetiredReports": Database.createBelayDevice_RetiredReports_Tables,
                "TBelayDeviceVisPlasticSelects": Database.createBelayDevicesVisPlasticSelTables,
                "TBelayDevicePhysPlasticSelects": Database.createBelayDevicesPhysPlasticSelTables,
                "TBelayDeviceFunctPlasticSelects": Database.createBelayDevicesFunctPlasticSelTables,
            }
            
            # Dictionary mapping views to functions to be executed if those tables are missing
            viewList = {
                "vAutoBelayFailedResults": Database.dbLoad_New_ABFail_View,
                "vAutoBelayMonitorResults": Database.dbLoad_New_ABMonitor_View,
                "vAutoBelayFailNowResults": Database.dbLoad_Now_ABFail_View,
                "vAutoBelayMonitorNowResults": Database.dbLoad_Now_ABMonitor_View,
                "vAutoBelayReserviceReports": Database.dbLoad_AutoBelay_Reservice_Reports,
                "vRopes": Database.dbLoad_vRopes,
                "vRopeDates": Database.dbLoad_vRopeDates,
                "vRopeLastNextInspectDates": Database.dbLoad_vRopeLastNextInspectDates,
                "vRopeInspectors": Database.dbLoad_vRopeInspectors,
                "vRopeWallLocations": Database.dbLoad_vRopeWallLocations,
                "vRopeWallLocationInspectDates": Database.dbLoad_vRopeWallLocationInspectDates,
                "vRopeInspectResults": Database.dbLoad_vRopeInspectResults,
                "vRopeFailedResults": Database.dbLoad_vRopeFailedResults,
                "vRopeMonitorResults": Database.dbLoad_vRopeMonitorResults,
                "vRopeFailNowResults": Database.dbLoad_vRopeFailNowResults,
                "vConnectors": Database.dbLoad_vConnectors,
                "vConnectorDates": Database.dbLoad_vConnectorDates,
                "vConnectorLastNextInspectDates": Database.dbLoad_vConnectorLastNextInspectDates,
                "vConnectorInspectors": Database.dbLoad_vConnectorInspectors,                
                "vConnectorWallLocations": Database.dbLoad_vConnectorWallLocations,
                "vConnectorWallLocationInspectDates": Database.dbLoad_vConnectorWallLocationInspectDates,
                "vConnectorInspectResults": Database.dbLoad_vConnectorInspectResults,
                "vConnectorFailedResults": Database.dbLoad_vConnectorFailedResults,
                "vConnectorMonitorResults": Database.dbLoad_vConnectorMonitorResults,
                "vConnectorFailNowResults": Database.dbLoad_vConnectorFailNowResults,
                "vBelayDevices": Database.dbLoad_vBelayDevices,
                "vBelayDeviceDates": Database.dbLoad_vBelayDeviceDates,
                "vBelayDeviceLastNextInspectDates": Database.dbLoad_vBelayDeviceLastNextInspectDates,
                "vBelayDeviceInspectors": Database.dbLoad_vBelayDeviceInspectors,
                "vBelayDeviceWallLocations": Database.dbLoad_vBelayDeviceWallLocations,
                "vBelayDeviceWallLocationInspectDates": Database.dbLoad_vBelayDeviceWallLocationInspectDates,
                "vBelayDeviceInspectResults": Database.dbLoad_vBelayDeviceInspectResults,
                "vBelayDeviceFailedResults": Database.dbLoad_vBelayDeviceFailedResults,
                "vBelayDeviceMonitorResults": Database.dbLoad_vBelayDeviceMonitorResults,
                "vBelayDeviceFailNowResults": Database.dbLoad_vBelayDeviceFailNowResults,
                "vABInspectID_OutForReservice": Database.dbLoad_ABInspectID_OutForReservice_View,
            }         
            
            # foreignKeyList = [
            #     Database.createUsrLog_Location_AutoBelay_ForeignKeys(self),
            #     Database.createRopes_ForeignKeys(self),
            #     Database.createConnector_ForeignKeys(self),
            #     Database.createBelayDevices_ForeignKeys(self),
            # ]
            
            for table, action in loadList.items():
                # Check if the table does not exist in the main (temporary) database
                if not self.Check_Table_Exists(cursor, table):
                    # Now, check if the table exists in the backup database
                    if not self.Check_Table_Exists(cursor, table, "backup_db") or self.Check_Table_Exists(cursor, table, "bundled_db"):
                        # Execute the function linked to the table name in loadList
                        action(Database)  
            
            # Load views from viewList
            for view, action in viewList.items():
                # If the view does not exist in the main database or is outdated
                if not self.Check_View_Exists(cursor, view, "main") or (not self.Check_View_Exists(cursor, view, "backup_db")) or (not self.Check_View_Exists(cursor, view, "bundled_db")):
                    # Create or update the view using the action function in the main database
                    action(Database)
                    
            # for view, action in viewList.items():
            #     # Check if the view does not exist in the main (temporary) database
            #     if not self.Check_View_Exists(cursor, view): # Make sure to switch to back to 'if not'
            #         # If the view exists in bundled_db but not in backup_db, or if it's outdated in backup_db
            #         if self.Check_View_Exists(cursor, view, "bundled_db") and (not self.Check_View_Exists(cursor, view, "backup_db") or self.Is_View_Outdated(cursor, view, "backup_db")):
            #             # Drop the view in backup_db if it exists
            #             if self.Check_View_Exists(cursor, view, "backup_db"):
            #                 self.Drop_View(cursor, view, "backup_db")

            #             # Create or update the view using the action function
            #             action(Database)                 
                
            # Commit any changes and detach the backup database
            self.conn.commit()
            cursor.execute("DETACH DATABASE 'backup_db'")
            cursor.execute("DETACH DATABASE 'bundled_db'")
        except sqlite3.Error as e:
            print(f"An error occurred during merging: {e}")
            self.conn.rollback()
        finally:
            cursor.close()
            self.conn.close()

    def Is_View_Outdated(self, cursor, view_name, target_db):
        """
        Function Name: Is_View_Outdated
        Function Purpose: Checks if a view in the target database is outdated compared to the main database.

        :param cursor: The database cursor.
        :param view_name: The name of the view to check.
        :param target_db: The name of the target database (e.g., 'backup_db').
        :return: True if the view is outdated, False otherwise.
        """
        # Retrieve the view definition from the main database
        main_view_def = self.Get_View_Definition(cursor, view_name, "main")

        # Retrieve the view definition from the target database
        target_view_def = self.Get_View_Definition(cursor, view_name, target_db)

        # Compare the two definitions
        return main_view_def != target_view_def

    def Get_View_Definition(self, cursor, view_name, db_name):
        """
        Function Name: Get_View_Definition
        Function Purpose: Retrieves the SQL definition of a view from the specified database.

        :param cursor: The database cursor.
        :param view_name: The name of the view.
        :param db_name: The name of the database (e.g., 'main', 'backup_db').
        :return: The SQL definition of the view as a string.
        """
        try:
            # Query to retrieve the view definition
            cursor.execute(f"SELECT sql FROM {db_name}.sqlite_master WHERE type='view' AND name=?", (view_name,))
            result = cursor.fetchone()
            return result[0] if result else None
        except sqlite3.Error as e:
            print(f"Error retrieving view definition from {db_name}: {e}")
            return None

    def Drop_View(self, cursor, view_name):
        """
        Function Name: Drop_View
        Function Purpose: Drops a view from the database.

        :param cursor: The database cursor.
        :param view_name: The name of the view to drop.
        """
        try:
            # SQL command to drop the view
            cursor.execute(f"DROP VIEW IF EXISTS {view_name}")
            print(f"View '{view_name}' dropped successfully.")
        except sqlite3.Error as e:
            print(f"Error dropping view '{view_name}': {e}")
            
    def Config_File_Path(start_path=None):
        """
        Function Name: Config_File_Path
        Function Purpose: Find a file path with the given name in the directory tree starting at start_path. 
        Returns the absolute path of the file if found in Main.
        """
        if 'Main' not in start_path:
            # Insert the 'Main' folder before the filename
            directory, filename = os.path.split(start_path)
            start_path = os.path.join(directory, 'Main', filename)
                    
        return start_path   
    
    def find_file(file_name, start_path=None):
        """
        Function Name: find_file
        Function Purpose: Find a file with the given name in the directory tree starting at start_path. 
        Returns the absolute path of the file if found, otherwise None.
        """
        for dirpath, dirnames, filenames in os.walk(start_path):
            if file_name in filenames:
                return os.path.abspath(os.path.join(dirpath, file_name))
        return None         
            
    def Get_Current_User(self):
            """
            Function Name: Get_Current_User
            Function Purpose: This function used to get the current user name and return to the database or other modules
            to store the name of the current user.
            """   
            # Select the dropbox item 
            currentUser = getpass.getuser()
            self.currentUser = currentUser
    
    def dbExeQuery(self, query, params=None):
        """ 
        Function Name: dbExeQuery
        Function Purpose: Execute queries against the database
        """       
        # # Check if the database connection is available
        # if Database.conn is None:
        #     print("Database is not connected.")
        #     return
        
        # Connect to the database
        Database.dbConnect(Database)
        
        # Assign cursor object
        cursor = self.conn.cursor()

        # Check if the parameters are assigned and execute the query
        if params is None:
            cursor.execute(query)
        else:
            cursor.execute(query, params)

        # Fetch all of the results from the query and close the cursor
        result = cursor.fetchall()
        cursor.close()

        # Return the results
        return result

    def dbExeStatement(self, sql):
        """
        Function Name: dbExeStatement
        Function Purpose: Execute the given SQL statement
        """
        # Check if the database connection is available
        if self.conn is None:
            print("Database is not connected.")
            return

        cursor = None
        try:
            cursor = self.conn.cursor()
            cursor.executescript(sql)
            # print("SQL executed successfully:", sql)
        except sqlite3.Error as e:
            print(f"Error executing SQL statement: {e}")
        finally:
            # Ensure the cursor is closed in any situation
            if cursor:
                cursor.close()
            
    def dbExeFunctStatement(self, sql, params=None):
        """
        Function Name: dbExeFunctStatement
        Function Purpose: Execute the given SQL statement
        """
        try:
            # Create the cursor object
            cursor = self.conn.cursor()
            
            # Check if the parameter is already defined or not
            if params is not None:
                cursor.execute(sql, params)              
            else:
                cursor.execute(sql)

            # Fetch the result
            result = cursor.fetchall()  
            
            # Close the cursor
            cursor.close()

            # Extract and return the inspection status if found, or None otherwise
            if result:
                return result
            else:
                return None     
            
        except sqlite3.Error as e:
            print(f"Error executing SQL statement: {e}")      
                                        
    def setNocount_On(self):
        """
        Function Name: setNocount_On
        Function Purpose: Set NOCOUNT ON for the SQLite database
        """
        sql = "PRAGMA count_changes = OFF;"
        self.dbExeStatement(sql)  # Removed the unnecessary self

    def setForeignKey_On(self):
        """
        Function Name: setForeignKey_On
        Function Purpose: Set Foreign Keys ON for the SQLite database
        """
        # Enable foreign keys support
        sql = "PRAGMA foreign_keys = ON;"
        self.dbExeStatement(sql)  # Removed the unnecessary self
        
    def dbInitialize(self):
        """ 
        Function Name: dbInitialize
        Function Purpose: Initialize the database by setting NOCOUNT ON and XACT_ABORT ON
        """        
        self.setNocount_On()  # Removed the unnecessary self
        self.setForeignKey_On()  # Removed the unnecessary self

    def dbInstance(self, ):
        """ 
        Function Name: dbInstance
        Function Purpose: Create an instance of the database to set the script 
        """        
        self.setNocount_On()  # Removed the unnecessary self
        self.setForeignKey_On()  # Removed the unnecessary self

    def dbCreateScript(self):
        """ 
        Function Name: dbCreateScript
        Function Purpose: Create the script for the sqlite database
        """        
        # Get the current user 
        self.Get_Current_User()
        
        # Connect to the database
        self.dbConnect()        
        
        # Initialize the database
        self.dbInitialize()
        
        # Create the tables and insert the base data into the database
        self.dbLoadTables()

        # Create the functions
        # self.dbLoadFunctions()
        
        # Load the foreign key constraints
        # self.dbLoadForeignKeys()
        
        # Create the views
        self.dbLoadViews()
        
    def dbBackup_Volume(dbPath):
        """ 
        Function Name: dbBackup_Volume
        Function Purpose: Back up the previously created database volume and append the new volume to the database
        """      
        # Ensure DatabaseDir exists in the CWD
        backup_dir = os.path.join(BackupDb.dbBackup_path, "BackupDir")
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)

        # Get the latest backup file
        latestBackup = max(glob.glob(f"{backup_dir}/*.bak"), key=os.path.getctime) if glob.glob(f"{backup_dir}/*.bak") else None

        try:
            # Create a backup file path with a timestamp
            if backup_dir is None:
                backupFile = f"{BackupDb.dbBackup_path}/{os.path.basename(dbPath)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.bak"
            else:
                # Use the given backup path
                backupFile = f"{backup_dir}/{os.path.basename(dbPath)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.bak"

            if latestBackup:
                # Copy the latest backup file to the new backup path
                shutil.copy2(latestBackup, backupFile)
            else:
                # Copy the database file to the backup path if no previous backups exist
                shutil.copy2(dbPath, backupFile)

            # Zip and encrypt the database back up file
            zip_file_path = os.path.join(backup_dir, f"{os.path.basename(dbPath)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip")
            Database.Zip_Encrypt(backupFile, zip_file_path, Database.db_password)

            # Remove the .bak file to save only the zip file
            os.remove(backupFile)
                    
        except Exception as e:
            print(f"Backup failed. Error: {str(e)}")

    def Zip_Encrypt(directory_path, zip_file_path, password):
        """ 
        Function Name: Zip_Encrypt
        Function Purpose: This function zips and encrypts the backup directory
        Parameters:
        - directory_path (str): The path to the directory to be zipped.
        - zip_file_path (str): The path where the zip file will be saved.
        - password (str): The password to encrypt the zip file.
        """
        with pyzipper.AESZipFile(zip_file_path, 'w', compression=pyzipper.ZIP_DEFLATED, encryption=pyzipper.WZ_AES) as zipf:
            if password:
                zipf.setpassword(bytes(password, 'utf-8'))
            zipf.write(directory_path, os.path.basename(directory_path))

    def Unzip_Decrypt(zip_file_path, dest_dir, password):
        """ 
        Function Name: Unzip_Decrypt
        Function Purpose: This function unzips and decrypts the backup directory
        Parameters:
        - zip_file_path (str): The path to the zip file to be unzipped.
        - dest_dir (str): The destination directory where the unzipped files will be saved.
        - password (str): The password to decrypt the zip file.
        """
        if not os.path.exists(zip_file_path):
            print(f"Zip file not found at: {zip_file_path}")
            return None

        if not os.path.exists(dest_dir):
            print(f"Destination directory not found. Creating: {dest_dir}")
            os.makedirs(dest_dir)

        try:
            with pyzipper.AESZipFile(zip_file_path) as zip_ref:
                if password:
                    zip_ref.pwd = bytes(password, 'utf-8')

                # Extract all files
                zip_ref.extractall(path=dest_dir)

                # Find the .bak file from extracted items
                extracted_files = [name for name in zip_ref.namelist()]
                bak_file = next((file for file in extracted_files if file.endswith('.bak')), None)

                if bak_file:
                    return os.path.join(dest_dir, bak_file)
                else:
                    print(f"No .bak file found among extracted items in {zip_file_path}")
                    return None

        except Exception as e:
            print(f"Failed to decrypt or unzip the directory. Error: {str(e)}")
            return None
                    
    def Create_XL_NewSheet(self, strSQLStringQuery, file):
        connection_established = False
        try:
            # Validate file path
            if not isinstance(file, (str, bytes, os.PathLike)):
                raise ValueError(f"Expected a path-like object for 'file', but got {type(file)}: {file}")

            # # Print file path for debugging
            # print(f"File path: {file}")

            # Load the excel workbook object
            wb = openpyxl.load_workbook(file)

            # Create the worksheet object with the name "New Sheet"
            ws = wb.create_sheet("New Sheet")

            # Check if connection is already established
            if not hasattr(self, 'conn') or self.conn is None:
                Database.dbConnect(self)
                connection_established = True

            # Ensure that the connection is valid
            if not self.conn:
                raise ValueError("Database connection is not established.")

            # Execute a query to store the data into the dataframe object
            df = pd.read_sql(strSQLStringQuery, self.conn)  # Replaced Database.conn with self.conn

            # Write data to new sheet
            for _, row in df.iterrows():
                ws.append(row.tolist())

            # Save the workbook to the file
            wb.save(file)

        # Display error message
        except Exception as err:
            print("Exception occurred while fetching the records:", err)
            return

        # # If all is good, close the connections
        # finally:
        #     # Close the connection if it was established in this method
        #     if connection_established:
        #         self.conn.close()  # Replaced Database.conn.close() with self.conn.close()

    def Create_XL_Write(self, strSQLStringQuery, file):
        connection_established = False
        try:
            # Check if connection is already established
            if not hasattr(self, 'conn') or self.conn is None:  # Use _conn if it's the internal variable
                # Connect to the database
                Database.dbConnect(self)  # Use self, assuming the method belongs to the same class
                connection_established = True

            # Validate file path
            if not isinstance(file, (str, bytes, os.PathLike)):
                raise ValueError(f"Expected a path-like object for 'file', but got {type(file)}: {file}")

            # Ensure self.conn returns the correct SQLite connection object
            connection = self.conn  # This is where you ensure the property is correct
            if not connection:
                raise ValueError("Database connection is not established or not correctly defined.")

            # Execute a query to store the data into the dataframe object
            df = pd.read_sql(strSQLStringQuery, connection)
            
            # Export the dataframe to the excel file
            df.to_excel(file, index=False)

        # Display error message
        except Exception as err:
            print("Exception occurred while fetching the records:", err)

        # # If all is good, close the connections
        # finally:
        #     # Close the connection if it was established in this method
        #     if connection_established:
        #         self.conn.close()

    def Save_Views_At_Start(self, fileName, fPath, curDir, sqlQuery): 
        """ 
        Function Name: Save_Views_At_Start
        Function Purpose: This function executes every day when the program is either running or not. This function
        pulls from the db specific views and downloads the views to a desired directory. Each file is saved as an excel
        file. 
        """            
        # Change to the desire directory
        os.chdir(fPath)
        
        # Check if the file exists
        if os.path.exists(fileName): 
            # Open and load the file
            if sys.platform.startswith('darwin'):
                # Create the excel file by passing in the sql query from the db for UNIX
                Database.Create_XL_NewSheet(self, sqlQuery, fileName)
                
            elif sys.platform.startswith('win32'):
                # Create the excel file by passing in the sql query from the db for Win
                Database.Create_XL_NewSheet(self, sqlQuery, fileName)
            
            elif sys.platform.startswith('linux'):
                # Create the excel file by passing in the sql query from the db for Linux
                Database.Create_XL_NewSheet(self, sqlQuery, fileName)   

        else:
            # Open and load the file
            if sys.platform.startswith('darwin'):
                # Create the excel file by passing in the sql query from the db for UNIX
                Database.Create_XL_Write(self, sqlQuery, fileName)
                
            elif sys.platform.startswith('win32'):
                # Create the excel file by passing in the sql query from the db for Win
                Database.Create_XL_Write(self, sqlQuery, fileName)

            elif sys.platform.startswith('linux'):
                # Create the excel file by passing in the sql query from the db for Linux
                Database.Create_XL_Write(self, sqlQuery, fileName)         
                # shutil.move(fileName, fPath)       

        # Close the file path and return to the parent directory
        os.chdir(curDir)      

    def Download_Files(self, user_triggered=True): 
        """ 
        Function Name: Download_Files
        Function Purpose: This function executes whenever the user clicks the 'Download Reports' button or 
                        daily when the program is either running or not. 
        Parameters:
        - user_triggered (bool): Whether this function was triggered by the user or automatically.
        """    
        # SQL Queries
        sqlParentQuery = "SELECT * FROM vAutoBelays"
        sqlStandInspectQuery = "SELECT * FROM vAutoBelayInspectResults"
        sqlWallLocateQuery = "SELECT * FROM vABWallLocationInspectDates"
        defaultextension = ".xlsx"

        # Get today's date
        dtmToday = datetime.date(datetime.now())
        dtmToday = dtmToday.strftime('%m_%d_%Y')

        # Set directory paths based on the desktop's location
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        records_dir = os.path.join(desktop_path, "RecordsDir")
        
        # Directory Paths
        fileABUnitDataDirPath = os.path.join(records_dir, 'ABUnitData')
        fileStandardInspectReportDirPath = os.path.join(records_dir, 'ABStandardInspectReports')
        fileWallLocationReportDirPath = os.path.join(records_dir, 'ABWallLocationReports')

        # Ensure directories exist
        for dir_path in [fileABUnitDataDirPath, fileStandardInspectReportDirPath, fileWallLocationReportDirPath]:
            os.makedirs(dir_path, exist_ok=True)

        # New File Names
        fNewParentFile = f'AutoBelayData{dtmToday}{defaultextension}'
        fNewStandardInspectReportFile = f'ABStandardInspectResults{dtmToday}{defaultextension}'
        fNewWallLocationReportFile = f'ABWallLocationResults{dtmToday}{defaultextension}'

        # File Paths
        ABUnitPath = os.path.join(fileABUnitDataDirPath, fNewParentFile)
        StandInspectPath = os.path.join(fileStandardInspectReportDirPath, fNewStandardInspectReportFile)
        WallLocatePath = os.path.join(fileWallLocationReportDirPath, fNewWallLocationReportFile)

        try:
            # Save the files
            Database.Save_Views_At_Start(self, fNewParentFile, fileABUnitDataDirPath, records_dir, sqlParentQuery)
            Database.Save_Views_At_Start(self, fNewStandardInspectReportFile, fileStandardInspectReportDirPath, records_dir, sqlStandInspectQuery)
            Database.Save_Views_At_Start(self, fNewWallLocationReportFile, fileWallLocationReportDirPath, records_dir, sqlWallLocateQuery)

            return [ABUnitPath, StandInspectPath, WallLocatePath]

        except Exception as err:
            if user_triggered:
                messagebox.showwarning(title='ERROR', message=f'Exception Error because {err}')
            else:
                # Log error or handle in some other way if not triggered by user
                pass

    def Download_Failed_Item_Files(self, user_triggered=True, sqlItemViewQuery=None, item_name=None): 
        """ 
        Function Name: Download_Failed_AB_Files
        Function Purpose: This function executes whenever the are any failed units/ppe equipment. 
        Parameters:
        - user_triggered (bool): Whether this function was triggered by the user or automatically.
        """    
        # SQL Queries
        sqlParentQuery = f"SELECT * FROM {sqlItemViewQuery}" 
        defaultextension = ".xlsx"

        # Get today's date
        dtmToday = datetime.date(datetime.now())
        dtmToday = dtmToday.strftime('%m_%d_%Y')

        # Set directory paths based on the desktop's location
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        records_dir = os.path.join(desktop_path, "RecordsDir")
        
        # Directory Paths
        fileFailedItemDataDirPath = os.path.join(records_dir, 'FailedItemsData')

        # Ensure directories exist
        for dir_path in [fileFailedItemDataDirPath]:
            os.makedirs(dir_path, exist_ok=True)

        # New File Names
        fNewParentFile = f'{item_name}{dtmToday}{defaultextension}'

        # File Paths
        ItemFailPath = os.path.join(fileFailedItemDataDirPath, fNewParentFile)

        try:
            # Save the files
            Database.Save_Views_At_Start(Database, fNewParentFile, fileFailedItemDataDirPath, records_dir, sqlParentQuery)
            return [ItemFailPath]

        except Exception as err:
            if user_triggered:
                messagebox.showwarning(title='ERROR', message=f'Exception Error because {err}')
            else:
                # Log error or handle in some other way if not triggered by user
                pass
            
    def Download_Failed_AB_Files(self, user_triggered=True): 
        """ 
        Function Name: Download_Failed_AB_Files
        Function Purpose: This function executes whenever the are any failed units/ppe equipment. 
        Parameters:
        - user_triggered (bool): Whether this function was triggered by the user or automatically.
        """    
        # SQL Queries
        sqlParentQuery = "SELECT * FROM vAutoBelayFailedResults" 
        defaultextension = ".xlsx"

        # Get today's date
        dtmToday = datetime.date(datetime.now())
        dtmToday = dtmToday.strftime('%m_%d_%Y')

        # Set directory paths based on the desktop's location
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        records_dir = os.path.join(desktop_path, "RecordsDir")
        
        # Directory Paths
        fileABFailedUnitDataDirPath = os.path.join(records_dir, 'ABFailedUnitData')

        # Ensure directories exist
        for dir_path in [fileABFailedUnitDataDirPath]:
            os.makedirs(dir_path, exist_ok=True)

        # New File Names
        fNewParentFile = f'AutoBelayFailStatusData{dtmToday}{defaultextension}'

        # File Paths
        ABFailUnitPath = os.path.join(fileABFailedUnitDataDirPath, fNewParentFile)

        try:
            # Save the files
            Database.Save_Views_At_Start(Database, fNewParentFile, fileABFailedUnitDataDirPath, records_dir, sqlParentQuery)
            return [ABFailUnitPath]

        except Exception as err:
            if user_triggered:
                messagebox.showwarning(title='ERROR', message=f'Exception Error because {err}')
            else:
                # Log error or handle in some other way if not triggered by user
                pass

    def Download_Reserviced_AB_Files(self, user_triggered=True): 
        """ 
        Function Name: Download_Reserviced_AB_Files
        Function Purpose: This function executes whenever the are any reserviced equipment. 
        Parameters:
        - user_triggered (bool): Whether this function was triggered by the user or automatically.
        """    
        # SQL Queries
        sqlParentQuery = "SELECT * FROM vAutoBelayReserviceReports"
        defaultextension = ".xlsx"

        # Get today's date
        dtmToday = datetime.date(datetime.now())
        dtmToday = dtmToday.strftime('%m_%d_%Y')

        # Set directory paths based on the desktop's location
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        records_dir = os.path.join(desktop_path, "RecordsDir")
        
        # Directory Paths
        fileABOutForReserviceDataDirPath = os.path.join(records_dir, 'ABOutForReserviceUnitData')

        # Ensure directories exist
        for dir_path in [fileABOutForReserviceDataDirPath]:
            os.makedirs(dir_path, exist_ok=True)

        # New File Names
        fNewParentFile = f'AutoBelayOutForReserviceData{dtmToday}{defaultextension}'

        # File Paths
        ABFailUnitPath = os.path.join(fileABOutForReserviceDataDirPath, fNewParentFile)

        try:
            # Save the files
            Database.Save_Views_At_Start(Database, fNewParentFile, fileABOutForReserviceDataDirPath, records_dir, sqlParentQuery)
            return [ABFailUnitPath]

        except Exception as err:
            if user_triggered:
                messagebox.showwarning(title='ERROR', message=f'Exception Error because {err}')
            else:
                # Log error or handle in some other way if not triggered by user
                pass
                        
    def dbDropTables(self):
        """ 
        Function Name: dbDropTables
        Function Purpose: Create the drop tables for the database
        """           
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Execute the drop tables    
            sql = """
            DROP TABLE IF EXISTS TInspectionStatus;
            DROP TABLE IF EXISTS TInspectionTypes;
            DROP TABLE IF EXISTS TMetallicInspections;
            DROP TABLE IF EXISTS TTextileInspections;
            DROP TABLE IF EXISTS TPlasticInspections;
            DROP TABLE IF EXISTS TCarabiners;
            DROP TABLE IF EXISTS TCarabVisMetalSelects;
            DROP TABLE IF EXISTS TCarabinerVisualInspections;
            DROP TABLE IF EXISTS TCarabPhysMetalSelects;
            DROP TABLE IF EXISTS TCarabinerPhysicalInspections;
            DROP TABLE IF EXISTS TCarabinerFunctions;
            DROP TABLE IF EXISTS TCarabFunctSelects;
            DROP TABLE IF EXISTS TCarabinerFunctionInspections;
            DROP TABLE IF EXISTS TStandardCarabinerInspections;
            DROP TABLE IF EXISTS TDeviceHandles;
            DROP TABLE IF EXISTS THandleVisMetalSelects;
            DROP TABLE IF EXISTS THandleVisualInspections;
            DROP TABLE IF EXISTS THandlePhysMetalSelects;
            DROP TABLE IF EXISTS THandlePhysicalInspections;
            DROP TABLE IF EXISTS TStandardHandleInspections;
            DROP TABLE IF EXISTS TCaseHousings;
            DROP TABLE IF EXISTS TCaseCompSelects;
            DROP TABLE IF EXISTS TCaseVisMetalSelects;
            DROP TABLE IF EXISTS TCaseHousingVisualInspections;
            DROP TABLE IF EXISTS TCasePhysMetalSelects;
            DROP TABLE IF EXISTS TCaseHousingPhysicalInspections;
            DROP TABLE IF EXISTS TStandardCaseHousingInspections;
            DROP TABLE IF EXISTS TBrakeHousings;
            DROP TABLE IF EXISTS TBrakeCompSelects;
            DROP TABLE IF EXISTS TBrakeVisMetalSelects;
            DROP TABLE IF EXISTS TBrakeHousingVisualInspections;
            DROP TABLE IF EXISTS TBrakePhysMetalSelects;
            DROP TABLE IF EXISTS TBrakeHousingPhysicalInspections;
            DROP TABLE IF EXISTS TStandardBrakeHousingInspections;
            DROP TABLE IF EXISTS TLanyards;
            DROP TABLE IF EXISTS TLanVisTextSelects;
            DROP TABLE IF EXISTS TLanyardVisualInspections;
            DROP TABLE IF EXISTS TLanPhysTextSelects;
            DROP TABLE IF EXISTS TLanyardPhysicalInspections;
            DROP TABLE IF EXISTS TRetractFunctions;
            DROP TABLE IF EXISTS TRetractFunctSelects;
            DROP TABLE IF EXISTS TLanyardRetractFunctionInspections;
            DROP TABLE IF EXISTS TStandardLanyardInspections;
            DROP TABLE IF EXISTS TStates;
            DROP TABLE IF EXISTS TInspectors;
            DROP TABLE IF EXISTS TLogins;
            DROP TABLE IF EXISTS TUserLogins;
            DROP TABLE IF EXISTS TAdminUsers;
            DROP TABLE IF EXISTS TGymLocations;
            DROP TABLE IF EXISTS TAutoBelays;
            DROP TABLE IF EXISTS TWallLocations;
            DROP TABLE IF EXISTS TLocations;
            DROP TABLE IF EXISTS TAutoBelayInspectors;
            DROP TABLE IF EXISTS TAutoBelayWallLocations;
            DROP TABLE IF EXISTS TStandardInspections;
            DROP TABLE IF EXISTS TAutoBelayInspections;
            DROP TABLE IF EXISTS TUserSentReports;
            DROP TABLE IF EXISTS TRopes;
            DROP TABLE IF EXISTS TRopeVisTextSelects;
            DROP TABLE IF EXISTS TRopeVisualInspections;
            DROP TABLE IF EXISTS TRopePhysTextSelects;
            DROP TABLE IF EXISTS TRopePhysicalInspections;
            DROP TABLE IF EXISTS TStandardRopeInspections;
            DROP TABLE IF EXISTS TRopeInspections;
            DROP TABLE IF EXISTS TRopeWallLocations;
            DROP TABLE IF EXISTS TRopeInspectors;
            DROP TABLE IF EXISTS TConnectors;
            DROP TABLE IF EXISTS TConnectorVisMetalSelects;
            DROP TABLE IF EXISTS TConnectorVisualInspections;
            DROP TABLE IF EXISTS TConnectorPhysMetalSelects;
            DROP TABLE IF EXISTS TConnectorPhysicalInspections;
            DROP TABLE IF EXISTS TConnectorFunctSelects;
            DROP TABLE IF EXISTS TConnectorFunctionInspections;
            DROP TABLE IF EXISTS TStandardConnectorInspections;
            DROP TABLE IF EXISTS TConnectorInspections;
            DROP TABLE IF EXISTS TConnectorWallLocations;
            DROP TABLE IF EXISTS TConnectorInspectors;
            DROP TABLE IF EXISTS TBelayDevices;
            DROP TABLE IF EXISTS TBelayDevicePhysMetalSelects;
            DROP TABLE IF EXISTS TBelayDeviceVisualInspections;
            DROP TABLE IF EXISTS TBelayDevicePhysMetalSelects;
            DROP TABLE IF EXISTS TBelayDevicePhysicalInspections;
            DROP TABLE IF EXISTS TBelayDeviceFunctSelects;
            DROP TABLE IF EXISTS TBelayDeviceFunctionInspections;
            DROP TABLE IF EXISTS TStandardBelayDeviceInspections;
            DROP TABLE IF EXISTS TBelayDeviceInspections;
            DROP TABLE IF EXISTS TBelayDeviceWallLocations;
            DROP TABLE IF EXISTS TBelayDeviceInspectors;            
            """
            # Execute the script 
            Database.dbExeStatement(sql)            
            print("Tables dropped successfully.")

        except sqlite3.Error as e:
            print(f"Error creating table: {e}")          

    def dbDropProcedures(self):
        """ 
        Function Name: dbDropProcedures
        Function Purpose: Create the drop stored procedures for the database
        """             
        try:            
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
            
            # Execute the drop procedure        
            sql = """
            DROP PROCEDURE IF EXISTS uspAddNewInspector;
            DROP PROCEDURE IF EXISTS uspAddNewAutoBelay;
            DROP PROCEDURE IF EXISTS uspAddNewWallLocation;
            DROP PROCEDURE IF EXISTS uspAddNewCarabVisSelection;
            DROP PROCEDURE IF EXISTS uspAddNewCarabinerVisualInspection;
            DROP PROCEDURE IF EXISTS uspAddNewCarabPhysSelection;
            DROP PROCEDURE IF EXISTS uspAddNewCarabinerPhysicalInspection;
            DROP PROCEDURE IF EXISTS uspAddNewCarabFunctSelection;
            DROP PROCEDURE IF EXISTS uspAddNewCarabinerFunctionInspection;
            DROP PROCEDURE IF EXISTS uspAddNewStandardCarabinerInspection;
            DROP PROCEDURE IF EXISTS uspAddNewHandleVisSelection;
            DROP PROCEDURE IF EXISTS uspAddNewHandleVisualInspection;
            DROP PROCEDURE IF EXISTS uspAddNewHandlePhysSelection;
            DROP PROCEDURE IF EXISTS uspAddNewHandlePhysicalInspection;
            DROP PROCEDURE IF EXISTS uspAddNewStandardHandleInspection;
            DROP PROCEDURE IF EXISTS uspAddNewCaseCompSelection;
            DROP PROCEDURE IF EXISTS uspAddNewCaseVisSelection;
            DROP PROCEDURE IF EXISTS uspAddNewCaseHousingVisualInspection;
            DROP PROCEDURE IF EXISTS uspAddNewCasePhysSelection;
            DROP PROCEDURE IF EXISTS uspAddNewCaseHousingPhysicalInspection;
            DROP PROCEDURE IF EXISTS uspAddNewStandardCaseHousingInspection;
            DROP PROCEDURE IF EXISTS uspAddNewBrakeCompSelection;
            DROP PROCEDURE IF EXISTS uspAddNewBrakeVisSelection;
            DROP PROCEDURE IF EXISTS uspAddNewBrakeHousingVisualInspection;
            DROP PROCEDURE IF EXISTS uspAddNewBrakePhysSelection;
            DROP PROCEDURE IF EXISTS uspAddNewBrakeHousingPhysicalInspection;
            DROP PROCEDURE IF EXISTS uspAddNewStandardBrakeHousingInspection;
            DROP PROCEDURE IF EXISTS uspAddNewLanyardLen;
            DROP PROCEDURE IF EXISTS uspAddNewLanyardVisSelection;
            DROP PROCEDURE IF EXISTS uspAddNewLanyardVisualInspection;
            DROP PROCEDURE IF EXISTS uspAddNewLanyardPhysSelection;
            DROP PROCEDURE IF EXISTS uspAddNewLanyardPhysicalInspection;
            DROP PROCEDURE IF EXISTS uspAddNewRetractFunctSelection;
            DROP PROCEDURE IF EXISTS uspAddNewLanyardFunctionInspection;
            DROP PROCEDURE IF EXISTS uspAddNewStandardLanyardInspection;
            DROP PROCEDURE IF EXISTS uspAddNewStandardInspection;
            DROP PROCEDURE IF EXISTS uspAddNewAutoBelayInspection;
            DROP PROCEDURE IF EXISTS uspAddNewState;
            DROP PROCEDURE IF EXISTS uspAddNewLogin;
            DROP PROCEDURE IF EXISTS uspAddNewPassword;
            DROP PROCEDURE IF EXISTS uspAddNewUserLogin;
            DROP PROCEDURE IF EXISTS uspAddNewAdminUser;
            DROP PROCEDURE IF EXISTS uspUpdateAutoBelayDevice;
            DROP PROCEDURE IF EXISTS uspUpdateAutoBelayDates;
            DROP PROCEDURE IF EXISTS uspAddNewAutoBelayWallLocation;
            DROP PROCEDURE IF EXISTS uspAddNewAutoBelayInspector;
            DROP PROCEDURE IF EXISTS uspAddNewUserSentReport;
            DROP PROCEDURE IF EXISTS uspUpdateAdminUserCred;
            DROP PROCEDURE IF EXISTS uspRemoveAdminUser;
            DROP PROCEDURE IF EXISTS uspBackupDatabase;            
            """
            # Execute the script 
            Database.dbExeStatement(sql)    
            print("Procedure dropped successfully.")
        
        except sqlite3.Error as e:
            print(f"Error creating procedures: {e}")    
            
    def dbDrop_Z_Tables(self):
        """ 
        Function Name: dbDrop_Z_Tables
        Function Purpose: Create the drop audit tables for the database
        """             
        try:         
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Execute the drop audit tables        
            sql = """
            DROP TABLE IF EXISTS Z_TInspectionStatus;
            DROP TABLE IF EXISTS Z_TInspectionTypes;
            DROP TABLE IF EXISTS Z_TMetallicInspections;
            DROP TABLE IF EXISTS Z_TTextileInspections;
            DROP TABLE IF EXISTS Z_TPlasticInspections;
            DROP TABLE IF EXISTS Z_TCarabiners;
            DROP TABLE IF EXISTS Z_TCarabVisMetalSelects;
            DROP TABLE IF EXISTS Z_TCarabinerVisualInspections;
            DROP TABLE IF EXISTS Z_TCarabPhysMetalSelects;
            DROP TABLE IF EXISTS Z_TCarabinerPhysicalInspections;
            DROP TABLE IF EXISTS Z_TCarabinerFunctions;
            DROP TABLE IF EXISTS Z_TCarabFunctSelects;
            DROP TABLE IF EXISTS Z_TCarabinerFunctionInspections;
            DROP TABLE IF EXISTS Z_TStandardCarabinerInspections;
            DROP TABLE IF EXISTS Z_TDeviceHandles;
            DROP TABLE IF EXISTS Z_THandleVisMetalSelects;
            DROP TABLE IF EXISTS Z_THandleVisualInspections;
            DROP TABLE IF EXISTS Z_THandlePhysMetalSelects;
            DROP TABLE IF EXISTS Z_THandlePhysicalInspections;
            DROP TABLE IF EXISTS Z_TStandardHandleInspections;
            DROP TABLE IF EXISTS Z_TCaseHousings;
            DROP TABLE IF EXISTS Z_TCaseCompSelects;
            DROP TABLE IF EXISTS Z_TCaseVisMetalSelects;
            DROP TABLE IF EXISTS Z_TCaseHousingVisualInspections;
            DROP TABLE IF EXISTS Z_TCasePhysMetalSelects;
            DROP TABLE IF EXISTS Z_TCaseHousingPhysicalInspections;
            DROP TABLE IF EXISTS Z_TStandardCaseHousingInspections;
            DROP TABLE IF EXISTS Z_TBrakeHousings;
            DROP TABLE IF EXISTS Z_TBrakeCompSelects;
            DROP TABLE IF EXISTS Z_TBrakeVisMetalSelects;
            DROP TABLE IF EXISTS Z_TBrakeHousingVisualInspections;
            DROP TABLE IF EXISTS Z_TBrakePhysMetalSelects;
            DROP TABLE IF EXISTS Z_TBrakeHousingPhysicalInspections;
            DROP TABLE IF EXISTS Z_TStandardBrakeHousingInspections;
            DROP TABLE IF EXISTS Z_TLanyards;
            DROP TABLE IF EXISTS Z_TLanVisTextSelects;
            DROP TABLE IF EXISTS Z_TLanyardVisualInspections;
            DROP TABLE IF EXISTS Z_TLanPhysTextSelects;
            DROP TABLE IF EXISTS Z_TLanyardPhysicalInspections;
            DROP TABLE IF EXISTS Z_TRetractFunctions;
            DROP TABLE IF EXISTS Z_TRetractFunctSelects;
            DROP TABLE IF EXISTS Z_TLanyardRetractFunctionInspections;
            DROP TABLE IF EXISTS Z_TStandardLanyardInspections;
            DROP TABLE IF EXISTS Z_TStates;
            DROP TABLE IF EXISTS Z_TInspectors;
            DROP TABLE IF EXISTS Z_TLogins;
            DROP TABLE IF EXISTS Z_TUserLogins;
            DROP TABLE IF EXISTS Z_TAdminUsers;
            DROP TABLE IF EXISTS Z_TGymLocations;
            DROP TABLE IF EXISTS Z_TAutoBelays;
            DROP TABLE IF EXISTS Z_TWallLocations;
            DROP TABLE IF EXISTS Z_TLocations;
            DROP TABLE IF EXISTS Z_TAutoBelayInspectors;
            DROP TABLE IF EXISTS Z_TAutoBelayWallLocations;
            DROP TABLE IF EXISTS Z_TStandardInspections;
            DROP TABLE IF EXISTS Z_TAutoBelayInspections;
            DROP TABLE IF EXISTS Z_TUserSentReports;
            DROP TABLE IF EXISTS Z_TRopes;
            DROP TABLE IF EXISTS Z_TRopeVisTextSelects;
            DROP TABLE IF EXISTS Z_TRopeVisualInspections;
            DROP TABLE IF EXISTS Z_TRopePhysTextSelects;
            DROP TABLE IF EXISTS Z_TRopePhysicalInspections;
            DROP TABLE IF EXISTS Z_TStandardRopeInspections;
            DROP TABLE IF EXISTS Z_TRopeInspections;
            DROP TABLE IF EXISTS Z_TRopeWallLocations;
            DROP TABLE IF EXISTS Z_TRopeInspectors;
            DROP TABLE IF EXISTS Z_TConnectors;
            DROP TABLE IF EXISTS Z_TConnectorVisMetalSelects;
            DROP TABLE IF EXISTS Z_TConnectorVisualInspections;
            DROP TABLE IF EXISTS Z_TConnectorPhysMetalSelects;
            DROP TABLE IF EXISTS Z_TConnectorPhysicalInspections;
            DROP TABLE IF EXISTS Z_TConnectorFunctSelects;
            DROP TABLE IF EXISTS Z_TConnectorFunctionInspections;
            DROP TABLE IF EXISTS Z_TStandardConnectorInspections;
            DROP TABLE IF EXISTS Z_TConnectorInspections;
            DROP TABLE IF EXISTS Z_TConnectorWallLocations;
            DROP TABLE IF EXISTS Z_TConnectorInspectors;     
            DROP TABLE IF EXISTS Z_TBelayDevices;
            DROP TABLE IF EXISTS Z_TBelayDevicePhysMetalSelects;
            DROP TABLE IF EXISTS Z_TBelayDeviceVisualInspections;
            DROP TABLE IF EXISTS Z_TBelayDevicePhysMetalSelects;
            DROP TABLE IF EXISTS Z_TBelayDevicePhysicalInspections;
            DROP TABLE IF EXISTS Z_TBelayDeviceFunctSelects;
            DROP TABLE IF EXISTS Z_TBelayDeviceFunctionInspections;
            DROP TABLE IF EXISTS Z_TStandardBelayDeviceInspections;
            DROP TABLE IF EXISTS Z_TBelayDeviceInspections;
            DROP TABLE IF EXISTS Z_TBelayDeviceWallLocations;
            DROP TABLE IF EXISTS Z_TBelayDeviceInspectors;                   
            """
            # Execute the script 
            Database.dbExeStatement(sql)             
            print("Audit Tables and procedure dropped successfully.")

        except sqlite3.Error as e:
            print(f"Error creating table: {e}")            

    def dbDropViews(self, target_db="main"):
        """ 
        Function Name: dbDropViews
        Function Purpose: Create the drop views for the database
        """             
        # Create the view names
        sqlViewName = ['vInspectors', 
                        'vInspectStatus',
                        'vInspectTypes',
                        'vMetallicInspect',
                        'vPlasticInspect',
                        'vTextileInspect',
                        'vCarabinerType',
                        'vCarabinerFunct',
                        'vDeviceHandleComp',
                        'vCaseHouseComp',
                        'vBrakeHouseComp',
                        'vRetractFunctions',
                        'vWallLocations',
                        'vLogins',
                        'vAutoBelays',
                        'vAutoBelayDates',
                        'vAutoBelayLastNextInspectDates',
                        'vAutoBelayInspectors',
                        'vAutoBelayWallLocations',
                        'vABWallLocationInspectDates',
                        'vABServiceDates',
                        'vAutoBelayInspectResults',
                        'vAutoBelayFailedResults',
                        'vAutoBelayMonitorResults',
                        'vAutoBelayFailNowResults',
                        'vAutoBelayReserviceReports',
                        'vRopes',
                        'vRopeDates',
                        'vRopeLastNextInspectDates',
                        'vRopeInspectors',
                        'vRopeWallLocations',
                        'vRopeWallLocationInspectDates',
                        'vRopeInspectResults',
                        'vRopeFailedResults',
                        'vRopeMonitorResults',
                        'vRopeFailNowResults',
                        'vConnectors',
                        'vConnectorDates',
                        'vConnectorLastNextInspectDates',
                        'vConnectorInspectors',
                        'vConnectorWallLocations',
                        'vConnectorWallLocationInspectDates',
                        'vConnectorInspectResults',
                        'vConnectorFailedResults',
                        'vConnectorMonitorResults',
                        'vConnectorFailNowResults',
                        'vBelayDevices',
                        'vBelayDeviceDates',
                        'vBelayDeviceLastNextInspectDates',
                        'vBelayDeviceInspectors',
                        'vBelayDeviceWallLocations',
                        'vBelayDeviceWallLocationInspectDates',
                        'vBelayDeviceInspectResults',
                        'vBelayDeviceFailedResults',
                        'vBelayDeviceMonitorResults',
                        'vBelayDeviceFailNowResults',
                        'vABInspectID_OutForReservice',
                        ]
        
        # Build the SQL command to drop views
        sql = ";\n".join(f"DROP VIEW IF EXISTS {target_db + '.' if target_db != 'main' else ''}{view}" for view in sqlViewName) + ";"

        try:
            # Check if connected to the database
            if not self.conn:
                raise Exception("Database is not connected.")

            # Execute the SQL statement
            if target_db == "main":
                # Use dbExeStatement for the main database
                self.dbExeStatement(sql)
            else:
                # Direct execution for other databases
                cursor = self.conn.cursor()
                try:
                    # Execute the drop view statement
                    cursor.executescript(sql)
                    
                    # print(f"SQL executed successfully for {target_db} database:", sql)
                except sqlite3.Error as e:
                    print(f"Error executing SQL statement: {e}")
                finally:
                    cursor.close()

            print(f"Views dropped successfully in {target_db} database.")

        except sqlite3.Error as e:
            print(f"Error dropping views in {target_db} database: {e}")

    def dropAll_New_ViewsInDatabase(self, db_alias, cursor):
        """
        Drops all views in the specified database.
        
        :param db_alias: Alias of the database where views will be dropped.
        :param cursor: Cursor object for executing SQL commands.
        """
        # Create the view names
        sqlViewName = ['vAutoBelayFailedResults',
                        'vAutoBelayMonitorResults',
                        'vAutoBelayFailNowResults',
                        'vAutoBelayReserviceReports',
                        'vRopes',
                        'vRopeDates',
                        'vRopeLastNextInspectDates',
                        'vRopeInspectors',
                        'vRopeWallLocations',
                        'vRopeWallLocationInspectDates',
                        'vRopeInspectResults',
                        'vRopeFailedResults',
                        'vRopeMonitorResults',
                        'vRopeFailNowResults',
                        'vConnectors',
                        'vConnectorDates',
                        'vConnectorLastNextInspectDates',
                        'vConnectorInspectors',
                        'vConnectorWallLocations',
                        'vConnectorWallLocationInspectDates',
                        'vConnectorInspectResults',
                        'vConnectorFailedResults',
                        'vConnectorMonitorResults',
                        'vConnectorFailNowResults',
                        'vBelayDevices',
                        'vBelayDeviceDates',
                        'vBelayDeviceLastNextInspectDates',
                        'vBelayDeviceInspectors',
                        'vBelayDeviceWallLocations',
                        'vBelayDeviceWallLocationInspectDates',
                        'vBelayDeviceInspectResults',
                        'vBelayDeviceFailedResults',
                        'vBelayDeviceMonitorResults',
                        'vBelayDeviceFailNowResults',
                        'vABInspectID_OutForReservice',
                        ]        
        try:
            # Drop each view in the list
            for view in sqlViewName:
                cursor.execute(f"DROP VIEW IF EXISTS {db_alias}.{view};")

        except sqlite3.Error as e:
            print(f"Error dropping views in database {db_alias}: {e}")
            
    def dbDropFunctions(self):
        """ 
        Function Name: dbDropFunctions
        Function Purpose: Drop the functions from the database
        """             
        try:        
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
            
            # Execute the drop functions
            sql = """
            DROP FUNCTION IF EXISTS fn_GetInspectionStatus;
            DROP FUNCTION IF EXISTS fn_GetInspectionType;
            DROP FUNCTION IF EXISTS fn_GetMetallicInspection;
            DROP FUNCTION IF EXISTS fn_GetTextileInspection;
            DROP FUNCTION IF EXISTS fn_GetPlasticInspection;
            DROP FUNCTION IF EXISTS fn_GetCarabiner;
            DROP FUNCTION IF EXISTS fn_GetCarabVisSelection;
            DROP FUNCTION IF EXISTS fn_GetCarabinerVisualResult;
            DROP FUNCTION IF EXISTS fn_GetCarabPhysSelection;
            DROP FUNCTION IF EXISTS fn_GetCarabinerPhysicalResult;
            DROP FUNCTION IF EXISTS fn_GetCarabFunctSelection;
            DROP FUNCTION IF EXISTS fn_GetCarabinerFunctionResult;
            DROP FUNCTION IF EXISTS fn_GetDeviceHandle;
            DROP FUNCTION IF EXISTS fn_GetHandVisSelection;
            DROP FUNCTION IF EXISTS fn_GetHandleVisualResult;
            DROP FUNCTION IF EXISTS fn_GetHandPhysSelection;
            DROP FUNCTION IF EXISTS fn_GetHandlePhysicalResult;
            DROP FUNCTION IF EXISTS fn_GetCaseCompSelection;
            DROP FUNCTION IF EXISTS fn_GetCaseHousingComponent;
            DROP FUNCTION IF EXISTS fn_GetCaseVisSelection;
            DROP FUNCTION IF EXISTS fn_GetCaseHousingVisualResult;
            DROP FUNCTION IF EXISTS fn_GetCasePhysSelection;
            DROP FUNCTION IF EXISTS fn_GetCaseHousingPhysicalResult;
            DROP FUNCTION IF EXISTS fn_GetBreakCompSelection;
            DROP FUNCTION IF EXISTS fn_GetBrakeHousingComponent;
            DROP FUNCTION IF EXISTS fn_GetBrakeVisSelection;
            DROP FUNCTION IF EXISTS fn_GetBrakeHousingVisualResult;
            DROP FUNCTION IF EXISTS fn_GetBrakePhysSelection;
            DROP FUNCTION IF EXISTS fn_GetBrakeHousingPhysicalResult;
            DROP FUNCTION IF EXISTS fn_GetLanyardLength;
            DROP FUNCTION IF EXISTS fn_GetLanVisSelection;
            DROP FUNCTION IF EXISTS fn_GetLanyardVisualResult;
            DROP FUNCTION IF EXISTS fn_GetLanPhysSelection;
            DROP FUNCTION IF EXISTS fn_GetLanyardPhysicalResult;
            DROP FUNCTION IF EXISTS fn_GetRetractFunctSelection;
            DROP FUNCTION IF EXISTS fn_GetLanyardRetractFunctionResult;
            DROP FUNCTION IF EXISTS fn_GetInspectorName;
            DROP FUNCTION IF EXISTS fn_GetInspectorEmail;
            DROP FUNCTION IF EXISTS fn_GetUserLoginName;
            DROP FUNCTION IF EXISTS fn_GetUserPassword;
            DROP FUNCTION IF EXISTS fn_ValidateUserPassword;
            DROP FUNCTION IF EXISTS fn_GetWallLocationName;
            DROP FUNCTION IF EXISTS fn_GetAutoBelayDeviceName;
            DROP FUNCTION IF EXISTS fn_GetAutoBelaySerialNum;  
            DROP FUNCTION IF EXISTS fn_GetAutoBelayManuDate;      
            DROP FUNCTION IF EXISTS fn_GetAutoBelayReserviceDate;   
            DROP FUNCTION IF EXISTS fn_GetAutoBelayInstallDate;      
            DROP FUNCTION IF EXISTS fn_GetAutoBelayLastInspectDate;      
            DROP FUNCTION IF EXISTS fn_GetAutoBelayNextInspectDate;      
            DROP FUNCTION IF EXISTS fn_GetAutoBelayDeviceInUseStatus;      
            DROP FUNCTION IF EXISTS fn_GetSentReportsDate;        
            DROP FUNCTION IF EXISTS fn_GetSentReportsFileName;        
            DROP FUNCTION IF EXISTS fn_GetSentReportsReceiveEmail;        
            DROP FUNCTION IF EXISTS fn_GetAutoBelayInspectComment;
            """ 
            # Execute the script 
            Database.dbExeStatement(sql)                
            print("Functions dropped successfully.")
 
        except sqlite3.Error as e:
            print(f"Error creating table: {e}")    
             
    def createInspectionStatusTable(self):
        """ 
        Function Name: createInspectionStatusTable
        Function Purpose: Create the Inspection Status Table inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
            
            # Create the table
            sqlTable = """
                -- Create Inspection Status Table
                CREATE TABLE IF NOT EXISTS TInspectionStatus 
                (
                    intInspectionStatusID           INTEGER NOT NULL
                    ,strInspectionStatusDesc        VARCHAR(255) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT TInspectionStatus_PK PRIMARY KEY (intInspectionStatusID)
                );
                """
            # Create the audit table 
            sqlAudit = f"""                
                -- Create Z Table: Inspection Status Table
                CREATE TABLE IF NOT EXISTS Z_TInspectionStatus 
                (
                    intInspectionStatusAuditID      INTEGER NOT NULL
                    ,intInspectionStatusID          INTEGER NOT NULL
                    ,strInspectionStatusDesc        VARCHAR(255) NOT NULL
                    ,strUpdatedBy                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}' 
                    ,dtmUpdatedOn                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                      VARCHAR(1) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT Z_TInspectionStatus_PK PRIMARY KEY (intInspectionStatusAuditID)
                );                
                """
            # Create the table trigger      
            sqlTrigger = f"""
                -- Create Trigger: Inspection Status Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TInspectionStatus_InsertTrigger
                AFTER INSERT ON TInspectionStatus
                BEGIN
                    INSERT INTO Z_TInspectionStatus 
                    (
                        intInspectionStatusID
                        ,strInspectionStatusDesc
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intInspectionStatusID
                        ,NEW.strInspectionStatusDesc
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Inspection Status Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TInspectionStatus_UpdateTrigger
                AFTER UPDATE ON TInspectionStatus
                BEGIN
                    INSERT INTO Z_TInspectionStatus 
                    (
                        intInspectionStatusID
                        ,strInspectionStatusDesc
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intInspectionStatusID
                        ,NEW.strInspectionStatusDesc
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Inspection Status Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TInspectionStatus_DeleteTrigger
                AFTER DELETE ON TInspectionStatus
                BEGIN
                    INSERT INTO Z_TInspectionStatus 
                    (
                        intInspectionStatusID
                        ,strInspectionStatusDesc
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intInspectionStatusID
                        ,OLD.strInspectionStatusDesc
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating inspection status tables: {e}")            
            
    def createInspectionTypesTables(self):
        """ 
        Function Name: createInspectionTypesTables
        Function Purpose: Create the Inspection Type Table inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
                    
            # Create the table
            sqlTable = """
                -- Create Inspection Type Table
                CREATE TABLE IF NOT EXISTS TInspectionTypes 
                (
                    intInspectionTypeID             INTEGER NOT NULL
                    ,strInspectionTypeDesc          VARCHAR(255) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT TInspectionTypes_PK PRIMARY KEY (intInspectionTypeID)
                );
                """
            # Create the audit table     
            sqlAudit = f"""                
                -- Create Z Table: Inspection Types Table
                CREATE TABLE IF NOT EXISTS Z_TInspectionTypes 
                (
                    intInspectionTypeAuditID        INTEGER NOT NULL
                    ,intInspectionTypeID            INTEGER NOT NULL
                    ,strInspectionTypeDesc          VARCHAR(255) NOT NULL
                    ,strUpdatedBy                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                      VARCHAR(1) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT Z_TInspectionTypes_PK PRIMARY KEY (intInspectionTypeAuditID)
                );             
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Inspection Types Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TInspectionTypes_InsertTrigger
                AFTER INSERT ON TInspectionTypes
                BEGIN
                    INSERT INTO Z_TInspectionTypes 
                    (
                        intInspectionTypeID
                        ,strInspectionTypeDesc
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intInspectionTypeID
                        ,NEW.strInspectionTypeDesc
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Inspection Types Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TInspectionTypes_UpdateTrigger
                AFTER UPDATE ON TInspectionTypes
                BEGIN
                    INSERT INTO Z_TInspectionTypes 
                    (
                        intInspectionTypeID
                        ,strInspectionTypeDesc
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intInspectionTypeID
                        ,NEW.strInspectionTypeDesc
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Inspection Types Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TInspectionTypes_DeleteTrigger
                AFTER DELETE ON TInspectionTypes
                BEGIN
                    INSERT INTO Z_TInspectionTypes 
                    (
                        intInspectionTypeID
                        ,strInspectionTypeDesc
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intInspectionTypeID
                        ,OLD.strInspectionTypeDesc
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)      
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating inspection type tables: {e}")                        

    def createMetallicInspectionsTable(self):
        """ 
        Function Name: createMetallicInspectionsTable
        Function Purpose: Create the Metallic Inspection Table inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
                    
            # Create the table
            sqlTable = f"""
                -- Create Metallic Inspection Table
                CREATE TABLE IF NOT EXISTS TMetallicInspections 
                (
                    intMetallicInspectionID         INTEGER NOT NULL
                    ,strMetallicInspectionDesc      VARCHAR(255) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT TMetallicInspections_PK PRIMARY KEY (intMetallicInspectionID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Metallic Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TMetallicInspections 
                (
                    intMetallicInspectionAuditID    INTEGER NOT NULL
                    ,intMetallicInspectionID        INTEGER NOT NULL
                    ,strMetallicInspectionDesc      VARCHAR(255) NOT NULL
                    ,strUpdatedBy                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                      VARCHAR(1) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT Z_TMetallicInspections_PK PRIMARY KEY (intMetallicInspectionAuditID)
                );       
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Metallic Inspection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TMetallicInspections_InsertTrigger
                AFTER INSERT ON TMetallicInspections
                BEGIN
                    INSERT INTO Z_TMetallicInspections 
                    (
                        intMetallicInspectionID
                        ,strMetallicInspectionDesc
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intMetallicInspectionID
                        ,NEW.strMetallicInspectionDesc
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Metallic Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TMetallicInspections_UpdateTrigger
                AFTER UPDATE ON TMetallicInspections
                BEGIN
                    INSERT INTO Z_TMetallicInspections 
                    (
                        intMetallicInspectionID
                        ,strMetallicInspectionDesc
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intMetallicInspectionID
                        ,NEW.strMetallicInspectionDesc
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Metallic Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TMetallicInspections_DeleteTrigger
                AFTER DELETE ON TMetallicInspections
                BEGIN
                    INSERT INTO Z_TMetallicInspections 
                    (
                        intMetallicInspectionID
                        ,strMetallicInspectionDesc
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intMetallicInspectionID
                        ,OLD.strMetallicInspectionDesc
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)     
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Metallic Component tables: {e}")              
            
    def createTextileInspectionsTable(self):
        """ 
        Function Name: createTextileInspectionsTable
        Function Purpose: Create the Textile Inspection Table inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
                    
            # Create the table
            sqlTable = """
                -- Create Textile Inspection Table
                CREATE TABLE IF NOT EXISTS TTextileInspections 
                (
                    intTextileInspectionID          INTEGER NOT NULL
                    ,strTextileInspectionDesc       VARCHAR(255) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT TTextileInspections_PK PRIMARY KEY (intTextileInspectionID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Textile Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TTextileInspections 
                (
                    intTextileInspectionAuditID     INTEGER NOT NULL
                    ,intTextileInspectionID         INTEGER NOT NULL
                    ,strTextileInspectionDesc       VARCHAR(255) NOT NULL
                    ,strUpdatedBy                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                      VARCHAR(1) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT Z_TTextileInspections_PK PRIMARY KEY (intTextileInspectionAuditID)
                );    
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Textile Inspection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TTextileInspections_InsertTrigger
                AFTER INSERT ON TTextileInspections
                BEGIN
                    INSERT INTO Z_TTextileInspections 
                    (
                        intTextileInspectionID
                        ,strTextileInspectionDesc
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intTextileInspectionID
                        ,NEW.strTextileInspectionDesc
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Textile Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TTextileInspections_UpdateTrigger
                AFTER UPDATE ON TTextileInspections
                BEGIN
                    INSERT INTO Z_TTextileInspections 
                    (
                        intTextileInspectionID
                        ,strTextileInspectionDesc
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intTextileInspectionID
                        ,NEW.strTextileInspectionDesc
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Textile Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TTextileInspections_DeleteTrigger
                AFTER DELETE ON TTextileInspections
                BEGIN
                    INSERT INTO Z_TTextileInspections 
                    (
                        intTextileInspectionID
                        ,strTextileInspectionDesc
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intTextileInspectionID
                        ,OLD.strTextileInspectionDesc
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)       
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Textile Component tables: {e}")              

    def createPlasticInspectionsTable(self):
        """ 
        Function Name: createPlasticInspectionsTable
        Function Purpose: Create the Plastic Inspection Table inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
                    
            # Create the table
            sqlTable = """
                -- Create Plastic Inspection Table
                CREATE TABLE IF NOT EXISTS TPlasticInspections 
                (
                    intPlasticInspectionID          INTEGER NOT NULL
                    ,strPlasticInspectionDesc       VARCHAR(255) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT TPlasticInspections_PK PRIMARY KEY (intPlasticInspectionID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Plastic Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TPlasticInspections 
                (
                    intPlasticInspectionAuditID     INTEGER NOT NULL
                    ,intPlasticInspectionID         INTEGER NOT NULL
                    ,strPlasticInspectionDesc       VARCHAR(255) NOT NULL
                    ,strUpdatedBy                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                      VARCHAR(1) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT Z_TPlasticInspections_PK PRIMARY KEY (intPlasticInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Plastic Inspection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TPlasticInspections_InsertTrigger
                AFTER INSERT ON TPlasticInspections
                BEGIN
                    INSERT INTO Z_TPlasticInspections 
                    (
                        intPlasticInspectionID
                        ,strPlasticInspectionDesc
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intPlasticInspectionID
                        ,NEW.strPlasticInspectionDesc
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Plastic Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TPlasticInspections_UpdateTrigger
                AFTER UPDATE ON TPlasticInspections
                BEGIN
                    INSERT INTO Z_TPlasticInspections 
                    (
                        intPlasticInspectionID
                        ,strPlasticInspectionDesc
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intPlasticInspectionID
                        ,NEW.strPlasticInspectionDesc
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Plastic Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TPlasticInspections_DeleteTrigger
                AFTER DELETE ON TPlasticInspections
                BEGIN
                    INSERT INTO Z_TPlasticInspections 
                    (
                        intPlasticInspectionID
                        ,strPlasticInspectionDesc
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intPlasticInspectionID
                        ,OLD.strPlasticInspectionDesc
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)      
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Plastic Component tables: {e}")                  
            
    def createCarabinerTables(self):
        """ 
        Function Name: createCarabinerTables
        Function Purpose: Create the Carabiner Type and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
                    
            # Create the table
            sqlTable = """
                -- Create Carabiner Type Table
                CREATE TABLE IF NOT EXISTS TCarabiners 
                (
                    intCarabinerID              INTEGER NOT NULL
                    ,strCarabinerType           VARCHAR(255) NOT NULL
                    ,strModifiedReason          VARCHAR(1000)
                    ,CONSTRAINT TCarabiners_PK PRIMARY KEY (intCarabinerID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Carabiner Type Table
                CREATE TABLE IF NOT EXISTS Z_TCarabiners 
                (
                    intCarabinerAuditID         INTEGER NOT NULL
                    ,intCarabinerID             INTEGER NOT NULL
                    ,strCarabinerType           VARCHAR(255) NOT NULL
                    ,strUpdatedBy               VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn               DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                  VARCHAR(1) NOT NULL
                    ,strModifiedReason          VARCHAR(1000)
                    ,CONSTRAINT Z_TCarabiners_PK PRIMARY KEY (intCarabinerAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Carabiner Type Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TCarabiners_InsertTrigger
                AFTER INSERT ON TCarabiners
                BEGIN
                    INSERT INTO Z_TCarabiners 
                    (
                        intCarabinerID
                        ,strCarabinerType
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCarabinerID
                        ,NEW.strCarabinerType
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Carabiner Type Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TCarabiners_UpdateTrigger
                AFTER UPDATE ON TCarabiners
                BEGIN
                    INSERT INTO Z_TCarabiners 
                    (
                        intCarabinerID
                        ,strCarabinerType
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCarabinerID
                        ,NEW.strCarabinerType
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Carabiner Type Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TCarabiners_DeleteTrigger
                AFTER DELETE ON TCarabiners
                BEGIN
                    INSERT INTO Z_TCarabiners 
                    (
                        intCarabinerID
                        ,strCarabinerType
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intCarabinerID
                        ,OLD.strCarabinerType
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)      
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Carabiner Type tables: {e}")                              
            
    def createCarabVisMetSelTables(self):
        """ 
        Function Name: createCarabVisMetSelTables
        Function Purpose: Create the Carabiner Visual Metallic Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
                    
            # Create the table
            sqlTable = """
                -- Create Carabiner Visual Metallic Selection Table
                CREATE TABLE IF NOT EXISTS TCarabVisMetalSelects 
                (
                    intCarabVisMetalSelectID            INTEGER NOT NULL
                    ,strCarabVisMetSelect               VARCHAR(1000) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TCarabVisMetalSelects_PK PRIMARY KEY (intCarabVisMetalSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Carabiner Visual Metallic Selection Table
                CREATE TABLE IF NOT EXISTS Z_TCarabVisMetalSelects 
                (
                    intCarabVisMetalSelectAuditID       INTEGER NOT NULL
                    ,intCarabVisMetalSelectID           INTEGER NOT NULL
                    ,strCarabVisMetSelect               VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}' DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TCarabVisMetalSelects_PK PRIMARY KEY (intCarabVisMetalSelectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Carabiner Visual Metallic Selection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TCarabVisMetalSelects_InsertTrigger
                AFTER INSERT ON TCarabVisMetalSelects
                BEGIN
                    INSERT INTO Z_TCarabVisMetalSelects 
                    (
                        intCarabVisMetalSelectID
                        ,strCarabVisMetSelect
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCarabVisMetalSelectID
                        ,NEW.strCarabVisMetSelect
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Carabiner Visual Metallic Selection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TCarabVisMetalSelects_UpdateTrigger
                AFTER UPDATE ON TCarabVisMetalSelects
                BEGIN
                    INSERT INTO Z_TCarabVisMetalSelects 
                    (
                        intCarabVisMetalSelectID
                        ,strCarabVisMetSelect
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCarabVisMetalSelectID
                        ,NEW.strCarabVisMetSelect
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Carabiner Visual Metallic Selection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TCarabVisMetalSelects_DeleteTrigger
                AFTER DELETE ON TCarabVisMetalSelects
                BEGIN
                    INSERT INTO Z_TCarabVisMetalSelects 
                    (
                        intCarabVisMetalSelectID
                        ,strCarabVisMetSelect
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intCarabVisMetalSelectID
                        ,OLD.strCarabVisMetSelect
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)   
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Carabiner Visual Selection tables: {e}")                 

    def createCarabVisInspectTables(self):
        """ 
        Function Name: createCarabVisInspectTables
        Function Purpose: Create the Carabiner Visual Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
                    
            # Create the table
            sqlTable = """
                -- Create Carabiner Visual Inspection Table
                CREATE TABLE IF NOT EXISTS TCarabinerVisualInspections 
                (
                    intCarabinerVisualID                INTEGER NOT NULL
                    ,intCarabinerID                     INTEGER NOT NULL
                    ,intInspectionTypeID                INTEGER NOT NULL
                    ,intCarabVisMetalSelectID           INTEGER NOT NULL
                    ,intInspectionStatusID              INTEGER NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TCarabinerVisualInspections_PK PRIMARY KEY (intCarabinerVisualID)
                    
                    ,FOREIGN KEY ( intCarabinerID ) REFERENCES TCarabiners ( intCarabinerID )
                    ,FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )
                    ,FOREIGN KEY ( intCarabVisMetalSelectID ) REFERENCES TCarabVisMetalSelects ( intCarabVisMetalSelectID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )
                );    
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Carabiner Visual Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TCarabinerVisualInspections
                (
                    intCarabinerVisualInspectionAuditID INTEGER NOT NULL
                    ,intCarabinerVisualID               INTEGER NOT NULL
                    ,intCarabinerID                     INTEGER NOT NULL
                    ,intInspectionTypeID                INTEGER NOT NULL
                    ,intCarabVisMetalSelectID           INTEGER NOT NULL
                    ,intInspectionStatusID              INTEGER NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}' DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TCarabinerVisualInspections_PK PRIMARY KEY (intCarabinerVisualInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Carabiner Visual Inspection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TCarabinerVisualInspections_InsertTrigger
                AFTER INSERT ON TCarabinerVisualInspections
                BEGIN
                    INSERT INTO Z_TCarabinerVisualInspections 
                    (
                        intCarabinerVisualID
                        ,intCarabinerID
                        ,intInspectionTypeID
                        ,intCarabVisMetalSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCarabinerVisualID
                        ,NEW.intCarabinerID
                        ,NEW.intInspectionTypeID
                        ,NEW.intCarabVisMetalSelectID
                        ,NEW.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Carabiner Visual Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TCarabinerVisualInspections_UpdateTrigger
                AFTER UPDATE ON TCarabinerVisualInspections
                BEGIN
                    INSERT INTO Z_TCarabinerVisualInspections 
                    (
                        intCarabinerVisualID
                        ,intCarabinerID
                        ,intInspectionTypeID
                        ,intCarabVisMetalSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCarabinerVisualID
                        ,NEW.intCarabinerID
                        ,NEW.intInspectionTypeID
                        ,NEW.intCarabVisMetalSelectID
                        ,NEW.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Carabiner Visual Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TCarabinerVisualInspections_DeleteTrigger
                AFTER DELETE ON TCarabinerVisualInspections
                BEGIN
                    INSERT INTO Z_TCarabinerVisualInspections 
                    (
                        intCarabinerVisualID
                        ,intCarabinerID
                        ,intInspectionTypeID
                        ,intCarabVisMetalSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intCarabinerVisualID
                        ,OLD.intCarabinerID
                        ,OLD.intInspectionTypeID
                        ,OLD.intCarabVisMetalSelectID
                        ,OLD.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)   
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Carabiner Visual Inspection tables: {e}")
            
    def createCarabPhysMetSelTables(self):
        """ 
        Function Name: createCarabPhysMetSelTables
        Function Purpose: Create the Carabiner Physical Metal Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
                    
            # Create the table
            sqlTable = """
                -- Create Carabiner Physical Metal Selection Table
                CREATE TABLE IF NOT EXISTS TCarabPhysMetalSelects 
                (
                    intCarabPhysMetalSelectID           INTEGER NOT NULL
                    ,strCarabPhysMetSelect              VARCHAR(1000) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TCarabPhysMetalSelects_PK PRIMARY KEY (intCarabPhysMetalSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Carabiner Physical Metal Selection Table
                CREATE TABLE IF NOT EXISTS Z_TCarabPhysMetalSelects 
                (
                    intCarabPhysMetalSelectAuditID      INTEGER NOT NULL
                    ,intCarabPhysMetalSelectID          INTEGER NOT NULL
                    ,strCarabPhysMetSelect              VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}' DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TCarabPhysMetalSelects_PK PRIMARY KEY (intCarabPhysMetalSelectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Carabiner Physical Metal Selection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TCarabPhysMetalSelects_InsertTrigger
                AFTER INSERT ON TCarabPhysMetalSelects
                BEGIN
                    INSERT INTO Z_TCarabPhysMetalSelects 
                    (
                        intCarabPhysMetalSelectID
                        ,strCarabPhysMetSelect
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCarabPhysMetalSelectID
                        ,NEW.strCarabPhysMetSelect
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Carabiner Physical Metal Selection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TCarabPhysMetalSelects_UpdateTrigger
                AFTER UPDATE ON TCarabPhysMetalSelects
                BEGIN
                    INSERT INTO Z_TCarabPhysMetalSelects 
                    (
                        intCarabPhysMetalSelectID
                        ,strCarabPhysMetSelect
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCarabPhysMetalSelectID
                        ,NEW.strCarabPhysMetSelect
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Carabiner Physical Metal Selection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TCarabPhysMetalSelects_DeleteTrigger
                AFTER DELETE ON TCarabPhysMetalSelects
                BEGIN
                    INSERT INTO Z_TCarabPhysMetalSelects 
                    (
                        intCarabPhysMetalSelectID
                        ,strCarabPhysMetSelect
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intCarabPhysMetalSelectID
                        ,OLD.strCarabPhysMetSelect
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Carabiner Physical Selection tables: {e}")            
            
    def createCarabPhysInspectTables(self):
        """ 
        Function Name: createCarabPhysInspectTables
        Function Purpose: Create the Carabiner Physical Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
                    
            # Create the table
            sqlTable = """
                -- Create Carabiner Physical Inspection Table
                CREATE TABLE IF NOT EXISTS TCarabinerPhysicalInspections 
                (
                    intCarabinerPhysicalID                  INTEGER NOT NULL
                    ,intCarabinerID                         INTEGER NOT NULL
                    ,intInspectionTypeID                    INTEGER NOT NULL
                    ,intCarabPhysMetalSelectID              INTEGER NOT NULL
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT TCarabinerPhysicalInspections_PK PRIMARY KEY (intCarabinerPhysicalID)
                    
                    ,FOREIGN KEY ( intCarabinerID ) REFERENCES TCarabiners ( intCarabinerID )
                    ,FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )
                    ,FOREIGN KEY ( intCarabPhysMetalSelectID ) REFERENCES TCarabPhysMetalSelects ( intCarabPhysMetalSelectID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )                    
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Carabiner Physical Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TCarabinerPhysicalInspections 
                (
                    intCarabinerPhysicalInspectionAuditID   INTEGER NOT NULL
                    ,intCarabinerPhysicalID                 INTEGER NOT NULL
                    ,intCarabinerID                         INTEGER NOT NULL
                    ,intInspectionTypeID                    INTEGER NOT NULL
                    ,intCarabPhysMetalSelectID              INTEGER NOT NULL
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}' DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TCarabinerPhysicalInspections_PK PRIMARY KEY (intCarabinerPhysicalInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Carabiner Physical Inspection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TCarabinerPhysicalInspections_InsertTrigger
                AFTER INSERT ON TCarabinerPhysicalInspections
                BEGIN
                    INSERT INTO Z_TCarabinerPhysicalInspections 
                    (
                        intCarabinerPhysicalID,
                        intCarabinerID,
                        intInspectionTypeID,
                        intCarabPhysMetalSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCarabinerPhysicalID,
                        NEW.intCarabinerID,
                        NEW.intInspectionTypeID,
                        NEW.intCarabPhysMetalSelectID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Carabiner Physical Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TCarabinerPhysicalInspections_UpdateTrigger
                AFTER UPDATE ON TCarabinerPhysicalInspections
                BEGIN
                    INSERT INTO Z_TCarabinerPhysicalInspections 
                    (
                        intCarabinerPhysicalID,
                        intCarabinerID,
                        intInspectionTypeID,
                        intCarabPhysMetalSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCarabinerPhysicalID,
                        NEW.intCarabinerID,
                        NEW.intInspectionTypeID,
                        NEW.intCarabPhysMetalSelectID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Carabiner Physical Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TCarabinerPhysicalInspections_DeleteTrigger
                AFTER DELETE ON TCarabinerPhysicalInspections
                BEGIN
                    INSERT INTO Z_TCarabinerPhysicalInspections 
                    (
                        intCarabinerPhysicalID,
                        intCarabinerID,
                        intInspectionTypeID,
                        intCarabPhysMetalSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intCarabinerPhysicalID,
                        OLD.intCarabinerID,
                        OLD.intInspectionTypeID,
                        OLD.intCarabPhysMetalSelectID,
                        OLD.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)      
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Carabiner Physical Inspection tables: {e}")
            
    def createCarabinerFunctionsTables(self):
        """ 
        Function Name: createCarabinerFunctionsTables
        Function Purpose: Create the Carabiner Functions and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
                    
            # Create the table
            sqlTable = """
                -- Create Carabiner Function Table
                CREATE TABLE IF NOT EXISTS TCarabinerFunctions 
                (
                    intCarabinerFunctionID              INTEGER NOT NULL
                    ,strCarabinerFunctionDesc           VARCHAR(255) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TCarabinerFunctions_PK PRIMARY KEY (intCarabinerFunctionID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Carabiner Function Table
                CREATE TABLE IF NOT EXISTS Z_TCarabinerFunctions 
                (
                    intCarabinerFunctionAuditID         INTEGER NOT NULL
                    ,intCarabinerFunctionID             INTEGER NOT NULL
                    ,strCarabinerFunctionDesc           VARCHAR(255) NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TCarabinerFunctions_PK PRIMARY KEY (intCarabinerFunctionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Carabiner Function Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TCarabinerFunctions_InsertTrigger
                AFTER INSERT ON TCarabinerFunctions
                BEGIN
                    INSERT INTO Z_TCarabinerFunctions 
                    (
                        intCarabinerFunctionID,
                        strCarabinerFunctionDesc,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCarabinerFunctionID,
                        NEW.strCarabinerFunctionDesc,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Carabiner Function Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TCarabinerFunctions_UpdateTrigger
                AFTER UPDATE ON TCarabinerFunctions
                BEGIN
                    INSERT INTO Z_TCarabinerFunctions 
                    (
                        intCarabinerFunctionID,
                        strCarabinerFunctionDesc,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCarabinerFunctionID,
                        NEW.strCarabinerFunctionDesc,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Carabiner Function Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TCarabinerFunctions_DeleteTrigger
                AFTER DELETE ON TCarabinerFunctions
                BEGIN
                    INSERT INTO Z_TCarabinerFunctions 
                    (
                        intCarabinerFunctionID,
                        strCarabinerFunctionDesc,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intCarabinerFunctionID,
                        OLD.strCarabinerFunctionDesc,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)     
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Carabiner Functional tables: {e}")              
            
    def createCarabFunctSelTables(self):
        """ 
        Function Name: createCarabFunctSelTables
        Function Purpose: Create the Carabiner Function Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Carabiner Function Selection Table
                CREATE TABLE IF NOT EXISTS TCarabFunctSelects 
                (
                    intCarabFunctSelectID               INTEGER NOT NULL
                    ,strCarabFunctSelect                VARCHAR(1000) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TCarabFunctSelects_PK PRIMARY KEY (intCarabFunctSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Carabiner Function Selection Table
                CREATE TABLE IF NOT EXISTS Z_TCarabFunctSelects 
                (
                    intCarabFunctSelectAuditID          INTEGER NOT NULL
                    ,intCarabFunctSelectID              INTEGER NOT NULL
                    ,strCarabFunctSelect                VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TCarabFunctSelects_PK PRIMARY KEY (intCarabFunctSelectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Carabiner Function Selection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TCarabFunctSelects_InsertTrigger
                AFTER INSERT ON TCarabFunctSelects
                BEGIN
                    INSERT INTO Z_TCarabFunctSelects 
                    (
                        intCarabFunctSelectID,
                        strCarabFunctSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCarabFunctSelectID,
                        NEW.strCarabFunctSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Carabiner Function Selection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TCarabFunctSelects_UpdateTrigger
                AFTER UPDATE ON TCarabFunctSelects
                BEGIN
                    INSERT INTO Z_TCarabFunctSelects 
                    (
                        intCarabFunctSelectID,
                        strCarabFunctSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCarabFunctSelectID,
                        NEW.strCarabFunctSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Carabiner Function Selection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TCarabFunctSelects_DeleteTrigger
                AFTER DELETE ON TCarabFunctSelects
                BEGIN
                    INSERT INTO Z_TCarabFunctSelects 
                    (
                        intCarabFunctSelectID,
                        strCarabFunctSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intCarabFunctSelectID,
                        OLD.strCarabFunctSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Carabiner Functional Selection tables: {e}")                  
            
    def createCarabFunctInspectTables(self):
        """ 
        Function Name: createCarabFunctInspectTables
        Function Purpose: Create the Carabiner Function Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Carabiner Function Inspection Table
                CREATE TABLE IF NOT EXISTS TCarabinerFunctionInspections 
                (
                    intCarabinerFunctionInspectID       INTEGER NOT NULL
                    ,intCarabinerID                     INTEGER NOT NULL
                    ,intInspectionTypeID                INTEGER NOT NULL
                    ,intCarabFunctSelectID              INTEGER NOT NULL
                    ,intInspectionStatusID              INTEGER NOT NULL
                    ,strModifiedReason VARCHAR(1000)
                    ,CONSTRAINT TCarabinerFunctionInspections_PK PRIMARY KEY (intCarabinerFunctionInspectID)

                    ,FOREIGN KEY ( intCarabinerID ) REFERENCES TCarabiners ( intCarabinerID )
                    ,FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )
                    ,FOREIGN KEY ( intCarabFunctSelectID ) REFERENCES TCarabFunctSelects ( intCarabFunctSelectID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )                    
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Carabiner Function Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TCarabinerFunctionInspection 
                (
                    intCarabinerFunctionInspectAuditID  INTEGER NOT NULL
                    ,intCarabinerFunctionInspectID      INTEGER NOT NULL
                    ,intCarabinerID                     INTEGER NOT NULL
                    ,intInspectionTypeID                INTEGER NOT NULL
                    ,intCarabFunctSelectID              INTEGER NOT NULL
                    ,intInspectionStatusID              INTEGER NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TCarabinerFunctionInspection_PK PRIMARY KEY (intCarabinerFunctionInspectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Carabiner Function Inspection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TCarabinerFunctionInspection_InsertTrigger
                AFTER INSERT ON TCarabinerFunctionInspections
                BEGIN
                    INSERT INTO Z_TCarabinerFunctionInspection 
                    (
                        intCarabinerFunctionInspectID,
                        intCarabinerID,
                        intInspectionTypeID,
                        intCarabFunctSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCarabinerFunctionInspectID,
                        NEW.intCarabinerID,
                        NEW.intInspectionTypeID,
                        NEW.intCarabFunctSelectID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Carabiner Function Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TCarabinerFunctionInspection_UpdateTrigger
                AFTER UPDATE ON TCarabinerFunctionInspections
                BEGIN
                    INSERT INTO Z_TCarabinerFunctionInspection 
                    (
                        intCarabinerFunctionInspectID,
                        intCarabinerID,
                        intInspectionTypeID,
                        intCarabFunctSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCarabinerFunctionInspectID,
                        NEW.intCarabinerID,
                        NEW.intInspectionTypeID,
                        NEW.intCarabFunctSelectID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Carabiner Function Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TCarabinerFunctionInspection_DeleteTrigger
                AFTER DELETE ON TCarabinerFunctionInspections
                BEGIN
                    INSERT INTO Z_TCarabinerFunctionInspection 
                    (
                        intCarabinerFunctionInspectID,
                        intCarabinerID,
                        intInspectionTypeID,
                        intCarabFunctSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intCarabinerFunctionInspectID,
                        OLD.intCarabinerID,
                        OLD.intInspectionTypeID,
                        OLD.intCarabFunctSelectID,
                        OLD.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Carabiner Functional Inspection tables: {e}")
            
    def createStandCarabInspectTables(self):
        """ 
        Function Name: createStandCarabInspectTables
        Function Purpose: Create the Standard Carabiner Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Standard Carabiner Inspection Table
                CREATE TABLE IF NOT EXISTS TStandardCarabinerInspections 
                (
                    intStandardCarabinerInspectionID        INTEGER NOT NULL
                    ,intCarabinerVisualID                   INTEGER NOT NULL
                    ,intCarabinerPhysicalID                 INTEGER NOT NULL
                    ,intCarabinerFunctionInspectID          INTEGER NOT NULL
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT TStandardCarabinerInspections_PK PRIMARY KEY (intStandardCarabinerInspectionID)
                    
                    ,FOREIGN KEY ( intCarabinerVisualID ) REFERENCES TCarabinerVisualInspections ( intCarabinerVisualID )
                    ,FOREIGN KEY ( intCarabinerPhysicalID ) REFERENCES TCarabinerPhysicalInspections ( intCarabinerPhysicalID )
                    ,FOREIGN KEY ( intCarabinerFunctionInspectID ) REFERENCES TCarabinerFunctionInspections ( intCarabinerFunctionInspectID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Standard Carabiner Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TStandardCarabinerInspections 
                (
                    intStandardCarabinerInspectionAuditID   INTEGER NOT NULL
                    ,intStandardCarabinerInspectionID       INTEGER NOT NULL
                    ,intCarabinerVisualID                   INTEGER NOT NULL
                    ,intCarabinerPhysicalID                 INTEGER NOT NULL
                    ,intCarabinerFunctionInspectID          INTEGER NOT NULL
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TStandardCarabinerInspections_PK PRIMARY KEY (intStandardCarabinerInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Standard Carabiner Inspection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TStandardCarabinerInspections_InsertTrigger
                AFTER INSERT ON TStandardCarabinerInspections
                BEGIN
                    INSERT INTO Z_TStandardCarabinerInspections 
                    (
                        intStandardCarabinerInspectionID,
                        intCarabinerVisualID,
                        intCarabinerPhysicalID,
                        intCarabinerFunctionInspectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intStandardCarabinerInspectionID,
                        NEW.intCarabinerVisualID,
                        NEW.intCarabinerPhysicalID,
                        NEW.intCarabinerFunctionInspectID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Standard Carabiner Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TStandardCarabinerInspections_UpdateTrigger
                AFTER UPDATE ON TStandardCarabinerInspections
                BEGIN
                    INSERT INTO Z_TStandardCarabinerInspections 
                    (
                        intStandardCarabinerInspectionID,
                        intCarabinerVisualID,
                        intCarabinerPhysicalID,
                        intCarabinerFunctionInspectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intStandardCarabinerInspectionID,
                        NEW.intCarabinerVisualID,
                        NEW.intCarabinerPhysicalID,
                        NEW.intCarabinerFunctionInspectID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Standard Carabiner Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TStandardCarabinerInspections_DeleteTrigger
                AFTER DELETE ON TStandardCarabinerInspections
                BEGIN
                    INSERT INTO Z_TStandardCarabinerInspections 
                    (
                        intStandardCarabinerInspectionID,
                        intCarabinerVisualID,
                        intCarabinerPhysicalID,
                        intCarabinerFunctionInspectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intStandardCarabinerInspectionID,
                        OLD.intCarabinerVisualID,
                        OLD.intCarabinerPhysicalID,
                        OLD.intCarabinerFunctionInspectID,
                        OLD.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)     
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Standard Carabiner Inspection tables: {e}")                 
            
    def insertStatus_Types_MetTextPlast_Carab_Data(self):
        """ 
        Function Name: insertStatus_Types_MetTextPlast_Carab_Data
        Function Purpose: Insert the Inspection Status/Types, Metal, Textile, Plastic, and Carabiner data into the tables
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlInsert = [
            """
            -- Add Data To Table: InspectionStatus
            INSERT INTO TInspectionStatus (intInspectionStatusID, strInspectionStatusDesc)
            VALUES (1, 'Good'),
                (2, 'Monitor'),
                (3, 'Fail'),
                (4, 'Replaced')
            """,
            """
            -- Add Data To Table: InspectionTypes
            INSERT INTO TInspectionTypes (intInspectionTypeID, strInspectionTypeDesc)
            VALUES (1, 'Visual Inspection'),
                (2, 'Physical Inspection'),
                (3, 'Functional Inspection'),
                (4, 'Entire Standard Inspection')
            """,
            """
            -- Add Data To Table: Metallic Inspections
            INSERT INTO TMetallicInspections (intMetallicInspectionID, strMetallicInspectionDesc)
            VALUES (1, 'No Metallic Issues'),
                (2, 'Deformation'),
                (3, 'Corrosion'),
                (4, 'Missing Components'),
                (5, 'Heat Exposure'),
                (6, 'Chemical Exposure'),
                (7, 'Cracks/Fracturing'),
                (8, 'Sharp Edges'),
                (9, 'Loss of Integrity'),
                (10, 'Other Metallic Issues')
            """,
            """
            -- Add Data To Table: Textile Inspections
            INSERT INTO TTextileInspections (intTextileInspectionID, strTextileInspectionDesc)
            VALUES (1, 'No Textile Issues'),
                (2, 'Sheath Slippage'),
                (3, 'Fraying'),
                (4, 'Missing Components'),
                (5, 'Glazing'),
                (6, 'Chemical Exposure'),
                (7, 'Hard Spots'),
                (8, 'Soft Spots'),
                (9, 'Excessive Age'),
                (10, 'Loss of Integrity'),
                (11, 'Other Textile Issues')
            """,
            """
            -- Add Data To Table: Plastic Inspections
            INSERT INTO TPlasticInspections (intPlasticInspectionID, strPlasticInspectionDesc)
            VALUES (1, 'No Plastic Issues'),
                (2, 'Deformation'),
                (3, 'Corrosion'),
                (4, 'Missing Components'),
                (5, 'Heat Exposure'),
                (6, 'Chemical Exposure'),
                (7, 'Cracks/Fracturing'),
                (8, 'Sharp Edges'),
                (9, 'Loss of Integrity'),
                (10, 'Other Plastic Issues')
            """,
            """
            -- Add Data To Table: Carabiner Component
            INSERT INTO TCarabiners (intCarabinerID, strCarabinerType)
            VALUES (1, 'Triact-Lock'),
                (2, 'Biact-Lock'),
                (3, 'Screw-Lock'),
                (4, 'Pin-Lock'),
                (5, 'Ball-Lock'),
                (6, 'Non-Lock')
            """,
            """
            -- Add Data To Table: Carabiner Functionality
            INSERT INTO TCarabinerFunctions (intCarabinerFunctionID, strCarabinerFunctionDesc)
            VALUES (1, 'No Functional Issues'),
                (2, 'Gate Misalignment'),
                (3, 'Slow Closure'),
                (4, 'Lock Malfunction'),
                (5, 'Unlock Malfunction'),
                (6, 'Sleeve Grime'),
                (7, 'Spring Issue'),
                (8, 'Other Functional Issues')
            """]
            
            # Execute the SQL statements
            for sql in sqlInsert:
                Database.dbExeStatement(self, sql)          
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")            

    def createDeviceHandleTables(self):
        """ 
        Function Name: createDeviceHandleTables
        Function Purpose: Create the Handle Type Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Handle Type Table
                CREATE TABLE IF NOT EXISTS TDeviceHandles 
                (
                    intDeviceHandleID               INTEGER NOT NULL
                    ,strHandleType                  VARCHAR(255) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT TDeviceHandles_PK PRIMARY KEY (intDeviceHandleID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Handle Type Table
                CREATE TABLE IF NOT EXISTS Z_TDeviceHandles 
                (
                    intDeviceHandleAuditID          INTEGER NOT NULL
                    ,intDeviceHandleID              INTEGER NOT NULL
                    ,strHandleType                  VARCHAR(255) NOT NULL
                    ,strUpdatedBy                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                      VARCHAR(1) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT Z_TDeviceHandles_PK PRIMARY KEY (intDeviceHandleAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Handle Type Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TDeviceHandles_InsertTrigger
                AFTER INSERT ON TDeviceHandles
                BEGIN
                    INSERT INTO Z_TDeviceHandles 
                    (
                        intDeviceHandleID,
                        strHandleType,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intDeviceHandleID,
                        NEW.strHandleType,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Handle Type Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TDeviceHandles_UpdateTrigger
                AFTER UPDATE ON TDeviceHandles
                BEGIN
                    INSERT INTO Z_TDeviceHandles 
                    (
                        intDeviceHandleID,
                        strHandleType,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intDeviceHandleID,
                        NEW.strHandleType,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Handle Type Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TDeviceHandles_DeleteTrigger
                AFTER DELETE ON TDeviceHandles
                BEGIN
                    INSERT INTO Z_TDeviceHandles 
                    (
                        intDeviceHandleID,
                        strHandleType,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intDeviceHandleID,
                        OLD.strHandleType,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)     
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Device Handle tables: {e}")              
            
    def createHandVisMetSelTables(self):
        """ 
        Function Name: createHandVisMetSelTables
        Function Purpose: Create the Handle Visual Metallic Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Handle Visual Metallic Selection Table
                CREATE TABLE IF NOT EXISTS THandleVisMetalSelects 
                (
                    intHandVisMetalSelectID             INTEGER NOT NULL
                    ,strHandVisMetSelect                VARCHAR(1000) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT THandleVisMetalSelects_PK PRIMARY KEY (intHandVisMetalSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Handle Visual Metallic Selection Table
                CREATE TABLE IF NOT EXISTS Z_THandleVisMetalSelects
                (
                    intHandVisMetalSelectAuditID        INTEGER NOT NULL   
                    ,intHandVisMetalSelectID            INTEGER NOT NULL                      
                    ,strHandVisMetSelect                VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_THandleVisMetalSelects_PK PRIMARY KEY (intHandVisMetalSelectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Handle Visual Metallic Selection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_THandleVisMetalSelects_InsertTrigger
                AFTER INSERT ON THandleVisMetalSelects
                BEGIN
                    INSERT INTO Z_THandleVisMetalSelects 
                    (
                        intHandVisMetalSelectID,
                        strHandVisMetSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intHandVisMetalSelectID,
                        NEW.strHandVisMetSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Handle Visual Metallic Selection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_THandleVisMetalSelects_UpdateTrigger
                AFTER UPDATE ON THandleVisMetalSelects
                BEGIN
                    INSERT INTO Z_THandleVisMetalSelects 
                    (
                        intHandVisMetalSelectID,
                        strHandVisMetSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intHandVisMetalSelectID,
                        NEW.strHandVisMetSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Handle Visual Metallic Selection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_THandleVisMetalSelects_DeleteTrigger
                AFTER DELETE ON THandleVisMetalSelects
                BEGIN
                    INSERT INTO Z_THandleVisMetalSelects 
                    (
                        intHandVisMetalSelectID,
                        strHandVisMetSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intHandVisMetalSelectID,
                        OLD.strHandVisMetSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)     
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Device Handle Visual Selection tables: {e}")            

    def createHandVisInspectTables(self):
        """ 
        Function Name: createHandVisInspectTables
        Function Purpose: Create the Handle Visual Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Handle Visual Inspection Table
                CREATE TABLE IF NOT EXISTS THandleVisualInspections
                (
                    intHandleVisualID                   INTEGER NOT NULL
                    ,intDeviceHandleID                  INTEGER NOT NULL
                    ,intInspectionTypeID                INTEGER NOT NULL
                    ,intHandVisMetalSelectID            INTEGER NOT NULL
                    ,intInspectionStatusID              INTEGER NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT THandleVisualInspections_PK PRIMARY KEY (intHandleVisualID)

                    ,FOREIGN KEY ( intDeviceHandleID ) REFERENCES TDeviceHandles ( intDeviceHandleID )
                    ,FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )
                    ,FOREIGN KEY ( intHandVisMetalSelectID ) REFERENCES THandleVisMetalSelects ( intHandVisMetalSelectID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )                    
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Handle Visual Inspection Table
                CREATE TABLE IF NOT EXISTS Z_THandleVisualInspections 
                (
                    intHandleVisualInspectionAuditID    INTEGER NOT NULL
                    ,intHandleVisualID                  INTEGER NOT NULL
                    ,intDeviceHandleID                  INTEGER NOT NULL
                    ,intInspectionTypeID                INTEGER NOT NULL            
                    ,intHandVisMetalSelectID            INTEGER NOT NULL
                    ,intInspectionStatusID              INTEGER NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_THandleVisualInspections_PK PRIMARY KEY (intHandleVisualInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Handle Visual Inspection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_THandleVisualInspections_InsertTrigger
                AFTER INSERT ON THandleVisualInspections
                BEGIN
                    INSERT INTO Z_THandleVisualInspections 
                    (
                        intHandleVisualID,
                        intDeviceHandleID,
                        intInspectionTypeID,
                        intHandVisMetalSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intHandleVisualID,
                        NEW.intDeviceHandleID,
                        NEW.intInspectionTypeID,
                        NEW.intHandVisMetalSelectID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Handle Visual Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_THandleVisualInspections_UpdateTrigger
                AFTER UPDATE ON THandleVisualInspections
                BEGIN
                    INSERT INTO Z_THandleVisualInspections 
                    (
                        intHandleVisualID,
                        intDeviceHandleID,
                        intInspectionTypeID,
                        intHandVisMetalSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intHandleVisualID,
                        NEW.intDeviceHandleID,
                        NEW.intInspectionTypeID,
                        NEW.intHandVisMetalSelectID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Handle Visual Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_THandleVisualInspections_DeleteTrigger
                AFTER DELETE ON THandleVisualInspections
                BEGIN
                    INSERT INTO Z_THandleVisualInspections 
                    (
                        intHandleVisualID,
                        intDeviceHandleID,
                        intInspectionTypeID,
                        intHandVisMetalSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intHandleVisualID,
                        OLD.intDeviceHandleID,
                        OLD.intInspectionTypeID,
                        OLD.intHandVisMetalSelectID,
                        OLD.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)     
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Device Handle Visual Inspection tables: {e}")            

    def createHandPhysMetSelTables(self):
        """ 
        Function Name: createHandPhysMetSelTables
        Function Purpose: Create the Handle Physical Metal Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Handle Physical Metal Selection Table
                CREATE TABLE IF NOT EXISTS THandlePhysMetalSelects
                (
                    intHandPhysMetalSelectID                INTEGER NOT NULL
                    ,strHandPhysMetSelect                   VARCHAR(1000) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT THandlePhysMetalSelects_PK PRIMARY KEY (intHandPhysMetalSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Handle Physical Metal Selection Table
                CREATE TABLE IF NOT EXISTS Z_THandlePhysMetalSelects
                (
                    intHandPhysMetalSelectAuditID           INTEGER NOT NULL
                    ,intHandPhysMetalSelectID               INTEGER NOT NULL
                    ,strHandPhysMetSelect                   VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_THandlePhysMetalSelects_PK PRIMARY KEY (intHandPhysMetalSelectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Handle Physical Metal Selection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_THandlePhysMetalSelects_InsertTrigger
                AFTER INSERT ON THandlePhysMetalSelects
                BEGIN
                    INSERT INTO Z_THandlePhysMetalSelects 
                    (
                        intHandPhysMetalSelectID,
                        strHandPhysMetSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intHandPhysMetalSelectID,
                        NEW.strHandPhysMetSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Handle Physical Metal Selection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_THandlePhysMetalSelects_UpdateTrigger
                AFTER UPDATE ON THandlePhysMetalSelects
                BEGIN
                    INSERT INTO Z_THandlePhysMetalSelects 
                    (
                        intHandPhysMetalSelectID,
                        strHandPhysMetSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intHandPhysMetalSelectID,
                        NEW.strHandPhysMetSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Handle Physical Metal Selection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_THandlePhysMetalSelects_DeleteTrigger
                AFTER DELETE ON THandlePhysMetalSelects
                BEGIN
                    INSERT INTO Z_THandlePhysMetalSelects 
                    (
                        intHandPhysMetalSelectID,
                        strHandPhysMetSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intHandPhysMetalSelectID,
                        OLD.strHandPhysMetSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Device Handle Physical Selection tables: {e}")            
            
    def createHandPhysInspectTables(self):
        """ 
        Function Name: createHandPhysInspectTables
        Function Purpose: Create the Handle Physical Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Handle Physical Inspection Table
                CREATE TABLE IF NOT EXISTS THandlePhysicalInspections 
                (
                    intHandlePhysicalID                             INTEGER NOT NULL
                    ,intDeviceHandleID                              INTEGER NOT NULL
                    ,intInspectionTypeID                            INTEGER NOT NULL            
                    ,intHandPhysMetalSelectID                       INTEGER NOT NULL 
                    ,intInspectionStatusID                          INTEGER NOT NULL
                    ,strModifiedReason                              VARCHAR(1000)
                    ,CONSTRAINT THandlePhysicalInspections_PK PRIMARY KEY (intHandlePhysicalID)

                    ,FOREIGN KEY ( intDeviceHandleID ) REFERENCES TDeviceHandles ( intDeviceHandleID )
                    ,FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )
                    ,FOREIGN KEY ( intHandPhysMetalSelectID ) REFERENCES THandlePhysMetalSelects ( intHandPhysMetalSelectID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )                     
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Handle Physical Inspection Table
                CREATE TABLE IF NOT EXISTS Z_THandlePhysicalInspections 
                (
                    intHandlePhysicalAuditID                        INTEGER NOT NULL
                    ,intHandlePhysicalID                            INTEGER NOT NULL
                    ,intDeviceHandleID                              INTEGER NOT NULL
                    ,intInspectionTypeID                            INTEGER NOT NULL            
                    ,intHandPhysMetalSelectID                       INTEGER NOT NULL
                    ,intInspectionStatusID                          INTEGER NOT NULL
                    ,strUpdatedBy                                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                                      VARCHAR(1) NOT NULL
                    ,strModifiedReason                              VARCHAR(1000)
                    ,CONSTRAINT Z_THandlePhysicalInspections_PK PRIMARY KEY (intHandlePhysicalAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Handle Physical Inspection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_THandlePhysicalInspections_InsertTrigger
                AFTER INSERT ON THandlePhysicalInspections
                BEGIN
                    INSERT INTO Z_THandlePhysicalInspections 
                    (
                        intHandlePhysicalID,
                        intDeviceHandleID,
                        intInspectionTypeID,
                        intHandPhysMetalSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intHandlePhysicalID,
                        NEW.intDeviceHandleID,
                        NEW.intInspectionTypeID,
                        NEW.intHandPhysMetalSelectID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Handle Physical Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_THandlePhysicalInspections_UpdateTrigger
                AFTER UPDATE ON THandlePhysicalInspections
                BEGIN
                    INSERT INTO Z_THandlePhysicalInspections 
                    (
                        intHandlePhysicalID,
                        intDeviceHandleID,
                        intInspectionTypeID,
                        intHandPhysMetalSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intHandlePhysicalID,
                        NEW.intDeviceHandleID,
                        NEW.intInspectionTypeID,
                        NEW.intHandPhysMetalSelectID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Handle Physical Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_THandlePhysicalInspections_DeleteTrigger
                AFTER DELETE ON THandlePhysicalInspections
                BEGIN
                    INSERT INTO Z_THandlePhysicalInspections 
                    (
                        intHandlePhysicalID,
                        intDeviceHandleID,
                        intInspectionTypeID,
                        intHandPhysMetalSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intHandlePhysicalID,
                        OLD.intDeviceHandleID,
                        OLD.intInspectionTypeID,
                        OLD.intHandPhysMetalSelectID,
                        OLD.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Device Handle Physical Inspection tables: {e}")            
            
    def createStandHandInspectTables(self):
        """ 
        Function Name: createStandHandInspectTables
        Function Purpose: Create the Standard Handle Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Standard Handle Inspection Table
                CREATE TABLE IF NOT EXISTS TStandardHandleInspections 
                (
                    intStandardHandleInspectionID               INTEGER NOT NULL
                    ,intHandleVisualID                          INTEGER NOT NULL 
                    ,intHandlePhysicalID                        INTEGER NOT NULL
                    ,intInspectionStatusID                      INTEGER NOT NULL
                    ,strModifiedReason                          VARCHAR(1000)
                    ,CONSTRAINT TStandardHandleInspections_PK PRIMARY KEY (intStandardHandleInspectionID)

                    ,FOREIGN KEY ( intHandleVisualID ) REFERENCES THandleVisualInspections ( intHandleVisualID )
                    ,FOREIGN KEY ( intHandlePhysicalID ) REFERENCES THandlePhysicalInspections ( intHandlePhysicalID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )                     
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Standard Handle Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TStandardHandleInspections 
                (
                    intStandardHandleInspectionAuditID          INTEGER NOT NULL
                    ,intStandardHandleInspectionID              INTEGER NOT NULL
                    ,intHandleVisualID                          INTEGER NOT NULL            
                    ,intHandlePhysicalID                        INTEGER NOT NULL
                    ,intInspectionStatusID                      INTEGER NOT NULL
                    ,strUpdatedBy                               VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                               DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                                  VARCHAR(1) NOT NULL
                    ,strModifiedReason                          VARCHAR(1000)
                    ,CONSTRAINT Z_TStandardHandleInspections_PK PRIMARY KEY (intStandardHandleInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Standard Handle Inspection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TStandardHandleInspections_InsertTrigger
                AFTER INSERT ON TStandardHandleInspections
                BEGIN
                    INSERT INTO Z_TStandardHandleInspections 
                    (
                        intStandardHandleInspectionID,
                        intHandleVisualID,
                        intHandlePhysicalID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intStandardHandleInspectionID,
                        NEW.intHandleVisualID,
                        NEW.intHandlePhysicalID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Standard Handle Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TStandardHandleInspections_UpdateTrigger
                AFTER UPDATE ON TStandardHandleInspections
                BEGIN
                    INSERT INTO Z_TStandardHandleInspections 
                    (
                        intStandardHandleInspectionID,
                        intHandleVisualID,
                        intHandlePhysicalID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intStandardHandleInspectionID,
                        NEW.intHandleVisualID,
                        NEW.intHandlePhysicalID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Standard Handle Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TStandardHandleInspections_DeleteTrigger
                AFTER DELETE ON TStandardHandleInspections
                BEGIN
                    INSERT INTO Z_TStandardHandleInspections 
                    (
                        intStandardHandleInspectionID,
                        intHandleVisualID,
                        intHandlePhysicalID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intStandardHandleInspectionID,
                        OLD.intHandleVisualID,
                        OLD.intHandlePhysicalID,
                        OLD.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Standard Device Handle Inspection tables: {e}")            

    def insertHandData(self):
        """ 
        Function Name: insertHandData
        Function Purpose: Insert the Handle data into the tables
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlInsert = """
                -- Add Data To Table: TDeviceHandles
                INSERT INTO TDeviceHandles (intDeviceHandleID, strHandleType)
                VALUES   (1, 'Single Installation Handle')
                        ,(2, 'Double Installation Handle')
                        ,(3, 'Multiple Installation Handle');            
                """
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlInsert)       
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error Inserting Device Handle Data: {e}")
            
    def createCaseTables(self):
        """ 
        Function Name: createCaseTables
        Function Purpose: Create the Case Housing and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Case Housing Table
                CREATE TABLE IF NOT EXISTS TCaseHousings 
                (
                    intCaseHousingID                 INTEGER NOT NULL
                    ,strCaseComponentDesc            VARCHAR(225) NOT NULL
                    ,strModifiedReason               VARCHAR(1000)
                    ,CONSTRAINT TCaseHousings_PK PRIMARY KEY (intCaseHousingID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Case Housing Table
                CREATE TABLE IF NOT EXISTS Z_TCaseHousings 
                (
                    intCaseHousingAuditID           INTEGER NOT NULL
                    ,intCaseHousingID               INTEGER NOT NULL
                    ,strCaseComponentDesc           VARCHAR(255) NOT NULL
                    ,strUpdatedBy                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                      VARCHAR(1) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT Z_TCaseHousings_PK PRIMARY KEY (intCaseHousingAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Case Housing Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TCaseHousings_InsertTrigger
                AFTER INSERT ON TCaseHousings
                BEGIN
                    INSERT INTO Z_TCaseHousings 
                    (
                        intCaseHousingID,
                        strCaseComponentDesc,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCaseHousingID,
                        NEW.strCaseComponentDesc,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Case Housing Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TCaseHousings_UpdateTrigger
                AFTER UPDATE ON TCaseHousings
                BEGIN
                    INSERT INTO Z_TCaseHousings 
                    (
                        intCaseHousingID,
                        strCaseComponentDesc,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCaseHousingID,
                        NEW.strCaseComponentDesc,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Case Housing Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TCaseHousings_DeleteTrigger
                AFTER DELETE ON TCaseHousings
                BEGIN
                    INSERT INTO Z_TCaseHousings 
                    (
                        intCaseHousingID,
                        strCaseComponentDesc,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intCaseHousingID,
                        OLD.strCaseComponentDesc,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Case Housing Component tables: {e}")            

    def createCaseCompSelTables(self):
        """ 
        Function Name: createCaseCompSelTables
        Function Purpose: Create the Case Component Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Case Component Selection Table
                CREATE TABLE IF NOT EXISTS TCaseCompSelects 
                (
                    intCaseCompSelectID                 INTEGER NOT NULL         
                    ,strCaseCompSelect                  VARCHAR(1000) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TCaseCompSelects_PK PRIMARY KEY (intCaseCompSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Case Component Selection Table
                CREATE TABLE IF NOT EXISTS Z_TCaseCompSelects
                (
                    intCaseCompSelectAuditID             INTEGER NOT NULL   
                    ,intCaseCompSelectID                 INTEGER NOT NULL                      
                    ,strCaseCompSelect                   VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                        VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                        DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                           VARCHAR(1) NOT NULL
                    ,strModifiedReason                   VARCHAR(1000)
                    ,CONSTRAINT Z_TCaseCompSelects_PK PRIMARY KEY (intCaseCompSelectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Case Component Selection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TCaseCompSelects_InsertTrigger
                AFTER INSERT ON TCaseCompSelects
                BEGIN
                    INSERT INTO Z_TCaseCompSelects 
                    (
                        intCaseCompSelectID,
                        strCaseCompSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCaseCompSelectID,
                        NEW.strCaseCompSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Case Component Selection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TCaseCompSelects_UpdateTrigger
                AFTER UPDATE ON TCaseCompSelects
                BEGIN
                    INSERT INTO Z_TCaseCompSelects 
                    (
                        intCaseCompSelectID,
                        strCaseCompSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCaseCompSelectID,
                        NEW.strCaseCompSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Case Component Selection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TCaseCompSelects_DeleteTrigger
                AFTER DELETE ON TCaseCompSelects
                BEGIN
                    INSERT INTO Z_TCaseCompSelects 
                    (
                        intCaseCompSelectID,
                        strCaseCompSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intCaseCompSelectID,
                        OLD.strCaseCompSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Case Component Selection tables: {e}")             

    def createCaseVisMetSelTables(self):
        """ 
        Function Name: createCaseVisMetSelTables
        Function Purpose: Create the Case Visual Metallic Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Case Visual Metallic Selection Table
                CREATE TABLE IF NOT EXISTS TCaseVisMetalSelects 
                (
                    intCaseVisMetalSelectID             INTEGER NOT NULL
                    ,strCaseVisMetSelect                VARCHAR(1000) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TCaseVisMetalSelects_PK PRIMARY KEY (intCaseVisMetalSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Case Visual Metallic Selection Table
                CREATE TABLE IF NOT EXISTS Z_TCaseVisMetalSelects
                (
                    intCaseVisMetalSelectAuditID        INTEGER NOT NULL   
                    ,intCaseVisMetalSelectID            INTEGER NOT NULL                      
                    ,strCaseVisMetSelect                VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TCaseVisMetalSelects_PK PRIMARY KEY (intCaseVisMetalSelectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Case Visual Metallic Selection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TCaseVisMetalSelects_InsertTrigger
                AFTER INSERT ON TCaseVisMetalSelects
                BEGIN
                    INSERT INTO Z_TCaseVisMetalSelects 
                    (
                        intCaseVisMetalSelectID,
                        strCaseVisMetSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCaseVisMetalSelectID,
                        NEW.strCaseVisMetSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Case Visual Metallic Selection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TCaseVisMetalSelects_UpdateTrigger
                AFTER UPDATE ON TCaseVisMetalSelects
                BEGIN
                    INSERT INTO Z_TCaseVisMetalSelects 
                    (
                        intCaseVisMetalSelectID,
                        strCaseVisMetSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCaseVisMetalSelectID,
                        NEW.strCaseVisMetSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Case Visual Metallic Selection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TCaseVisMetalSelects_DeleteTrigger
                AFTER DELETE ON TCaseVisMetalSelects
                BEGIN
                    INSERT INTO Z_TCaseVisMetalSelects 
                    (
                        intCaseVisMetalSelectID,
                        strCaseVisMetSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intCaseVisMetalSelectID,
                        OLD.strCaseVisMetSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)      
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Case Visual Selection tables: {e}")            

    def createCaseVisInspectTables(self):
        """ 
        Function Name: createCaseVisInspectTables
        Function Purpose: Create the Case Visual Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Case Visual Inspection Table
                CREATE TABLE IF NOT EXISTS TCaseHousingVisualInspections
                (
                    intCaseHousingVisualInspectionID        INTEGER NOT NULL
                    ,intCaseCompSelectID                    INTEGER NOT NULL
                    ,intInspectionTypeID                    INTEGER NOT NULL
                    ,intCaseVisMetalSelectID                INTEGER NOT NULL
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT TCaseHousingVisualInspections_PK PRIMARY KEY (intCaseHousingVisualInspectionID)

                    ,FOREIGN KEY ( intCaseCompSelectID ) REFERENCES TCaseCompSelects ( intCaseCompSelectID )
                    ,FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )
                    ,FOREIGN KEY ( intCaseVisMetalSelectID ) REFERENCES TCaseVisMetalSelects ( intCaseVisMetalSelectID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )                     
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Case Visual Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TCaseHousingVisualInspections 
                (
                    intCaseHousingVisualInspectionAuditID   INTEGER NOT NULL
                    ,intCaseHousingVisualInspectionID       INTEGER NOT NULL
                    ,intCaseCompSelectID                    INTEGER NOT NULL
                    ,intInspectionTypeID                    INTEGER NOT NULL            
                    ,intCaseVisMetalSelectID                INTEGER NOT NULL
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TCaseHousingVisualInspections_PK PRIMARY KEY (intCaseHousingVisualInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Case Visual Inspection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TCaseHousingVisualInspections_InsertTrigger
                AFTER INSERT ON TCaseHousingVisualInspections
                BEGIN
                    INSERT INTO Z_TCaseHousingVisualInspections 
                    (
                        intCaseHousingVisualInspectionID,
                        intCaseCompSelectID,
                        intInspectionTypeID,
                        intCaseVisMetalSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCaseHousingVisualInspectionID,
                        NEW.intCaseCompSelectID,
                        NEW.intInspectionTypeID,
                        NEW.intCaseVisMetalSelectID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Case Visual Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TCaseHousingVisualInspections_UpdateTrigger
                AFTER UPDATE ON TCaseHousingVisualInspections
                BEGIN
                    INSERT INTO Z_TCaseHousingVisualInspections 
                    (
                        intCaseHousingVisualInspectionID,
                        intCaseCompSelectID,
                        intInspectionTypeID,
                        intCaseVisMetalSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCaseHousingVisualInspectionID,
                        NEW.intCaseCompSelectID,
                        NEW.intInspectionTypeID,
                        NEW.intCaseVisMetalSelectID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Case Visual Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TCaseHousingVisualInspections_DeleteTrigger
                AFTER DELETE ON TCaseHousingVisualInspections
                BEGIN
                    INSERT INTO Z_TCaseHousingVisualInspections 
                    (
                        intCaseHousingVisualInspectionID,
                        intCaseCompSelectID,
                        intInspectionTypeID,
                        intCaseVisMetalSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intCaseHousingVisualInspectionID,
                        OLD.intCaseCompSelectID,
                        OLD.intInspectionTypeID,
                        OLD.intCaseVisMetalSelectID,
                        OLD.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Case Visual Inspection tables: {e}")            

    def createCasePhysMetSelTables(self):
        """ 
        Function Name: createCasePhysMetSelTables
        Function Purpose: Create the Case Physical Metal Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Case Physical Metal Selection Table
                CREATE TABLE IF NOT EXISTS TCasePhysMetalSelects
                (
                    intCasePhysMetalSelectID                INTEGER NOT NULL
                    ,strCasePhysMetSelect                   VARCHAR(1000) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT TCasePhysMetalSelects_PK PRIMARY KEY (intCasePhysMetalSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Case Physical Metal Selection Table
                CREATE TABLE IF NOT EXISTS Z_TCasePhysMetalSelects
                (
                    intCasePhysMetalSelectAuditID           INTEGER NOT NULL
                    ,intCasePhysMetalSelectID               INTEGER NOT NULL
                    ,strCasePhysMetSelect                   VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TCasePhysMetalSelects_PK PRIMARY KEY (intCasePhysMetalSelectAuditID)
                );
                """
            # Create the table trigger
            sqlTrigger = f"""
                -- Create Trigger: Case Physical Metal Selection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TCasePhysMetalSelects_InsertTrigger
                AFTER INSERT ON TCasePhysMetalSelects
                BEGIN
                    INSERT INTO Z_TCasePhysMetalSelects 
                    (
                        intCasePhysMetalSelectID,
                        strCasePhysMetSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCasePhysMetalSelectID,
                        NEW.strCasePhysMetSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Case Physical Metal Selection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TCasePhysMetalSelects_UpdateTrigger
                AFTER UPDATE ON TCasePhysMetalSelects
                BEGIN
                    INSERT INTO Z_TCasePhysMetalSelects 
                    (
                        intCasePhysMetalSelectID,
                        strCasePhysMetSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCasePhysMetalSelectID,
                        NEW.strCasePhysMetSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Case Physical Metal Selection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TCasePhysMetalSelects_DeleteTrigger
                AFTER DELETE ON TCasePhysMetalSelects
                BEGIN
                    INSERT INTO Z_TCasePhysMetalSelects 
                    (
                        intCasePhysMetalSelectID,
                        strCasePhysMetSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intCasePhysMetalSelectID,
                        OLD.strCasePhysMetSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)     
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Case Physical Selection tables: {e}")            

    def createCasePhysInspectTables(self):
        """ 
        Function Name: createCasePhysInspectTables
        Function Purpose: Create the Case Physical Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Case Physical Inspection Table
                CREATE TABLE IF NOT EXISTS TCaseHousingPhysicalInspections 
                (
                    intCaseHousingPhysicalInspectionID              INTEGER NOT NULL
                    ,intCaseCompSelectID                            INTEGER NOT NULL
                    ,intInspectionTypeID                            INTEGER NOT NULL            
                    ,intCasePhysMetalSelectID                       INTEGER NOT NULL 
                    ,intInspectionStatusID                          INTEGER NOT NULL
                    ,strModifiedReason                              VARCHAR(1000)
                    ,CONSTRAINT TCaseHousingPhysicalInspections_PK PRIMARY KEY (intCaseHousingPhysicalInspectionID)

                    ,FOREIGN KEY ( intCaseCompSelectID ) REFERENCES TCaseCompSelects ( intCaseCompSelectID )
                    ,FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )
                    ,FOREIGN KEY ( intCasePhysMetalSelectID ) REFERENCES TCasePhysMetalSelects ( intCasePhysMetalSelectID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )                     
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Case Physical Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TCaseHousingPhysicalInspections 
                (
                    intCaseHousingPhysicalInspectionAuditID         INTEGER NOT NULL
                    ,intCaseHousingPhysicalInspectionID             INTEGER NOT NULL
                    ,intCaseCompSelectID                            INTEGER NOT NULL
                    ,intInspectionTypeID                            INTEGER NOT NULL            
                    ,intCasePhysMetalSelectID                       INTEGER NOT NULL
                    ,intInspectionStatusID                          INTEGER NOT NULL
                    ,strUpdatedBy                                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                                      VARCHAR(1) NOT NULL
                    ,strModifiedReason                              VARCHAR(1000)
                    ,CONSTRAINT Z_TCaseHousingPhysicalInspections_PK PRIMARY KEY (intCaseHousingPhysicalInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Case Physical Inspection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TCaseHousingPhysicalInspections_InsertTrigger
                AFTER INSERT ON TCaseHousingPhysicalInspections
                BEGIN
                    INSERT INTO Z_TCaseHousingPhysicalInspections 
                    (
                        intCaseHousingPhysicalInspectionID,
                        intCaseCompSelectID,
                        intInspectionTypeID,
                        intCasePhysMetalSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCaseHousingPhysicalInspectionID,
                        NEW.intCaseCompSelectID,
                        NEW.intInspectionTypeID,
                        NEW.intCasePhysMetalSelectID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Case Physical Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TCaseHousingPhysicalInspections_UpdateTrigger
                AFTER UPDATE ON TCaseHousingPhysicalInspections
                BEGIN
                    INSERT INTO Z_TCaseHousingPhysicalInspections 
                    (
                        intCaseHousingPhysicalInspectionID,
                        intCaseCompSelectID,
                        intInspectionTypeID,
                        intCasePhysMetalSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intCaseHousingPhysicalInspectionID,
                        NEW.intCaseCompSelectID,
                        NEW.intInspectionTypeID,
                        NEW.intCasePhysMetalSelectID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Case Physical Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TCaseHousingPhysicalInspections_DeleteTrigger
                AFTER DELETE ON TCaseHousingPhysicalInspections
                BEGIN
                    INSERT INTO Z_TCaseHousingPhysicalInspections 
                    (
                        intCaseHousingPhysicalInspectionID,
                        intCaseCompSelectID,
                        intInspectionTypeID,
                        intCasePhysMetalSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intCaseHousingPhysicalInspectionID,
                        OLD.intCaseCompSelectID,
                        OLD.intInspectionTypeID,
                        OLD.intCasePhysMetalSelectID,
                        OLD.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Case Physical Inspection tables: {e}")            

    def createStandCaseInspectTables(self):
        """ 
        Function Name: createStandCaseInspectTables
        Function Purpose: Create the Standard Case Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Standard Case Inspection Table
                CREATE TABLE IF NOT EXISTS TStandardCaseHousingInspections 
                (
                    intStandardCaseHousingInspectionID          INTEGER NOT NULL
                    ,intCaseHousingVisualInspectionID           INTEGER NOT NULL 
                    ,intCaseHousingPhysicalInspectionID         INTEGER NOT NULL
                    ,intInspectionStatusID                      INTEGER NOT NULL
                    ,strModifiedReason                          VARCHAR(1000)
                    ,CONSTRAINT TStandardCaseHousingInspections_PK PRIMARY KEY (intStandardCaseHousingInspectionID)

                    ,FOREIGN KEY ( intCaseHousingVisualInspectionID ) REFERENCES TCaseHousingVisualInspections ( intCaseHousingVisualInspectionID )
                    ,FOREIGN KEY ( intCaseHousingPhysicalInspectionID ) REFERENCES TCaseHousingPhysicalInspections ( intCaseHousingPhysicalInspectionID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Standard Case Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TStandardCaseHousingInspections 
                (
                    intStandardCaseHousingInspectionAuditID     INTEGER NOT NULL
                    ,intStandardCaseHousingInspectionID         INTEGER NOT NULL
                    ,intCaseHousingVisualInspectionID           INTEGER NOT NULL            
                    ,intCaseHousingPhysicalInspectionID         INTEGER NOT NULL
                    ,intInspectionStatusID                      INTEGER NOT NULL
                    ,strUpdatedBy                               VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                               DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                                  VARCHAR(1) NOT NULL
                    ,strModifiedReason                          VARCHAR(1000)
                    ,CONSTRAINT Z_TStandardCaseHousingInspections_PK PRIMARY KEY (intStandardCaseHousingInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Standard Case Inspection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TStandardCaseHousingInspections_InsertTrigger
                AFTER INSERT ON TStandardCaseHousingInspections
                BEGIN
                    INSERT INTO Z_TStandardCaseHousingInspections 
                    (
                        intStandardCaseHousingInspectionID,
                        intCaseHousingVisualInspectionID,
                        intCaseHousingPhysicalInspectionID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intStandardCaseHousingInspectionID,
                        NEW.intCaseHousingVisualInspectionID,
                        NEW.intCaseHousingPhysicalInspectionID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Standard Case Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TStandardCaseHousingInspections_UpdateTrigger
                AFTER UPDATE ON TStandardCaseHousingInspections
                BEGIN
                    INSERT INTO Z_TStandardCaseHousingInspections 
                    (
                        intStandardCaseHousingInspectionID,
                        intCaseHousingVisualInspectionID,
                        intCaseHousingPhysicalInspectionID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intStandardCaseHousingInspectionID,
                        NEW.intCaseHousingVisualInspectionID,
                        NEW.intCaseHousingPhysicalInspectionID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Standard Case Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TStandardCaseHousingInspections_DeleteTrigger
                AFTER DELETE ON TStandardCaseHousingInspections
                BEGIN
                    INSERT INTO Z_TStandardCaseHousingInspections 
                    (
                        intStandardCaseHousingInspectionID,
                        intCaseHousingVisualInspectionID,
                        intCaseHousingPhysicalInspectionID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intStandardCaseHousingInspectionID,
                        OLD.intCaseHousingVisualInspectionID,
                        OLD.intCaseHousingPhysicalInspectionID,
                        OLD.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Standard Case Inspection tables: {e}")                        

    def insertCaseData(self):
        """ 
        Function Name: insertCaseData
        Function Purpose: Insert the Case data into the tables
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlInsert = """
                -- Add Data To Table: TCaseHousings
                INSERT INTO TCaseHousings (intCaseHousingID, strCaseComponentDesc)
                VALUES   (1, 'No Component Issues')
                        ,(2, 'Front Case')
                        ,(3, 'Back Case')
                        ,(4, 'Housing Fasteners')
                        ,(5, 'Axel Nut')
                        ,(6, 'Other Component Issues');           
                """
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlInsert)       
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error Inserting Case Data: {e}")            

    def createBrakeTables(self):
        """ 
        Function Name: createBrakeTables
        Function Purpose: Create the Brake Housing and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Brake Housing Table
                CREATE TABLE IF NOT EXISTS TBrakeHousings 
                (
                    intBrakeHousingID                INTEGER NOT NULL
                    ,strBrakeComponentDesc           VARCHAR(225) NOT NULL
                    ,strModifiedReason               VARCHAR(1000)
                    ,CONSTRAINT TBrakeHousings_PK PRIMARY KEY (intBrakeHousingID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Brake Housing Table
                CREATE TABLE IF NOT EXISTS Z_TBrakeHousings 
                (
                    intBrakeHousingAuditID          INTEGER NOT NULL
                    ,intBrakeHousingID              INTEGER NOT NULL
                    ,strBrakeComponentDesc          VARCHAR(255) NOT NULL
                    ,strUpdatedBy                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                      VARCHAR(1) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT Z_TBrakeHousings_PK PRIMARY KEY (intBrakeHousingAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Brake Housing Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TBrakeHousings_InsertTrigger
                AFTER INSERT ON TBrakeHousings
                BEGIN
                    INSERT INTO Z_TBrakeHousings 
                    (
                        intBrakeHousingID,
                        strBrakeComponentDesc,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBrakeHousingID,
                        NEW.strBrakeComponentDesc,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Brake Housing Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TBrakeHousings_UpdateTrigger
                AFTER UPDATE ON TBrakeHousings
                BEGIN
                    INSERT INTO Z_TBrakeHousings 
                    (
                        intBrakeHousingID,
                        strBrakeComponentDesc,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBrakeHousingID,
                        NEW.strBrakeComponentDesc,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Brake Housing Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TBrakeHousings_DeleteTrigger
                AFTER DELETE ON TBrakeHousings
                BEGIN
                    INSERT INTO Z_TBrakeHousings 
                    (
                        intBrakeHousingID,
                        strBrakeComponentDesc,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intBrakeHousingID,
                        OLD.strBrakeComponentDesc,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Brake Housing Component tables: {e}")            

    def createBrakeCompSelTables(self):
        """ 
        Function Name: createBrakeCompSelTables
        Function Purpose: Create the Brake Component Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Brake Component Selection Table
                CREATE TABLE IF NOT EXISTS TBrakeCompSelects 
                (
                    intBrakeCompSelectID                INTEGER NOT NULL         
                    ,strBrakeCompSelect                 VARCHAR(1000) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TBrakeCompSelects_PK PRIMARY KEY (intBrakeCompSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Brake Component Selection Table
                CREATE TABLE IF NOT EXISTS Z_TBrakeCompSelects
                (
                    intBrakeCompSelectAuditID            INTEGER NOT NULL   
                    ,intBrakeCompSelectID                INTEGER NOT NULL                      
                    ,strBrakeCompSelect                  VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                        VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                        DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                           VARCHAR(1) NOT NULL
                    ,strModifiedReason                   VARCHAR(1000)
                    ,CONSTRAINT Z_TBrakeCompSelects_PK PRIMARY KEY (intBrakeCompSelectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Brake Component Selection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TBrakeCompSelects_InsertTrigger
                AFTER INSERT ON TBrakeCompSelects
                BEGIN
                    INSERT INTO Z_TBrakeCompSelects 
                    (
                        intBrakeCompSelectID,
                        strBrakeCompSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBrakeCompSelectID,
                        NEW.strBrakeCompSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Brake Component Selection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TBrakeCompSelects_UpdateTrigger
                AFTER UPDATE ON TBrakeCompSelects
                BEGIN
                    INSERT INTO Z_TBrakeCompSelects 
                    (
                        intBrakeCompSelectID,
                        strBrakeCompSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBrakeCompSelectID,
                        NEW.strBrakeCompSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Brake Component Selection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TBrakeCompSelects_DeleteTrigger
                AFTER DELETE ON TBrakeCompSelects
                BEGIN
                    INSERT INTO Z_TBrakeCompSelects 
                    (
                        intBrakeCompSelectID,
                        strBrakeCompSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intBrakeCompSelectID,
                        OLD.strBrakeCompSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Brake Component Selection tables: {e}")            

    def createBrakeVisMetSelTables(self):
        """ 
        Function Name: createBrakeVisMetSelTables
        Function Purpose: Create the Brake Visual Metallic Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Brake Visual Metallic Selection Table
                CREATE TABLE IF NOT EXISTS TBrakeVisMetalSelects 
                (
                    intBrakeVisMetalSelectID            INTEGER NOT NULL
                    ,strBrakeVisMetSelect               VARCHAR(1000) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TBrakeVisMetalSelects_PK PRIMARY KEY (intBrakeVisMetalSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Brake Visual Metallic Selection Table
                CREATE TABLE IF NOT EXISTS Z_TBrakeVisMetalSelects
                (
                    intBrakeVisMetalSelectAuditID       INTEGER NOT NULL   
                    ,intBrakeVisMetalSelectID           INTEGER NOT NULL                      
                    ,strBrakeVisMetSelect               VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TBrakeVisMetalSelects_PK PRIMARY KEY (intBrakeVisMetalSelectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Brake Visual Metallic Selection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TBrakeVisMetalSelects_InsertTrigger
                AFTER INSERT ON TBrakeVisMetalSelects
                BEGIN
                    INSERT INTO Z_TBrakeVisMetalSelects 
                    (
                        intBrakeVisMetalSelectID,
                        strBrakeVisMetSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBrakeVisMetalSelectID,
                        NEW.strBrakeVisMetSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Brake Visual Metallic Selection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TBrakeVisMetalSelects_UpdateTrigger
                AFTER UPDATE ON TBrakeVisMetalSelects
                BEGIN
                    INSERT INTO Z_TBrakeVisMetalSelects 
                    (
                        intBrakeVisMetalSelectID,
                        strBrakeVisMetSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBrakeVisMetalSelectID,
                        NEW.strBrakeVisMetSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Brake Visual Metallic Selection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TBrakeVisMetalSelects_DeleteTrigger
                AFTER DELETE ON TBrakeVisMetalSelects
                BEGIN
                    INSERT INTO Z_TBrakeVisMetalSelects 
                    (
                        intBrakeVisMetalSelectID,
                        strBrakeVisMetSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intBrakeVisMetalSelectID,
                        OLD.strBrakeVisMetSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Brake Visual Selection tables: {e}")            

    def createBrakeVisInspectTables(self):
        """ 
        Function Name: createBrakeVisInspectTables
        Function Purpose: Create the Brake Visual Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Brake Visual Inspection Table
                CREATE TABLE IF NOT EXISTS TBrakeHousingVisualInspections
                (
                    intBrakeHousingVisualInspectionID       INTEGER NOT NULL
                    ,intBrakeCompSelectID                   INTEGER NOT NULL
                    ,intInspectionTypeID                    INTEGER NOT NULL
                    ,intBrakeVisMetalSelectID               INTEGER NOT NULL
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT TBrakeHousingVisualInspections_PK PRIMARY KEY (intBrakeHousingVisualInspectionID)

                    ,FOREIGN KEY ( intBrakeCompSelectID ) REFERENCES TBrakeCompSelects ( intBrakeCompSelectID )
                    ,FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )
                    ,FOREIGN KEY ( intBrakeVisMetalSelectID ) REFERENCES TBrakeVisMetalSelects ( intBrakeVisMetalSelectID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )                    
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Brake Visual Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TBrakeHousingVisualInspections 
                (
                    intBrakeHousingVisualInspectionAuditID  INTEGER NOT NULL
                    ,intBrakeHousingVisualInspectionID      INTEGER NOT NULL
                    ,intBrakeCompSelectID                   INTEGER NOT NULL
                    ,intInspectionTypeID                    INTEGER NOT NULL            
                    ,intBrakeVisMetalSelectID               INTEGER NOT NULL
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TBrakeHousingVisualInspections_PK PRIMARY KEY (intBrakeHousingVisualInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Brake Housing Visual Inspection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TBrakeHousingVisualInspections_InsertTrigger
                AFTER INSERT ON TBrakeHousingVisualInspections
                BEGIN
                    INSERT INTO Z_TBrakeHousingVisualInspections 
                    (
                        intBrakeHousingVisualInspectionID,
                        intBrakeCompSelectID,
                        intInspectionTypeID,
                        intBrakeVisMetalSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBrakeHousingVisualInspectionID,
                        NEW.intBrakeCompSelectID,
                        NEW.intInspectionTypeID,
                        NEW.intBrakeVisMetalSelectID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Brake Housing Visual Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TBrakeHousingVisualInspections_UpdateTrigger
                AFTER UPDATE ON TBrakeHousingVisualInspections
                BEGIN
                    INSERT INTO Z_TBrakeHousingVisualInspections 
                    (
                        intBrakeHousingVisualInspectionID,
                        intBrakeCompSelectID,
                        intInspectionTypeID,
                        intBrakeVisMetalSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBrakeHousingVisualInspectionID,
                        NEW.intBrakeCompSelectID,
                        NEW.intInspectionTypeID,
                        NEW.intBrakeVisMetalSelectID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Brake Housing Visual Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TBrakeHousingVisualInspections_DeleteTrigger
                AFTER DELETE ON TBrakeHousingVisualInspections
                BEGIN
                    INSERT INTO Z_TBrakeHousingVisualInspections 
                    (
                        intBrakeHousingVisualInspectionID,
                        intBrakeCompSelectID,
                        intInspectionTypeID,
                        intBrakeVisMetalSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intBrakeHousingVisualInspectionID,
                        OLD.intBrakeCompSelectID,
                        OLD.intInspectionTypeID,
                        OLD.intBrakeVisMetalSelectID,
                        OLD.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """


            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Brake Visual Inspection tables: {e}")            

    def createBrakePhysMetSelTables(self):
        """ 
        Function Name: createBrakePhysMetSelTables
        Function Purpose: Create the Brake Physical Metal Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Brake Physical Metal Selection Table
                CREATE TABLE IF NOT EXISTS TBrakePhysMetalSelects
                (
                    intBrakePhysMetalSelectID               INTEGER NOT NULL
                    ,strBrakePhysMetSelect                  VARCHAR(1000) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT TBrakePhysMetalSelects_PK PRIMARY KEY (intBrakePhysMetalSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Brake Physical Metal Selection Table
                CREATE TABLE IF NOT EXISTS Z_TBrakePhysMetalSelects
                (
                    intBrakePhysMetalSelectAuditID          INTEGER NOT NULL
                    ,intBrakePhysMetalSelectID              INTEGER NOT NULL
                    ,strBrakePhysMetSelect                  VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TBrakePhysMetalSelects_PK PRIMARY KEY (intBrakePhysMetalSelectAuditID)
                );
                """
            # Create the table trigger
            sqlTrigger = f"""
                -- Create Trigger: Brake Physical Metal Selection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TBrakePhysMetalSelects_InsertTrigger
                AFTER INSERT ON TBrakePhysMetalSelects
                BEGIN
                    INSERT INTO Z_TBrakePhysMetalSelects 
                    (
                        intBrakePhysMetalSelectID,
                        strBrakePhysMetSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBrakePhysMetalSelectID,
                        NEW.strBrakePhysMetSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Brake Physical Metal Selection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TBrakePhysMetalSelects_UpdateTrigger
                AFTER UPDATE ON TBrakePhysMetalSelects
                BEGIN
                    INSERT INTO Z_TBrakePhysMetalSelects 
                    (
                        intBrakePhysMetalSelectID,
                        strBrakePhysMetSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBrakePhysMetalSelectID,
                        NEW.strBrakePhysMetSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Brake Physical Metal Selection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TBrakePhysMetalSelects_DeleteTrigger
                AFTER DELETE ON TBrakePhysMetalSelects
                BEGIN
                    INSERT INTO Z_TBrakePhysMetalSelects 
                    (
                        intBrakePhysMetalSelectID,
                        strBrakePhysMetSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intBrakePhysMetalSelectID,
                        OLD.strBrakePhysMetSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Brake Physical Selection tables: {e}")            

    def createBrakePhysInspectTables(self):
        """ 
        Function Name: createCasePhysInspectTables
        Function Purpose: Create the Brake Physical Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Brake Physical Inspection Table
                CREATE TABLE IF NOT EXISTS TBrakeHousingPhysicalInspections 
                (
                    intBrakeHousingPhysicalInspectionID             INTEGER NOT NULL
                    ,intBrakeCompSelectID                           INTEGER NOT NULL
                    ,intInspectionTypeID                            INTEGER NOT NULL            
                    ,intBrakePhysMetalSelectID                      INTEGER NOT NULL 
                    ,intInspectionStatusID                          INTEGER NOT NULL
                    ,strModifiedReason                              VARCHAR(1000)
                    ,CONSTRAINT TBrakeHousingPhysicalInspections_PK PRIMARY KEY (intBrakeHousingPhysicalInspectionID)

                    ,FOREIGN KEY ( intBrakeCompSelectID ) REFERENCES TBrakeCompSelects ( intBrakeCompSelectID )
                    ,FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )
                    ,FOREIGN KEY ( intBrakePhysMetalSelectID ) REFERENCES TBrakePhysMetalSelects ( intBrakePhysMetalSelectID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )                      
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Brake Physical Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TBrakeHousingPhysicalInspections 
                (
                    intBrakeHousingPhysicalInspectionAuditID        INTEGER NOT NULL
                    ,intBrakeHousingPhysicalInspectionID            INTEGER NOT NULL
                    ,intBrakeCompSelectID                           INTEGER NOT NULL
                    ,intInspectionTypeID                            INTEGER NOT NULL            
                    ,intBrakePhysMetalSelectID                      INTEGER NOT NULL
                    ,intInspectionStatusID                          INTEGER NOT NULL
                    ,strUpdatedBy                                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                                      VARCHAR(1) NOT NULL
                    ,strModifiedReason                              VARCHAR(1000)
                    ,CONSTRAINT Z_TBrakeHousingPhysicalInspections_PK PRIMARY KEY (intBrakeHousingPhysicalInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Brake Housing Physical Inspections Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TBrakeHousingPhysicalInspections_InsertTrigger
                AFTER INSERT ON TBrakeHousingPhysicalInspections
                BEGIN
                    INSERT INTO Z_TBrakeHousingPhysicalInspections 
                    (
                        intBrakeHousingPhysicalInspectionID,
                        intBrakeCompSelectID,
                        intInspectionTypeID,
                        intBrakePhysMetalSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBrakeHousingPhysicalInspectionID,
                        NEW.intBrakeCompSelectID,
                        NEW.intInspectionTypeID,
                        NEW.intBrakePhysMetalSelectID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Brake Housing Physical Inspections Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TBrakeHousingPhysicalInspections_UpdateTrigger
                AFTER UPDATE ON TBrakeHousingPhysicalInspections
                BEGIN
                    INSERT INTO Z_TBrakeHousingPhysicalInspections 
                    (
                        intBrakeHousingPhysicalInspectionID,
                        intBrakeCompSelectID,
                        intInspectionTypeID,
                        intBrakePhysMetalSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBrakeHousingPhysicalInspectionID,
                        NEW.intBrakeCompSelectID,
                        NEW.intInspectionTypeID,
                        NEW.intBrakePhysMetalSelectID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Brake Housing Physical Inspections Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TBrakeHousingPhysicalInspections_DeleteTrigger
                AFTER DELETE ON TBrakeHousingPhysicalInspections
                BEGIN
                    INSERT INTO Z_TBrakeHousingPhysicalInspections 
                    (
                        intBrakeHousingPhysicalInspectionID,
                        intBrakeCompSelectID,
                        intInspectionTypeID,
                        intBrakePhysMetalSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intBrakeHousingPhysicalInspectionID,
                        OLD.intBrakeCompSelectID,
                        OLD.intInspectionTypeID,
                        OLD.intBrakePhysMetalSelectID,
                        OLD.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Brake Physical Inspection tables: {e}")            

    def createStandBrakeInspectTables(self):
        """ 
        Function Name: createStandBrakeInspectTables
        Function Purpose: Create the Standard Brake Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Standard Brake Inspection Table
                CREATE TABLE IF NOT EXISTS TStandardBrakeHousingInspections 
                (
                    intStandardBrakeHousingInspectionID         INTEGER NOT NULL
                    ,intBrakeHousingVisualInspectionID          INTEGER NOT NULL 
                    ,intBrakeHousingPhysicalInspectionID        INTEGER NOT NULL
                    ,intInspectionStatusID                      INTEGER NOT NULL
                    ,strModifiedReason                          VARCHAR(1000)
                    ,CONSTRAINT TStandardBrakeHousingInspections_PK PRIMARY KEY (intStandardBrakeHousingInspectionID)
                    
                    ,FOREIGN KEY ( intBrakeHousingVisualInspectionID ) REFERENCES TBrakeHousingVisualInspections ( intBrakeHousingVisualInspectionID )
                    ,FOREIGN KEY ( intBrakeHousingPhysicalInspectionID ) REFERENCES TBrakeHousingPhysicalInspections ( intBrakeHousingPhysicalInspectionID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Standard Brake Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TStandardBrakeHousingInspections 
                (
                    intStandardBrakeHousingInspectionAuditID    INTEGER NOT NULL
                    ,intStandardBrakeHousingInspectionID        INTEGER NOT NULL
                    ,intBrakeHousingVisualInspectionID          INTEGER NOT NULL            
                    ,intBrakeHousingPhysicalInspectionID        INTEGER NOT NULL
                    ,intInspectionStatusID                      INTEGER NOT NULL
                    ,strUpdatedBy                               VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                               DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                                  VARCHAR(1) NOT NULL
                    ,strModifiedReason                          VARCHAR(1000)
                    ,CONSTRAINT Z_TStandardBrakeHousingInspections_PK PRIMARY KEY (intStandardBrakeHousingInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Standard Brake Inspection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TStandardBrakeHousingInspections_InsertTrigger
                AFTER INSERT ON TStandardBrakeHousingInspections
                BEGIN
                    INSERT INTO Z_TStandardBrakeHousingInspections 
                    (
                        intStandardBrakeHousingInspectionID,
                        intBrakeHousingVisualInspectionID,
                        intBrakeHousingPhysicalInspectionID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intStandardBrakeHousingInspectionID,
                        NEW.intBrakeHousingVisualInspectionID,
                        NEW.intBrakeHousingPhysicalInspectionID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Standard Brake Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TStandardBrakeHousingInspections_UpdateTrigger
                AFTER UPDATE ON TStandardBrakeHousingInspections
                BEGIN
                    INSERT INTO Z_TStandardBrakeHousingInspections 
                    (
                        intStandardBrakeHousingInspectionID,
                        intBrakeHousingVisualInspectionID,
                        intBrakeHousingPhysicalInspectionID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intStandardBrakeHousingInspectionID,
                        NEW.intBrakeHousingVisualInspectionID,
                        NEW.intBrakeHousingPhysicalInspectionID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Standard Brake Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TStandardBrakeHousingInspections_DeleteTrigger
                AFTER DELETE ON TStandardBrakeHousingInspections
                BEGIN
                    INSERT INTO Z_TStandardBrakeHousingInspections 
                    (
                        intStandardBrakeHousingInspectionID,
                        intBrakeHousingVisualInspectionID,
                        intBrakeHousingPhysicalInspectionID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intStandardBrakeHousingInspectionID,
                        OLD.intBrakeHousingVisualInspectionID,
                        OLD.intBrakeHousingPhysicalInspectionID,
                        OLD.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)  
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Standard Brake Inspection tables: {e}")                        

    def insertBrakeData(self):
        """ 
        Function Name: insertBrakeData
        Function Purpose: Insert the Brake data into the tables
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlInsert = """
                -- Add Data To Table: TBrakeHousings
                INSERT INTO TBrakeHousings (intBrakeHousingID, strBrakeComponentDesc)
                VALUES   (1, 'No Component Issues')
                        ,(2, 'Brake Cover Screws')
                        ,(3, 'Brake Cover')
                        ,(4, 'Brake Drum')
                        ,(5, 'Nozzle Pin')
                        ,(6, 'Other Component Issues');
                """
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlInsert)       
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error Inserting Brake Data: {e}")              

    def createLanyardTables(self):
        """ 
        Function Name: createLanyardTables
        Function Purpose: Create the Lanyard Length and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Lanyard Length Table
                CREATE TABLE IF NOT EXISTS TLanyards 
                (
                    intLanyardID                INTEGER NOT NULL
                    ,strLanyardLength           VARCHAR(255) NOT NULL
                    ,strModifiedReason          VARCHAR(1000)
                    ,CONSTRAINT TLanyards_PK PRIMARY KEY (intLanyardID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Lanyard Length Table
                CREATE TABLE IF NOT EXISTS Z_TLanyards 
                (
                    intLanyardAuditID           INTEGER NOT NULL
                    ,intLanyardID               INTEGER NOT NULL
                    ,strLanyardLength           VARCHAR(255) NOT NULL
                    ,strUpdatedBy               VARCHAR(225) NOT NULL
                    ,dtmUpdatedOn               DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                  VARCHAR(1) NOT NULL
                    ,strModifiedReason          VARCHAR(1000)
                    ,CONSTRAINT Z_TLanyards_PK PRIMARY KEY (intLanyardAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Lanyard Length Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TLanyards_InsertTrigger
                AFTER INSERT ON TLanyards
                BEGIN
                    INSERT INTO Z_TLanyards 
                    (
                        intLanyardID,
                        strLanyardLength,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intLanyardID,
                        NEW.strLanyardLength,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Lanyard Length Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TLanyards_UpdateTrigger
                AFTER UPDATE ON TLanyards
                BEGIN
                    INSERT INTO Z_TLanyards 
                    (
                        intLanyardID,
                        strLanyardLength,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intLanyardID,
                        NEW.strLanyardLength,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Lanyard Length Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TLanyards_DeleteTrigger
                AFTER DELETE ON TLanyards
                BEGIN
                    INSERT INTO Z_TLanyards 
                    (
                        intLanyardID,
                        strLanyardLength,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intLanyardID,
                        OLD.strLanyardLength,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Lanyard Length tables: {e}")            

    def createLanVisTextSelTables(self):
        """ 
        Function Name: createLanVisTextSelTables
        Function Purpose: Create the Lanyard Visual Textile Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Lanyard Visual Textile Selection Table
                CREATE TABLE IF NOT EXISTS TLanVisTextSelects 
                (
                    intLanVisTextSelectID               INTEGER NOT NULL
                    ,strLanVisTextSelect                VARCHAR(1000) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TLanVisTextSelects_PK PRIMARY KEY (intLanVisTextSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Lanyard Visual Textile Selection Table
                CREATE TABLE IF NOT EXISTS Z_TLanVisTextSelects 
                (
                    intLanVisTextSelectAuditID          INTEGER NOT NULL
                    ,intLanVisTextSelectID              INTEGER NOT NULL
                    ,strLanVisTextSelect                VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TLanVisTextSelects_PK PRIMARY KEY (intLanVisTextSelectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Lanyard Visual Textile Selection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TLanVisTextSelects_InsertTrigger
                AFTER INSERT ON TLanVisTextSelects
                BEGIN
                    INSERT INTO Z_TLanVisTextSelects 
                    (
                        intLanVisTextSelectID,
                        strLanVisTextSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intLanVisTextSelectID,
                        NEW.strLanVisTextSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Lanyard Visual Textile Selection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TLanVisTextSelects_UpdateTrigger
                AFTER UPDATE ON TLanVisTextSelects
                BEGIN
                    INSERT INTO Z_TLanVisTextSelects 
                    (
                        intLanVisTextSelectID,
                        strLanVisTextSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intLanVisTextSelectID,
                        NEW.strLanVisTextSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Lanyard Visual Textile Selection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TLanVisTextSelects_DeleteTrigger
                AFTER DELETE ON TLanVisTextSelects
                BEGIN
                    INSERT INTO Z_TLanVisTextSelects 
                    (
                        intLanVisTextSelectID,
                        strLanVisTextSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intLanVisTextSelectID,
                        OLD.strLanVisTextSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Lanyard Visual Textile Selection tables: {e}")             

    def createLanVisInspectTables(self):
        """ 
        Function Name: createLanVisInspectTables
        Function Purpose: Create the Lanyard Visual Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Lanyard Visual Inspection Table
                CREATE TABLE IF NOT EXISTS TLanyardVisualInspections 
                (
                    intLanyardVisualInspectionID       INTEGER NOT NULL
                    ,intLanyardID                       INTEGER NOT NULL
                    ,intInspectionTypeID                INTEGER NOT NULL
                    ,intLanVisTextSelectID              INTEGER NOT NULL
                    ,intInspectionStatusID              INTEGER NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TLanyardVisualInspections_PK PRIMARY KEY (intLanyardVisualInspectionID)
                    
                    ,FOREIGN KEY ( intLanyardID ) REFERENCES TLanyards ( intLanyardID )
                    ,FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )
                    ,FOREIGN KEY ( intLanVisTextSelectID ) REFERENCES TLanVisTextSelects ( intLanVisTextSelectID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )                     
                );    
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Lanyard Visual Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TLanyardVisualInspections
                (
                    intLanyardVisualInspectionAuditID   INTEGER NOT NULL
                    ,intLanyardVisualInspectionID       INTEGER NOT NULL
                    ,intLanyardID                       INTEGER NOT NULL
                    ,intInspectionTypeID                INTEGER NOT NULL
                    ,intLanVisTextSelectID              INTEGER NOT NULL
                    ,intInspectionStatusID              INTEGER NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TLanyardVisualInspections_PK PRIMARY KEY (intLanyardVisualInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Lanyard Visual Inspection Table
                CREATE TRIGGER IF NOT EXISTS Z_TLanyardVisualInspections_InsertTrigger
                AFTER INSERT ON TLanyardVisualInspections
                BEGIN
                    INSERT INTO Z_TLanyardVisualInspections 
                    (
                        intLanyardVisualInspectionID
                        ,intLanyardID
                        ,intInspectionTypeID
                        ,intLanVisTextSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intLanyardVisualInspectionID
                        ,NEW.intLanyardID
                        ,NEW.intInspectionTypeID
                        ,NEW.intLanVisTextSelectID
                        ,NEW.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Lanyard Visual Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TLanyardVisualInspections_UpdateTrigger
                AFTER UPDATE ON TLanyardVisualInspections
                BEGIN
                    INSERT INTO Z_TLanyardVisualInspections 
                    (
                        intLanyardVisualInspectionID
                        ,intLanyardID
                        ,intInspectionTypeID
                        ,intLanVisTextSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intLanyardVisualInspectionID
                        ,NEW.intLanyardID
                        ,NEW.intInspectionTypeID
                        ,NEW.intLanVisTextSelectID
                        ,NEW.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Lanyard Visual Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TLanyardVisualInspections_DeleteTrigger
                AFTER DELETE ON TLanyardVisualInspections
                BEGIN
                    INSERT INTO Z_TLanyardVisualInspections 
                    (
                        intLanyardVisualInspectionID
                        ,intLanyardID
                        ,intInspectionTypeID
                        ,intLanVisTextSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intLanyardVisualInspectionID
                        ,OLD.intLanyardID
                        ,OLD.intInspectionTypeID
                        ,OLD.intLanVisTextSelectID
                        ,OLD.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,datetime('now')
                        ,'D' -- Delete
                        ,NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Lanyard Visual Inspection tables: {e}")            

    def createLanPhysTextSelTables(self):
        """ 
        Function Name: createLanPhysTextSelTables
        Function Purpose: Create the Lanyard Physical Textile Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Lanyard Physical Textile Selection Table
                CREATE TABLE IF NOT EXISTS TLanPhysTextSelects 
                (
                    intLanPhysTextSelectID              INTEGER NOT NULL
                    ,strLanPhysTextSelect               VARCHAR(1000) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TLanPhysTextSelects_PK PRIMARY KEY (intLanPhysTextSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Lanyard Physical Textile Selection Table
                CREATE TABLE IF NOT EXISTS Z_TLanPhysTextSelects 
                (
                    intLanPhysTextSelectAuditID         INTEGER NOT NULL
                    ,intLanPhysTextSelectID             INTEGER NOT NULL
                    ,strLanPhysTextSelect               VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TLanPhysTextSelects_PK PRIMARY KEY (intLanPhysTextSelectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Lanyard Physical Textile Selection Table
                CREATE TRIGGER IF NOT EXISTS Z_TLanPhysTextSelects_InsertTrigger
                AFTER INSERT ON TLanPhysTextSelects
                BEGIN
                    INSERT INTO Z_TLanPhysTextSelects 
                    (
                        intLanPhysTextSelectID
                        ,strLanPhysTextSelect
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intLanPhysTextSelectID
                        ,NEW.strLanPhysTextSelect
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Lanyard Physical Textile Selection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TLanPhysTextSelects_UpdateTrigger
                AFTER UPDATE ON TLanPhysTextSelects
                BEGIN
                    INSERT INTO Z_TLanPhysTextSelects 
                    (
                        intLanPhysTextSelectID
                        ,strLanPhysTextSelect
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intLanPhysTextSelectID
                        ,NEW.strLanPhysTextSelect
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Lanyard Physical Textile Selection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TLanPhysTextSelects_DeleteTrigger
                AFTER DELETE ON TLanPhysTextSelects
                BEGIN
                    INSERT INTO Z_TLanPhysTextSelects 
                    (
                        intLanPhysTextSelectID
                        ,strLanPhysTextSelect
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intLanPhysTextSelectID
                        ,OLD.strLanPhysTextSelect
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Lanyard Physical Textile Selection tables: {e}")            

    def createLanPhysInspectTables(self):
        """ 
        Function Name: createLanPhysInspectTables
        Function Purpose: Create the Lanyard Physical Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Lanyard Physical Inspection Table
                CREATE TABLE IF NOT EXISTS TLanyardPhysicalInspections 
                (
                    intLanyardPhysicalInspectionID          INTEGER NOT NULL
                    ,intLanyardID                           INTEGER NOT NULL
                    ,intInspectionTypeID                    INTEGER NOT NULL
                    ,intLanPhysTextSelectID                 INTEGER NOT NULL
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT TLanyardPhysicalInspections_PK PRIMARY KEY (intLanyardPhysicalInspectionID)
                    
                    ,FOREIGN KEY ( intLanyardID ) REFERENCES TLanyards ( intLanyardID )
                    ,FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )
                    ,FOREIGN KEY ( intLanPhysTextSelectID ) REFERENCES TLanPhysTextSelects ( intLanPhysTextSelectID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )                     
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Lanyard Physical Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TLanyardPhysicalInspections 
                (
                    intLanyardPhysicalInspectionAuditID     INTEGER NOT NULL
                    ,intLanyardPhysicalInspectionID         INTEGER NOT NULL
                    ,intLanyardID                           INTEGER NOT NULL
                    ,intInspectionTypeID                    INTEGER NOT NULL
                    ,intLanPhysTextSelectID                 INTEGER NOT NULL
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TLanyardPhysicalInspections_PK PRIMARY KEY (intLanyardPhysicalInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Lanyard Physical Inspection Table
                CREATE TRIGGER IF NOT EXISTS Z_TLanyardPhysicalInspections_InsertTrigger
                AFTER INSERT ON TLanyardPhysicalInspections
                BEGIN
                    INSERT INTO Z_TLanyardPhysicalInspections 
                    (
                        intLanyardPhysicalInspectionID
                        ,intLanyardID
                        ,intInspectionTypeID
                        ,intLanPhysTextSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intLanyardPhysicalInspectionID
                        ,NEW.intLanyardID
                        ,NEW.intInspectionTypeID
                        ,NEW.intLanPhysTextSelectID
                        ,NEW.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- INSERT
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Lanyard Physical Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TLanyardPhysicalInspections_UpdateTrigger
                AFTER UPDATE ON TLanyardPhysicalInspections
                BEGIN
                    INSERT INTO Z_TLanyardPhysicalInspections 
                    (
                        intLanyardPhysicalInspectionID
                        ,intLanyardID
                        ,intInspectionTypeID
                        ,intLanPhysTextSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intLanyardPhysicalInspectionID
                        ,NEW.intLanyardID
                        ,NEW.intInspectionTypeID
                        ,NEW.intLanPhysTextSelectID
                        ,NEW.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Lanyard Physical Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TLanyardPhysicalInspections_DeleteTrigger
                AFTER DELETE ON TLanyardPhysicalInspections
                BEGIN
                    INSERT INTO Z_TLanyardPhysicalInspections 
                    (
                        intLanyardPhysicalInspectionID
                        ,intLanyardID
                        ,intInspectionTypeID
                        ,intLanPhysTextSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intLanyardPhysicalInspectionID
                        ,OLD.intLanyardID
                        ,OLD.intInspectionTypeID
                        ,OLD.intLanPhysTextSelectID
                        ,OLD.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Lanyard Physical Inspection tables: {e}")            

    def createRetractFunctionsTables(self):
        """ 
        Function Name: createRetractFunctionsTables
        Function Purpose: Create the Retract Functions and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Retract Function Table
                CREATE TABLE IF NOT EXISTS TRetractFunctions 
                (
                    intRetractFunctionID                INTEGER NOT NULL
                    ,strRetractFunctionDesc             VARCHAR(255) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TRetractFunctions_PK PRIMARY KEY (intRetractFunctionID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Retract Function Table
                CREATE TABLE IF NOT EXISTS Z_TRetractFunctions 
                (
                    intRetractFunctionAuditID           INTEGER NOT NULL
                    ,intRetractFunctionID               INTEGER NOT NULL
                    ,strRetractFunctionDesc             VARCHAR(255) NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TRetractFunctions_PK PRIMARY KEY (intRetractFunctionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Retract Function Table
                CREATE TRIGGER IF NOT EXISTS Z_TRetractFunctions_InsertTrigger
                AFTER INSERT ON TRetractFunctions
                BEGIN
                    INSERT INTO Z_TRetractFunctions 
                    (
                        intRetractFunctionID,
                        strRetractFunctionDesc,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intRetractFunctionID,
                        NEW.strRetractFunctionDesc,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Retract Function Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TRetractFunctions_UpdateTrigger
                AFTER UPDATE ON TRetractFunctions
                BEGIN
                    INSERT INTO Z_TRetractFunctions 
                    (
                        intRetractFunctionID,
                        strRetractFunctionDesc,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intRetractFunctionID,
                        NEW.strRetractFunctionDesc,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Retract Function Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TRetractFunctions_DeleteTrigger
                AFTER DELETE ON TRetractFunctions
                BEGIN
                    INSERT INTO Z_TRetractFunctions 
                    (
                        intRetractFunctionID,
                        strRetractFunctionDesc,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intRetractFunctionID,
                        OLD.strRetractFunctionDesc,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Retract Functional tables: {e}")            

    def createRetractFunctSelTables(self):
        """ 
        Function Name: createLanFunctSelTables
        Function Purpose: Create the Retract Function Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Retract Function Selection Table
                CREATE TABLE IF NOT EXISTS TRetractFunctSelects 
                (
                    intRetractFunctSelectID             INTEGER NOT NULL
                    ,strRetractFunctSelect              VARCHAR(1000) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TRetractFunctSelects_PK PRIMARY KEY (intRetractFunctSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Retract Function Selection Table
                CREATE TABLE IF NOT EXISTS Z_TRetractFunctSelects 
                (
                    intRetractFunctSelectAuditID        INTEGER NOT NULL
                    ,intRetractFunctSelectID            INTEGER NOT NULL
                    ,strRetractFunctSelect              VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TRetractFunctSelects_PK PRIMARY KEY (intRetractFunctSelectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Retract Function Selection Table
                CREATE TRIGGER IF NOT EXISTS Z_TRetractFunctSelects_AuditTrigger
                AFTER INSERT ON TRetractFunctSelects
                BEGIN
                    INSERT INTO Z_TRetractFunctSelects 
                    (
                        intRetractFunctSelectID,
                        strRetractFunctSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intRetractFunctSelectID,
                        NEW.strRetractFunctSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                CREATE TRIGGER IF NOT EXISTS Z_TRetractFunctSelects_AuditTrigger
                AFTER UPDATE ON TRetractFunctSelects
                BEGIN
                    INSERT INTO Z_TRetractFunctSelects 
                    (
                        intRetractFunctSelectID,
                        strRetractFunctSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intRetractFunctSelectID,
                        NEW.strRetractFunctSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                CREATE TRIGGER IF NOT EXISTS Z_TRetractFunctSelects_AuditTrigger
                AFTER DELETE ON TRetractFunctSelects
                BEGIN
                    INSERT INTO Z_TRetractFunctSelects 
                    (
                        intRetractFunctSelectID,
                        strRetractFunctSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intRetractFunctSelectID,
                        OLD.strRetractFunctSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)     
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Retract Functional Selection tables: {e}")            

    def createRetractFunctInspectTables(self):
        """ 
        Function Name: createRetractFunctInspectTables
        Function Purpose: Create the Retract Function Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Retract Function Inspection Table
                CREATE TABLE IF NOT EXISTS TLanyardRetractFunctionInspections 
                (
                    intLanyardRetractFunctionInspectionID       INTEGER NOT NULL
                    ,intLanyardID                               INTEGER NOT NULL
                    ,intInspectionTypeID                        INTEGER NOT NULL
                    ,intRetractFunctSelectID                    INTEGER NOT NULL
                    ,intInspectionStatusID                      INTEGER NOT NULL
                    ,strModifiedReason VARCHAR(1000)
                    ,CONSTRAINT TLanyardRetractFunctionInspections_PK PRIMARY KEY (intLanyardRetractFunctionInspectionID)
                    
                    ,FOREIGN KEY ( intLanyardID ) REFERENCES TLanyards ( intLanyardID )
                    ,FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )
                    ,FOREIGN KEY ( intRetractFunctSelectID ) REFERENCES TRetractFunctSelects ( intRetractFunctSelectID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )                      
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Retract Function Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TLanyardRetractFunctionInspections 
                (
                    intLanyardRetractFunctionInspectAuditID     INTEGER NOT NULL
                    ,intLanyardRetractFunctionInspectionID      INTEGER NOT NULL
                    ,intLanyardID                               INTEGER NOT NULL
                    ,intInspectionTypeID                        INTEGER NOT NULL
                    ,intRetractFunctSelectID                    INTEGER NOT NULL
                    ,intInspectionStatusID                      INTEGER NOT NULL
                    ,strUpdatedBy                               VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                               DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                                  VARCHAR(1) NOT NULL
                    ,strModifiedReason                          VARCHAR(1000)
                    ,CONSTRAINT Z_TLanyardRetractFunctionInspections_PK PRIMARY KEY (intLanyardRetractFunctionInspectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Retract Function Inspection Table (Insert)
                CREATE TRIGGER IF NOT EXISTS Z_TLanyardRetractFunctionInspections_Insert_AuditTrigger
                AFTER INSERT ON TLanyardRetractFunctionInspections
                BEGIN
                    INSERT INTO Z_TLanyardRetractFunctionInspections 
                    (
                        intLanyardRetractFunctionInspectionID,
                        intLanyardID,
                        intInspectionTypeID,
                        intRetractFunctSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intLanyardRetractFunctionInspectionID,
                        NEW.intLanyardID,
                        NEW.intInspectionTypeID,
                        NEW.intRetractFunctSelectID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I',
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Retract Function Inspection Table (Update)
                CREATE TRIGGER IF NOT EXISTS Z_TLanyardRetractFunctionInspections_Update_AuditTrigger
                AFTER UPDATE ON TLanyardRetractFunctionInspections
                BEGIN
                    INSERT INTO Z_TLanyardRetractFunctionInspections 
                    (
                        intLanyardRetractFunctionInspectionID,
                        intLanyardID,
                        intInspectionTypeID,
                        intRetractFunctSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intLanyardRetractFunctionInspectionID,
                        NEW.intLanyardID,
                        NEW.intInspectionTypeID,
                        NEW.intRetractFunctSelectID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U',
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Retract Function Inspection Table (Delete)
                CREATE TRIGGER IF NOT EXISTS Z_TLanyardRetractFunctionInspections_Delete_AuditTrigger
                AFTER DELETE ON TLanyardRetractFunctionInspections
                BEGIN
                    INSERT INTO Z_TLanyardRetractFunctionInspections 
                    (
                        intLanyardRetractFunctionInspectionID,
                        intLanyardID,
                        intInspectionTypeID,
                        intRetractFunctSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intLanyardRetractFunctionInspectionID,
                        OLD.intLanyardID,
                        OLD.intInspectionTypeID,
                        OLD.intRetractFunctSelectID,
                        OLD.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D',
                        OLD.strModifiedReason
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Retract Functional Inspection tables: {e}")            

    def createStandLanyInspectTables(self):
        """ 
        Function Name: createStandLanyInspectTables
        Function Purpose: Create the Standard Lanyard/Retract Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Standard Lanyard/Retract Inspection Table
                CREATE TABLE IF NOT EXISTS TStandardLanyardInspections 
                (
                    intStandardLanyardInspectionID          INTEGER NOT NULL
                    ,intLanyardVisualInspectionID           INTEGER NOT NULL
                    ,intLanyardPhysicalInspectionID         INTEGER NOT NULL
                    ,intLanyardRetractFunctionInspectionID  INTEGER NOT NULL
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT TStandardLanyardInspections_PK PRIMARY KEY (intStandardLanyardInspectionID)
                    
                    ,FOREIGN KEY ( intLanyardVisualInspectionID ) REFERENCES TLanyardVisualInspections ( intLanyardVisualInspectionID )
                    ,FOREIGN KEY ( intLanyardPhysicalInspectionID ) REFERENCES TLanyardPhysicalInspections ( intLanyardPhysicalInspectionID )
                    ,FOREIGN KEY ( intLanyardRetractFunctionInspectionID ) REFERENCES TLanyardRetractFunctionInspections ( intLanyardRetractFunctionInspectionID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )                     
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Standard Lanyard/Retract Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TStandardLanyardInspections 
                (
                    intStandardLanyardInspectionAuditID     INTEGER NOT NULL
                    ,intStandardLanyardInspectionID         INTEGER NOT NULL
                    ,intLanyardVisualInspectionID           INTEGER NOT NULL
                    ,intLanyardPhysicalInspectionID         INTEGER NOT NULL
                    ,intLanyardRetractFunctionInspectionID  INTEGER NOT NULL
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TStandardLanyardInspections_PK PRIMARY KEY (intStandardLanyardInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Standard Lanyard/Retract Inspection Table (Insert)
                CREATE TRIGGER IF NOT EXISTS Z_TStandardLanyardInspections_Insert_AuditTrigger
                AFTER INSERT ON TStandardLanyardInspections
                BEGIN
                    INSERT INTO Z_TStandardLanyardInspections 
                    (
                        intStandardLanyardInspectionID,
                        intLanyardVisualInspectionID,
                        intLanyardPhysicalInspectionID,
                        intLanyardRetractFunctionInspectionID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intStandardLanyardInspectionID,
                        NEW.intLanyardVisualInspectionID,
                        NEW.intLanyardPhysicalInspectionID,
                        NEW.intLanyardRetractFunctionInspectionID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I',
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Standard Lanyard/Retract Inspection Table (Update)
                CREATE TRIGGER IF NOT EXISTS Z_TStandardLanyardInspections_Update_AuditTrigger
                AFTER UPDATE ON TStandardLanyardInspections
                BEGIN
                    INSERT INTO Z_TStandardLanyardInspections 
                    (
                        intStandardLanyardInspectionID,
                        intLanyardVisualInspectionID,
                        intLanyardPhysicalInspectionID,
                        intLanyardRetractFunctionInspectionID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intStandardLanyardInspectionID,
                        NEW.intLanyardVisualInspectionID,
                        NEW.intLanyardPhysicalInspectionID,
                        NEW.intLanyardRetractFunctionInspectionID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U',
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Standard Lanyard/Retract Inspection Table (Delete)
                CREATE TRIGGER IF NOT EXISTS Z_TStandardLanyardInspections_Delete_AuditTrigger
                AFTER DELETE ON TStandardLanyardInspections
                BEGIN
                    INSERT INTO Z_TStandardLanyardInspections 
                    (
                        intStandardLanyardInspectionID,
                        intLanyardVisualInspectionID,
                        intLanyardPhysicalInspectionID,
                        intLanyardRetractFunctionInspectionID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intStandardLanyardInspectionID,
                        OLD.intLanyardVisualInspectionID,
                        OLD.intLanyardPhysicalInspectionID,
                        OLD.intLanyardRetractFunctionInspectionID,
                        OLD.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D',
                        OLD.strModifiedReason
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Standard Lanyard/Retract Inspection tables: {e}")                        

    def insertRetractData(self):
        """ 
        Function Name: insertRetractData
        Function Purpose: Insert the Retract data into the tables
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlInsert = """
                -- Add Data To Table: TRetractFunctions
                INSERT INTO TRetractFunctions (intRetractFunctionID, strRetractFunctionDesc)
                VALUES   (1, 'No Functional Issues')
                        ,(2, 'Fast Extraction')
                        ,(3, 'Slow Extraction')
                        ,(4, 'Fast Retraction')
                        ,(5, 'Slow Retraction')
                        ,(6, 'No Extraction')
                        ,(7, 'No Retraction')
                        ,(8, 'Other Functional Issues');                
                """
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlInsert)       
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error Inserting Retract Data: {e}")            

    def createStateTables(self):
        """ 
        Function Name: createStateTables
        Function Purpose: Create the State Tables and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create State Table
                CREATE TABLE IF NOT EXISTS TStates 
                (
                    intStateID                      INTEGER NOT NULL
                    ,strStateName                   VARCHAR(255) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT TStates_PK PRIMARY KEY (intStateID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: State Table
                CREATE TABLE IF NOT EXISTS Z_TStates 
                (
                    intStateAuditID                 INTEGER NOT NULL
                    ,intStateID                     INTEGER NOT NULL
                    ,strStateName                   VARCHAR(255) NOT NULL
                    ,strUpdatedBy                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                      VARCHAR(1) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT Z_TStates_PK PRIMARY KEY (intStateAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: State Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TStates_InsertTrigger
                AFTER INSERT ON TStates
                BEGIN
                    INSERT INTO Z_TStates 
                    (
                        intStateID,
                        strStateName,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intStateID,
                        NEW.strStateName,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: State Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TStates_UpdateTrigger
                AFTER UPDATE ON TStates
                BEGIN
                    INSERT INTO Z_TStates 
                    (
                        intStateID,
                        strStateName,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intStateID,
                        NEW.strStateName,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: State Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TStates_DeleteTrigger
                AFTER DELETE ON TStates
                BEGIN
                    INSERT INTO Z_TStates 
                    (
                        intStateID,
                        strStateName,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intStateID,
                        OLD.strStateName,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating State tables: {e}")            

    def createInspectorTables(self):
        """ 
        Function Name: createInspectorTables
        Function Purpose: Create the Inspector Tables and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Inspector Table
                CREATE TABLE IF NOT EXISTS TInspectors 
                (
                    intInspectorID                  INTEGER NOT NULL
                    ,strFirstName                   VARCHAR(255) NOT NULL
                    ,strLastName                    VARCHAR(255) NOT NULL
                    ,strEmail                       VARCHAR(255) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT TInspectors_PK PRIMARY KEY (intInspectorID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Inspector Table
                CREATE TABLE IF NOT EXISTS Z_TInspectors 
                (
                    intInspectorAuditID             INTEGER NOT NULL
                    ,intInspectorID                 INTEGER NOT NULL
                    ,strFirstName                   VARCHAR(255) NOT NULL
                    ,strLastName                    VARCHAR(255) NOT NULL
                    ,strEmail                       VARCHAR(255) NOT NULL
                    ,strUpdatedBy                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                      VARCHAR(1) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT Z_TInspectors_PK PRIMARY KEY (intInspectorAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Inspector Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TInspectors_AuditTrigger_Insert
                AFTER INSERT ON TInspectors
                BEGIN
                    INSERT INTO Z_TInspectors 
                    (
                        intInspectorID
                        ,strFirstName
                        ,strLastName
                        ,strEmail
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intInspectorID
                        ,NEW.strFirstName
                        ,NEW.strLastName
                        ,NEW.strEmail
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Inspector Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TInspectors_AuditTrigger_Update
                AFTER UPDATE ON TInspectors
                BEGIN
                    INSERT INTO Z_TInspectors 
                    (
                        intInspectorID
                        ,strFirstName
                        ,strLastName
                        ,strEmail
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intInspectorID
                        ,NEW.strFirstName
                        ,NEW.strLastName
                        ,NEW.strEmail
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Inspector Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TInspectors_AuditTrigger_Delete
                AFTER DELETE ON TInspectors
                BEGIN
                    INSERT INTO Z_TInspectors 
                    (
                        intInspectorID
                        ,strFirstName
                        ,strLastName
                        ,strEmail
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intInspectorID
                        ,OLD.strFirstName
                        ,OLD.strLastName
                        ,OLD.strEmail
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Inspector tables: {e}")                 

    def createLoginTables(self):
        """ 
        Function Name: createLoginTables
        Function Purpose: Create the Login Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Login Table
                CREATE TABLE IF NOT EXISTS TLogins 
                (
                    intLoginID                      INTEGER NOT NULL
                    ,strLoginName                   VARCHAR(255) NOT NULL
                    ,strPassword                    VARCHAR(255) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT TLogins_PK PRIMARY KEY (intLoginID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Login Table
                CREATE TABLE IF NOT EXISTS Z_TLogins 
                (
                    intLoginAuditID                 INTEGER NOT NULL
                    ,intLoginID                     INTEGER NOT NULL
                    ,strLoginName                   VARCHAR(255) NOT NULL
                    ,strPassword                    VARCHAR(255) NOT NULL
                    ,strUpdatedBy                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                      VARCHAR(1) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT Z_TLogins_PK PRIMARY KEY (intLoginAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Login Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TLogins_AuditTrigger_Insert
                AFTER INSERT ON TLogins
                BEGIN
                    INSERT INTO Z_TLogins 
                    (
                        intLoginID
                        ,strLoginName
                        ,strPassword
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intLoginID
                        ,NEW.strLoginName
                        ,NEW.strPassword
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Login Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TLogins_AuditTrigger_Update
                AFTER UPDATE ON TLogins
                BEGIN
                    INSERT INTO Z_TLogins 
                    (
                        intLoginID
                        ,strLoginName
                        ,strPassword
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intLoginID
                        ,NEW.strLoginName
                        ,NEW.strPassword
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Login Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TLogins_AuditTrigger_Delete
                AFTER DELETE ON TLogins
                BEGIN
                    INSERT INTO Z_TLogins 
                    (
                        intLoginID
                        ,strLoginName
                        ,strPassword
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intLoginID
                        ,OLD.strLoginName
                        ,OLD.strPassword
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Login tables: {e}")            

    def createUserLoginTables(self):
        """ 
        Function Name: createUserLoginTables
        Function Purpose: Create the User Login Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create User Login Table
                CREATE TABLE IF NOT EXISTS TUserLogins 
                (
                    intUserLoginID                  INTEGER NOT NULL
                    ,intInspectorID                 INTEGER NOT NULL
                    ,intLoginID                     INTEGER NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT TUserLogins_UQ UNIQUE (intInspectorID, intLoginID)
                    ,CONSTRAINT TUserLogins_PK PRIMARY KEY (intUserLoginID)

                    ,FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID )
                    ,FOREIGN KEY ( intLoginID ) REFERENCES TLogins ( intLoginID )                  
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: User Login Table
                CREATE TABLE IF NOT EXISTS Z_TUserLogins 
                (
                    intUserLoginAuditID             INTEGER NOT NULL
                    ,intUserLoginID                 INTEGER NOT NULL
                    ,intInspectorID                 INTEGER NOT NULL
                    ,intLoginID                     INTEGER NOT NULL
                    ,strUpdatedBy                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                      VARCHAR(1) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT Z_TUserLogins_PK PRIMARY KEY (intUserLoginAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: User Login Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TUserLogins_AuditTrigger_Insert
                AFTER INSERT ON TUserLogins
                BEGIN
                    INSERT INTO Z_TUserLogins 
                    (
                        intUserLoginID
                        ,intInspectorID
                        ,intLoginID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intUserLoginID
                        ,NEW.intInspectorID
                        ,NEW.intLoginID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: User Login Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TUserLogins_AuditTrigger_Update
                AFTER UPDATE ON TUserLogins
                BEGIN
                    INSERT INTO Z_TUserLogins 
                    (
                        intUserLoginID
                        ,intInspectorID
                        ,intLoginID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intUserLoginID
                        ,NEW.intInspectorID
                        ,NEW.intLoginID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: User Login Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TUserLogins_AuditTrigger_Delete
                AFTER DELETE ON TUserLogins
                BEGIN
                    INSERT INTO Z_TUserLogins 
                    (
                        intUserLoginID
                        ,intInspectorID
                        ,intLoginID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intUserLoginID
                        ,OLD.intInspectorID
                        ,OLD.intLoginID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating User Login tables: {e}")            

    def createAdminUserTables(self):
        """ 
        Function Name: createAdminUserTables
        Function Purpose: Create the Admin User Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Admin User Table
                CREATE TABLE IF NOT EXISTS TAdminUsers 
                (
                    intAdminUserID                  INTEGER NOT NULL
                    ,intInspectorID                 INTEGER NOT NULL
                    ,intLoginID                     INTEGER NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT TAdminUsers_UQ UNIQUE (intInspectorID, intLoginID)
                    ,CONSTRAINT TAdminUsers_PK PRIMARY KEY (intAdminUserID)

                    ,FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID )
                    ,FOREIGN KEY ( intLoginID ) REFERENCES TLogins ( intLoginID )                     
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Admin User Table
                CREATE TABLE IF NOT EXISTS Z_TAdminUsers 
                (
                    intAdminUserAuditID             INTEGER NOT NULL
                    ,intAdminUserID                 INTEGER NOT NULL
                    ,intInspectorID                 INTEGER NOT NULL
                    ,intLoginID                     INTEGER NOT NULL
                    ,strUpdatedBy                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                      VARCHAR(1) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT Z_TAdminUsers_PK PRIMARY KEY (intAdminUserAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Admin User Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TAdminUsers_AuditTrigger_Insert
                AFTER INSERT ON TAdminUsers
                BEGIN
                    INSERT INTO Z_TAdminUsers 
                    (
                        intAdminUserID
                        ,intInspectorID
                        ,intLoginID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intAdminUserID
                        ,NEW.intInspectorID
                        ,NEW.intLoginID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Admin User Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TAdminUsers_AuditTrigger_Update
                AFTER UPDATE ON TAdminUsers
                BEGIN
                    INSERT INTO Z_TAdminUsers 
                    (
                        intAdminUserID
                        ,intInspectorID
                        ,intLoginID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intAdminUserID
                        ,NEW.intInspectorID
                        ,NEW.intLoginID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Admin User Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TAdminUsers_AuditTrigger_Delete
                AFTER DELETE ON TAdminUsers
                BEGIN
                    INSERT INTO Z_TAdminUsers 
                    (
                        intAdminUserID
                        ,intInspectorID
                        ,intLoginID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intAdminUserID
                        ,OLD.intInspectorID
                        ,OLD.intLoginID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Admin User tables: {e}")                
            
    def createGymLocationTables(self):
        """ 
        Function Name: createGymLocationTables
        Function Purpose: Create the Gym Location Tables and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Gym Location Table
                CREATE TABLE IF NOT EXISTS TGymLocations 
                (
                    intGymLocationID                INTEGER NOT NULL
                    ,strGymName                     VARCHAR(255) NOT NULL
                    ,strAddress                     VARCHAR(255) NOT NULL
                    ,intStateID                     INTEGER NOT NULL
                    ,strZip                         VARCHAR(255) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT TGymLocations_PK PRIMARY KEY (intGymLocationID)
                    
                    ,FOREIGN KEY ( intStateID ) REFERENCES TStates ( intStateID )
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Gym Location Table
                CREATE TABLE IF NOT EXISTS Z_TGymLocations 
                (
                    intGymLocationAuditID           INTEGER NOT NULL
                    ,intGymLocationID               INTEGER NOT NULL
                    ,strGymName                     VARCHAR(255) NOT NULL
                    ,strAddress                     VARCHAR(255) NOT NULL
                    ,intStateID                     INTEGER NOT NULL
                    ,strZip                         VARCHAR(255) NOT NULL
                    ,strUpdatedBy                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                      VARCHAR(1) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT Z_TGymLocations_PK PRIMARY KEY (intGymLocationAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Gym Location Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TGymLocations_AuditTrigger_Insert
                AFTER INSERT ON TGymLocations
                BEGIN
                    INSERT INTO Z_TGymLocations 
                    (
                        intGymLocationID
                        ,strGymName
                        ,strAddress
                        ,intStateID
                        ,strZip
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intGymLocationID
                        ,NEW.strGymName
                        ,NEW.strAddress
                        ,NEW.intStateID
                        ,NEW.strZip
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Gym Location Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TGymLocations_AuditTrigger_Update
                AFTER UPDATE ON TGymLocations
                BEGIN
                    INSERT INTO Z_TGymLocations 
                    (
                        intGymLocationID
                        ,strGymName
                        ,strAddress
                        ,intStateID
                        ,strZip
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intGymLocationID
                        ,NEW.strGymName
                        ,NEW.strAddress
                        ,NEW.intStateID
                        ,NEW.strZip
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Gym Location Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TGymLocations_AuditTrigger_Delete
                AFTER DELETE ON TGymLocations
                BEGIN
                    INSERT INTO Z_TGymLocations 
                    (
                        intGymLocationID
                        ,strGymName
                        ,strAddress
                        ,intStateID
                        ,strZip
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intGymLocationID
                        ,OLD.strGymName
                        ,OLD.strAddress
                        ,OLD.intStateID
                        ,OLD.strZip
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Gym Location tables: {e}")
            
    def createAutoBelayTables(self):
        """ 
        Function Name: createAutoBelayTables
        Function Purpose: Create the Auto Belay Tables and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Auto Belay Table
                CREATE TABLE IF NOT EXISTS TAutoBelays 
                (
                    intAutoBelayID                  INTEGER NOT NULL
                    ,strDeviceName                  VARCHAR(255) NOT NULL
                    ,strSerialNum                   VARCHAR(255) NOT NULL
                    ,strBumperNum                   VARCHAR(255) 
                    ,dtmManufactureDate             DATETIME 
                    ,dtmServiceDate                 DATETIME NOT NULL   
                    ,dtmReserviceDate               DATETIME NOT NULL   
                    ,dtmInstallationDate            DATETIME NOT NULL   
                    ,dtmLastInspectionDate          DATETIME NOT NULL   
                    ,dtmNextInspectionDate          DATETIME NOT NULL       
                    ,blnDeviceInUse                 VARCHAR(255) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT TAutoBelays_PK PRIMARY KEY (intAutoBelayID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Auto Belay Table
                CREATE TABLE IF NOT EXISTS Z_TAutoBelays 
                (
                    intAutoBelayAuditID             INTEGER NOT NULL
                    ,intAutoBelayID                 INTEGER NOT NULL
                    ,strDeviceName                  VARCHAR(255) NOT NULL
                    ,strSerialNum                   VARCHAR(255) NOT NULL
                    ,strBumperNum                   VARCHAR(255) NOT NULL
                    ,dtmManufactureDate             DATETIME 
                    ,dtmServiceDate                 DATETIME NOT NULL   
                    ,dtmReserviceDate               DATETIME NOT NULL   
                    ,dtmInstallationDate            DATETIME NOT NULL   
                    ,dtmLastInspectionDate          DATETIME NOT NULL   
                    ,dtmNextInspectionDate          DATETIME NOT NULL       
                    ,blnDeviceInUse                 VARCHAR(255) NOT NULL
                    ,strUpdatedBy                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                      VARCHAR(1) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT Z_TAutoBelays_PK PRIMARY KEY (intAutoBelayAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Auto Belay Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TAutoBelays_AuditTrigger_Insert
                AFTER INSERT ON TAutoBelays
                BEGIN
                    INSERT INTO Z_TAutoBelays 
                    (
                        intAutoBelayID
                        ,strDeviceName
                        ,strSerialNum
                        ,strBumperNum
                        ,dtmManufactureDate
                        ,dtmServiceDate
                        ,dtmReserviceDate
                        ,dtmInstallationDate
                        ,dtmLastInspectionDate
                        ,dtmNextInspectionDate
                        ,blnDeviceInUse
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intAutoBelayID
                        ,NEW.strDeviceName
                        ,NEW.strSerialNum
                        ,NEW.strBumperNum
                        ,NEW.dtmManufactureDate
                        ,NEW.dtmServiceDate
                        ,NEW.dtmReserviceDate
                        ,NEW.dtmInstallationDate
                        ,NEW.dtmLastInspectionDate
                        ,NEW.dtmNextInspectionDate
                        ,NEW.blnDeviceInUse
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Auto Belay Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TAutoBelays_AuditTrigger_Update
                AFTER UPDATE ON TAutoBelays
                BEGIN
                    INSERT INTO Z_TAutoBelays 
                    (
                        intAutoBelayID
                        ,strDeviceName
                        ,strSerialNum
                        ,strBumperNum
                        ,dtmManufactureDate
                        ,dtmServiceDate
                        ,dtmReserviceDate
                        ,dtmInstallationDate
                        ,dtmLastInspectionDate
                        ,dtmNextInspectionDate
                        ,blnDeviceInUse
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intAutoBelayID
                        ,NEW.strDeviceName
                        ,NEW.strSerialNum
                        ,NEW.strBumperNum
                        ,NEW.dtmManufactureDate
                        ,NEW.dtmServiceDate
                        ,NEW.dtmReserviceDate
                        ,NEW.dtmInstallationDate
                        ,NEW.dtmLastInspectionDate
                        ,NEW.dtmNextInspectionDate
                        ,NEW.blnDeviceInUse
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Auto Belay Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TAutoBelays_AuditTrigger_Delete
                AFTER DELETE ON TAutoBelays
                BEGIN
                    INSERT INTO Z_TAutoBelays 
                    (
                        intAutoBelayID
                        ,strDeviceName
                        ,strSerialNum
                        ,strBumperNum
                        ,dtmManufactureDate
                        ,dtmServiceDate
                        ,dtmReserviceDate
                        ,dtmInstallationDate
                        ,dtmLastInspectionDate
                        ,dtmNextInspectionDate
                        ,blnDeviceInUse
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intAutoBelayID
                        ,OLD.strDeviceName
                        ,OLD.strSerialNum
                        ,OLD.strBumperNum
                        ,OLD.dtmManufactureDate
                        ,OLD.dtmServiceDate
                        ,OLD.dtmReserviceDate
                        ,OLD.dtmInstallationDate
                        ,OLD.dtmLastInspectionDate
                        ,OLD.dtmNextInspectionDate
                        ,OLD.blnDeviceInUse
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Auto Belay tables: {e}")                        

    def createWallLocationTables(self):
        """ 
        Function Name: createWallLocationTables
        Function Purpose: Create the Wall Location Tables and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Wall Location Table
                CREATE TABLE IF NOT EXISTS TWallLocations 
                (
                    intWallLocationID               INTEGER NOT NULL
                    ,strLocationName                VARCHAR(255) NOT NULL
                    ,strWallLocationDesc            VARCHAR(255) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT TWallLocations_PK PRIMARY KEY (intWallLocationID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Wall Location Table
                CREATE TABLE IF NOT EXISTS Z_TWallLocations 
                (
                    intWallLocationAuditID          INTEGER NOT NULL
                    ,intWallLocationID              INTEGER NOT NULL
                    ,strLocationName                VARCHAR(255) NOT NULL
                    ,strWallLocationDesc            VARCHAR(255) NOT NULL
                    ,strUpdatedBy                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                      VARCHAR(1) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT Z_TWallLocations_PK PRIMARY KEY (intWallLocationAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Wall Location Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TWallLocations_AuditTrigger_Insert
                AFTER INSERT ON TWallLocations
                BEGIN
                    INSERT INTO Z_TWallLocations 
                    (
                        intWallLocationID
                        ,strLocationName
                        ,strWallLocationDesc
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intWallLocationID
                        ,NEW.strLocationName
                        ,NEW.strWallLocationDesc
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Wall Location Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TWallLocations_AuditTrigger_Update
                AFTER UPDATE ON TWallLocations
                BEGIN
                    INSERT INTO Z_TWallLocations 
                    (
                        intWallLocationID
                        ,strLocationName
                        ,strWallLocationDesc
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intWallLocationID
                        ,NEW.strLocationName
                        ,NEW.strWallLocationDesc
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Wall Location Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TWallLocations_AuditTrigger_Delete
                AFTER DELETE ON TWallLocations
                BEGIN
                    INSERT INTO Z_TWallLocations 
                    (
                        intWallLocationID
                        ,strLocationName
                        ,strWallLocationDesc
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intWallLocationID
                        ,OLD.strLocationName
                        ,OLD.strWallLocationDesc
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Wall Location tables: {e}")            

    def createLocationTables(self):
        """ 
        Function Name: createLocationTables
        Function Purpose: Create the Location Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Location Table
                CREATE TABLE IF NOT EXISTS TLocations 
                (
                    intLocationID                   INTEGER NOT NULL
                    ,strLocationName                VARCHAR(255) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT TLocations_PK PRIMARY KEY (intLocationID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Location Table
                CREATE TABLE IF NOT EXISTS Z_TLocations 
                (
                    intLocationAuditID              INTEGER NOT NULL
                    ,intLocationID                  INTEGER NOT NULL
                    ,strLocationName                VARCHAR(255) NOT NULL
                    ,strUpdatedBy                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                      VARCHAR(1) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT Z_TLocations_PK PRIMARY KEY (intLocationAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Location Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TLocations_AuditTrigger_Insert
                AFTER INSERT ON TLocations
                BEGIN
                    INSERT INTO Z_TLocations 
                    (
                        intLocationID
                        ,strLocationName
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intLocationID
                        ,NEW.strLocationName
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Location Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TLocations_AuditTrigger_Update
                AFTER UPDATE ON TLocations
                BEGIN
                    INSERT INTO Z_TLocations 
                    (
                        intLocationID
                        ,strLocationName
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intLocationID
                        ,NEW.strLocationName
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Location Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TLocations_AuditTrigger_Delete
                AFTER DELETE ON TLocations
                BEGIN
                    INSERT INTO Z_TLocations 
                    (
                        intLocationID
                        ,strLocationName
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intLocationID
                        ,OLD.strLocationName
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Location tables: {e}")
            
    def createAutoBelay_InspectorTables(self):
        """ 
        Function Name: createAutoBelay_InspectorTables
        Function Purpose: Create the Auto Belay Inspector Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Auto Belay Inspector Table
                CREATE TABLE IF NOT EXISTS TAutoBelayInspectors 
                (
                    intAutoBelayInspectorID         INTEGER NOT NULL
                    ,intInspectorID                 INTEGER NOT NULL
                    ,intAutoBelayID                 INTEGER NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT TAutoBelayInspectors_PK PRIMARY KEY (intAutoBelayInspectorID)
                    
                    ,FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID )
                    ,FOREIGN KEY ( intAutoBelayID ) REFERENCES TAutoBelays ( intAutoBelayID )
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Auto Belay Inspector Table
                CREATE TABLE IF NOT EXISTS Z_TAutoBelayInspectors 
                (
                    intAutoBelayInspectorAuditID    INTEGER NOT NULL
                    ,intAutoBelayInspectorID        INTEGER NOT NULL
                    ,intInspectorID                 INTEGER NOT NULL
                    ,intAutoBelayID                 INTEGER NOT NULL
                    ,strUpdatedBy                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                      VARCHAR(1) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT Z_TAutoBelayInspectors_PK PRIMARY KEY (intAutoBelayInspectorAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Auto Belay Inspector Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TAutoBelayInspectors_AuditTrigger_Insert
                AFTER INSERT ON TAutoBelayInspectors
                BEGIN
                    INSERT INTO Z_TAutoBelayInspectors 
                    (
                        intAutoBelayInspectorID
                        ,intInspectorID
                        ,intAutoBelayID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intAutoBelayInspectorID
                        ,NEW.intInspectorID
                        ,NEW.intAutoBelayID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Auto Belay Inspector Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TAutoBelayInspectors_AuditTrigger_Update
                AFTER UPDATE ON TAutoBelayInspectors
                BEGIN
                    INSERT INTO Z_TAutoBelayInspectors 
                    (
                        intAutoBelayInspectorID
                        ,intInspectorID
                        ,intAutoBelayID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intAutoBelayInspectorID
                        ,NEW.intInspectorID
                        ,NEW.intAutoBelayID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Auto Belay Inspector Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TAutoBelayInspectors_AuditTrigger_Delete
                AFTER DELETE ON TAutoBelayInspectors
                BEGIN
                    INSERT INTO Z_TAutoBelayInspectors 
                    (
                        intAutoBelayInspectorID
                        ,intInspectorID
                        ,intAutoBelayID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intAutoBelayInspectorID
                        ,OLD.intInspectorID
                        ,OLD.intAutoBelayID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Auto Belay Inspector tables: {e}")            

    def createAutoBelay_WallLocationTables(self):
        """ 
        Function Name: createAutoBelay_WallLocationTables
        Function Purpose: Create the Auto Wall Location Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Auto Belay Wall Location Table
                CREATE TABLE IF NOT EXISTS TAutoBelayWallLocations 
                (
                    intAutoBelayWallLocationID          INTEGER NOT NULL
                    ,intWallLocationID                  INTEGER NOT NULL
                    ,intAutoBelayID                     INTEGER NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,FOREIGN KEY ( intWallLocationID ) REFERENCES TWallLocations ( intWallLocationID )
                    ,FOREIGN KEY ( intAutoBelayID ) REFERENCES TAutoBelays ( intAutoBelayID )
                    
                    ,CONSTRAINT TAutoBelayWallLocations_PK PRIMARY KEY (intAutoBelayWallLocationID)
                    ,CONSTRAINT TAutoBelayWallLocations_UQ UNIQUE (intWallLocationID, intAutoBelayID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Auto Belay Wall Location Table
                CREATE TABLE IF NOT EXISTS Z_TAutoBelayWallLocations 
                (
                    intAutoBelayWallLocationAuditID     INTEGER NOT NULL
                    ,intAutoBelayWallLocationID         INTEGER NOT NULL
                    ,intWallLocationID                  INTEGER NOT NULL
                    ,intAutoBelayID                     INTEGER NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TAutoBelayWallLocations_PK PRIMARY KEY (intAutoBelayWallLocationAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Auto Belay Wall Location Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TAutoBelayWallLocations_AuditTrigger_Insert
                AFTER INSERT ON TAutoBelayWallLocations
                BEGIN
                    INSERT INTO Z_TAutoBelayWallLocations 
                    (
                        intAutoBelayWallLocationID
                        ,intWallLocationID
                        ,intAutoBelayID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intAutoBelayWallLocationID
                        ,NEW.intWallLocationID
                        ,NEW.intAutoBelayID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Auto Belay Wall Location Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TAutoBelayWallLocations_AuditTrigger_Update
                AFTER UPDATE ON TAutoBelayWallLocations
                BEGIN
                    INSERT INTO Z_TAutoBelayWallLocations 
                    (
                        intAutoBelayWallLocationID
                        ,intWallLocationID
                        ,intAutoBelayID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intAutoBelayWallLocationID
                        ,NEW.intWallLocationID
                        ,NEW.intAutoBelayID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Auto Belay Wall Location Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TAutoBelayWallLocations_AuditTrigger_Delete
                AFTER DELETE ON TAutoBelayWallLocations
                BEGIN
                    INSERT INTO Z_TAutoBelayWallLocations 
                    (
                        intAutoBelayWallLocationID
                        ,intWallLocationID
                        ,intAutoBelayID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intAutoBelayWallLocationID
                        ,OLD.intWallLocationID
                        ,OLD.intAutoBelayID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Auto Belay Wall Location tables: {e}")            

    def createStandardInspectionTables(self):
        """ 
        Function Name: createStandardInspectionTables
        Function Purpose: Create the Standard Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Standard Inspection Table
                CREATE TABLE IF NOT EXISTS TStandardInspections 
                (
                    intStandardInspectionID                 INTEGER NOT NULL
                    ,intStandardCarabinerInspectionID       INTEGER NOT NULL
                    ,intStandardHandleInspectionID          INTEGER NOT NULL
                    ,intStandardCaseHousingInspectionID     INTEGER NOT NULL
                    ,intStandardBrakeHousingInspectionID    INTEGER NOT NULL
                    ,intStandardLanyardInspectionID         INTEGER NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT TStandardInspections_PK PRIMARY KEY (intStandardInspectionID)
                    
                    ,FOREIGN KEY ( intStandardCarabinerInspectionID ) REFERENCES TStandardCarabinerInspections ( intStandardCarabinerInspectionID )
                    ,FOREIGN KEY ( intStandardHandleInspectionID ) REFERENCES TStandardHandleInspections ( intStandardHandleInspectionID )
                    ,FOREIGN KEY ( intStandardCaseHousingInspectionID ) REFERENCES TStandardCaseHousingInspections ( intStandardCaseHousingInspectionID )
                    ,FOREIGN KEY ( intStandardBrakeHousingInspectionID ) REFERENCES TStandardBrakeHousingInspections ( intStandardBrakeHousingInspectionID )
                    ,FOREIGN KEY ( intStandardLanyardInspectionID ) REFERENCES TStandardLanyardInspections ( intStandardLanyardInspectionID )
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Standard Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TStandardInspections 
                (
                    intStandardInspectionAuditID            INTEGER NOT NULL
                    ,intStandardInspectionID                INTEGER NOT NULL
                    ,intStandardCarabinerInspectionID       INTEGER NOT NULL
                    ,intStandardHandleInspectionID          INTEGER NOT NULL
                    ,intStandardCaseHousingInspectionID     INTEGER NOT NULL
                    ,intStandardBrakeHousingInspectionID    INTEGER NOT NULL
                    ,intStandardLanyardInspectionID         INTEGER NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TStandardInspections_PK PRIMARY KEY (intStandardInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Standard Inspection Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TStandardInspections_AuditTrigger_Insert
                AFTER INSERT ON TStandardInspections
                BEGIN
                    INSERT INTO Z_TStandardInspections 
                    (
                        intStandardInspectionID
                        ,intStandardCarabinerInspectionID
                        ,intStandardHandleInspectionID
                        ,intStandardCaseHousingInspectionID
                        ,intStandardBrakeHousingInspectionID
                        ,intStandardLanyardInspectionID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intStandardInspectionID
                        ,NEW.intStandardCarabinerInspectionID
                        ,NEW.intStandardHandleInspectionID
                        ,NEW.intStandardCaseHousingInspectionID
                        ,NEW.intStandardBrakeHousingInspectionID
                        ,NEW.intStandardLanyardInspectionID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Standard Inspection Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TStandardInspections_AuditTrigger_Update
                AFTER UPDATE ON TStandardInspections
                BEGIN
                    INSERT INTO Z_TStandardInspections 
                    (
                        intStandardInspectionID
                        ,intStandardCarabinerInspectionID
                        ,intStandardHandleInspectionID
                        ,intStandardCaseHousingInspectionID
                        ,intStandardBrakeHousingInspectionID
                        ,intStandardLanyardInspectionID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intStandardInspectionID
                        ,NEW.intStandardCarabinerInspectionID
                        ,NEW.intStandardHandleInspectionID
                        ,NEW.intStandardCaseHousingInspectionID
                        ,NEW.intStandardBrakeHousingInspectionID
                        ,NEW.intStandardLanyardInspectionID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Standard Inspection Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TStandardInspections_AuditTrigger_Delete
                AFTER DELETE ON TStandardInspections
                BEGIN
                    INSERT INTO Z_TStandardInspections 
                    (
                        intStandardInspectionID
                        ,intStandardCarabinerInspectionID
                        ,intStandardHandleInspectionID
                        ,intStandardCaseHousingInspectionID
                        ,intStandardBrakeHousingInspectionID
                        ,intStandardLanyardInspectionID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intStandardInspectionID
                        ,OLD.intStandardCarabinerInspectionID
                        ,OLD.intStandardHandleInspectionID
                        ,OLD.intStandardCaseHousingInspectionID
                        ,OLD.intStandardBrakeHousingInspectionID
                        ,OLD.intStandardLanyardInspectionID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Standard Inspection tables: {e}")              

    def createAutoBelay_InspectionTables(self):
        """ 
        Function Name: createAutoBelay_InspectionTables
        Function Purpose: Create the Auto Belay Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Auto Belay Inspection Table
                CREATE TABLE IF NOT EXISTS TAutoBelayInspections 
                (
                    intAutoBelayInspectionID            INTEGER NOT NULL
                    ,intAutoBelayID                     INTEGER NOT NULL
                    ,intWallLocationID                  INTEGER NOT NULL
                    ,intStandardInspectionID            INTEGER NOT NULL
                    ,intInspectorID                     INTEGER NOT NULL
                    ,intInspectionStatusID              INTEGER NOT NULL
                    ,dtmLastInspectionDate              DATETIME NOT NULL
                    ,dtmNextInspectionDate              DATETIME NOT NULL
                    ,strComment                         VARCHAR(225) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TAutoBelayInspections_PK PRIMARY KEY (intAutoBelayInspectionID)
                    
                    ,FOREIGN KEY ( intAutoBelayID ) REFERENCES TAutoBelays ( intAutoBelayID )
                    ,FOREIGN KEY ( intWallLocationID ) REFERENCES TWallLocations ( intWallLocationID ) 
                    ,FOREIGN KEY ( intStandardInspectionID ) REFERENCES TStandardInspections ( intStandardInspectionID ) 
                    ,FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID ) 
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID ) 
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Auto Belay Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TAutoBelayInspections 
                (
                    intAutoBelayInspectionAuditID       INTEGER NOT NULL
                    ,intAutoBelayInspectionID           INTEGER NOT NULL
                    ,intAutoBelayID                     INTEGER NOT NULL
                    ,intWallLocationID                  INTEGER NOT NULL
                    ,intStandardInspectionID            INTEGER NOT NULL
                    ,intInspectorID                     INTEGER NOT NULL
                    ,intInspectionStatusID              INTEGER NOT NULL
                    ,dtmLastInspectionDate              DATETIME NOT NULL
                    ,dtmNextInspectionDate              DATETIME NOT NULL
                    ,strComment                         VARCHAR(225) NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TAutoBelayInspections_PK PRIMARY KEY (intAutoBelayInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Auto Belay Inspection Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TAutoBelayInspections_AuditTrigger_Insert
                AFTER INSERT ON TAutoBelayInspections
                BEGIN
                    INSERT INTO Z_TAutoBelayInspections 
                    (
                        intAutoBelayInspectionID
                        ,intAutoBelayID
                        ,intWallLocationID
                        ,intStandardInspectionID
                        ,intInspectorID
                        ,intInspectionStatusID
                        ,dtmLastInspectionDate
                        ,dtmNextInspectionDate
                        ,strComment
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intAutoBelayInspectionID
                        ,NEW.intAutoBelayID
                        ,NEW.intWallLocationID
                        ,NEW.intStandardInspectionID
                        ,NEW.intInspectorID
                        ,NEW.intInspectionStatusID
                        ,NEW.dtmLastInspectionDate
                        ,NEW.dtmNextInspectionDate
                        ,NEW.strComment
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Auto Belay Inspection Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TAutoBelayInspections_AuditTrigger_Update
                AFTER UPDATE ON TAutoBelayInspections
                BEGIN
                    INSERT INTO Z_TAutoBelayInspections 
                    (
                        intAutoBelayInspectionID
                        ,intAutoBelayID
                        ,intWallLocationID
                        ,intStandardInspectionID
                        ,intInspectorID
                        ,intInspectionStatusID
                        ,dtmLastInspectionDate
                        ,dtmNextInspectionDate
                        ,strComment
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intAutoBelayInspectionID
                        ,NEW.intAutoBelayID
                        ,NEW.intWallLocationID
                        ,NEW.intStandardInspectionID
                        ,NEW.intInspectorID
                        ,NEW.intInspectionStatusID
                        ,NEW.dtmLastInspectionDate
                        ,NEW.dtmNextInspectionDate
                        ,NEW.strComment
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Auto Belay Inspection Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TAutoBelayInspections_AuditTrigger_Delete
                AFTER DELETE ON TAutoBelayInspections
                BEGIN
                    INSERT INTO Z_TAutoBelayInspections 
                    (
                        intAutoBelayInspectionID
                        ,intAutoBelayID
                        ,intWallLocationID
                        ,intStandardInspectionID
                        ,intInspectorID
                        ,intInspectionStatusID
                        ,dtmLastInspectionDate
                        ,dtmNextInspectionDate
                        ,strComment
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intAutoBelayInspectionID
                        ,OLD.intAutoBelayID
                        ,OLD.intWallLocationID
                        ,OLD.intStandardInspectionID
                        ,OLD.intInspectorID
                        ,OLD.intInspectionStatusID
                        ,OLD.dtmLastInspectionDate
                        ,OLD.dtmNextInspectionDate
                        ,OLD.strComment
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Auto Belay Inspection tables: {e}")                 

    def createAutoBelay_ReserviceReports_Tables(self):
        """ 
        Function Name: createAutoBelay_ReserviceReports_Tables
        Function Purpose: Create the Auto Belay Reservice Reports Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Auto Belay Inspection Table
                CREATE TABLE IF NOT EXISTS TAutoBelayReserviceReports 
                (
                    intAutoBelayReserviceReportID           INTEGER NOT NULL
                    ,intAutoBelayInspectionID               INTEGER NOT NULL
                    ,intInspectorID                         INTEGER NOT NULL
                    ,dtmReportDate                          DATETIME NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT TAutoBelayReserviceReports_PK PRIMARY KEY (intAutoBelayReserviceReportID)
                    
                    ,FOREIGN KEY ( intAutoBelayInspectionID ) REFERENCES TAutoBelayInspections ( intAutoBelayInspectionID )
                    ,FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID ) 
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Auto Belay Reservice Report Table
                CREATE TABLE IF NOT EXISTS Z_TAutoBelayReserviceReports 
                (
                    intAutoBelayReserviceReportAuditID      INTEGER NOT NULL
                    ,intAutoBelayReserviceReportID          INTEGER NOT NULL
                    ,intAutoBelayInspectionID               INTEGER NOT NULL
                    ,intInspectorID                         INTEGER NOT NULL
                    ,dtmReportDate                          DATETIME NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TAutoBelayReserviceReports_PK PRIMARY KEY (intAutoBelayReserviceReportAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Auto Belay Reservice Report Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TAutoBelayReserviceReports_AuditTrigger_Insert
                AFTER INSERT ON TAutoBelayReserviceReports
                BEGIN
                    INSERT INTO Z_TAutoBelayReserviceReports 
                    (
                        intAutoBelayReserviceReportID
                        ,intAutoBelayInspectionID
                        ,intInspectorID
                        ,dtmReportDate
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intAutoBelayReserviceReportID
                        ,NEW.intAutoBelayInspectionID
                        ,NEW.intInspectorID
                        ,NEW.dtmReportDate
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Auto Belay Reservice Report Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TAutoBelayReserviceReports_AuditTrigger_Update
                AFTER UPDATE ON TAutoBelayReserviceReports
                BEGIN
                    INSERT INTO Z_TAutoBelayReserviceReports 
                    (
                        intAutoBelayReserviceReportID
                        ,intAutoBelayInspectionID
                        ,intInspectorID
                        ,dtmReportDate
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intAutoBelayReserviceReportID
                        ,NEW.intAutoBelayInspectionID
                        ,NEW.intInspectorID
                        ,NEW.dtmReportDate
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Auto Belay Inspection Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TAutoBelayReserviceReports_AuditTrigger_Delete
                AFTER DELETE ON TAutoBelayReserviceReports
                BEGIN
                    INSERT INTO Z_TAutoBelayReserviceReports 
                    (
                        intAutoBelayReserviceReportID
                        ,intAutoBelayInspectionID
                        ,intInspectorID
                        ,dtmReportDate
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intAutoBelayReserviceReportID
                        ,OLD.intAutoBelayInspectionID
                        ,OLD.intInspectorID
                        ,OLD.dtmReportDate
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Auto Belay Inspection tables: {e}")
            
    def createUsrLog_Location_AutoBelay_ForeignKeys(self):
        """ 
        Function Name: createUsrLog_Location_AutoBelay_ForeignKeys
        Function Purpose: Create the User Logins, Locations, AutoBelays and Standard Inspection foreign keys inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the foreign keys
            sqlForeignKeys = """
                -- -------------------------------------------------------------------------------------------------------
                -- Establish Referential Integrity - Inserts
                -- -------------------------------------------------------------------------------------------------------

                -- #           Child                           Parent                              Column
                -- --         -------                         --------                            --------
                -- 1. TUserLogins                            TInspectors                         intInspectorID
                -- 2. TUserLogins                            TLogins                             intLoginID
                -- 3. TAdminUsers                            TInspectors                         intInspectorID
                -- 4. TAdminUsers                            TLogins                             intLoginID
                -- 5. TGymLocations                          TStates                             intStateID
                -- 6. TAutoBelayInspectors                   TInspectors                         intInspectorID
                -- 7. TAutoBelayInspectors                   TAutoBelays                         intAutoBelayID
                -- 8. TAutoBelayWallLocations                TWallLocations                      intWallLocationID
                -- 9. TAutoBelayWallLocations                TAutoBelays                         intAutoBelayID
                -- 10. TStandardInspections                  TStandardCarabinerInspections       intStandardCarabinerInspectionID
                -- 11. TStandardInspections                  TStandardHandleInspections          intStandardHandleInspectionID
                -- 12. TStandardInspections                  TStandardCaseHousingInspections     intStandardCaseHousingInspectionID
                -- 13. TStandardInspections                  TStandardBrakeHousingInspections    intStandardBrakeHousingInspectionID
                -- 14. TStandardInspections                  TStandardLanyardInspections         intStandardLanyardInspectionID
                -- 15. TAutoBelayInspections                 TAutoBelays                         intAutoBelayID
                -- 16. TAutoBelayInspections                 TWallLocations                      intWallLocationID
                -- 17. TAutoBelayInspections                 TStandardInspections                intStandardInspectionID
                -- 18. TAutoBelayInspections                 TInspectors                         intInspectorID 
                -- 19. TAutoBelayInspections                 TInspectionStatus                   intInspectionStatusID
                -- 20. TAutoBelayReserviceReports            TAutoBelayInspections               intAutoBelayInspectionID
                -- 21. TAutoBelayReserviceReports            TInspectors                         intInspectorID

                -- -------------------------------------------------------------------------------------------------------
                -- Create Foreign Keys
                -- -------------------------------------------------------------------------------------------------------
                -- 1.
                ALTER TABLE TUserLogins ADD CONSTRAINT TUserLogins_TInspectors_FK
                FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID )

                -- 2.
                ALTER TABLE TUserLogins ADD CONSTRAINT TUserLogins_TLogins_FK
                FOREIGN KEY ( intLoginID ) REFERENCES TLogins ( intLoginID )

                -- 3.
                ALTER TABLE TAdminUsers ADD CONSTRAINT TAdminUsers_TInspectors_FK
                FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID )

                -- 4.
                ALTER TABLE TAdminUsers ADD CONSTRAINT TAdminUsers_TLogins_FK
                FOREIGN KEY ( intLoginID ) REFERENCES TLogins ( intLoginID )

                -- -------------------------------------------------------------------------------------------------------
                -- Ensure No Duplicated Data
                -- -------------------------------------------------------------------------------------------------------
                ALTER TABLE TUserLogins ADD CONSTRAINT TUserLogins_UNIQUE
                UNIQUE (intInspectorID, intLoginID)

                ALTER TABLE TAdminUsers ADD CONSTRAINT TAdminUsers_UNIQUE
                UNIQUE (intInspectorID, intLoginID)

                -- -------------------------------------------------------------------------------------------------------
                -- Create Foreign Keys
                -- -------------------------------------------------------------------------------------------------------
                -- 5.
                ALTER TABLE TGymLocations ADD CONSTRAINT TGymLocations_TStates_FK
                FOREIGN KEY ( intStateID ) REFERENCES TStates ( intStateID )

                -- 6.
                ALTER TABLE TAutoBelayInspectors ADD CONSTRAINT TAutoBelayInspectors_TInspectors_FK
                FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID )

                -- 7.
                ALTER TABLE TAutoBelayInspectors ADD CONSTRAINT TAutoBelayInspectors_TAutoBelays_FK
                FOREIGN KEY ( intAutoBelayID ) REFERENCES TAutoBelays ( intAutoBelayID )

                -- 8. 
                ALTER TABLE TAutoBelayWallLocations ADD CONSTRAINT TAutoBelayWallLocations_TWallLocations_FK
                FOREIGN KEY ( intWallLocationID ) REFERENCES TWallLocations ( intWallLocationID )

                -- 9. 
                ALTER TABLE TAutoBelayWallLocations ADD CONSTRAINT TAutoBelayWallLocations_TAutoBelays_FK
                FOREIGN KEY ( intAutoBelayID ) REFERENCES TAutoBelays ( intAutoBelayID )

                -- 10.
                ALTER TABLE TStandardInspections ADD CONSTRAINT TStandardInspections_TStandardCarabinerInspections_FK
                FOREIGN KEY ( intStandardCarabinerInspectionID ) REFERENCES TStandardCarabinerInspections ( intStandardCarabinerInspectionID )

                -- 11.
                ALTER TABLE TStandardInspections ADD CONSTRAINT TStandardInspections_TStandardHandleInspections_FK
                FOREIGN KEY ( intStandardHandleInspectionID ) REFERENCES TStandardHandleInspections ( intStandardHandleInspectionID )

                -- 12.
                ALTER TABLE TStandardInspections ADD CONSTRAINT TStandardInspections_TStandardCaseHousingInspections_FK
                FOREIGN KEY ( intStandardCaseHousingInspectionID ) REFERENCES TStandardCaseHousingInspections ( intStandardCaseHousingInspectionID )

                -- 13.
                ALTER TABLE TStandardInspections ADD CONSTRAINT TStandardInspections_TStandardBrakeHousingInspections_FK
                FOREIGN KEY ( intStandardBrakeHousingInspectionID ) REFERENCES TStandardBrakeHousingInspections ( intStandardBrakeHousingInspectionID )

                -- 14.
                ALTER TABLE TStandardInspections ADD CONSTRAINT TStandardInspections_TStandardLanyardInspections_FK
                FOREIGN KEY ( intStandardLanyardInspectionID ) REFERENCES TStandardLanyardInspections ( intStandardLanyardInspectionID )
                
                -- 15.
                ALTER TABLE TAutoBelayInspections ADD CONSTRAINT TAutoBelayInspections_TAutoBelays_FK
                FOREIGN KEY ( intAutoBelayID ) REFERENCES TAutoBelays ( intAutoBelayID )

                -- 16.
                ALTER TABLE TAutoBelayInspections ADD CONSTRAINT TAutoBelayInspections_TWallLocations_FK
                FOREIGN KEY ( intWallLocationID ) REFERENCES TWallLocations ( intWallLocationID ) 

                -- 17.
                ALTER TABLE TAutoBelayInspections ADD CONSTRAINT TAutoBelayInspections_TStandardInspections_FK
                FOREIGN KEY ( intStandardInspectionID ) REFERENCES TStandardInspections ( intStandardInspectionID ) 

                -- 18.
                ALTER TABLE TAutoBelayInspections ADD CONSTRAINT TAutoBelayInspections_TInspectors_FK
                FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID ) 

                -- 19.                  
                ALTER TABLE TAutoBelayInspections ADD CONSTRAINT TAutoBelayInspections_TInspectionStatus_FK
                FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )

                -- 20.                  
                ALTER TABLE TAutoBelayReserviceReports ADD CONSTRAINT TAutoBelayReserviceReports_TAutoBelayInspections_FK
                FOREIGN KEY ( intAutoBelayInspectionID ) REFERENCES TAutoBelayInspections ( intAutoBelayInspectionID )

                -- 21.                  
                ALTER TABLE TAutoBelayReserviceReports ADD CONSTRAINT TAutoBelayReserviceReports_TInspectors_FK
                FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID )                                 
                """            
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlForeignKeys)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Foreign Keys: {e}")            

    def createRopeTables(self):
        """ 
        Function Name: createRopeTables
        Function Purpose: Create the Ropes Tables and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Ropes Table
                CREATE TABLE IF NOT EXISTS TRopes 
                (
                    intRopeID                       INTEGER NOT NULL
                    ,strSerialNum                   VARCHAR(255) NOT NULL
                    ,strBumperNum                   VARCHAR(255) 
                    ,strRopeLength                  VARCHAR(225) NOT NULL 
                    ,strDiameter                    VARCHAR(255) NOT NULL
                    ,strElasticity                  VARCHAR(255) NOT NULL
                    ,strManufactureName             VARCHAR(255) NOT NULL
                    ,dtmManufactureDate             DATETIME 
                    ,dtmInstallationDate            DATETIME NOT NULL   
                    ,dtmLastInspectionDate          DATETIME NOT NULL   
                    ,dtmNextInspectionDate          DATETIME NOT NULL       
                    ,strEquipInUse                  VARCHAR(255) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT TRopes_PK PRIMARY KEY (intRopeID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Ropes Table
                CREATE TABLE IF NOT EXISTS Z_TRopes 
                (
                    intRopeAuditID                  INTEGER NOT NULL
                    ,intRopeID                      INTEGER NOT NULL
                    ,strSerialNum                   VARCHAR(255) NOT NULL
                    ,strBumperNum                   VARCHAR(255) 
                    ,strRopeLength                  VARCHAR(225) NOT NULL 
                    ,strDiameter                    VARCHAR(255) NOT NULL
                    ,strElasticity                  VARCHAR(255) NOT NULL
                    ,strManufactureName             VARCHAR(255) NOT NULL
                    ,dtmManufactureDate             DATETIME 
                    ,dtmInstallationDate            DATETIME NOT NULL   
                    ,dtmLastInspectionDate          DATETIME NOT NULL   
                    ,dtmNextInspectionDate          DATETIME NOT NULL       
                    ,strEquipInUse                  VARCHAR(255) NOT NULL
                    ,strUpdatedBy                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                      VARCHAR(1) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT Z_TRopes_PK PRIMARY KEY (intRopeAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Ropes Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TRopes_AuditTrigger_Insert
                AFTER INSERT ON TRopes
                BEGIN
                    INSERT INTO Z_TRopes 
                    (
                        intRopeID                        
                        ,strSerialNum
                        ,strBumperNum
                        ,strRopeLength
                        ,strDiameter
                        ,strElasticity
                        ,strManufactureName
                        ,dtmManufactureDate
                        ,dtmInstallationDate
                        ,dtmLastInspectionDate
                        ,dtmNextInspectionDate
                        ,strEquipInUse
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intRopeID                        
                        ,NEW.strSerialNum
                        ,NEW.strBumperNum
                        ,NEW.strRopeLength
                        ,NEW.strDiameter
                        ,NEW.strElasticity
                        ,NEW.strManufactureName
                        ,NEW.dtmManufactureDate
                        ,NEW.dtmInstallationDate
                        ,NEW.dtmLastInspectionDate
                        ,NEW.dtmNextInspectionDate
                        ,NEW.strEquipInUse
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Ropes Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TRopes_AuditTrigger_Update
                AFTER UPDATE ON TRopes
                BEGIN
                    INSERT INTO Z_TRopes 
                    (
                        intRopeID                        
                        ,strSerialNum
                        ,strBumperNum
                        ,strRopeLength
                        ,strDiameter
                        ,strElasticity
                        ,strManufactureName
                        ,dtmManufactureDate
                        ,dtmInstallationDate
                        ,dtmLastInspectionDate
                        ,dtmNextInspectionDate
                        ,strEquipInUse
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intRopeID                        
                        ,NEW.strSerialNum
                        ,NEW.strBumperNum
                        ,NEW.strRopeLength
                        ,NEW.strDiameter
                        ,NEW.strElasticity
                        ,NEW.strManufactureName
                        ,NEW.dtmManufactureDate
                        ,NEW.dtmInstallationDate
                        ,NEW.dtmLastInspectionDate
                        ,NEW.dtmNextInspectionDate
                        ,NEW.strEquipInUse
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Ropes Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TRopes_AuditTrigger_Delete
                AFTER DELETE ON TRopes
                BEGIN
                    INSERT INTO Z_TRopes 
                    (
                        intRopeID                        
                        ,strSerialNum
                        ,strBumperNum
                        ,strRopeLength
                        ,strDiameter
                        ,strElasticity
                        ,strManufactureName
                        ,dtmManufactureDate
                        ,dtmInstallationDate
                        ,dtmLastInspectionDate
                        ,dtmNextInspectionDate
                        ,strEquipInUse
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intRopeID                        
                        ,OLD.strSerialNum
                        ,OLD.strBumperNum
                        ,OLD.strRopeLength
                        ,OLD.strDiameter
                        ,OLD.strElasticity
                        ,OLD.strManufactureName
                        ,OLD.dtmManufactureDate
                        ,OLD.dtmInstallationDate
                        ,OLD.dtmLastInspectionDate
                        ,OLD.dtmNextInspectionDate
                        ,OLD.strEquipInUse
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Ropes tables: {e}")

    def createRopeVisTextSelTables(self):
        """ 
        Function Name: createRopeVisTextSelTables
        Function Purpose: Create the Rope Visual Textile Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Rope Visual Textile Selection Table
                CREATE TABLE IF NOT EXISTS TRopeVisTextSelects 
                (
                    intRopeVisTextSelectID               INTEGER NOT NULL
                    ,strRopeVisTextSelect                VARCHAR(1000) NOT NULL
                    ,strModifiedReason                   VARCHAR(1000)
                    ,CONSTRAINT TRopeVisTextSelects_PK PRIMARY KEY (intRopeVisTextSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Rope Visual Textile Selection Table
                CREATE TABLE IF NOT EXISTS Z_TRopeVisTextSelects 
                (
                    intRopeVisTextSelectAuditID          INTEGER NOT NULL
                    ,intRopeVisTextSelectID              INTEGER NOT NULL
                    ,strRopeVisTextSelect                VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                        VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                        DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                           VARCHAR(1) NOT NULL
                    ,strModifiedReason                   VARCHAR(1000)
                    ,CONSTRAINT Z_TRopeVisTextSelects_PK PRIMARY KEY (intRopeVisTextSelectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Rope Visual Textile Selection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TRopeVisTextSelects_InsertTrigger
                AFTER INSERT ON TRopeVisTextSelects
                BEGIN
                    INSERT INTO Z_TRopeVisTextSelects 
                    (
                        intRopeVisTextSelectID,
                        strRopeVisTextSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intRopeVisTextSelectID,
                        NEW.strRopeVisTextSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Rope Visual Textile Selection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TRopeVisTextSelects_UpdateTrigger
                AFTER UPDATE ON TRopeVisTextSelects
                BEGIN
                    INSERT INTO Z_TRopeVisTextSelects 
                    (
                        intRopeVisTextSelectID,
                        strRopeVisTextSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intRopeVisTextSelectID,
                        NEW.strRopeVisTextSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Rope Visual Textile Selection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TRopeVisTextSelects_DeleteTrigger
                AFTER DELETE ON TRopeVisTextSelects
                BEGIN
                    INSERT INTO Z_TRopeVisTextSelects 
                    (
                        intRopeVisTextSelectID,
                        strRopeVisTextSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intRopeVisTextSelectID,
                        OLD.strRopeVisTextSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Rope Visual Textile Selection tables: {e}")             

    def createRopeVisInspectTables(self):
        """ 
        Function Name: createRopeVisInspectTables
        Function Purpose: Create the Rope Visual Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Rope Visual Inspection Table
                CREATE TABLE IF NOT EXISTS TRopeVisualInspections 
                (
                    intRopeVisualInspectionID           INTEGER NOT NULL
                    ,intRopeID                          INTEGER NOT NULL
                    ,intInspectionTypeID                INTEGER NOT NULL
                    ,intRopeVisTextSelectID             INTEGER NOT NULL
                    ,intInspectionStatusID              INTEGER NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TRopeVisualInspections_PK PRIMARY KEY (intRopeVisualInspectionID)
                    
                    ,FOREIGN KEY ( intRopeID ) REFERENCES TRopes ( intRopeID )
                    ,FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )
                    ,FOREIGN KEY ( intRopeVisTextSelectID ) REFERENCES TRopeVisTextSelects ( intRopeVisTextSelectID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )                     
                );    
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Rope Visual Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TRopeVisualInspections
                (
                    intRopeVisualInspectionAuditID      INTEGER NOT NULL
                    ,intRopeVisualInspectionID          INTEGER NOT NULL
                    ,intRopeID                          INTEGER NOT NULL
                    ,intInspectionTypeID                INTEGER NOT NULL
                    ,intRopeVisTextSelectID             INTEGER NOT NULL
                    ,intInspectionStatusID              INTEGER NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TRopeVisualInspections_PK PRIMARY KEY (intRopeVisualInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Rope Visual Inspection Table
                CREATE TRIGGER IF NOT EXISTS Z_TRopeVisualInspections_InsertTrigger
                AFTER INSERT ON TRopeVisualInspections
                BEGIN
                    INSERT INTO Z_TRopeVisualInspections 
                    (
                        intRopeVisualInspectionID
                        ,intRopeID
                        ,intInspectionTypeID
                        ,intRopeVisTextSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intRopeVisualInspectionID
                        ,NEW.intRopeID
                        ,NEW.intInspectionTypeID
                        ,NEW.intRopeVisTextSelectID
                        ,NEW.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Rope Visual Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TRopeVisualInspections_UpdateTrigger
                AFTER UPDATE ON TRopeVisualInspections
                BEGIN
                    INSERT INTO Z_TRopeVisualInspections 
                    (
                        intRopeVisualInspectionID
                        ,intRopeID
                        ,intInspectionTypeID
                        ,intRopeVisTextSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intRopeVisualInspectionID
                        ,NEW.intRopeID
                        ,NEW.intInspectionTypeID
                        ,NEW.intRopeVisTextSelectID
                        ,NEW.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Rope Visual Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TRopeVisualInspections_DeleteTrigger
                AFTER DELETE ON TRopeVisualInspections
                BEGIN
                    INSERT INTO Z_TRopeVisualInspections 
                    (
                        intRopeVisualInspectionID
                        ,intRopeID
                        ,intInspectionTypeID
                        ,intRopeVisTextSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intRopeVisualInspectionID
                        ,OLD.intRopeID
                        ,OLD.intInspectionTypeID
                        ,OLD.intRopeVisTextSelectID
                        ,OLD.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,datetime('now')
                        ,'D' -- Delete
                        ,NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Rope Visual Inspection tables: {e}")            

    def createRopePhysTextSelTables(self):
        """ 
        Function Name: createRopePhysTextSelTables
        Function Purpose: Create the Rope Physical Textile Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Rope Physical Textile Selection Table
                CREATE TABLE IF NOT EXISTS TRopePhysTextSelects 
                (
                    intRopePhysTextSelectID             INTEGER NOT NULL
                    ,strRopePhysTextSelect              VARCHAR(1000) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TRopePhysTextSelects_PK PRIMARY KEY (intRopePhysTextSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Rope Physical Textile Selection Table
                CREATE TABLE IF NOT EXISTS Z_TRopePhysTextSelects 
                (
                    intRopePhysTextSelectAuditID        INTEGER NOT NULL
                    ,intRopePhysTextSelectID            INTEGER NOT NULL
                    ,strRopePhysTextSelect              VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TRopePhysTextSelects_PK PRIMARY KEY (intRopePhysTextSelectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Rope Physical Textile Selection Table
                CREATE TRIGGER IF NOT EXISTS Z_TRopePhysTextSelects_InsertTrigger
                AFTER INSERT ON TRopePhysTextSelects
                BEGIN
                    INSERT INTO Z_TRopePhysTextSelects 
                    (
                        intRopePhysTextSelectID
                        ,strRopePhysTextSelect
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intRopePhysTextSelectID
                        ,NEW.strRopePhysTextSelect
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Rope Physical Textile Selection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TRopePhysTextSelects_UpdateTrigger
                AFTER UPDATE ON TRopePhysTextSelects
                BEGIN
                    INSERT INTO Z_TRopePhysTextSelects 
                    (
                        intRopePhysTextSelectID
                        ,strRopePhysTextSelect
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intRopePhysTextSelectID
                        ,NEW.strRopePhysTextSelect
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Rope Physical Textile Selection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TRopePhysTextSelects_DeleteTrigger
                AFTER DELETE ON TRopePhysTextSelects
                BEGIN
                    INSERT INTO Z_TRopePhysTextSelects 
                    (
                        intRopePhysTextSelectID
                        ,strRopePhysTextSelect
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intRopePhysTextSelectID
                        ,OLD.strRopePhysTextSelect
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Rope Physical Textile Selection tables: {e}")            

    def createRopePhysInspectTables(self):
        """ 
        Function Name: createRopePhysInspectTables
        Function Purpose: Create the Rope Physical Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Rope Physical Inspection Table
                CREATE TABLE IF NOT EXISTS TRopePhysicalInspections 
                (
                    intRopePhysicalInspectionID             INTEGER NOT NULL
                    ,intRopeID                              INTEGER NOT NULL
                    ,intInspectionTypeID                    INTEGER NOT NULL
                    ,intRopePhysTextSelectID                INTEGER NOT NULL
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT TRopePhysicalInspections_PK PRIMARY KEY (intRopePhysicalInspectionID)
                    
                    ,FOREIGN KEY ( intRopeID ) REFERENCES TRopes ( intRopeID )
                    ,FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )
                    ,FOREIGN KEY ( intRopePhysTextSelectID ) REFERENCES TRopePhysTextSelects ( intRopePhysTextSelectID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )                     
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Rope Physical Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TRopePhysicalInspections 
                (
                    intRopePhysicalInspectionAuditID        INTEGER NOT NULL
                    ,intRopePhysicalInspectionID            INTEGER NOT NULL
                    ,intRopeID                              INTEGER NOT NULL
                    ,intInspectionTypeID                    INTEGER NOT NULL
                    ,intRopePhysTextSelectID                INTEGER NOT NULL
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TRopePhysicalInspections_PK PRIMARY KEY (intRopePhysicalInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Rope Physical Inspection Table
                CREATE TRIGGER IF NOT EXISTS Z_TRopePhysicalInspections_InsertTrigger
                AFTER INSERT ON TRopePhysicalInspections
                BEGIN
                    INSERT INTO Z_TRopePhysicalInspections 
                    (
                        intRopePhysicalInspectionID
                        ,intRopeID
                        ,intInspectionTypeID
                        ,intRopePhysTextSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intRopePhysicalInspectionID
                        ,NEW.intRopeID
                        ,NEW.intInspectionTypeID
                        ,NEW.intRopePhysTextSelectID
                        ,NEW.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- INSERT
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Rope Physical Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TRopePhysicalInspections_UpdateTrigger
                AFTER UPDATE ON TRopePhysicalInspections
                BEGIN
                    INSERT INTO Z_TRopePhysicalInspections 
                    (
                        intRopePhysicalInspectionID
                        ,intRopeID
                        ,intInspectionTypeID
                        ,intRopePhysTextSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intRopePhysicalInspectionID
                        ,NEW.intRopeID
                        ,NEW.intInspectionTypeID
                        ,NEW.intRopePhysTextSelectID
                        ,NEW.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Rope Physical Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TRopePhysicalInspections_DeleteTrigger
                AFTER DELETE ON TRopePhysicalInspections
                BEGIN
                    INSERT INTO Z_TRopePhysicalInspections 
                    (
                        intRopePhysicalInspectionID
                        ,intRopeID
                        ,intInspectionTypeID
                        ,intRopePhysTextSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intRopePhysicalInspectionID
                        ,OLD.intRopeID
                        ,OLD.intInspectionTypeID
                        ,OLD.intRopePhysTextSelectID
                        ,OLD.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Rope Physical Inspection tables: {e}")            

    def createStandRopeInspectTables(self):
        """ 
        Function Name: createStandRopeInspectTables
        Function Purpose: Create the Standard Rope Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Standard Rope Inspection Table
                CREATE TABLE IF NOT EXISTS TStandardRopeInspections 
                (
                    intStandardRopeInspectionID             INTEGER NOT NULL
                    ,intRopeVisualInspectionID              INTEGER NOT NULL
                    ,intRopePhysicalInspectionID            INTEGER NOT NULL                    
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT TStandardRopeInspections_PK PRIMARY KEY (intStandardRopeInspectionID)
                    
                    ,FOREIGN KEY ( intRopeVisualInspectionID ) REFERENCES TRopeVisualInspections ( intRopeVisualInspectionID )
                    ,FOREIGN KEY ( intRopePhysicalInspectionID ) REFERENCES TRopePhysicalInspections ( intRopePhysicalInspectionID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )                     
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Standard Rope Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TStandardRopeInspections 
                (
                    intStandardRopeInspectionAuditID        INTEGER NOT NULL
                    ,intStandardRopeInspectionID            INTEGER NOT NULL
                    ,intRopeVisualInspectionID              INTEGER NOT NULL
                    ,intRopePhysicalInspectionID            INTEGER NOT NULL                    
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TStandardRopeInspections_PK PRIMARY KEY (intStandardRopeInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Standard Rope Inspection Table (Insert)
                CREATE TRIGGER IF NOT EXISTS Z_TStandardRopeInspections_Insert_AuditTrigger
                AFTER INSERT ON TStandardRopeInspections
                BEGIN
                    INSERT INTO Z_TStandardRopeInspections 
                    (
                        intStandardRopeInspectionID,
                        intRopeVisualInspectionID,
                        intRopePhysicalInspectionID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intStandardRopeInspectionID,
                        NEW.intRopeVisualInspectionID,
                        NEW.intRopePhysicalInspectionID,                    
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I',
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Standard Rope Inspection Table (Update)
                CREATE TRIGGER IF NOT EXISTS Z_TStandardRopeInspections_Update_AuditTrigger
                AFTER UPDATE ON TStandardRopeInspections
                BEGIN
                    INSERT INTO Z_TStandardRopeInspections 
                    (
                        intStandardRopeInspectionID,
                        intRopeVisualInspectionID,
                        intRopePhysicalInspectionID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intStandardRopeInspectionID,
                        NEW.intRopeVisualInspectionID,
                        NEW.intRopePhysicalInspectionID,                    
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U',
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Standard Rope Inspection Table (Delete)
                CREATE TRIGGER IF NOT EXISTS Z_TStandardRopeInspections_Delete_AuditTrigger
                AFTER DELETE ON TStandardRopeInspections
                BEGIN
                    INSERT INTO Z_TStandardRopeInspections 
                    (
                        intStandardRopeInspectionID,
                        intRopeVisualInspectionID,
                        intRopePhysicalInspectionID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intStandardRopeInspectionID,
                        OLD.intRopeVisualInspectionID,
                        OLD.intRopePhysicalInspectionID,                    
                        OLD.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D',
                        OLD.strModifiedReason
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Standard Rope Inspection tables: {e}")                                    

    def createRope_InspectionTables(self):
        """ 
        Function Name: createRope_InspectionTables
        Function Purpose: Create the Rope Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Rope Inspection Table
                CREATE TABLE IF NOT EXISTS TRopeInspections 
                (
                    intRopeInspectionID                 INTEGER NOT NULL
                    ,intRopeID                          INTEGER NOT NULL
                    ,intWallLocationID                  INTEGER NOT NULL
                    ,intStandardRopeInspectionID        INTEGER NOT NULL
                    ,intInspectorID                     INTEGER NOT NULL
                    ,intInspectionStatusID              INTEGER NOT NULL
                    ,dtmLastInspectionDate              DATETIME NOT NULL
                    ,dtmNextInspectionDate              DATETIME NOT NULL
                    ,strComment                         VARCHAR(225) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TRopeInspections_PK PRIMARY KEY (intRopeInspectionID)
                    
                    ,FOREIGN KEY ( intRopeID ) REFERENCES TRopes ( intRopeID )
                    ,FOREIGN KEY ( intWallLocationID ) REFERENCES TWallLocations ( intWallLocationID ) 
                    ,FOREIGN KEY ( intStandardRopeInspectionID ) REFERENCES TStandardRopeInspections ( intStandardRopeInspectionID ) 
                    ,FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID ) 
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID ) 
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Rope Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TRopeInspections 
                (
                    intRopeInspectionAuditID            INTEGER NOT NULL
                    ,intRopeInspectionID                INTEGER NOT NULL
                    ,intRopeID                          INTEGER NOT NULL
                    ,intWallLocationID                  INTEGER NOT NULL
                    ,intStandardRopeInspectionID        INTEGER NOT NULL
                    ,intInspectorID                     INTEGER NOT NULL
                    ,intInspectionStatusID              INTEGER NOT NULL
                    ,dtmLastInspectionDate              DATETIME NOT NULL
                    ,dtmNextInspectionDate              DATETIME NOT NULL
                    ,strComment                         VARCHAR(225) NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TRopeInspections_PK PRIMARY KEY (intRopeInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Rope Inspection Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TRopeInspections_AuditTrigger_Insert
                AFTER INSERT ON TRopeInspections
                BEGIN
                    INSERT INTO Z_TRopeInspections 
                    (
                        intRopeInspectionID
                        ,intRopeID
                        ,intWallLocationID
                        ,intStandardRopeInspectionID
                        ,intInspectorID
                        ,intInspectionStatusID
                        ,dtmLastInspectionDate
                        ,dtmNextInspectionDate
                        ,strComment
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intRopeInspectionID
                        ,NEW.intRopeID
                        ,NEW.intWallLocationID
                        ,NEW.intStandardRopeInspectionID
                        ,NEW.intInspectorID
                        ,NEW.intInspectionStatusID
                        ,NEW.dtmLastInspectionDate
                        ,NEW.dtmNextInspectionDate
                        ,NEW.strComment
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Rope Inspection Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TRopeInspections_AuditTrigger_Update
                AFTER UPDATE ON TRopeInspections
                BEGIN
                    INSERT INTO Z_TRopeInspections 
                    (
                        intRopeInspectionID
                        ,intRopeID
                        ,intWallLocationID
                        ,intStandardRopeInspectionID
                        ,intInspectorID
                        ,intInspectionStatusID
                        ,dtmLastInspectionDate
                        ,dtmNextInspectionDate
                        ,strComment
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intRopeInspectionID
                        ,NEW.intRopeID
                        ,NEW.intWallLocationID
                        ,NEW.intStandardRopeInspectionID
                        ,NEW.intInspectorID
                        ,NEW.intInspectionStatusID
                        ,NEW.dtmLastInspectionDate
                        ,NEW.dtmNextInspectionDate
                        ,NEW.strComment
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Rope Inspection Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TRopeInspections_AuditTrigger_Delete
                AFTER DELETE ON TRopeInspections
                BEGIN
                    INSERT INTO Z_TRopeInspections 
                    (
                        intRopeInspectionID
                        ,intRopeID
                        ,intWallLocationID
                        ,intStandardRopeInspectionID
                        ,intInspectorID
                        ,intInspectionStatusID
                        ,dtmLastInspectionDate
                        ,dtmNextInspectionDate
                        ,strComment
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intRopeInspectionID
                        ,OLD.intRopeID
                        ,OLD.intWallLocationID
                        ,OLD.intStandardRopeInspectionID
                        ,OLD.intInspectorID
                        ,OLD.intInspectionStatusID
                        ,OLD.dtmLastInspectionDate
                        ,OLD.dtmNextInspectionDate
                        ,OLD.strComment
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Rope Inspection tables: {e}")            

    def createRope_WallLocationTables(self):
        """ 
        Function Name: createRope_WallLocationTables
        Function Purpose: Create the Rope Location Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Rope Wall Location Table
                CREATE TABLE IF NOT EXISTS TRopeWallLocations 
                (
                    intRopeWallLocationID               INTEGER NOT NULL
                    ,intWallLocationID                  INTEGER NOT NULL
                    ,intRopeID                          INTEGER NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,FOREIGN KEY ( intWallLocationID ) REFERENCES TWallLocations ( intWallLocationID )
                    ,FOREIGN KEY ( intRopeID ) REFERENCES TRopes ( intRopeID )
                    
                    ,CONSTRAINT TRopeWallLocations_PK PRIMARY KEY (intRopeWallLocationID)
                    ,CONSTRAINT TRopeWallLocations_UQ UNIQUE (intWallLocationID, intRopeID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Rope Wall Location Table
                CREATE TABLE IF NOT EXISTS Z_TRopeWallLocations 
                (
                    intRopeWallLocationAuditID          INTEGER NOT NULL
                    ,intRopeWallLocationID              INTEGER NOT NULL
                    ,intWallLocationID                  INTEGER NOT NULL
                    ,intRopeID                          INTEGER NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TRopeWallLocations_PK PRIMARY KEY (intRopeWallLocationAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Rope Wall Location Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TRopeWallLocations_AuditTrigger_Insert
                AFTER INSERT ON TRopeWallLocations
                BEGIN
                    INSERT INTO Z_TRopeWallLocations 
                    (
                        intRopeWallLocationID
                        ,intWallLocationID
                        ,intRopeID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intRopeWallLocationID
                        ,NEW.intWallLocationID
                        ,NEW.intRopeID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Rope Wall Location Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TRopeWallLocations_AuditTrigger_Update
                AFTER UPDATE ON TRopeWallLocations
                BEGIN
                    INSERT INTO Z_TRopeWallLocations 
                    (
                        intRopeWallLocationID
                        ,intWallLocationID
                        ,intRopeID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intRopeWallLocationID
                        ,NEW.intWallLocationID
                        ,NEW.intRopeID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Rope Wall Location Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TRopeWallLocations_AuditTrigger_Delete
                AFTER DELETE ON TRopeWallLocations
                BEGIN
                    INSERT INTO Z_TRopeWallLocations 
                    (
                        intRopeWallLocationID
                        ,intWallLocationID
                        ,intRopeID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intRopeWallLocationID
                        ,OLD.intWallLocationID
                        ,OLD.intRopeID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Rope Wall Location tables: {e}")            

    def createRope_InspectorTables(self):
        """ 
        Function Name: createRope_InspectorTables
        Function Purpose: Create the Rope Inspector Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Rope Inspector Table
                CREATE TABLE IF NOT EXISTS TRopeInspectors 
                (
                    intRopeInspectorID              INTEGER NOT NULL
                    ,intInspectorID                 INTEGER NOT NULL
                    ,intRopeID                      INTEGER NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT TRopeInspectors_PK PRIMARY KEY (intRopeInspectorID)
                    
                    ,FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID )
                    ,FOREIGN KEY ( intRopeID ) REFERENCES TRopes ( intRopeID )
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Rope Inspector Table
                CREATE TABLE IF NOT EXISTS Z_TRopeInspectors 
                (
                    intRopeInspectorAuditID         INTEGER NOT NULL
                    ,intRopeInspectorID             INTEGER NOT NULL
                    ,intInspectorID                 INTEGER NOT NULL
                    ,intRopeID                      INTEGER NOT NULL
                    ,strUpdatedBy                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                      VARCHAR(1) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT Z_TRopeInspectors_PK PRIMARY KEY (intRopeInspectorAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Rope Inspector Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TRopeInspectors_AuditTrigger_Insert
                AFTER INSERT ON TRopeInspectors
                BEGIN
                    INSERT INTO Z_TRopeInspectors 
                    (
                        intRopeInspectorID
                        ,intInspectorID
                        ,intRopeID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intRopeInspectorID
                        ,NEW.intInspectorID
                        ,NEW.intRopeID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Rope Inspector Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TRopeInspectors_AuditTrigger_Update
                AFTER UPDATE ON TRopeInspectors
                BEGIN
                    INSERT INTO Z_TRopeInspectors 
                    (
                        intRopeInspectorID
                        ,intInspectorID
                        ,intRopeID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intRopeInspectorID
                        ,NEW.intInspectorID
                        ,NEW.intRopeID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Rope Inspector Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TRopeInspectors_AuditTrigger_Delete
                AFTER DELETE ON TRopeInspectors
                BEGIN
                    INSERT INTO Z_TRopeInspectors 
                    (
                        intRopeInspectorID
                        ,intInspectorID
                        ,intRopeID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intRopeInspectorID
                        ,OLD.intInspectorID
                        ,OLD.intRopeID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Rope Inspector tables: {e}")            

    def createRope_RetiredReports_Tables(self):
        """ 
        Function Name: createRope_RetiredReports_Tables
        Function Purpose: Create the Rope Retired Reports Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Rope Inspection Table
                CREATE TABLE IF NOT EXISTS TRopeRetiredReports 
                (
                    intRopeRetiredReportID                  INTEGER NOT NULL
                    ,intRopeInspectionID                    INTEGER NOT NULL
                    ,intInspectorID                         INTEGER NOT NULL
                    ,dtmReportDate                          DATETIME NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT TRopeRetiredReports_PK PRIMARY KEY (intRopeRetiredReportID)
                    
                    ,FOREIGN KEY ( intRopeInspectionID ) REFERENCES TRopeInspections ( intRopeInspectionID )
                    ,FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID ) 
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Rope Retired Report Table
                CREATE TABLE IF NOT EXISTS Z_TRopeRetiredReports 
                (
                    intRopeRetiredReportAuditID             INTEGER NOT NULL
                    ,intRopeRetiredReportID                 INTEGER NOT NULL
                    ,intRopeInspectionID                    INTEGER NOT NULL
                    ,intInspectorID                         INTEGER NOT NULL
                    ,dtmReportDate                          DATETIME NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TRopeRetiredReports_PK PRIMARY KEY (intRopeRetiredReportAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Rope Retired Report Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TRopeRetiredReports_AuditTrigger_Insert
                AFTER INSERT ON TRopeRetiredReports
                BEGIN
                    INSERT INTO Z_TRopeRetiredReports 
                    (
                        intRopeRetiredReportID
                        ,intRopeInspectionID
                        ,intInspectorID
                        ,dtmReportDate
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intRopeRetiredReportID
                        ,NEW.intRopeInspectionID
                        ,NEW.intInspectorID
                        ,NEW.dtmReportDate
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Rope Retired Report Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TRopeRetiredReports_AuditTrigger_Update
                AFTER UPDATE ON TRopeRetiredReports
                BEGIN
                    INSERT INTO Z_TRopeRetiredReports 
                    (
                        intRopeRetiredReportID
                        ,intRopeInspectionID
                        ,intInspectorID
                        ,dtmReportDate
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intRopeRetiredReportID
                        ,NEW.intRopeInspectionID
                        ,NEW.intInspectorID
                        ,NEW.dtmReportDate
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Rope Inspection Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TRopeRetiredReports_AuditTrigger_Delete
                AFTER DELETE ON TRopeRetiredReports
                BEGIN
                    INSERT INTO Z_TRopeRetiredReports 
                    (
                        intRopeRetiredReportID
                        ,intRopeInspectionID
                        ,intInspectorID
                        ,dtmReportDate
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intRopeRetiredReportID
                        ,OLD.intRopeInspectionID
                        ,OLD.intInspectorID
                        ,OLD.dtmReportDate
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Rope Inspection tables: {e}")
                        
    def createRopes_ForeignKeys(self):
        """ 
        Function Name: createRopes_ForeignKeys
        Function Purpose: Create the Ropes Visual, Physical, Standard, and Inspection foreign keys inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the foreign keys
            sqlForeignKeys = """
                -- -------------------------------------------------------------------------------------------------------
                -- Establish Referential Integrity - Inserts
                -- -------------------------------------------------------------------------------------------------------

                -- #           Child                           Parent                                Column
                -- --         -------                         --------                              --------
                -- 1. TRopeVisualInspections                 TRopes                                 intRopeID
                -- 2. TRopeVisualInspections                 TInspectionTypes                       intInspectionTypeID
                -- 3. TRopeVisualInspections                 TRopeVisTextSelects                    intRopeVisTextSelectID
                -- 4. TRopeVisualInspections                 TInspectionStatus                      intInspectionStatusID
                -- 5. TRopePhysicalInspections               TRopes                                 intRopeID
                -- 6. TRopePhysicalInspections               TInspectionTypes                       intInspectionTypeID
                -- 7. TRopePhysicalInspections               TRopePhysTextSelects                   intRopePhysTextSelectID
                -- 8. TRopePhysicalInspections               TInspectionStatus                      intInspectionStatusID
                -- 9. TStandardRopeInspections               TRopeVisualInspections                 intRopeVisualInspectionID
                -- 10. TStandardRopeInspections              TRopePhysicalInspections               intRopePhysicalInspectionID
                -- 11. TStandardRopeInspections              TInspectionStatus                      intInspectionStatusID
                -- 12. TRopeInspections                      TRopes                                 intRopeID
                -- 13. TRopeInspections                      TWallLocations                         intWallLocationID
                -- 14. TRopeInspections                      TStandardRopeInspections               intStandardRopeInspectionID
                -- 15. TRopeInspections                      TInspectors                            intInspectorID
                -- 16. TRopeInspections                      TInspectionStatus                      intInspectionStatusID
                -- 17. TRopeRetiredReports                   TRopeInspections                       intRopeInspectionID
                -- 18. TRopeRetiredReports                   TInspectors                            intInspectorID

                -- -------------------------------------------------------------------------------------------------------
                -- Create Foreign Keys
                -- -------------------------------------------------------------------------------------------------------
                -- 1.
                ALTER TABLE TRopeVisualInspections ADD CONSTRAINT TRopeVisualInspections_TRopes_FK
                FOREIGN KEY ( intRopeID ) REFERENCES TRopes ( intRopeID )

                -- 2.
                ALTER TABLE TRopeVisualInspections ADD CONSTRAINT TRopeVisualInspections_TInspectionTypes_FK
                FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )

                -- 3.
                ALTER TABLE TRopeVisualInspections ADD CONSTRAINT TRopeVisualInspections_TRopeVisTextSelects_FK
                FOREIGN KEY ( intRopeVisTextSelectID ) REFERENCES TRopeVisTextSelects ( intRopeVisTextSelectID )

                -- 4.
                ALTER TABLE TRopeVisualInspections ADD CONSTRAINT TRopeVisualInspections_TInspectionStatus_FK
                FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )

                -- -------------------------------------------------------------------------------------------------------
                -- Ensure No Duplicated Data
                -- -------------------------------------------------------------------------------------------------------
                ALTER TABLE TRopeWallLocations ADD CONSTRAINT TRopeWallLocations_UNIQUE
                UNIQUE (intWallLocationID, intRopeID)

                -- -------------------------------------------------------------------------------------------------------
                -- Create Foreign Keys
                -- -------------------------------------------------------------------------------------------------------
                -- 5.
                ALTER TABLE TRopePhysicalInspections ADD CONSTRAINT TRopePhysicalInspections_TRopes_FK
                FOREIGN KEY ( intRopeID ) REFERENCES TRopes ( intRopeID )

                -- 6.
                ALTER TABLE TRopePhysicalInspections ADD CONSTRAINT TRopePhysicalInspections_TInspectionTypes_FK
                FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )

                -- 7.
                ALTER TABLE TRopePhysicalInspections ADD CONSTRAINT TRopePhysicalInspections_TRopePhysTextSelects_FK
                FOREIGN KEY ( intRopePhysTextSelectID ) REFERENCES TRopePhysTextSelects ( intRopePhysTextSelectID )

                -- 8. 
                ALTER TABLE TRopePhysicalInspections ADD CONSTRAINT TRopePhysicalInspections_TInspectionStatus_FK
                FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )

                -- 9. 
                ALTER TABLE TStandardRopeInspections ADD CONSTRAINT TStandardRopeInspections_TRopeVisualInspections_FK
                FOREIGN KEY ( intRopeVisualInspectionID ) REFERENCES TRopeVisualInspections ( intRopeVisualInspectionID )

                -- 10.
                ALTER TABLE TStandardRopeInspections ADD CONSTRAINT TStandardRopeInspections_TRopePhysicalInspections_FK
                FOREIGN KEY ( intRopePhysicalInspectionID ) REFERENCES TRopePhysicalInspections ( intRopePhysicalInspectionID )

                -- 11.
                ALTER TABLE TStandardRopeInspections ADD CONSTRAINT TStandardRopeInspections_TInspectionStatus_FK
                FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )

                -- 12.
                ALTER TABLE TRopeInspections ADD CONSTRAINT TRopeInspections_TRopes_FK
                FOREIGN KEY ( intRopeID ) REFERENCES TRopes ( intRopeID )

                -- 13.
                ALTER TABLE TRopeInspections ADD CONSTRAINT TRopeInspections_TWallLocations_FK
                FOREIGN KEY ( intWallLocationID ) REFERENCES TWallLocations ( intWallLocationID )

                -- 14.
                ALTER TABLE TRopeInspections ADD CONSTRAINT TRopeInspections_TStandardRopeInspections_FK
                FOREIGN KEY ( intStandardRopeInspectionID ) REFERENCES TStandardRopeInspections ( intStandardRopeInspectionID )
                
                -- 15.
                ALTER TABLE TRopeInspections ADD CONSTRAINT TRopeInspections_TInspectors_FK
                FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID )

                -- 16.
                ALTER TABLE TRopeInspections ADD CONSTRAINT TRopeInspections_TInspectionStatus_FK
                FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )

                -- 17.                  
                ALTER TABLE TRopeRetiredReports ADD CONSTRAINT TRopeRetiredReports_TRopeInspections_FK
                FOREIGN KEY ( intRopeInspectionID ) REFERENCES TRopeInspections ( intRopeInspectionID )

                -- 18.                  
                ALTER TABLE TRopeRetiredReports ADD CONSTRAINT TRopeRetiredReports_TInspectors_FK
                FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID ) 
                """            
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlForeignKeys)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Foreign Keys: {e}")            

    def createConnectorTables(self):
        """ 
        Function Name: createConnectorTables
        Function Purpose: Create the Connectors Tables and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Connectors Table
                CREATE TABLE IF NOT EXISTS TConnectors 
                (
                    intConnectorID                  INTEGER NOT NULL
                    ,strSerialNum                   VARCHAR(255) NOT NULL
                    ,strBumperNum                   VARCHAR(255) 
                    ,strManufactureName             VARCHAR(255) NOT NULL
                    ,dtmManufactureDate             DATETIME 
                    ,dtmInstallationDate            DATETIME NOT NULL   
                    ,dtmLastInspectionDate          DATETIME NOT NULL   
                    ,dtmNextInspectionDate          DATETIME NOT NULL       
                    ,strDeviceType                  VARCHAR(255) NOT NULL
                    ,strEquipInUse                  VARCHAR(255) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT TConnectors_PK PRIMARY KEY (intConnectorID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Connectors Table
                CREATE TABLE IF NOT EXISTS Z_TConnectors 
                (
                    intConnectorAuditID             INTEGER NOT NULL
                    ,intConnectorID                 INTEGER NOT NULL
                    ,strSerialNum                   VARCHAR(255) NOT NULL
                    ,strBumperNum                   VARCHAR(255) 
                    ,strManufactureName             VARCHAR(255) NOT NULL
                    ,dtmManufactureDate             DATETIME 
                    ,dtmInstallationDate            DATETIME NOT NULL   
                    ,dtmLastInspectionDate          DATETIME NOT NULL   
                    ,dtmNextInspectionDate          DATETIME NOT NULL       
                    ,strDeviceType                  VARCHAR(255) NOT NULL
                    ,strEquipInUse                  VARCHAR(255) NOT NULL
                    ,strUpdatedBy                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                      VARCHAR(1) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT Z_TConnectors_PK PRIMARY KEY (intConnectorAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Connectors Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TConnectors_AuditTrigger_Insert
                AFTER INSERT ON TConnectors
                BEGIN
                    INSERT INTO Z_TConnectors 
                    (
                        intConnectorID                        
                        ,strSerialNum
                        ,strBumperNum
                        ,strManufactureName
                        ,dtmManufactureDate
                        ,dtmInstallationDate
                        ,dtmLastInspectionDate
                        ,dtmNextInspectionDate
                        ,strDeviceType
                        ,strEquipInUse
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intConnectorID                        
                        ,NEW.strSerialNum
                        ,NEW.strBumperNum
                        ,NEW.strManufactureName
                        ,NEW.dtmManufactureDate
                        ,NEW.dtmInstallationDate
                        ,NEW.dtmLastInspectionDate
                        ,NEW.dtmNextInspectionDate
                        ,NEW.strDeviceType
                        ,NEW.strEquipInUse
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Connectors Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TConnectors_AuditTrigger_Update
                AFTER UPDATE ON TConnectors
                BEGIN
                    INSERT INTO Z_TConnectors 
                    (
                        intConnectorID                        
                        ,strSerialNum
                        ,strBumperNum
                        ,strManufactureName
                        ,dtmManufactureDate
                        ,dtmInstallationDate
                        ,dtmLastInspectionDate
                        ,dtmNextInspectionDate
                        ,strDeviceType
                        ,strEquipInUse
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intConnectorID                        
                        ,NEW.strSerialNum
                        ,NEW.strBumperNum
                        ,NEW.strManufactureName
                        ,NEW.dtmManufactureDate
                        ,NEW.dtmInstallationDate
                        ,NEW.dtmLastInspectionDate
                        ,NEW.dtmNextInspectionDate
                        ,NEW.strDeviceType
                        ,NEW.strEquipInUse
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Connectors Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TConnectors_AuditTrigger_Delete
                AFTER DELETE ON TConnectors
                BEGIN
                    INSERT INTO Z_TConnectors 
                    (
                        intConnectorID                        
                        ,strSerialNum
                        ,strBumperNum
                        ,strManufactureName
                        ,dtmManufactureDate
                        ,dtmInstallationDate
                        ,dtmLastInspectionDate
                        ,dtmNextInspectionDate
                        ,strDeviceType
                        ,strEquipInUse
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intConnectorID                        
                        ,OLD.strSerialNum
                        ,OLD.strBumperNum
                        ,OLD.strManufactureName
                        ,OLD.dtmManufactureDate
                        ,OLD.dtmInstallationDate
                        ,OLD.dtmLastInspectionDate
                        ,OLD.dtmNextInspectionDate
                        ,OLD.strDeviceType
                        ,OLD.strEquipInUse
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Connectors tables: {e}")

    def createConnectorVisMetalSelTables(self):
        """ 
        Function Name: createConnectorVisMetalSelTables
        Function Purpose: Create the Connector Visual Metallic Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Connector Visual Metallic Selection Table
                CREATE TABLE IF NOT EXISTS TConnectorVisMetalSelects 
                (
                    intConnectorVisMetalSelectID         INTEGER NOT NULL
                    ,strConnectorVisMetalSelect          VARCHAR(1000) NOT NULL
                    ,strModifiedReason                   VARCHAR(1000)
                    ,CONSTRAINT TConnectorVisMetalSelects_PK PRIMARY KEY (intConnectorVisMetalSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Connector Visual Metallic Selection Table
                CREATE TABLE IF NOT EXISTS Z_TConnectorVisMetalSelects 
                (
                    intConnectorVisMetalSelectAuditID    INTEGER NOT NULL
                    ,intConnectorVisMetalSelectID        INTEGER NOT NULL
                    ,strConnectorVisMetalSelect          VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                        VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                        DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                           VARCHAR(1) NOT NULL
                    ,strModifiedReason                   VARCHAR(1000)
                    ,CONSTRAINT Z_TConnectorVisMetalSelects_PK PRIMARY KEY (intConnectorVisMetalSelectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Connector Visual Metallic Selection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorVisMetalSelects_InsertTrigger
                AFTER INSERT ON TConnectorVisMetalSelects
                BEGIN
                    INSERT INTO Z_TConnectorVisMetalSelects 
                    (
                        intConnectorVisMetalSelectID,
                        strConnectorVisMetalSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intConnectorVisMetalSelectID,
                        NEW.strConnectorVisMetalSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Connector Visual Metallic Selection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorVisMetalSelects_UpdateTrigger
                AFTER UPDATE ON TConnectorVisMetalSelects
                BEGIN
                    INSERT INTO Z_TConnectorVisMetalSelects 
                    (
                        intConnectorVisMetalSelectID,
                        strConnectorVisMetalSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intConnectorVisMetalSelectID,
                        NEW.strConnectorVisMetalSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Connector Visual Metallic Selection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorVisMetalSelects_DeleteTrigger
                AFTER DELETE ON TConnectorVisMetalSelects
                BEGIN
                    INSERT INTO Z_TConnectorVisMetalSelects 
                    (
                        intConnectorVisMetalSelectID,
                        strConnectorVisMetalSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intConnectorVisMetalSelectID,
                        OLD.strConnectorVisMetalSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Connector Visual Metallic Selection tables: {e}")             

    def createConnectorVisInspectTables(self):
        """ 
        Function Name: createConnectorVisInspectTables
        Function Purpose: Create the Connector Visual Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Connector Visual Inspection Table
                CREATE TABLE IF NOT EXISTS TConnectorVisualInspections 
                (
                    intConnectorVisualInspectionID      INTEGER NOT NULL
                    ,intConnectorID                     INTEGER NOT NULL
                    ,intInspectionTypeID                INTEGER NOT NULL
                    ,intConnectorVisMetalSelectID       INTEGER NOT NULL
                    ,intInspectionStatusID              INTEGER NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TConnectorVisualInspections_PK PRIMARY KEY (intConnectorVisualInspectionID)
                    
                    ,FOREIGN KEY ( intConnectorID ) REFERENCES TConnectors ( intConnectorID )
                    ,FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )
                    ,FOREIGN KEY ( intConnectorVisMetalSelectID ) REFERENCES TConnectorVisMetalSelects ( intConnectorVisMetalSelectID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )                     
                );    
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Connector Visual Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TConnectorVisualInspections
                (
                    intConnectorVisualInspectionAuditID INTEGER NOT NULL
                    ,intConnectorVisualInspectionID     INTEGER NOT NULL
                    ,intConnectorID                     INTEGER NOT NULL
                    ,intInspectionTypeID                INTEGER NOT NULL
                    ,intConnectorVisMetalSelectID       INTEGER NOT NULL
                    ,intInspectionStatusID              INTEGER NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TConnectorVisualInspections_PK PRIMARY KEY (intConnectorVisualInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Connector Visual Inspection Table
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorVisualInspections_InsertTrigger
                AFTER INSERT ON TConnectorVisualInspections
                BEGIN
                    INSERT INTO Z_TConnectorVisualInspections 
                    (
                        intConnectorVisualInspectionID
                        ,intConnectorID
                        ,intInspectionTypeID
                        ,intConnectorVisMetalSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intConnectorVisualInspectionID
                        ,NEW.intConnectorID
                        ,NEW.intInspectionTypeID
                        ,NEW.intConnectorVisMetalSelectID
                        ,NEW.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Connector Visual Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorVisualInspections_UpdateTrigger
                AFTER UPDATE ON TConnectorVisualInspections
                BEGIN
                    INSERT INTO Z_TConnectorVisualInspections 
                    (
                        intConnectorVisualInspectionID
                        ,intConnectorID
                        ,intInspectionTypeID
                        ,intConnectorVisMetalSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intConnectorVisualInspectionID
                        ,NEW.intConnectorID
                        ,NEW.intInspectionTypeID
                        ,NEW.intConnectorVisMetalSelectID
                        ,NEW.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Connector Visual Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorVisualInspections_DeleteTrigger
                AFTER DELETE ON TConnectorVisualInspections
                BEGIN
                    INSERT INTO Z_TConnectorVisualInspections 
                    (
                        intConnectorVisualInspectionID
                        ,intConnectorID
                        ,intInspectionTypeID
                        ,intConnectorVisMetalSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intConnectorVisualInspectionID
                        ,OLD.intConnectorID
                        ,OLD.intInspectionTypeID
                        ,OLD.intConnectorVisMetalSelectID
                        ,OLD.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,datetime('now')
                        ,'D' -- Delete
                        ,NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Connector Visual Inspection tables: {e}")            

    def createConnectorPhysMetalSelTables(self):
        """ 
        Function Name: createConnectorPhysMetalSelTables
        Function Purpose: Create the Connector Physical Metallic Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Connector Physical Metallic Selection Table
                CREATE TABLE IF NOT EXISTS TConnectorPhysMetalSelects 
                (
                    intConnectorPhysMetalSelectID       INTEGER NOT NULL
                    ,strConnectorPhysMetalSelect        VARCHAR(1000) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TConnectorPhysMetalSelects_PK PRIMARY KEY (intConnectorPhysMetalSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Connector Physical Metallic Selection Table
                CREATE TABLE IF NOT EXISTS Z_TConnectorPhysMetalSelects 
                (
                    intConnectorPhysMetalSelectAuditID  INTEGER NOT NULL
                    ,intConnectorPhysMetalSelectID      INTEGER NOT NULL
                    ,strConnectorPhysMetalSelect        VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TConnectorPhysMetalSelects_PK PRIMARY KEY (intConnectorPhysMetalSelectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Connector Physical Metallic Selection Table
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorPhysMetalSelects_InsertTrigger
                AFTER INSERT ON TConnectorPhysMetalSelects
                BEGIN
                    INSERT INTO Z_TConnectorPhysMetalSelects 
                    (
                        intConnectorPhysMetalSelectID
                        ,strConnectorPhysMetalSelect
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intConnectorPhysMetalSelectID
                        ,NEW.strConnectorPhysMetalSelect
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Connector Physical Metallic Selection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorPhysMetalSelects_UpdateTrigger
                AFTER UPDATE ON TConnectorPhysMetalSelects
                BEGIN
                    INSERT INTO Z_TConnectorPhysMetalSelects 
                    (
                        intConnectorPhysMetalSelectID
                        ,strConnectorPhysMetalSelect
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intConnectorPhysMetalSelectID
                        ,NEW.strConnectorPhysMetalSelect
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Connector Physical Metallic Selection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorPhysMetalSelects_DeleteTrigger
                AFTER DELETE ON TConnectorPhysMetalSelects
                BEGIN
                    INSERT INTO Z_TConnectorPhysMetalSelects 
                    (
                        intConnectorPhysMetalSelectID
                        ,strConnectorPhysMetalSelect
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intConnectorPhysMetalSelectID
                        ,OLD.strConnectorPhysMetalSelect
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Connector Physical Metallic Selection tables: {e}")            

    def createConnectorPhysInspectTables(self):
        """ 
        Function Name: createConnectorPhysInspectTables
        Function Purpose: Create the Connector Physical Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Connector Physical Inspection Table
                CREATE TABLE IF NOT EXISTS TConnectorPhysicalInspections 
                (
                    intConnectorPhysicalInspectionID        INTEGER NOT NULL
                    ,intConnectorID                         INTEGER NOT NULL
                    ,intInspectionTypeID                    INTEGER NOT NULL
                    ,intConnectorPhysMetalSelectID          INTEGER NOT NULL
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT TConnectorPhysicalInspections_PK PRIMARY KEY (intConnectorPhysicalInspectionID)
                    
                    ,FOREIGN KEY ( intConnectorID ) REFERENCES TConnectors ( intConnectorID )
                    ,FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )
                    ,FOREIGN KEY ( intConnectorPhysMetalSelectID ) REFERENCES TConnectorPhysMetalSelects ( intConnectorPhysMetalSelectID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )                     
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Connector Physical Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TConnectorPhysicalInspections 
                (
                    intConnectorPhysicalInspectionAuditID   INTEGER NOT NULL
                    ,intConnectorPhysicalInspectionID       INTEGER NOT NULL
                    ,intConnectorID                         INTEGER NOT NULL
                    ,intInspectionTypeID                    INTEGER NOT NULL
                    ,intConnectorPhysMetalSelectID          INTEGER NOT NULL
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TConnectorPhysicalInspections_PK PRIMARY KEY (intConnectorPhysicalInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Connector Physical Inspection Table
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorPhysicalInspections_InsertTrigger
                AFTER INSERT ON TConnectorPhysicalInspections
                BEGIN
                    INSERT INTO Z_TConnectorPhysicalInspections 
                    (
                        intConnectorPhysicalInspectionID
                        ,intConnectorID
                        ,intInspectionTypeID
                        ,intConnectorPhysMetalSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intConnectorPhysicalInspectionID
                        ,NEW.intConnectorID
                        ,NEW.intInspectionTypeID
                        ,NEW.intConnectorPhysMetalSelectID
                        ,NEW.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- INSERT
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Connector Physical Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorPhysicalInspections_UpdateTrigger
                AFTER UPDATE ON TConnectorPhysicalInspections
                BEGIN
                    INSERT INTO Z_TConnectorPhysicalInspections 
                    (
                        intConnectorPhysicalInspectionID
                        ,intConnectorID
                        ,intInspectionTypeID
                        ,intConnectorPhysMetalSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intConnectorPhysicalInspectionID
                        ,NEW.intConnectorID
                        ,NEW.intInspectionTypeID
                        ,NEW.intConnectorPhysMetalSelectID
                        ,NEW.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Connector Physical Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorPhysicalInspections_DeleteTrigger
                AFTER DELETE ON TConnectorPhysicalInspections
                BEGIN
                    INSERT INTO Z_TConnectorPhysicalInspections 
                    (
                        intConnectorPhysicalInspectionID
                        ,intConnectorID
                        ,intInspectionTypeID
                        ,intConnectorPhysMetalSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intConnectorPhysicalInspectionID
                        ,OLD.intConnectorID
                        ,OLD.intInspectionTypeID
                        ,OLD.intConnectorPhysMetalSelectID
                        ,OLD.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Connector Physical Inspection tables: {e}")            

    def createConnectorFunctionsTables(self):
        """ 
        Function Name: createConnectorFunctionsTables
        Function Purpose: Create the Connector Functions and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
                    
            # Create the table
            sqlTable = """
                -- Create Connector Function Table
                CREATE TABLE IF NOT EXISTS TConnectorFunctions 
                (
                    intConnectorFunctionID              INTEGER NOT NULL
                    ,strConnectorFunctionDesc           VARCHAR(255) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TConnectorFunctions_PK PRIMARY KEY (intConnectorFunctionID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Connector Function Table
                CREATE TABLE IF NOT EXISTS Z_TConnectorFunctions 
                (
                    intConnectorFunctionAuditID         INTEGER NOT NULL
                    ,intConnectorFunctionID             INTEGER NOT NULL
                    ,strConnectorFunctionDesc           VARCHAR(255) NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TConnectorFunctions_PK PRIMARY KEY (intConnectorFunctionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Connector Function Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorFunctions_InsertTrigger
                AFTER INSERT ON TConnectorFunctions
                BEGIN
                    INSERT INTO Z_TConnectorFunctions 
                    (
                        intConnectorFunctionID,
                        strConnectorFunctionDesc,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intConnectorFunctionID,
                        NEW.strConnectorFunctionDesc,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Connector Function Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorFunctions_UpdateTrigger
                AFTER UPDATE ON TConnectorFunctions
                BEGIN
                    INSERT INTO Z_TConnectorFunctions 
                    (
                        intConnectorFunctionID,
                        strConnectorFunctionDesc,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intConnectorFunctionID,
                        NEW.strConnectorFunctionDesc,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Connector Function Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorFunctions_DeleteTrigger
                AFTER DELETE ON TConnectorFunctions
                BEGIN
                    INSERT INTO Z_TConnectorFunctions 
                    (
                        intConnectorFunctionID,
                        strConnectorFunctionDesc,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intConnectorFunctionID,
                        OLD.strConnectorFunctionDesc,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)     
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Connector Functional tables: {e}")              
            
    def createConnectorFunctSelTables(self):
        """ 
        Function Name: createConnectorFunctSelTables
        Function Purpose: Create the Connector Function Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Connector Function Selection Table
                CREATE TABLE IF NOT EXISTS TConnectorFunctSelects 
                (
                    intConnectorFunctSelectID           INTEGER NOT NULL
                    ,strConnectorFunctSelect            VARCHAR(1000) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TConnectorFunctSelects_PK PRIMARY KEY (intConnectorFunctSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Connector Function Selection Table
                CREATE TABLE IF NOT EXISTS Z_TConnectorFunctSelects 
                (
                    intConnectorFunctSelectAuditID      INTEGER NOT NULL
                    ,intConnectorFunctSelectID          INTEGER NOT NULL
                    ,strConnectorFunctSelect            VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TConnectorFunctSelects_PK PRIMARY KEY (intConnectorFunctSelectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Connector Function Selection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorFunctSelects_InsertTrigger
                AFTER INSERT ON TConnectorFunctSelects
                BEGIN
                    INSERT INTO Z_TConnectorFunctSelects 
                    (
                        intConnectorFunctSelectID,
                        strConnectorFunctSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intConnectorFunctSelectID,
                        NEW.strConnectorFunctSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Connector Function Selection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorFunctSelects_UpdateTrigger
                AFTER UPDATE ON TConnectorFunctSelects
                BEGIN
                    INSERT INTO Z_TConnectorFunctSelects 
                    (
                        intConnectorFunctSelectID,
                        strConnectorFunctSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intConnectorFunctSelectID,
                        NEW.strConnectorFunctSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Connector Function Selection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorFunctSelects_DeleteTrigger
                AFTER DELETE ON TConnectorFunctSelects
                BEGIN
                    INSERT INTO Z_TConnectorFunctSelects 
                    (
                        intConnectorFunctSelectID,
                        strConnectorFunctSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intConnectorFunctSelectID,
                        OLD.strConnectorFunctSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Connector Functional Selection tables: {e}")                  
            
    def createConnectorFunctInspectTables(self):
        """ 
        Function Name: createConnectorFunctInspectTables
        Function Purpose: Create the Connector Function Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Connector Function Inspection Table
                CREATE TABLE IF NOT EXISTS TConnectorFunctionInspections 
                (
                    intConnectorFunctionInspectID       INTEGER NOT NULL
                    ,intConnectorID                     INTEGER NOT NULL
                    ,intInspectionTypeID                INTEGER NOT NULL
                    ,intConnectorFunctSelectID          INTEGER NOT NULL
                    ,intInspectionStatusID              INTEGER NOT NULL
                    ,strModifiedReason VARCHAR(1000)
                    ,CONSTRAINT TConnectorFunctionInspections_PK PRIMARY KEY (intConnectorFunctionInspectID)

                    ,FOREIGN KEY ( intConnectorID ) REFERENCES TConnectors ( intConnectorID )
                    ,FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )
                    ,FOREIGN KEY ( intConnectorFunctSelectID ) REFERENCES TConnectorFunctSelects ( intConnectorFunctSelectID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )                    
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Connector Function Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TConnectorFunctionInspection 
                (
                    intConnectorFunctionInspectAuditID  INTEGER NOT NULL
                    ,intConnectorFunctionInspectID      INTEGER NOT NULL
                    ,intConnectorID                     INTEGER NOT NULL
                    ,intInspectionTypeID                INTEGER NOT NULL
                    ,intConnectorFunctSelectID          INTEGER NOT NULL
                    ,intInspectionStatusID              INTEGER NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TConnectorFunctionInspection_PK PRIMARY KEY (intConnectorFunctionInspectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Connector Function Inspection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorFunctionInspection_InsertTrigger
                AFTER INSERT ON TConnectorFunctionInspections
                BEGIN
                    INSERT INTO Z_TConnectorFunctionInspection 
                    (
                        intConnectorFunctionInspectID,
                        intConnectorID,
                        intInspectionTypeID,
                        intConnectorFunctSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intConnectorFunctionInspectID,
                        NEW.intConnectorID,
                        NEW.intInspectionTypeID,
                        NEW.intConnectorFunctSelectID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Connector Function Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorFunctionInspection_UpdateTrigger
                AFTER UPDATE ON TConnectorFunctionInspections
                BEGIN
                    INSERT INTO Z_TConnectorFunctionInspection 
                    (
                        intConnectorFunctionInspectID,
                        intConnectorID,
                        intInspectionTypeID,
                        intConnectorFunctSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intConnectorFunctionInspectID,
                        NEW.intConnectorID,
                        NEW.intInspectionTypeID,
                        NEW.intConnectorFunctSelectID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Connector Function Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorFunctionInspection_DeleteTrigger
                AFTER DELETE ON TConnectorFunctionInspections
                BEGIN
                    INSERT INTO Z_TConnectorFunctionInspection 
                    (
                        intConnectorFunctionInspectID,
                        intConnectorID,
                        intInspectionTypeID,
                        intConnectorFunctSelectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intConnectorFunctionInspectID,
                        OLD.intConnectorID,
                        OLD.intInspectionTypeID,
                        OLD.intConnectorFunctSelectID,
                        OLD.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Connector Functional Inspection tables: {e}")
            
    def createStandConnectorInspectTables(self):
        """ 
        Function Name: createStandConnectorInspectTables
        Function Purpose: Create the Standard Connector Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Standard Connector Inspection Table
                CREATE TABLE IF NOT EXISTS TStandardConnectorInspections 
                (
                    intStandardConnectorInspectionID        INTEGER NOT NULL
                    ,intConnectorVisualInspectionID         INTEGER NOT NULL
                    ,intConnectorPhysicalInspectionID       INTEGER NOT NULL
                    ,intConnectorFunctionInspectID          INTEGER NOT NULL
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT TStandardConnectorInspections_PK PRIMARY KEY (intStandardConnectorInspectionID)
                    
                    ,FOREIGN KEY ( intConnectorVisualInspectionID ) REFERENCES TConnectorVisualInspections ( intConnectorVisualInspectionID )
                    ,FOREIGN KEY ( intConnectorPhysicalInspectionID ) REFERENCES TConnectorPhysicalInspections ( intConnectorPhysicalInspectionID )
                    ,FOREIGN KEY ( intConnectorFunctionInspectID ) REFERENCES TConnectorFunctionInspections ( intConnectorFunctionInspectID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Standard Connector Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TStandardConnectorInspections 
                (
                    intStandardConnectorInspectionAuditID   INTEGER NOT NULL
                    ,intStandardConnectorInspectionID       INTEGER NOT NULL
                    ,intConnectorVisualInspectionID                   INTEGER NOT NULL
                    ,intConnectorPhysicalInspectionID                 INTEGER NOT NULL
                    ,intConnectorFunctionInspectID          INTEGER NOT NULL
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TStandardConnectorInspections_PK PRIMARY KEY (intStandardConnectorInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Standard Connector Inspection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TStandardConnectorInspections_InsertTrigger
                AFTER INSERT ON TStandardConnectorInspections
                BEGIN
                    INSERT INTO Z_TStandardConnectorInspections 
                    (
                        intStandardConnectorInspectionID,
                        intConnectorVisualInspectionID,
                        intConnectorPhysicalInspectionID,
                        intConnectorFunctionInspectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intStandardConnectorInspectionID,
                        NEW.intConnectorVisualInspectionID,
                        NEW.intConnectorPhysicalInspectionID,
                        NEW.intConnectorFunctionInspectID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Standard Connector Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TStandardConnectorInspections_UpdateTrigger
                AFTER UPDATE ON TStandardConnectorInspections
                BEGIN
                    INSERT INTO Z_TStandardConnectorInspections 
                    (
                        intStandardConnectorInspectionID,
                        intConnectorVisualInspectionID,
                        intConnectorPhysicalInspectionID,
                        intConnectorFunctionInspectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intStandardConnectorInspectionID,
                        NEW.intConnectorVisualInspectionID,
                        NEW.intConnectorPhysicalInspectionID,
                        NEW.intConnectorFunctionInspectID,
                        NEW.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Standard Connector Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TStandardConnectorInspections_DeleteTrigger
                AFTER DELETE ON TStandardConnectorInspections
                BEGIN
                    INSERT INTO Z_TStandardConnectorInspections 
                    (
                        intStandardConnectorInspectionID,
                        intConnectorVisualInspectionID,
                        intConnectorPhysicalInspectionID,
                        intConnectorFunctionInspectID,
                        intInspectionStatusID,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intStandardConnectorInspectionID,
                        OLD.intConnectorVisualInspectionID,
                        OLD.intConnectorPhysicalInspectionID,
                        OLD.intConnectorFunctionInspectID,
                        OLD.intInspectionStatusID,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)     
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Standard Connector Inspection tables: {e}")

    def createConnector_InspectionTables(self):
        """ 
        Function Name: createConnector_InspectionTables
        Function Purpose: Create the Connector Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Connector Inspection Table
                CREATE TABLE IF NOT EXISTS TConnectorInspections 
                (
                    intConnectorInspectionID            INTEGER NOT NULL
                    ,intConnectorID                     INTEGER NOT NULL
                    ,intWallLocationID                  INTEGER NOT NULL
                    ,intStandardConnectorInspectionID   INTEGER NOT NULL
                    ,intInspectorID                     INTEGER NOT NULL
                    ,intInspectionStatusID              INTEGER NOT NULL
                    ,dtmLastInspectionDate              DATETIME NOT NULL
                    ,dtmNextInspectionDate              DATETIME NOT NULL
                    ,strComment                         VARCHAR(225) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TConnectorInspections_PK PRIMARY KEY (intConnectorInspectionID)
                    
                    ,FOREIGN KEY ( intConnectorID ) REFERENCES TConnectors ( intConnectorID )
                    ,FOREIGN KEY ( intWallLocationID ) REFERENCES TWallLocations ( intWallLocationID ) 
                    ,FOREIGN KEY ( intStandardConnectorInspectionID ) REFERENCES TStandardConnectorInspections ( intStandardConnectorInspectionID ) 
                    ,FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID ) 
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID ) 
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Connector Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TConnectorInspections 
                (
                    intConnectorInspectionAuditID       INTEGER NOT NULL
                    ,intConnectorInspectionID           INTEGER NOT NULL
                    ,intConnectorID                     INTEGER NOT NULL
                    ,intWallLocationID                  INTEGER NOT NULL
                    ,intStandardConnectorInspectionID   INTEGER NOT NULL
                    ,intInspectorID                     INTEGER NOT NULL
                    ,intInspectionStatusID              INTEGER NOT NULL
                    ,dtmLastInspectionDate              DATETIME NOT NULL
                    ,dtmNextInspectionDate              DATETIME NOT NULL
                    ,strComment                         VARCHAR(225) NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TConnectorInspections_PK PRIMARY KEY (intConnectorInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Connector Inspection Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorInspections_AuditTrigger_Insert
                AFTER INSERT ON TConnectorInspections
                BEGIN
                    INSERT INTO Z_TConnectorInspections 
                    (
                        intConnectorInspectionID
                        ,intConnectorID
                        ,intWallLocationID
                        ,intStandardConnectorInspectionID
                        ,intInspectorID
                        ,intInspectionStatusID
                        ,dtmLastInspectionDate
                        ,dtmNextInspectionDate
                        ,strComment
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intConnectorInspectionID
                        ,NEW.intConnectorID
                        ,NEW.intWallLocationID
                        ,NEW.intStandardConnectorInspectionID
                        ,NEW.intInspectorID
                        ,NEW.intInspectionStatusID
                        ,NEW.dtmLastInspectionDate
                        ,NEW.dtmNextInspectionDate
                        ,NEW.strComment
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Connector Inspection Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorInspections_AuditTrigger_Update
                AFTER UPDATE ON TConnectorInspections
                BEGIN
                    INSERT INTO Z_TConnectorInspections 
                    (
                        intConnectorInspectionID
                        ,intConnectorID
                        ,intWallLocationID
                        ,intStandardConnectorInspectionID
                        ,intInspectorID
                        ,intInspectionStatusID
                        ,dtmLastInspectionDate
                        ,dtmNextInspectionDate
                        ,strComment
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intConnectorInspectionID
                        ,NEW.intConnectorID
                        ,NEW.intWallLocationID
                        ,NEW.intStandardConnectorInspectionID
                        ,NEW.intInspectorID
                        ,NEW.intInspectionStatusID
                        ,NEW.dtmLastInspectionDate
                        ,NEW.dtmNextInspectionDate
                        ,NEW.strComment
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Connector Inspection Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorInspections_AuditTrigger_Delete
                AFTER DELETE ON TConnectorInspections
                BEGIN
                    INSERT INTO Z_TConnectorInspections 
                    (
                        intConnectorInspectionID
                        ,intConnectorID
                        ,intWallLocationID
                        ,intStandardConnectorInspectionID
                        ,intInspectorID
                        ,intInspectionStatusID
                        ,dtmLastInspectionDate
                        ,dtmNextInspectionDate
                        ,strComment
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intConnectorInspectionID
                        ,OLD.intConnectorID
                        ,OLD.intWallLocationID
                        ,OLD.intStandardConnectorInspectionID
                        ,OLD.intInspectorID
                        ,OLD.intInspectionStatusID
                        ,OLD.dtmLastInspectionDate
                        ,OLD.dtmNextInspectionDate
                        ,OLD.strComment
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Connector Inspection tables: {e}")            

    def createConnector_WallLocationTables(self):
        """ 
        Function Name: createConnector_WallLocationTables
        Function Purpose: Create the Connector Location Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Connector Wall Location Table
                CREATE TABLE IF NOT EXISTS TConnectorWallLocations 
                (
                    intConnectorWallLocationID          INTEGER NOT NULL
                    ,intWallLocationID                  INTEGER NOT NULL
                    ,intConnectorID                     INTEGER NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,FOREIGN KEY ( intWallLocationID ) REFERENCES TWallLocations ( intWallLocationID )
                    ,FOREIGN KEY ( intConnectorID ) REFERENCES TConnectors ( intConnectorID )
                    
                    ,CONSTRAINT TConnectorWallLocations_PK PRIMARY KEY (intConnectorWallLocationID)
                    ,CONSTRAINT TConnectorWallLocations_UQ UNIQUE (intWallLocationID, intConnectorID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Connector Wall Location Table
                CREATE TABLE IF NOT EXISTS Z_TConnectorWallLocations 
                (
                    intConnectorWallLocationAuditID     INTEGER NOT NULL
                    ,intConnectorWallLocationID         INTEGER NOT NULL
                    ,intWallLocationID                  INTEGER NOT NULL
                    ,intConnectorID                     INTEGER NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TConnectorWallLocations_PK PRIMARY KEY (intConnectorWallLocationAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Connector Wall Location Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorWallLocations_AuditTrigger_Insert
                AFTER INSERT ON TConnectorWallLocations
                BEGIN
                    INSERT INTO Z_TConnectorWallLocations 
                    (
                        intConnectorWallLocationID
                        ,intWallLocationID
                        ,intConnectorID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intConnectorWallLocationID
                        ,NEW.intWallLocationID
                        ,NEW.intConnectorID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Connector Wall Location Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorWallLocations_AuditTrigger_Update
                AFTER UPDATE ON TConnectorWallLocations
                BEGIN
                    INSERT INTO Z_TConnectorWallLocations 
                    (
                        intConnectorWallLocationID
                        ,intWallLocationID
                        ,intConnectorID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intConnectorWallLocationID
                        ,NEW.intWallLocationID
                        ,NEW.intConnectorID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Connector Wall Location Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorWallLocations_AuditTrigger_Delete
                AFTER DELETE ON TConnectorWallLocations
                BEGIN
                    INSERT INTO Z_TConnectorWallLocations 
                    (
                        intConnectorWallLocationID
                        ,intWallLocationID
                        ,intConnectorID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intConnectorWallLocationID
                        ,OLD.intWallLocationID
                        ,OLD.intConnectorID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Connector Wall Location tables: {e}")            

    def createConnector_InspectorTables(self):
        """ 
        Function Name: createConnector_InspectorTables
        Function Purpose: Create the Connector Inspector Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Connector Inspector Table
                CREATE TABLE IF NOT EXISTS TConnectorInspectors 
                (
                    intConnectorInspectorID         INTEGER NOT NULL
                    ,intInspectorID                 INTEGER NOT NULL
                    ,intConnectorID                 INTEGER NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT TConnectorInspectors_PK PRIMARY KEY (intConnectorInspectorID)
                    
                    ,FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID )
                    ,FOREIGN KEY ( intConnectorID ) REFERENCES TConnectors ( intConnectorID )
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Connector Inspector Table
                CREATE TABLE IF NOT EXISTS Z_TConnectorInspectors 
                (
                    intConnectorInspectorAuditID    INTEGER NOT NULL
                    ,intConnectorInspectorID        INTEGER NOT NULL
                    ,intInspectorID                 INTEGER NOT NULL
                    ,intConnectorID                 INTEGER NOT NULL
                    ,strUpdatedBy                   VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                   DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                      VARCHAR(1) NOT NULL
                    ,strModifiedReason              VARCHAR(1000)
                    ,CONSTRAINT Z_TConnectorInspectors_PK PRIMARY KEY (intConnectorInspectorAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Connector Inspector Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorInspectors_AuditTrigger_Insert
                AFTER INSERT ON TConnectorInspectors
                BEGIN
                    INSERT INTO Z_TConnectorInspectors 
                    (
                        intConnectorInspectorID
                        ,intInspectorID
                        ,intConnectorID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intConnectorInspectorID
                        ,NEW.intInspectorID
                        ,NEW.intConnectorID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Connector Inspector Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorInspectors_AuditTrigger_Update
                AFTER UPDATE ON TConnectorInspectors
                BEGIN
                    INSERT INTO Z_TConnectorInspectors 
                    (
                        intConnectorInspectorID
                        ,intInspectorID
                        ,intConnectorID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intConnectorInspectorID
                        ,NEW.intInspectorID
                        ,NEW.intConnectorID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Connector Inspector Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorInspectors_AuditTrigger_Delete
                AFTER DELETE ON TConnectorInspectors
                BEGIN
                    INSERT INTO Z_TConnectorInspectors 
                    (
                        intConnectorInspectorID
                        ,intInspectorID
                        ,intConnectorID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intConnectorInspectorID
                        ,OLD.intInspectorID
                        ,OLD.intConnectorID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Connector Inspector tables: {e}")            

    def createConnector_RetiredReports_Tables(self):
        """ 
        Function Name: createConnector_RetiredReports_Tables
        Function Purpose: Create the Connector Retired Reports Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Connector Inspection Table
                CREATE TABLE IF NOT EXISTS TConnectorRetiredReports 
                (
                    intConnectorRetiredReportID             INTEGER NOT NULL
                    ,intConnectorInspectionID               INTEGER NOT NULL
                    ,intInspectorID                         INTEGER NOT NULL
                    ,dtmReportDate                          DATETIME NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT TConnectorRetiredReports_PK PRIMARY KEY (intConnectorRetiredReportID)
                    
                    ,FOREIGN KEY ( intConnectorInspectionID ) REFERENCES TConnectorInspections ( intConnectorInspectionID )
                    ,FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID ) 
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Connector Retired Report Table
                CREATE TABLE IF NOT EXISTS Z_TConnectorRetiredReports 
                (
                    intConnectorRetiredReportAuditID        INTEGER NOT NULL
                    ,intConnectorRetiredReportID            INTEGER NOT NULL
                    ,intConnectorInspectionID               INTEGER NOT NULL
                    ,intInspectorID                         INTEGER NOT NULL
                    ,dtmReportDate                          DATETIME NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TConnectorRetiredReports_PK PRIMARY KEY (intConnectorRetiredReportAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Connector Retired Report Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorRetiredReports_AuditTrigger_Insert
                AFTER INSERT ON TConnectorRetiredReports
                BEGIN
                    INSERT INTO Z_TConnectorRetiredReports 
                    (
                        intConnectorRetiredReportID
                        ,intConnectorInspectionID
                        ,intInspectorID
                        ,dtmReportDate
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intConnectorRetiredReportID
                        ,NEW.intConnectorInspectionID
                        ,NEW.intInspectorID
                        ,NEW.dtmReportDate
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Connector Retired Report Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorRetiredReports_AuditTrigger_Update
                AFTER UPDATE ON TConnectorRetiredReports
                BEGIN
                    INSERT INTO Z_TConnectorRetiredReports 
                    (
                        intConnectorRetiredReportID
                        ,intConnectorInspectionID
                        ,intInspectorID
                        ,dtmReportDate
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intConnectorRetiredReportID
                        ,NEW.intConnectorInspectionID
                        ,NEW.intInspectorID
                        ,NEW.dtmReportDate
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Connector Inspection Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TConnectorRetiredReports_AuditTrigger_Delete
                AFTER DELETE ON TConnectorRetiredReports
                BEGIN
                    INSERT INTO Z_TConnectorRetiredReports 
                    (
                        intConnectorRetiredReportID
                        ,intConnectorInspectionID
                        ,intInspectorID
                        ,dtmReportDate
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intConnectorRetiredReportID
                        ,OLD.intConnectorInspectionID
                        ,OLD.intInspectorID
                        ,OLD.dtmReportDate
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Connector Inspection tables: {e}")
                        
    def createConnector_ForeignKeys(self):
        """ 
        Function Name: createConnector_ForeignKeys
        Function Purpose: Create the Connector Visual, Physical, Standard, and Inspection foreign keys inside the database
        """           
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the foreign keys
            sqlForeignKeys = """
                -- -------------------------------------------------------------------------------------------------------
                -- Establish Referential Integrity - Inserts
                -- -------------------------------------------------------------------------------------------------------

                -- #           Child                           Parent                                Column
                -- --         -------                         --------                              --------
                -- 1. TConnectorVisualInspections             TConnectors                           intConnectorID
                -- 2. TConnectorVisualInspections             TInspectionTypes                      intInspectionTypeID
                -- 3. TConnectorVisualInspections             TConnectorVisMetalSelects             intConnectorVisMetalSelectID
                -- 4. TConnectorVisualInspections             TInspectionStatus                     intInspectionStatusID
                -- 5. TConnectorPhysicalInspections           TConnectors                           intConnectorID
                -- 6. TConnectorPhysicalInspections           TInspectionTypes                      intInspectionTypeID
                -- 7. TConnectorPhysicalInspections           TConnectorPhysMetalSelects            intConnectorPhysMetalSelectID
                -- 8. TConnectorPhysicalInspections           TInspectionStatus                     intInspectionStatusID
                -- 9. TConnectorFunctionInspections           TConnectors                           intConnectorID
                -- 10. TConnectorFunctionInspections          TInspectionTypes                      intInspectionTypeID
                -- 11. TConnectorFunctionInspections          TConnectorFunctSelects                intConnectorFunctSelectID
                -- 12. TConnectorFunctionInspections          TInspectionStatus                     intInspectionStatusID
                -- 13. TStandardConnectorInspections          TConnectorVisualInspections           intConnectorVisualInspectionID
                -- 14. TStandardConnectorInspections          TConnectorPhysicalInspections         intConnectorPhysicalInspectionID
                -- 15. TStandardConnectorInspections          TConnectorFunctionInspections         intConnectorFunctionInspectID
                -- 16. TStandardConnectorInspections          TInspectionStatus                     intInspectionStatusID
                -- 17. TConnectorInspections                  TConnectors                           intConnectorID
                -- 18. TConnectorInspections                  TWallLocations                        intWallLocationID
                -- 19. TConnectorInspections                  TStandardConnectorInspections         intStandardConnectorInspectionID
                -- 20. TConnectorInspections                  TInspectors                           intInspectorID
                -- 21. TConnectorInspections                  TInspectionStatus                     intInspectionStatusID
                -- 22. TConnectorRetiredReports               TConnectorInspections                 intConnectorInspectionID
                -- 23. TConnectorRetiredReports               TInspectors                           intInspectorID

                -- -------------------------------------------------------------------------------------------------------
                -- Create Foreign Keys
                -- -------------------------------------------------------------------------------------------------------
                -- 1.
                ALTER TABLE TConnectorVisualInspections ADD CONSTRAINT TConnectorVisualInspections_TConnectors_FK
                FOREIGN KEY ( intConnectorID ) REFERENCES TConnectors ( intConnectorID )

                -- 2.
                ALTER TABLE TConnectorVisualInspections ADD CONSTRAINT TConnectorVisualInspections_TInspectionTypes_FK
                FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )

                -- 3.
                ALTER TABLE TConnectorVisualInspections ADD CONSTRAINT TConnectorVisualInspections_TConnectorVisMetalSelects_FK
                FOREIGN KEY ( intConnectorVisMetalSelectID ) REFERENCES TConnectorVisMetalSelects ( intConnectorVisMetalSelectID )

                -- 4.
                ALTER TABLE TConnectorVisualInspections ADD CONSTRAINT TConnectorVisualInspections_TInspectionStatus_FK
                FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )

                -- -------------------------------------------------------------------------------------------------------
                -- Ensure No Duplicated Data
                -- -------------------------------------------------------------------------------------------------------
                ALTER TABLE TConnectorWallLocations ADD CONSTRAINT TConnectorWallLocations_UNIQUE
                UNIQUE (intWallLocationID, intConnectorID)

                -- -------------------------------------------------------------------------------------------------------
                -- Create Foreign Keys
                -- -------------------------------------------------------------------------------------------------------
                -- 5.
                ALTER TABLE TConnectorPhysicalInspections ADD CONSTRAINT TConnectorPhysicalInspections_TConnectors_FK
                FOREIGN KEY ( intConnectorID ) REFERENCES TConnectors ( intConnectorID )

                -- 6.
                ALTER TABLE TConnectorPhysicalInspections ADD CONSTRAINT TConnectorPhysicalInspections_TInspectionTypes_FK
                FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )

                -- 7.
                ALTER TABLE TConnectorPhysicalInspections ADD CONSTRAINT TConnectorPhysicalInspections_TConnectorPhysMetalSelects_FK
                FOREIGN KEY ( intConnectorPhysMetalSelectID ) REFERENCES TConnectorPhysMetalSelects ( intConnectorPhysMetalSelectID )

                -- 8. 
                ALTER TABLE TConnectorPhysicalInspections ADD CONSTRAINT TConnectorPhysicalInspections_TInspectionStatus_FK
                FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )

                -- 9.
                ALTER TABLE TConnectorFunctionInspections ADD CONSTRAINT TConnectorFunctionInspections_TConnectors_FK
                FOREIGN KEY ( intConnectorID ) REFERENCES TConnectors ( intConnectorID )

                -- 10.
                ALTER TABLE TConnectorFunctionInspections ADD CONSTRAINT TConnectorFunctionInspections_TInspectionTypes_FK
                FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )

                -- 11.
                ALTER TABLE TConnectorFunctionInspections ADD CONSTRAINT TConnectorFunctionInspections_TConnectorFunctSelects_FK
                FOREIGN KEY ( intConnectorFunctSelectID ) REFERENCES TConnectorFunctSelects ( intConnectorFunctSelectID )

                -- 12. 
                ALTER TABLE TConnectorFunctionInspections ADD CONSTRAINT TConnectorFunctionInspections_TInspectionStatus_FK
                FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )
                
                -- 13. 
                ALTER TABLE TStandardConnectorInspections ADD CONSTRAINT TStandardConnectorInspections_TConnectorVisualInspections_FK
                FOREIGN KEY ( intConnectorVisualInspectionID ) REFERENCES TConnectorVisualInspections ( intConnectorVisualInspectionID )

                -- 14.
                ALTER TABLE TStandardConnectorInspections ADD CONSTRAINT TStandardConnectorInspections_TConnectorPhysicalInspections_FK
                FOREIGN KEY ( intConnectorPhysicalInspectionID ) REFERENCES TConnectorPhysicalInspections ( intConnectorPhysicalInspectionID )

                -- 15.
                ALTER TABLE TStandardConnectorInspections ADD CONSTRAINT TStandardConnectorInspections_TConnectorFunctionInspections_FK
                FOREIGN KEY ( intConnectorFunctionInspectID ) REFERENCES TConnectorFunctionInspections ( intConnectorFunctionInspectID )

                -- 16.
                ALTER TABLE TStandardConnectorInspections ADD CONSTRAINT TStandardConnectorInspections_TInspectionStatus_FK
                FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )

                -- 17.
                ALTER TABLE TConnectorInspections ADD CONSTRAINT TConnectorInspections_TConnectors_FK
                FOREIGN KEY ( intConnectorID ) REFERENCES TConnectors ( intConnectorID )

                -- 18.
                ALTER TABLE TConnectorInspections ADD CONSTRAINT TConnectorInspections_TWallLocations_FK
                FOREIGN KEY ( intWallLocationID ) REFERENCES TWallLocations ( intWallLocationID )

                -- 19.
                ALTER TABLE TConnectorInspections ADD CONSTRAINT TConnectorInspections_TStandardConnectorInspections_FK
                FOREIGN KEY ( intStandardConnectorInspectionID ) REFERENCES TStandardConnectorInspections ( intStandardConnectorInspectionID )
                
                -- 20.
                ALTER TABLE TConnectorInspections ADD CONSTRAINT TConnectorInspections_TInspectors_FK
                FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID )

                -- 21.
                ALTER TABLE TConnectorInspections ADD CONSTRAINT TConnectorInspections_TInspectionStatus_FK
                FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )          

                -- 22.
                ALTER TABLE TConnectorRetiredReports ADD CONSTRAINT TConnectorRetiredReports_TConnectorInspections_FK
                FOREIGN KEY ( intConnectorInspectionID ) REFERENCES TConnectorInspections ( intConnectorInspectionID )

                -- 23.
                ALTER TABLE TConnectorRetiredReports ADD CONSTRAINT TConnectorRetiredReports_TInspectors_FK
                FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID )
                """            
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlForeignKeys)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Foreign Keys: {e}") 

    def createBelayDevicesTables(self):
        """ 
        Function Name: createBelayDevicesTables
        Function Purpose: Create the BelayDevices Tables and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create BelayDevices Table
                CREATE TABLE IF NOT EXISTS TBelayDevices 
                (
                    intBelayDeviceID                    INTEGER NOT NULL
                    ,strBelayDeviceName                 VARCHAR(255) NOT NULL
                    ,strSerialNum                       VARCHAR(255) NOT NULL
                    ,strBumperNum                       VARCHAR(255) 
                    ,strManufactureName                 VARCHAR(255) NOT NULL
                    ,dtmManufactureDate                 DATETIME 
                    ,dtmInstallationDate                DATETIME NOT NULL   
                    ,dtmLastInspectionDate              DATETIME NOT NULL   
                    ,dtmNextInspectionDate              DATETIME NOT NULL       
                    ,strDeviceType                      VARCHAR(255) NOT NULL
                    ,strEquipInUse                      VARCHAR(255) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TBelayDevices_PK PRIMARY KEY (intBelayDeviceID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: BelayDevices Table
                CREATE TABLE IF NOT EXISTS Z_TBelayDevices 
                (
                    intBelayDeviceAuditID               INTEGER NOT NULL
                    ,intBelayDeviceID                   INTEGER NOT NULL
                    ,strBelayDeviceName                 VARCHAR(255) NOT NULL
                    ,strSerialNum                       VARCHAR(255) NOT NULL
                    ,strBumperNum                       VARCHAR(255) 
                    ,strManufactureName                 VARCHAR(255) NOT NULL
                    ,dtmManufactureDate                 DATETIME 
                    ,dtmInstallationDate                DATETIME NOT NULL   
                    ,dtmLastInspectionDate              DATETIME NOT NULL   
                    ,dtmNextInspectionDate              DATETIME NOT NULL       
                    ,strDeviceType                      VARCHAR(255) NOT NULL
                    ,strEquipInUse                      VARCHAR(255) NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TBelayDevices_PK PRIMARY KEY (intBelayDeviceAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: BelayDevices Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDevices_AuditTrigger_Insert
                AFTER INSERT ON TBelayDevices
                BEGIN
                    INSERT INTO Z_TBelayDevices 
                    (
                        intBelayDeviceID             
                        ,strBelayDeviceName           
                        ,strSerialNum
                        ,strBumperNum
                        ,strManufactureName
                        ,dtmManufactureDate
                        ,dtmInstallationDate
                        ,dtmLastInspectionDate
                        ,dtmNextInspectionDate
                        ,strDeviceType
                        ,strEquipInUse
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDeviceID       
                        ,NEW.strBelayDeviceName                 
                        ,NEW.strSerialNum
                        ,NEW.strBumperNum
                        ,NEW.strManufactureName
                        ,NEW.dtmManufactureDate
                        ,NEW.dtmInstallationDate
                        ,NEW.dtmLastInspectionDate
                        ,NEW.dtmNextInspectionDate
                        ,NEW.strDeviceType
                        ,NEW.strEquipInUse
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDevices_AuditTrigger_Update
                AFTER UPDATE ON TBelayDevices
                BEGIN
                    INSERT INTO Z_TBelayDevices 
                    (
                        intBelayDeviceID             
                        ,strBelayDeviceName           
                        ,strSerialNum
                        ,strBumperNum
                        ,strManufactureName
                        ,dtmManufactureDate
                        ,dtmInstallationDate
                        ,dtmLastInspectionDate
                        ,dtmNextInspectionDate
                        ,strDeviceType
                        ,strEquipInUse
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDeviceID       
                        ,NEW.strBelayDeviceName                 
                        ,NEW.strSerialNum
                        ,NEW.strBumperNum
                        ,NEW.strManufactureName
                        ,NEW.dtmManufactureDate
                        ,NEW.dtmInstallationDate
                        ,NEW.dtmLastInspectionDate
                        ,NEW.dtmNextInspectionDate
                        ,NEW.strDeviceType
                        ,NEW.strEquipInUse
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDevices_AuditTrigger_Delete
                AFTER DELETE ON TBelayDevices
                BEGIN
                    INSERT INTO Z_TBelayDevices 
                    (
                        intBelayDeviceID             
                        ,strBelayDeviceName           
                        ,strSerialNum
                        ,strBumperNum
                        ,strManufactureName
                        ,dtmManufactureDate
                        ,dtmInstallationDate
                        ,dtmLastInspectionDate
                        ,dtmNextInspectionDate
                        ,strDeviceType
                        ,strEquipInUse
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intBelayDeviceID                        
                        ,OLD.strBelayDeviceName  
                        ,OLD.strSerialNum
                        ,OLD.strBumperNum
                        ,OLD.strManufactureName
                        ,OLD.dtmManufactureDate
                        ,OLD.dtmInstallationDate
                        ,OLD.dtmLastInspectionDate
                        ,OLD.dtmNextInspectionDate
                        ,OLD.strDeviceType
                        ,OLD.strEquipInUse
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating BelayDevices tables: {e}")

    def createBelayDevicesVisMetalSelTables(self):
        """ 
        Function Name: createBelayDevicesVisMetalSelTables
        Function Purpose: Create the BelayDevices Visual Metallic Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create BelayDevices Visual Metallic Selection Table
                CREATE TABLE IF NOT EXISTS TBelayDeviceVisMetalSelects 
                (
                    intBelayDeviceVisMetalSelectID          INTEGER NOT NULL
                    ,strBelayDeviceVisMetSelect             VARCHAR(1000) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT TBelayDeviceVisMetalSelects_PK PRIMARY KEY (intBelayDeviceVisMetalSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: BelayDevices Visual Metallic Selection Table
                CREATE TABLE IF NOT EXISTS Z_TBelayDeviceVisMetalSelects 
                (
                    intBelayDeviceVisMetalSelectAuditID     INTEGER NOT NULL
                    ,intBelayDeviceVisMetalSelectID         INTEGER NOT NULL
                    ,strBelayDeviceVisMetSelect           VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TBelayDeviceVisMetalSelects_PK PRIMARY KEY (intBelayDeviceVisMetalSelectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: BelayDevices Visual Metallic Selection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceVisMetalSelects_InsertTrigger
                AFTER INSERT ON TBelayDeviceVisMetalSelects
                BEGIN
                    INSERT INTO Z_TBelayDeviceVisMetalSelects 
                    (
                        intBelayDeviceVisMetalSelectID,
                        strBelayDeviceVisMetSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDeviceVisMetalSelectID,
                        NEW.strBelayDeviceVisMetSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Visual Metallic Selection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceVisMetalSelects_UpdateTrigger
                AFTER UPDATE ON TBelayDeviceVisMetalSelects
                BEGIN
                    INSERT INTO Z_TBelayDeviceVisMetalSelects 
                    (
                        intBelayDeviceVisMetalSelectID,
                        strBelayDeviceVisMetSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDeviceVisMetalSelectID,
                        NEW.strBelayDeviceVisMetSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Visual Metallic Selection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceVisMetalSelects_DeleteTrigger
                AFTER DELETE ON TBelayDeviceVisMetalSelects
                BEGIN
                    INSERT INTO Z_TBelayDeviceVisMetalSelects 
                    (
                        intBelayDeviceVisMetalSelectID,
                        strBelayDeviceVisMetSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intBelayDeviceVisMetalSelectID,
                        OLD.strBelayDeviceVisMetSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating BelayDevices Visual Metallic Selection tables: {e}")             

    def createBelayDevicesVisInspectTables(self):
        """ 
        Function Name: createBelayDevicesVisInspectTables
        Function Purpose: Create the BelayDevices Visual Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create BelayDevices Visual Inspection Table
                CREATE TABLE IF NOT EXISTS TBelayDeviceVisualInspections 
                (
                    intBelayDeviceVisualInspectionID        INTEGER NOT NULL
                    ,intBelayDeviceID                       INTEGER NOT NULL
                    ,intInspectionTypeID                    INTEGER NOT NULL
                    ,intBelayDeviceVisMetalSelectID         INTEGER NOT NULL
                    ,intBelayDeviceVisPlasticSelectID       INTEGER NOT NULL
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT TBelayDeviceVisualInspections_PK PRIMARY KEY (intBelayDeviceVisualInspectionID)
                    
                    ,FOREIGN KEY ( intBelayDeviceID ) REFERENCES TBelayDevices ( intBelayDeviceID  )
                    ,FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )
                    ,FOREIGN KEY ( intBelayDeviceVisMetalSelectID ) REFERENCES TBelayDeviceVisMetalSelects ( intBelayDeviceVisMetalSelectID )
                    ,FOREIGN KEY ( intBelayDeviceVisPlasticSelectID ) REFERENCES TBelayDeviceVisPlasticSelects ( intBelayDeviceVisPlasticSelectID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )                     
                );    
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: BelayDevices Visual Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TBelayDeviceVisualInspections
                (
                    intBelayDevicesVisualInspectionAuditID  INTEGER NOT NULL
                    ,intBelayDeviceVisualInspectionID       INTEGER NOT NULL
                    ,intBelayDeviceID                       INTEGER NOT NULL
                    ,intInspectionTypeID                    INTEGER NOT NULL
                    ,intBelayDeviceVisMetalSelectID         INTEGER NOT NULL
                    ,intBelayDeviceVisPlasticSelectID       INTEGER NOT NULL
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TBelayDeviceVisualInspections_PK PRIMARY KEY (intBelayDevicesVisualInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: BelayDevices Visual Inspection Table
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceVisualInspections_InsertTrigger
                AFTER INSERT ON TBelayDeviceVisualInspections
                BEGIN
                    INSERT INTO Z_TBelayDeviceVisualInspections 
                    (
                        intBelayDeviceVisualInspectionID
                        ,intBelayDeviceID
                        ,intInspectionTypeID
                        ,intBelayDeviceVisMetalSelectID
                        ,intBelayDeviceVisPlasticSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDeviceVisualInspectionID
                        ,NEW.intBelayDeviceID
                        ,NEW.intInspectionTypeID
                        ,NEW.intBelayDeviceVisMetalSelectID
                        ,NEW.intBelayDeviceVisPlasticSelectID
                        ,NEW.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Visual Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceVisualInspections_UpdateTrigger
                AFTER UPDATE ON TBelayDeviceVisualInspections
                BEGIN
                    INSERT INTO Z_TBelayDeviceVisualInspections 
                    (
                        intBelayDeviceVisualInspectionID
                        ,intBelayDeviceID
                        ,intInspectionTypeID
                        ,intBelayDeviceVisMetalSelectID
                        ,intBelayDeviceVisPlasticSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDeviceVisualInspectionID
                        ,NEW.intBelayDeviceID
                        ,NEW.intInspectionTypeID
                        ,NEW.intBelayDeviceVisMetalSelectID
                        ,NEW.intBelayDeviceVisPlasticSelectID
                        ,NEW.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Visual Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceVisualInspections_DeleteTrigger
                AFTER DELETE ON TBelayDeviceVisualInspections
                BEGIN
                    INSERT INTO Z_TBelayDeviceVisualInspections 
                    (
                        intBelayDeviceVisualInspectionID
                        ,intBelayDeviceID
                        ,intInspectionTypeID
                        ,intBelayDeviceVisMetalSelectID
                        ,intBelayDeviceVisPlasticSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intBelayDeviceVisualInspectionID
                        ,OLD.intBelayDeviceID
                        ,OLD.intInspectionTypeID
                        ,OLD.intBelayDeviceVisMetalSelectID
                        ,OLD.intBelayDeviceVisPlasticSelectID
                        ,OLD.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,datetime('now')
                        ,'D' -- Delete
                        ,NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating BelayDevices Visual Inspection tables: {e}")            

    def createBelayDevicesPhysMetalSelTables(self):
        """ 
        Function Name: createBelayDevicesPhysMetalSelTables
        Function Purpose: Create the BelayDevices Physical Metallic Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create BelayDevices Physical Metallic Selection Table
                CREATE TABLE IF NOT EXISTS TBelayDevicePhysMetalSelects 
                (
                    intBelayDevicePhysMetalSelectID         INTEGER NOT NULL
                    ,strBelayDevicePhysMetSelect            VARCHAR(1000) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT TBelayDevicePhysMetalSelects_PK PRIMARY KEY (intBelayDevicePhysMetalSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: BelayDevices Physical Metallic Selection Table
                CREATE TABLE IF NOT EXISTS Z_TBelayDevicePhysMetalSelects 
                (
                    intBelayDevicesPhysMetalSelectAuditID   INTEGER NOT NULL
                    ,intBelayDevicePhysMetalSelectID        INTEGER NOT NULL
                    ,strBelayDevicePhysMetSelect            VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TBelayDevicePhysMetalSelects_PK PRIMARY KEY (intBelayDevicesPhysMetalSelectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: BelayDevices Physical Metallic Selection Table
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDevicePhysMetalSelects_InsertTrigger
                AFTER INSERT ON TBelayDevicePhysMetalSelects
                BEGIN
                    INSERT INTO Z_TBelayDevicePhysMetalSelects 
                    (
                        intBelayDevicePhysMetalSelectID 
                        ,strBelayDevicePhysMetSelect   
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDevicePhysMetalSelectID 
                        ,NEW.strBelayDevicePhysMetSelect   
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Physical Metallic Selection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDevicePhysMetalSelects_UpdateTrigger
                AFTER UPDATE ON TBelayDevicePhysMetalSelects
                BEGIN
                    INSERT INTO Z_TBelayDevicePhysMetalSelects 
                    (
                        intBelayDevicePhysMetalSelectID 
                        ,strBelayDevicePhysMetSelect   
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDevicePhysMetalSelectID 
                        ,NEW.strBelayDevicePhysMetSelect   
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Physical Metallic Selection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDevicePhysMetalSelects_DeleteTrigger
                AFTER DELETE ON TBelayDevicePhysMetalSelects
                BEGIN
                    INSERT INTO Z_TBelayDevicePhysMetalSelects 
                    (
                        intBelayDevicePhysMetalSelectID 
                        ,strBelayDevicePhysMetSelect   
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intBelayDevicePhysMetalSelectID 
                        ,OLD.strBelayDevicePhysMetSelect   
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating BelayDevices Physical Metallic Selection tables: {e}")            

    def createBelayDevicesPhysInspectTables(self):
        """ 
        Function Name: createBelayDevicesPhysInspectTables
        Function Purpose: Create the BelayDevices Physical Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create BelayDevices Physical Inspection Table
                CREATE TABLE IF NOT EXISTS TBelayDevicePhysicalInspections 
                (
                    intBelayDevicePhysicalInspectionID          INTEGER NOT NULL
                    ,intBelayDeviceID                           INTEGER NOT NULL
                    ,intInspectionTypeID                        INTEGER NOT NULL
                    ,intBelayDevicePhysMetalSelectID            INTEGER NOT NULL
                    ,intBelayDevicePhysPlasticSelectID          INTEGER NOT NULL
                    ,intInspectionStatusID                      INTEGER NOT NULL
                    ,strModifiedReason                          VARCHAR(1000)
                    ,CONSTRAINT TBelayDevicePhysicalInspections_PK PRIMARY KEY (intBelayDevicePhysicalInspectionID)
                    
                    ,FOREIGN KEY ( intBelayDeviceID ) REFERENCES TBelayDevices ( intBelayDeviceID )
                    ,FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )
                    ,FOREIGN KEY ( intBelayDevicePhysMetalSelectID  ) REFERENCES TBelayDevicePhysMetalSelects ( intBelayDevicePhysMetalSelectID )
                    ,FOREIGN KEY ( intBelayDevicePhysPlasticSelectID  ) REFERENCES TBelayDevicePhysPlasticSelects ( intBelayDevicePhysPlasticSelectID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )                     
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: BelayDevices Physical Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TBelayDevicePhysicalInspections 
                (
                    intBelayDevicesPhysicalInspectionAuditID    INTEGER NOT NULL
                    ,intBelayDevicePhysicalInspectionID         INTEGER NOT NULL
                    ,intBelayDeviceID                           INTEGER NOT NULL
                    ,intInspectionTypeID                        INTEGER NOT NULL
                    ,intBelayDevicePhysMetalSelectID            INTEGER NOT NULL
                    ,intBelayDevicePhysPlasticSelectID          INTEGER NOT NULL
                    ,intInspectionStatusID                      INTEGER NOT NULL
                    ,strUpdatedBy                               VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                               DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                                  VARCHAR(1) NOT NULL
                    ,strModifiedReason                          VARCHAR(1000)
                    ,CONSTRAINT Z_TBelayDevicePhysicalInspections_PK PRIMARY KEY (intBelayDevicesPhysicalInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: BelayDevices Physical Inspection Table
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDevicePhysicalInspections_InsertTrigger
                AFTER INSERT ON TBelayDevicePhysicalInspections
                BEGIN
                    INSERT INTO Z_TBelayDevicePhysicalInspections 
                    (
                        intBelayDevicePhysicalInspectionID
                        ,intBelayDeviceID
                        ,intInspectionTypeID
                        ,intBelayDevicePhysMetalSelectID 
                        ,intBelayDevicePhysPlasticSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDevicePhysicalInspectionID
                        ,NEW.intBelayDeviceID
                        ,NEW.intInspectionTypeID
                        ,NEW.intBelayDevicePhysMetalSelectID
                        ,NEW.intBelayDevicePhysPlasticSelectID 
                        ,NEW.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- INSERT
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Physical Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDevicePhysicalInspections_UpdateTrigger
                AFTER UPDATE ON TBelayDevicePhysicalInspections
                BEGIN
                    INSERT INTO Z_TBelayDevicePhysicalInspections 
                    (
                        intBelayDevicePhysicalInspectionID
                        ,intBelayDeviceID
                        ,intInspectionTypeID
                        ,intBelayDevicePhysMetalSelectID 
                        ,intBelayDevicePhysPlasticSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDevicePhysicalInspectionID
                        ,NEW.intBelayDeviceID
                        ,NEW.intInspectionTypeID
                        ,NEW.intBelayDevicePhysMetalSelectID 
                        ,NEW.intBelayDevicePhysPlasticSelectID
                        ,NEW.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Physical Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDevicePhysicalInspections_DeleteTrigger
                AFTER DELETE ON TBelayDevicePhysicalInspections
                BEGIN
                    INSERT INTO Z_TBelayDevicePhysicalInspections 
                    (
                        intBelayDevicePhysicalInspectionID
                        ,intBelayDeviceID
                        ,intInspectionTypeID
                        ,intBelayDevicePhysMetalSelectID 
                        ,intBelayDevicePhysPlasticSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intBelayDevicePhysicalInspectionID
                        ,OLD.intBelayDeviceID
                        ,OLD.intInspectionTypeID
                        ,OLD.intBelayDevicePhysMetalSelectID 
                        ,OLD.intBelayDevicePhysPlasticSelectID
                        ,OLD.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating BelayDevices Physical Inspection tables: {e}")            

    def createBelayDevicesFunctionsTables(self):
        """ 
        Function Name: createBelayDevicesFunctionsTables
        Function Purpose: Create the BelayDevices Functions and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
                    
            # Create the table
            sqlTable = """
                -- Create BelayDevices Function Table
                CREATE TABLE IF NOT EXISTS TBelayDeviceFunctions 
                (
                    intBelayDeviceFunctionID                INTEGER NOT NULL
                    ,strBelayDeviceFunctionDesc             VARCHAR(255) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT TBelayDeviceFunctions_PK PRIMARY KEY (intBelayDeviceFunctionID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: BelayDevices Function Table
                CREATE TABLE IF NOT EXISTS Z_TBelayDeviceFunctions 
                (
                    intBelayDeviceFunctionAuditID           INTEGER NOT NULL
                    ,intBelayDeviceFunctionID               INTEGER NOT NULL
                    ,strBelayDeviceFunctionDesc             VARCHAR(255) NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TBelayDeviceFunctions_PK PRIMARY KEY (intBelayDeviceFunctionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: BelayDevices Function Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceFunctions_InsertTrigger
                AFTER INSERT ON TBelayDeviceFunctions
                BEGIN
                    INSERT INTO Z_TBelayDeviceFunctions 
                    (
                        intBelayDeviceFunctionID,
                        strBelayDeviceFunctionDesc,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDeviceFunctionID,
                        NEW.strBelayDeviceFunctionDesc,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Function Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceFunctions_UpdateTrigger
                AFTER UPDATE ON TBelayDeviceFunctions
                BEGIN
                    INSERT INTO Z_TBelayDeviceFunctions 
                    (
                        intBelayDeviceFunctionID,
                        strBelayDeviceFunctionDesc,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDeviceFunctionID,
                        NEW.strBelayDeviceFunctionDesc,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Function Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceFunctions_DeleteTrigger
                AFTER DELETE ON TBelayDeviceFunctions
                BEGIN
                    INSERT INTO Z_TBelayDeviceFunctions 
                    (
                        intBelayDeviceFunctionID,
                        strBelayDeviceFunctionDesc,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intBelayDeviceFunctionID,
                        OLD.strBelayDeviceFunctionDesc,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)     
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating BelayDevices Functional tables: {e}")              
            
    def createBelayDevicesFunctSelTables(self):
        """ 
        Function Name: createBelayDevicesFunctSelTables
        Function Purpose: Create the BelayDevices Function Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create BelayDevices Function Selection Table
                CREATE TABLE IF NOT EXISTS TBelayDeviceFunctSelects 
                (
                    intBelayDeviceFunctSelectID             INTEGER NOT NULL
                    ,strBelayDeviceFunctSelect              VARCHAR(1000) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT TBelayDeviceFunctSelects_PK PRIMARY KEY (intBelayDeviceFunctSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: BelayDevices Function Selection Table
                CREATE TABLE IF NOT EXISTS Z_TBelayDeviceFunctSelects 
                (
                    intBelayDeviceFunctSelectAuditID        INTEGER NOT NULL
                    ,intBelayDeviceFunctSelectID            INTEGER NOT NULL
                    ,strBelayDeviceFunctSelect              VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TBelayDeviceFunctSelects_PK PRIMARY KEY (intBelayDeviceFunctSelectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: BelayDevices Function Selection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceFunctSelects_InsertTrigger
                AFTER INSERT ON TBelayDeviceFunctSelects
                BEGIN
                    INSERT INTO Z_TBelayDeviceFunctSelects 
                    (
                        intBelayDeviceFunctSelectID,
                        strBelayDeviceFunctSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDeviceFunctSelectID,
                        NEW.strBelayDeviceFunctSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Function Selection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceFunctSelects_UpdateTrigger
                AFTER UPDATE ON TBelayDeviceFunctSelects
                BEGIN
                    INSERT INTO Z_TBelayDeviceFunctSelects 
                    (
                        intBelayDeviceFunctSelectID,
                        strBelayDeviceFunctSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDeviceFunctSelectID,
                        NEW.strBelayDeviceFunctSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Function Selection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceFunctSelects_DeleteTrigger
                AFTER DELETE ON TBelayDeviceFunctSelects
                BEGIN
                    INSERT INTO Z_TBelayDeviceFunctSelects 
                    (
                        intBelayDeviceFunctSelectID,
                        strBelayDeviceFunctSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intBelayDeviceFunctSelectID,
                        OLD.strBelayDeviceFunctSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating BelayDevices Functional Selection tables: {e}")                  

    def createBelayDevicesFunctPlasticSelTables(self):
        """ 
        Function Name: createBelayDevicesFunctPlasticSelTables
        Function Purpose: Create the BelayDevices Functional Plastic Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create BelayDevices Functional Plastic Selection Table
                CREATE TABLE IF NOT EXISTS TBelayDeviceFunctPlasticSelects 
                (
                    intBelayDeviceFunctPlastSelectID            INTEGER NOT NULL
                    ,strBelayDeviceFunctPlastSelect             VARCHAR(1000) NOT NULL
                    ,strModifiedReason                          VARCHAR(1000)
                    ,CONSTRAINT TBelayDeviceFunctPlasticSelects_PK PRIMARY KEY (intBelayDeviceFunctPlastSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: BelayDevices Functional Plastic Selection Table
                CREATE TABLE IF NOT EXISTS Z_TBelayDeviceFunctPlasticSelects 
                (
                    intBelayDeviceFunctPlasticSelectAuditID     INTEGER NOT NULL
                    ,intBelayDeviceFunctPlastSelectID           INTEGER NOT NULL
                    ,strBelayDeviceFunctPlastSelect             VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                               VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                               DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                                  VARCHAR(1) NOT NULL
                    ,strModifiedReason                          VARCHAR(1000)
                    ,CONSTRAINT Z_TBelayDeviceFunctPlasticSelects_PK PRIMARY KEY (intBelayDeviceFunctPlasticSelectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: BelayDevices Functional Plastic Selection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceFunctPlasticSelects_InsertTrigger
                AFTER INSERT ON TBelayDeviceFunctPlasticSelects
                BEGIN
                    INSERT INTO Z_TBelayDeviceFunctPlasticSelects 
                    (
                        intBelayDeviceFunctPlastSelectID,
                        strBelayDeviceFunctPlastSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDeviceFunctPlastSelectID,
                        NEW.strBelayDeviceFunctPlastSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Functional Plastic Selection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceFunctPlasticSelects_UpdateTrigger
                AFTER UPDATE ON TBelayDeviceFunctPlasticSelects
                BEGIN
                    INSERT INTO Z_TBelayDeviceFunctPlasticSelects 
                    (
                        intBelayDeviceFunctPlastSelectID,
                        strBelayDeviceFunctPlastSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDeviceFunctPlastSelectID,
                        NEW.strBelayDeviceFunctPlastSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Functional Plastic Selection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceFunctPlasticSelects_DeleteTrigger
                AFTER DELETE ON TBelayDeviceFunctPlasticSelects
                BEGIN
                    INSERT INTO Z_TBelayDeviceFunctPlasticSelects 
                    (
                        intBelayDeviceFunctPlastSelectID,
                        strBelayDeviceFunctPlastSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intBelayDeviceFunctPlastSelectID,
                        OLD.strBelayDeviceFunctPlastSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating BelayDevices Functional Plastic Selection tables: {e}")
                        
    def createBelayDevicesFunctInspectTables(self):
        """ 
        Function Name: createBelayDevicesFunctInspectTables
        Function Purpose: Create the BelayDevices Function Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create BelayDevices Function Inspection Table
                CREATE TABLE IF NOT EXISTS TBelayDeviceFunctionInspections 
                (
                    intBelayDeviceFunctionInspectID         INTEGER NOT NULL
                    ,intBelayDeviceID                       INTEGER NOT NULL
                    ,intInspectionTypeID                    INTEGER NOT NULL
                    ,intBelayDeviceFunctSelectID            INTEGER NOT NULL
                    ,intBelayDeviceFunctPlastSelectID       INTEGER NOT NULL
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,strModifiedReason VARCHAR(1000)
                    ,CONSTRAINT TBelayDeviceFunctionInspections_PK PRIMARY KEY (intBelayDeviceFunctionInspectID)

                    ,FOREIGN KEY ( intBelayDeviceID ) REFERENCES TBelayDevices ( intBelayDeviceID )
                    ,FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )
                    ,FOREIGN KEY ( intBelayDeviceFunctSelectID ) REFERENCES TBelayDeviceFunctSelects ( intBelayDeviceFunctSelectID )
                    ,FOREIGN KEY ( intBelayDeviceFunctPlastSelectID ) REFERENCES TBelayDeviceFunctPlasticSelects ( intBelayDeviceFunctPlastSelectID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )                    
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: BelayDevices Function Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TBelayDeviceFunctionInspection 
                (
                    intBelayDeviceFunctionInspectAuditID    INTEGER NOT NULL
                    ,intBelayDeviceFunctionInspectID        INTEGER NOT NULL
                    ,intBelayDeviceID                       INTEGER NOT NULL
                    ,intInspectionTypeID                    INTEGER NOT NULL
                    ,intBelayDeviceFunctSelectID            INTEGER NOT NULL
                    ,intBelayDeviceFunctPlastSelectID       INTEGER NOT NULL
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TBelayDeviceFunctionInspection_PK PRIMARY KEY (intBelayDeviceFunctionInspectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: BelayDevices Function Inspection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceFunctionInspection_InsertTrigger
                AFTER INSERT ON TBelayDeviceFunctionInspections
                BEGIN
                    INSERT INTO Z_TBelayDeviceFunctionInspection 
                    (
                        intBelayDeviceFunctionInspectID
                        ,intBelayDeviceID
                        ,intInspectionTypeID
                        ,intBelayDeviceFunctSelectID
                        ,intBelayDeviceFunctPlastSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDeviceFunctionInspectID
                        ,NEW.intBelayDeviceID
                        ,NEW.intInspectionTypeID
                        ,NEW.intBelayDeviceFunctSelectID
                        ,NEW.intBelayDeviceFunctPlastSelectID
                        ,NEW.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Function Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceFunctionInspection_UpdateTrigger
                AFTER UPDATE ON TBelayDeviceFunctionInspections
                BEGIN
                    INSERT INTO Z_TBelayDeviceFunctionInspection 
                    (
                        intBelayDeviceFunctionInspectID
                        ,intBelayDeviceID
                        ,intInspectionTypeID
                        ,intBelayDeviceFunctSelectID
                        ,intBelayDeviceFunctPlastSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDeviceFunctionInspectID
                        ,NEW.intBelayDeviceID
                        ,NEW.intInspectionTypeID
                        ,NEW.intBelayDeviceFunctSelectID
                        ,NEW.intBelayDeviceFunctPlastSelectID
                        ,NEW.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Function Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceFunctionInspection_DeleteTrigger
                AFTER DELETE ON TBelayDeviceFunctionInspections
                BEGIN
                    INSERT INTO Z_TBelayDeviceFunctionInspection 
                    (
                        intBelayDeviceFunctionInspectID
                        ,intBelayDeviceID
                        ,intInspectionTypeID
                        ,intBelayDeviceFunctSelectID
                        ,intBelayDeviceFunctPlastSelectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intBelayDeviceFunctionInspectID
                        ,OLD.intBelayDeviceID
                        ,OLD.intInspectionTypeID
                        ,OLD.intBelayDeviceFunctSelectID
                        ,OlD.intBelayDeviceFunctPlastSelectID
                        ,OLD.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating BelayDevices Functional Inspection tables: {e}")
            
    def createStandBelayDevicesInspectTables(self):
        """ 
        Function Name: createStandBelayDevicesInspectTables
        Function Purpose: Create the Standard BelayDevices Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Standard BelayDevices Inspection Table
                CREATE TABLE IF NOT EXISTS TStandardBelayDeviceInspections 
                (
                    intStandardBelayDeviceInspectionID          INTEGER NOT NULL
                    ,intBelayDeviceVisualInspectionID           INTEGER NOT NULL
                    ,intBelayDevicePhysicalInspectionID         INTEGER NOT NULL
                    ,intBelayDeviceFunctionInspectID            INTEGER NOT NULL
                    ,intInspectionStatusID                      INTEGER NOT NULL
                    ,strModifiedReason                          VARCHAR(1000)
                    ,CONSTRAINT TStandardBelayDeviceInspections_PK PRIMARY KEY (intStandardBelayDeviceInspectionID)
                    
                    ,FOREIGN KEY ( intBelayDeviceVisualInspectionID ) REFERENCES TBelayDeviceVisualInspections ( intBelayDeviceVisualInspectionID )
                    ,FOREIGN KEY ( intBelayDevicePhysicalInspectionID ) REFERENCES TBelayDevicePhysicalInspections ( intBelayDevicePhysicalInspectionID )
                    ,FOREIGN KEY ( intBelayDeviceFunctionInspectID ) REFERENCES TBelayDeviceFunctionInspections ( intBelayDeviceFunctionInspectID )
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Standard BelayDevices Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TStandardBelayDeviceInspections 
                (
                    intStandardBelayDeviceInspectionAuditID     INTEGER NOT NULL
                    ,intStandardBelayDeviceInspectionID         INTEGER NOT NULL
                    ,intBelayDeviceVisualInspectionID           INTEGER NOT NULL
                    ,intBelayDevicePhysicalInspectionID         INTEGER NOT NULL
                    ,intBelayDeviceFunctionInspectID            INTEGER NOT NULL
                    ,intInspectionStatusID                      INTEGER NOT NULL
                    ,strUpdatedBy                               VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                               DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                                  VARCHAR(1) NOT NULL
                    ,strModifiedReason                          VARCHAR(1000)
                    ,CONSTRAINT Z_TStandardBelayDeviceInspections_PK PRIMARY KEY (intStandardBelayDeviceInspectionAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Standard BelayDevices Inspection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TStandardBelayDeviceInspections_InsertTrigger
                AFTER INSERT ON TStandardBelayDeviceInspections
                BEGIN
                    INSERT INTO Z_TStandardBelayDeviceInspections 
                    (
                        intStandardBelayDeviceInspectionID
                        ,intBelayDeviceVisualInspectionID
                        ,intBelayDevicePhysicalInspectionID
                        ,intBelayDeviceFunctionInspectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intStandardBelayDeviceInspectionID
                        ,NEW.intBelayDeviceVisualInspectionID
                        ,NEW.intBelayDevicePhysicalInspectionID
                        ,NEW.intBelayDeviceFunctionInspectID
                        ,NEW.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Standard BelayDevices Inspection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TStandardBelayDeviceInspections_UpdateTrigger
                AFTER UPDATE ON TStandardBelayDeviceInspections
                BEGIN
                    INSERT INTO Z_TStandardBelayDeviceInspections 
                    (
                        intStandardBelayDeviceInspectionID
                        ,intBelayDeviceVisualInspectionID
                        ,intBelayDevicePhysicalInspectionID
                        ,intBelayDeviceFunctionInspectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intStandardBelayDeviceInspectionID
                        ,NEW.intBelayDeviceVisualInspectionID
                        ,NEW.intBelayDevicePhysicalInspectionID
                        ,NEW.intBelayDeviceFunctionInspectID
                        ,NEW.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: Standard BelayDevices Inspection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TStandardBelayDeviceInspections_DeleteTrigger
                AFTER DELETE ON TStandardBelayDeviceInspections
                BEGIN
                    INSERT INTO Z_TStandardBelayDeviceInspections 
                    (
                        intStandardBelayDeviceInspectionID
                        ,intBelayDeviceVisualInspectionID
                        ,intBelayDevicePhysicalInspectionID
                        ,intBelayDeviceFunctionInspectID
                        ,intInspectionStatusID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intStandardBelayDeviceInspectionID
                        ,OLD.intBelayDeviceVisualInspectionID
                        ,OLD.intBelayDevicePhysicalInspectionID
                        ,OLD.intBelayDeviceFunctionInspectID
                        ,OLD.intInspectionStatusID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,NULL
                    );
                END;
                """

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)    
            Database.dbExeStatement(self, sqlAudit)    
            Database.dbExeStatement(self, sqlTrigger)     
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Standard BelayDevices Inspection tables: {e}")

    def createBelayDevices_InspectionTables(self):
        """ 
        Function Name: createBelayDevices_InspectionTables
        Function Purpose: Create the BelayDevices Inspection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create BelayDevices Inspection Table
                CREATE TABLE IF NOT EXISTS TBelayDeviceInspections 
                (
                    intBelayDeviceInspectionID              INTEGER NOT NULL
                    ,intBelayDeviceID                       INTEGER NOT NULL
                    ,intWallLocationID                      INTEGER NOT NULL
                    ,intStandardBelayDeviceInspectionID     INTEGER NOT NULL
                    ,intInspectorID                         INTEGER NOT NULL
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,dtmLastInspectionDate                  DATETIME NOT NULL
                    ,dtmNextInspectionDate                  DATETIME NOT NULL
                    ,strComment                             VARCHAR(225) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT TBelayDeviceInspections_PK PRIMARY KEY (intBelayDeviceInspectionID)
                    
                    ,FOREIGN KEY ( intBelayDeviceID ) REFERENCES TBelayDevices ( intBelayDeviceID  )
                    ,FOREIGN KEY ( intWallLocationID ) REFERENCES TWallLocations ( intWallLocationID ) 
                    ,FOREIGN KEY ( intStandardBelayDeviceInspectionID ) REFERENCES TStandardBelayDeviceInspections ( intStandardBelayDeviceInspectionID ) 
                    ,FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID ) 
                    ,FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID ) 
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: BelayDevices Inspection Table
                CREATE TABLE IF NOT EXISTS Z_TBelayDeviceInspections 
                (
                    intBelayDeviceInspectionAuditID         INTEGER NOT NULL
                    ,intBelayDeviceInspectionID             INTEGER NOT NULL
                    ,intBelayDeviceID                       INTEGER NOT NULL
                    ,intWallLocationID                      INTEGER NOT NULL
                    ,intStandardBelayDeviceInspectionID     INTEGER NOT NULL
                    ,intInspectorID                         INTEGER NOT NULL
                    ,intInspectionStatusID                  INTEGER NOT NULL
                    ,dtmLastInspectionDate                  DATETIME NOT NULL
                    ,dtmNextInspectionDate                  DATETIME NOT NULL
                    ,strComment                             VARCHAR(225) NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TBelayDeviceInspections_PK PRIMARY KEY ( intBelayDeviceInspectionAuditID )
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: BelayDevices Inspection Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceInspections_AuditTrigger_Insert
                AFTER INSERT ON TBelayDeviceInspections
                BEGIN
                    INSERT INTO Z_TBelayDeviceInspections 
                    (
                        intBelayDeviceInspectionID
                        ,intBelayDeviceID
                        ,intWallLocationID
                        ,intStandardBelayDeviceInspectionID
                        ,intInspectorID
                        ,intInspectionStatusID
                        ,dtmLastInspectionDate
                        ,dtmNextInspectionDate
                        ,strComment
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDeviceInspectionID
                        ,NEW.intBelayDeviceID
                        ,NEW.intWallLocationID
                        ,NEW.intStandardBelayDeviceInspectionID
                        ,NEW.intInspectorID
                        ,NEW.intInspectionStatusID
                        ,NEW.dtmLastInspectionDate
                        ,NEW.dtmNextInspectionDate
                        ,NEW.strComment
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Inspection Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceInspections_AuditTrigger_Update
                AFTER UPDATE ON TBelayDeviceInspections
                BEGIN
                    INSERT INTO Z_TBelayDeviceInspections 
                    (
                        intBelayDeviceInspectionID
                        ,intBelayDeviceID
                        ,intWallLocationID
                        ,intStandardBelayDeviceInspectionID
                        ,intInspectorID
                        ,intInspectionStatusID
                        ,dtmLastInspectionDate
                        ,dtmNextInspectionDate
                        ,strComment
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDeviceInspectionID
                        ,NEW.intBelayDeviceID
                        ,NEW.intWallLocationID
                        ,NEW.intStandardBelayDeviceInspectionID
                        ,NEW.intInspectorID
                        ,NEW.intInspectionStatusID
                        ,NEW.dtmLastInspectionDate
                        ,NEW.dtmNextInspectionDate
                        ,NEW.strComment
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Inspection Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceInspections_AuditTrigger_Delete
                AFTER DELETE ON TBelayDeviceInspections
                BEGIN
                    INSERT INTO Z_TBelayDeviceInspections 
                    (
                        intBelayDeviceInspectionID
                        ,intBelayDeviceID
                        ,intWallLocationID
                        ,intStandardBelayDeviceInspectionID
                        ,intInspectorID
                        ,intInspectionStatusID
                        ,dtmLastInspectionDate
                        ,dtmNextInspectionDate
                        ,strComment
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intBelayDeviceInspectionID
                        ,OLD.intBelayDeviceID
                        ,OLD.intWallLocationID
                        ,OLD.intStandardBelayDeviceInspectionID
                        ,OLD.intInspectorID
                        ,OLD.intInspectionStatusID
                        ,OLD.dtmLastInspectionDate
                        ,OLD.dtmNextInspectionDate
                        ,OLD.strComment
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating BelayDevices Inspection tables: {e}")            

    def createBelayDevices_WallLocationTables(self):
        """ 
        Function Name: createBelayDevices_WallLocationTables
        Function Purpose: Create the BelayDevices Location Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create BelayDevices Wall Location Table
                CREATE TABLE IF NOT EXISTS TBelayDeviceWallLocations 
                (
                    intBelayDeviceWallLocationID            INTEGER NOT NULL
                    ,intWallLocationID                      INTEGER NOT NULL
                    ,intBelayDeviceID                       INTEGER NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,FOREIGN KEY ( intWallLocationID ) REFERENCES TWallLocations ( intWallLocationID )
                    ,FOREIGN KEY ( intBelayDeviceID ) REFERENCES TBelayDevices ( intBelayDeviceID )
                    
                    ,CONSTRAINT TBelayDeviceWallLocations_PK PRIMARY KEY (intBelayDeviceWallLocationID)
                    ,CONSTRAINT TBelayDeviceWallLocations_UQ UNIQUE ( intWallLocationID, intBelayDeviceID )
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: BelayDevices Wall Location Table
                CREATE TABLE IF NOT EXISTS Z_TBelayDeviceWallLocations 
                (
                    intBelayDeviceWallLocationAuditID       INTEGER NOT NULL
                    ,intBelayDeviceWallLocationID           INTEGER NOT NULL
                    ,intWallLocationID                      INTEGER NOT NULL
                    ,intBelayDeviceID                       INTEGER NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TBelayDeviceWallLocations_PK PRIMARY KEY ( intBelayDeviceWallLocationAuditID )
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: BelayDevices Wall Location Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceWallLocations_AuditTrigger_Insert
                AFTER INSERT ON TBelayDeviceWallLocations
                BEGIN
                    INSERT INTO Z_TBelayDeviceWallLocations 
                    (
                        intBelayDeviceWallLocationID
                        ,intWallLocationID
                        ,intBelayDeviceID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDeviceWallLocationID
                        ,NEW.intWallLocationID
                        ,NEW.intBelayDeviceID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Wall Location Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceWallLocations_AuditTrigger_Update
                AFTER UPDATE ON TBelayDeviceWallLocations
                BEGIN
                    INSERT INTO Z_TBelayDeviceWallLocations 
                    (
                        intBelayDeviceWallLocationID
                        ,intWallLocationID
                        ,intBelayDeviceID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDeviceWallLocationID
                        ,NEW.intWallLocationID
                        ,NEW.intBelayDeviceID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Wall Location Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceWallLocations_AuditTrigger_Delete
                AFTER DELETE ON TBelayDeviceWallLocations
                BEGIN
                    INSERT INTO Z_TBelayDeviceWallLocations 
                    (
                        intBelayDeviceWallLocationID
                        ,intWallLocationID
                        ,intBelayDeviceID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intBelayDeviceWallLocationID
                        ,OLD.intWallLocationID
                        ,OLD.intBelayDeviceID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating BelayDevices Wall Location tables: {e}")            

    def createBelayDevices_InspectorTables(self):
        """ 
        Function Name: createBelayDevices_InspectorTables
        Function Purpose: Create the BelayDevices Inspector Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create BelayDevices Inspector Table
                CREATE TABLE IF NOT EXISTS TBelayDeviceInspectors 
                (
                    intBelayDeviceInspectorID           INTEGER NOT NULL
                    ,intInspectorID                     INTEGER NOT NULL
                    ,intBelayDeviceID                   INTEGER NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TBelayDeviceInspectors_PK PRIMARY KEY ( intBelayDeviceInspectorID )
                    
                    ,FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID )
                    ,FOREIGN KEY ( intBelayDeviceID ) REFERENCES TBelayDevices ( intBelayDeviceID )
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: BelayDevices Inspector Table
                CREATE TABLE IF NOT EXISTS Z_TBelayDeviceInspectors 
                (
                    intBelayDeviceInspectorAuditID      INTEGER NOT NULL
                    ,intBelayDeviceInspectorID          INTEGER NOT NULL
                    ,intInspectorID                     INTEGER NOT NULL
                    ,intBelayDeviceID                   INTEGER NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TBelayDeviceInspectors_PK PRIMARY KEY ( intBelayDeviceInspectorAuditID )
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: BelayDevices Inspector Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceInspectors_AuditTrigger_Insert
                AFTER INSERT ON TBelayDeviceInspectors
                BEGIN
                    INSERT INTO Z_TBelayDeviceInspectors 
                    (
                        intBelayDeviceInspectorID
                        ,intInspectorID
                        ,intBelayDeviceID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDeviceInspectorID
                        ,NEW.intInspectorID
                        ,NEW.intBelayDeviceID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Inspector Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceInspectors_AuditTrigger_Update
                AFTER UPDATE ON TBelayDeviceInspectors
                BEGIN
                    INSERT INTO Z_TBelayDeviceInspectors 
                    (
                        intBelayDeviceInspectorID
                        ,intInspectorID
                        ,intBelayDeviceID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDeviceInspectorID
                        ,NEW.intInspectorID
                        ,NEW.intBelayDeviceID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Inspector Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceInspectors_AuditTrigger_Delete
                AFTER DELETE ON TBelayDeviceInspectors
                BEGIN
                    INSERT INTO Z_TBelayDeviceInspectors 
                    (
                        intBelayDeviceInspectorID
                        ,intInspectorID
                        ,intBelayDeviceID
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intBelayDeviceInspectorID
                        ,OLD.intInspectorID
                        ,OLD.intBelayDeviceID
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating BelayDevices Inspector tables: {e}")            

    def createCustomRopeSystemsTables(self):
        """ 
        Function Name: createCustomRopeSystemsTables
        Function Purpose: Create the CustomRopeSystems Tables and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create Custom Rope Systems Table
                CREATE TABLE IF NOT EXISTS TCustomRopeSystems 
                (
                    intRopeSystemID                     INTEGER NOT NULL
                    ,strRopeSystemName                  VARCHAR(255) NOT NULL
                    ,strComplexity                      VARCHAR(255) NOT NULL
                    ,strPreTiedKnot                     VARCHAR(255) NOT NULL
                    ,intConnectorCount                  INTEGER NOT NULL 
                    ,strFirstConnectorType              VARCHAR(255) NOT NULL
                    ,strSecondConnectorType             VARCHAR(255) NOT NULL
                    ,strDeviceType                      VARCHAR(255) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT TCustomRopeSystems_PK PRIMARY KEY (intRopeSystemID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: Custom Rope Systems Table
                CREATE TABLE IF NOT EXISTS Z_TCustomRopeSystems 
                (
                    intRopeSystemAuditID                INTEGER NOT NULL
                    ,intRopeSystemID                    INTEGER NOT NULL
                    ,strRopeSystemName                  VARCHAR(255) NOT NULL
                    ,strComplexity                      VARCHAR(255) NOT NULL
                    ,strPreTiedKnot                     VARCHAR(255) NOT NULL
                    ,intConnectorCount                  INTEGER NOT NULL 
                    ,strFirstConnectorType              VARCHAR(255) NOT NULL
                    ,strSecondConnectorType             VARCHAR(255) NOT NULL
                    ,strDeviceType                      VARCHAR(255) NOT NULL
                    ,strUpdatedBy                       VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                       DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                          VARCHAR(1) NOT NULL
                    ,strModifiedReason                  VARCHAR(1000)
                    ,CONSTRAINT Z_TCustomRopeSystems_PK PRIMARY KEY (intRopeSystemAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: Custom Rope Systems Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TCustomRopeSystems_AuditTrigger_Insert
                AFTER INSERT ON TCustomRopeSystems
                BEGIN
                    INSERT INTO Z_TCustomRopeSystems 
                    (
                        intRopeSystemID             
                        ,strRopeSystemName           
                        ,strComplexity
                        ,strPreTiedKnot
                        ,intConnectorCount
                        ,strFirstConnectorType
                        ,strSecondConnectorType
                        ,strDeviceType
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intRopeSystemID       
                        ,NEW.strRopeSystemName                 
                        ,NEW.strComplexity
                        ,NEW.strPreTiedKnot
                        ,NEW.intConnectorCount
                        ,NEW.strFirstConnectorType
                        ,NEW.strSecondConnectorType
                        ,NEW.strDeviceType
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: CustomRopeSystems Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TCustomRopeSystems_AuditTrigger_Update
                AFTER UPDATE ON TCustomRopeSystems
                BEGIN
                    INSERT INTO Z_TCustomRopeSystems 
                    (
                        intRopeSystemID             
                        ,strRopeSystemName           
                        ,strComplexity
                        ,strPreTiedKnot
                        ,intConnectorCount
                        ,strFirstConnectorType
                        ,strSecondConnectorType
                        ,strDeviceType
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intRopeSystemID       
                        ,NEW.strRopeSystemName                 
                        ,NEW.strComplexity
                        ,NEW.strPreTiedKnot
                        ,NEW.intConnectorCount
                        ,NEW.strFirstConnectorType
                        ,NEW.strSecondConnectorType
                        ,NEW.strDeviceType
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: CustomRopeSystems Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TCustomRopeSystems_AuditTrigger_Delete
                AFTER DELETE ON TCustomRopeSystems
                BEGIN
                    INSERT INTO Z_TCustomRopeSystems 
                    (
                        intRopeSystemID             
                        ,strRopeSystemName           
                        ,strComplexity
                        ,strPreTiedKnot
                        ,intConnectorCount
                        ,strFirstConnectorType
                        ,strSecondConnectorType
                        ,strDeviceType
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intRopeSystemID                        
                        ,OLD.strRopeSystemName  
                        ,OLD.strComplexity
                        ,OLD.strPreTiedKnot
                        ,OLD.intConnectorCount
                        ,OLD.strFirstConnectorType
                        ,OLD.strSecondConnectorType
                        ,OLD.strDeviceType
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating CustomRopeSystems tables: {e}")

    def createBelayDevice_RetiredReports_Tables(self):
        """ 
        Function Name: createBelayDevice_RetiredReports_Tables
        Function Purpose: Create the BelayDevice Retired Reports Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create BelayDevice Inspection Table
                CREATE TABLE IF NOT EXISTS TBelayDeviceRetiredReports 
                (
                    intBelayDeviceRetiredReportID           INTEGER NOT NULL
                    ,intBelayDeviceInspectionID             INTEGER NOT NULL
                    ,intInspectorID                         INTEGER NOT NULL
                    ,dtmReportDate                          DATETIME NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT TBelayDeviceRetiredReports_PK PRIMARY KEY (intBelayDeviceRetiredReportID)
                    
                    ,FOREIGN KEY ( intBelayDeviceInspectionID ) REFERENCES TBelayDeviceInspections ( intBelayDeviceInspectionID )
                    ,FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID ) 
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: BelayDevice Retired Report Table
                CREATE TABLE IF NOT EXISTS Z_TBelayDeviceRetiredReports 
                (
                    intBelayDeviceRetiredReportAuditID      INTEGER NOT NULL
                    ,intBelayDeviceRetiredReportID          INTEGER NOT NULL
                    ,intBelayDeviceInspectionID             INTEGER NOT NULL
                    ,intInspectorID                         INTEGER NOT NULL
                    ,dtmReportDate                          DATETIME NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TBelayDeviceRetiredReports_PK PRIMARY KEY (intBelayDeviceRetiredReportAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: BelayDevice Retired Report Table - Insert Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceRetiredReports_AuditTrigger_Insert
                AFTER INSERT ON TBelayDeviceRetiredReports
                BEGIN
                    INSERT INTO Z_TBelayDeviceRetiredReports 
                    (
                        intBelayDeviceRetiredReportID
                        ,intBelayDeviceInspectionID
                        ,intInspectorID
                        ,dtmReportDate
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDeviceRetiredReportID
                        ,NEW.intBelayDeviceInspectionID
                        ,NEW.intInspectorID
                        ,NEW.dtmReportDate
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'I' -- Insert
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevice Retired Report Table - Update Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceRetiredReports_AuditTrigger_Update
                AFTER UPDATE ON TBelayDeviceRetiredReports
                BEGIN
                    INSERT INTO Z_TBelayDeviceRetiredReports 
                    (
                        intBelayDeviceRetiredReportID
                        ,intBelayDeviceInspectionID
                        ,intInspectorID
                        ,dtmReportDate
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDeviceRetiredReportID
                        ,NEW.intBelayDeviceInspectionID
                        ,NEW.intInspectorID
                        ,NEW.dtmReportDate
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'U' -- Update
                        ,NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevice Inspection Table - Delete Trigger
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceRetiredReports_AuditTrigger_Delete
                AFTER DELETE ON TBelayDeviceRetiredReports
                BEGIN
                    INSERT INTO Z_TBelayDeviceRetiredReports 
                    (
                        intBelayDeviceRetiredReportID
                        ,intBelayDeviceInspectionID
                        ,intInspectorID
                        ,dtmReportDate
                        ,strUpdatedBy
                        ,dtmUpdatedOn
                        ,strAction
                        ,strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intBelayDeviceRetiredReportID
                        ,OLD.intBelayDeviceInspectionID
                        ,OLD.intInspectorID
                        ,OLD.dtmReportDate
                        ,'{self.currentUser}'
                        ,DATETIME('now')
                        ,'D' -- Delete
                        ,OLD.strModifiedReason
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating BelayDevice Inspection tables: {e}")

    def createBelayDevicesVisPlasticSelTables(self):
        """ 
        Function Name: createBelayDevicesVisPlasticSelTables
        Function Purpose: Create the BelayDevices Visual Plastic Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create BelayDevices Visual Plastic Selection Table
                CREATE TABLE IF NOT EXISTS TBelayDeviceVisPlasticSelects 
                (
                    intBelayDeviceVisPlasticSelectID        INTEGER NOT NULL
                    ,strBelayDeviceVisPlastSelect           VARCHAR(1000) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT TBelayDeviceVisPlasticSelects_PK PRIMARY KEY (intBelayDeviceVisPlasticSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: BelayDevices Visual Plastic Selection Table
                CREATE TABLE IF NOT EXISTS Z_TBelayDeviceVisPlasticSelects 
                (
                    intBelayDeviceVisPlasticSelectAuditID   INTEGER NOT NULL
                    ,intBelayDeviceVisPlasticSelectID       INTEGER NOT NULL
                    ,strBelayDeviceVisPlastSelect           VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                           VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                           DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                              VARCHAR(1) NOT NULL
                    ,strModifiedReason                      VARCHAR(1000)
                    ,CONSTRAINT Z_TBelayDeviceVisPlasticSelects_PK PRIMARY KEY (intBelayDeviceVisPlasticSelectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: BelayDevices Visual Plastic Selection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceVisPlasticSelects_InsertTrigger
                AFTER INSERT ON TBelayDeviceVisPlasticSelects
                BEGIN
                    INSERT INTO Z_TBelayDeviceVisPlasticSelects 
                    (
                        intBelayDeviceVisPlasticSelectID,
                        strBelayDeviceVisPlastSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDeviceVisPlasticSelectID,
                        NEW.strBelayDeviceVisPlastSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Visual Plastic Selection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceVisPlasticSelects_UpdateTrigger
                AFTER UPDATE ON TBelayDeviceVisPlasticSelects
                BEGIN
                    INSERT INTO Z_TBelayDeviceVisPlasticSelects 
                    (
                        intBelayDeviceVisPlasticSelectID,
                        strBelayDeviceVisPlastSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDeviceVisPlasticSelectID,
                        NEW.strBelayDeviceVisPlastSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Visual Plastic Selection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDeviceVisPlasticSelects_DeleteTrigger
                AFTER DELETE ON TBelayDeviceVisPlasticSelects
                BEGIN
                    INSERT INTO Z_TBelayDeviceVisPlasticSelects 
                    (
                        intBelayDeviceVisPlasticSelectID,
                        strBelayDeviceVisPlastSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intBelayDeviceVisPlasticSelectID,
                        OLD.strBelayDeviceVisPlastSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating BelayDevices Visual Plastic Selection tables: {e}")

    def createBelayDevicesPhysPlasticSelTables(self):
        """ 
        Function Name: createBelayDevicesPhysPlasticSelTables
        Function Purpose: Create the BelayDevices Physical Plastic Selection Table and Z Tables inside the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlTable = """
                -- Create BelayDevices Physical Plastic Selection Table
                CREATE TABLE IF NOT EXISTS TBelayDevicePhysPlasticSelects 
                (
                    intBelayDevicePhysPlasticSelectID           INTEGER NOT NULL
                    ,strBelayDevicePhysPlastSelect              VARCHAR(1000) NOT NULL
                    ,strModifiedReason                          VARCHAR(1000)
                    ,CONSTRAINT TBelayDevicePhysPlasticSelects_PK PRIMARY KEY (intBelayDevicePhysPlasticSelectID)
                );
                """
            # Create the audit table    
            sqlAudit = f"""                
                -- Create Z Table: BelayDevices Physical Plastic Selection Table
                CREATE TABLE IF NOT EXISTS Z_TBelayDevicePhysPlasticSelects 
                (
                    intBelayDevicePhysPlasticSelectAuditID      INTEGER NOT NULL
                    ,intBelayDevicePhysPlasticSelectID          INTEGER NOT NULL
                    ,strBelayDevicePhysPlastSelect              VARCHAR(1000) NOT NULL
                    ,strUpdatedBy                               VARCHAR(225) NOT NULL DEFAULT '{self.currentUser}'
                    ,dtmUpdatedOn                               DATETIME DEFAULT CURRENT_TIMESTAMP
                    ,strAction                                  VARCHAR(1) NOT NULL
                    ,strModifiedReason                          VARCHAR(1000)
                    ,CONSTRAINT Z_TBelayDevicePhysPlasticSelects_PK PRIMARY KEY (intBelayDevicePhysPlasticSelectAuditID)
                );
                """
            # Create the table trigger    
            sqlTrigger = f"""
                -- Create Trigger: BelayDevices Physical Plastic Selection Table - Insert
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDevicePhysPlasticSelects_InsertTrigger
                AFTER INSERT ON TBelayDevicePhysPlasticSelects
                BEGIN
                    INSERT INTO Z_TBelayDevicePhysPlasticSelects 
                    (
                        intBelayDevicePhysPlasticSelectID,
                        strBelayDevicePhysPlastSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDevicePhysPlasticSelectID,
                        NEW.strBelayDevicePhysPlastSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'I', -- Insert
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Physical Plastic Selection Table - Update
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDevicePhysPlasticSelects_UpdateTrigger
                AFTER UPDATE ON TBelayDevicePhysPlasticSelects
                BEGIN
                    INSERT INTO Z_TBelayDevicePhysPlasticSelects 
                    (
                        intBelayDevicePhysPlasticSelectID,
                        strBelayDevicePhysPlastSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        NEW.intBelayDevicePhysPlasticSelectID,
                        NEW.strBelayDevicePhysPlastSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'U', -- Update
                        NEW.strModifiedReason
                    );
                END;

                -- Create Trigger: BelayDevices Physical Plastic Selection Table - Delete
                CREATE TRIGGER IF NOT EXISTS Z_TBelayDevicePhysPlasticSelects_DeleteTrigger
                AFTER DELETE ON TBelayDevicePhysPlasticSelects
                BEGIN
                    INSERT INTO Z_TBelayDevicePhysPlasticSelects 
                    (
                        intBelayDevicePhysPlasticSelectID,
                        strBelayDevicePhysPlastSelect,
                        strUpdatedBy,
                        dtmUpdatedOn,
                        strAction,
                        strModifiedReason
                    )
                    VALUES 
                    (
                        OLD.intBelayDevicePhysPlasticSelectID,
                        OLD.strBelayDevicePhysPlastSelect,
                        '{self.currentUser}',
                        DATETIME('now'),
                        'D', -- Delete
                        NULL
                    );
                END;
                """
                
            # Execute the SQL statements
            Database.dbExeStatement(self, sqlTable)
            Database.dbExeStatement(self, sqlAudit)
            Database.dbExeStatement(self, sqlTrigger)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating BelayDevices Physical Plastic Selection tables: {e}")
                                                
    def createBelayDevices_ForeignKeys(self):
        """ 
        Function Name: createBelayDevices_ForeignKeys
        Function Purpose: Create the BelayDevices Visual, Physical, Standard, and Inspection foreign keys inside the database
        """           
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the foreign keys
            sqlForeignKeys = """
                -- -------------------------------------------------------------------------------------------------------
                -- Establish Referential Integrity - Inserts
                -- -------------------------------------------------------------------------------------------------------

                -- #           Child                                 Parent                                 Column
                -- --         -------                               --------                               --------
                -- 1. TBelayDeviceVisualInspections              TBelayDevices                         intBelayDeviceID
                -- 2. TBelayDeviceVisualInspections              TInspectionTypes                      intInspectionTypeID
                -- 3. TBelayDeviceVisualInspections              TBelayDeviceVisMetalSelects           intBelayDeviceVisMetalSelectID
                -- 4. TBelayDeviceVisualInspections              TInspectionStatus                     intInspectionStatusID
                -- 5. TBelayDevicePhysicalInspections            TBelayDevices                         intBelayDeviceID
                -- 6. TBelayDevicePhysicalInspections            TInspectionTypes                      intInspectionTypeID
                -- 7. TBelayDevicePhysicalInspections            TBelayDevicePhysMetalSelects          intBelayDevicePhysMetalSelectID 
                -- 8. TBelayDevicePhysicalInspections            TInspectionStatus                     intInspectionStatusID
                -- 9. TBelayDeviceFunctionInspections            TBelayDevices                         intBelayDeviceID
                -- 10. TBelayDeviceFunctionInspections           TInspectionTypes                      intInspectionTypeID
                -- 11. TBelayDeviceFunctionInspections           TBelayDeviceFunctSelects              intBelayDeviceFunctSelectID
                -- 12. TBelayDeviceFunctionInspections           TInspectionStatus                     intInspectionStatusID
                -- 13. TStandardBelayDeviceInspections           TBelayDeviceVisualInspections         intBelayDeviceVisualInspectionID
                -- 14. TStandardBelayDeviceInspections           TBelayDevicePhysicalInspections       intBelayDevicePhysicalInspectionID
                -- 15. TStandardBelayDeviceInspections           TBelayDeviceFunctionInspections       intBelayDeviceFunctionInspectID
                -- 16. TStandardBelayDeviceInspections           TInspectionStatus                     intInspectionStatusID
                -- 17. TBelayDeviceInspections                   TBelayDevices                         intBelayDeviceID
                -- 18. TBelayDeviceInspections                   TWallLocations                        intWallLocationID
                -- 19. TBelayDeviceInspections                   TStandardBelayDeviceInspections       intStandardBelayDeviceInspectionID
                -- 20. TBelayDeviceInspections                   TInspectors                           intInspectorID
                -- 21. TBelayDeviceInspections                   TInspectionStatus                     intInspectionStatusID
                -- 22. TBelayDeviceRetiredReports                TBelayDeviceInspections               intBelayDeviceInspectionID
                -- 23. TBelayDeviceRetiredReports                TInspectors                           intInspectorID
                -- 24. TBelayDeviceVisualInspections             TBelayDeviceVisPlasticSelects         intBelayDeviceVisPlasticSelectID
                -- 25. TBelayDevicePhysicalInspections           TBelayDevicePhysPlasticSelects        intBelayDevicePhysPlasticSelectID
                -- 26. TBelayDeviceFunctionInspections           TBelayDeviceFunctPlasticSelects       intBelayDeviceFunctPlastSelectID
                
                -- -------------------------------------------------------------------------------------------------------
                -- Create Foreign Keys
                -- -------------------------------------------------------------------------------------------------------
                -- 1.
                ALTER TABLE TBelayDeviceVisualInspections ADD CONSTRAINT TBelayDeviceVisualInspections_TBelayDevices_FK
                FOREIGN KEY ( intBelayDeviceID ) REFERENCES TBelayDevices ( intBelayDeviceID  )

                -- 2.
                ALTER TABLE TBelayDeviceVisualInspections ADD CONSTRAINT TBelayDeviceVisualInspections_TInspectionTypes_FK
                FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )

                -- 3.
                ALTER TABLE TBelayDeviceVisualInspections ADD CONSTRAINT TBelayDeviceVisualInspections_TBelayDeviceVisMetalSelects_FK
                FOREIGN KEY ( intBelayDeviceVisMetalSelectID ) REFERENCES TBelayDeviceVisMetalSelects ( intBelayDeviceVisMetalSelectID )

                -- 4.
                ALTER TABLE TBelayDeviceVisualInspections ADD CONSTRAINT TBelayDeviceVisualInspections_TInspectionStatus_FK
                FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )

                -- -------------------------------------------------------------------------------------------------------
                -- Ensure No Duplicated Data
                -- -------------------------------------------------------------------------------------------------------
                ALTER TABLE TBelayDeviceWallLocations ADD CONSTRAINT TBelayDeviceWallLocations_UNIQUE
                UNIQUE ( intWallLocationID,  intBelayDeviceID )

                -- -------------------------------------------------------------------------------------------------------
                -- Create Foreign Keys
                -- -------------------------------------------------------------------------------------------------------
                -- 5.
                ALTER TABLE TBelayDevicePhysicalInspections ADD CONSTRAINT TBelayDevicePhysicalInspections_TBelayDevices_FK
                FOREIGN KEY ( intBelayDeviceID ) REFERENCES TBelayDevices ( intBelayDeviceID  )

                -- 6.
                ALTER TABLE TBelayDevicePhysicalInspections ADD CONSTRAINT TBelayDevicePhysicalInspections_TInspectionTypes_FK
                FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )

                -- 7.
                ALTER TABLE TBelayDevicePhysicalInspections ADD CONSTRAINT TBelayDevicePhysicalInspections_TBelayDevicePhysMetalSelects_FK
                FOREIGN KEY ( intBelayDevicePhysMetalSelectID ) REFERENCES TBelayDevicePhysMetalSelects ( intBelayDevicePhysMetalSelectID )

                -- 8. 
                ALTER TABLE TBelayDevicePhysicalInspections ADD CONSTRAINT TBelayDevicePhysicalInspections_TInspectionStatus_FK
                FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )

                -- 9.
                ALTER TABLE TBelayDeviceFunctionInspections ADD CONSTRAINT TBelayDeviceFunctionInspections_TBelayDevices_FK
                FOREIGN KEY ( intBelayDeviceID ) REFERENCES TBelayDevices ( intBelayDeviceID )

                -- 10.
                ALTER TABLE TBelayDeviceFunctionInspections ADD CONSTRAINT TBelayDeviceFunctionInspections_TInspectionTypes_FK
                FOREIGN KEY ( intInspectionTypeID ) REFERENCES TInspectionTypes ( intInspectionTypeID )

                -- 11.
                ALTER TABLE TBelayDeviceFunctionInspections ADD CONSTRAINT TBelayDeviceFunctionInspections_TBelayDeviceFunctSelects_FK
                FOREIGN KEY ( intBelayDeviceFunctSelectID ) REFERENCES TBelayDeviceFunctSelects ( intBelayDeviceFunctSelectID )

                -- 12. 
                ALTER TABLE TBelayDeviceFunctionInspections ADD CONSTRAINT TBelayDeviceFunctionInspections_TInspectionStatus_FK
                FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )
                
                -- 13. 
                ALTER TABLE TStandardBelayDeviceInspections ADD CONSTRAINT TStandardBelayDeviceInspections_TBelayDeviceVisualInspections_FK
                FOREIGN KEY ( intBelayDeviceVisualInspectionID ) REFERENCES TBelayDeviceVisualInspections ( intBelayDeviceVisualInspectionID )

                -- 14.
                ALTER TABLE TStandardBelayDeviceInspections ADD CONSTRAINT TStandardBelayDeviceInspections_TBelayDevicePhysicalInspections_FK
                FOREIGN KEY ( intBelayDevicePhysicalInspectionID ) REFERENCES TBelayDevicePhysicalInspections ( intBelayDevicePhysicalInspectionID )

                -- 15.
                ALTER TABLE TStandardBelayDeviceInspections ADD CONSTRAINT TStandardBelayDeviceInspections_TBelayDeviceFunctionInspections_FK
                FOREIGN KEY ( intBelayDeviceFunctionInspectID ) REFERENCES TBelayDeviceFunctionInspections ( intBelayDeviceFunctionInspectID )

                -- 16.
                ALTER TABLE TStandardBelayDeviceInspections ADD CONSTRAINT TStandardBelayDeviceInspections_TInspectionStatus_FK
                FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID )

                -- 17.
                ALTER TABLE TBelayDeviceInspections ADD CONSTRAINT TBelayDeviceInspections_TBelayDevices_FK
                FOREIGN KEY ( intBelayDeviceID ) REFERENCES TBelayDevices ( intBelayDeviceID  )

                -- 18.
                ALTER TABLE TBelayDeviceInspections ADD CONSTRAINT TBelayDeviceInspections_TWallLocations_FK
                FOREIGN KEY ( intWallLocationID ) REFERENCES TWallLocations ( intWallLocationID )

                -- 19.
                ALTER TABLE TBelayDeviceInspections ADD CONSTRAINT TBelayDeviceInspections_TStandardBelayDeviceInspections_FK
                FOREIGN KEY ( intStandardBelayDeviceInspectionID ) REFERENCES TStandardBelayDeviceInspections ( intStandardBelayDeviceInspectionID )
                
                -- 20.
                ALTER TABLE TBelayDeviceInspections ADD CONSTRAINT TBelayDeviceInspections_TInspectors_FK
                FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID )

                -- 21.
                ALTER TABLE TBelayDeviceInspections ADD CONSTRAINT TBelayDeviceInspections_TInspectionStatus_FK
                FOREIGN KEY ( intInspectionStatusID ) REFERENCES TInspectionStatus ( intInspectionStatusID ) 
                
                -- 22.
                ALTER TABLE TBelayDeviceRetiredReports ADD CONSTRAINT TBelayDeviceRetiredReports_TBelayDeviceInspections_FK
                FOREIGN KEY ( intBelayDeviceInspectionID ) REFERENCES TBelayDeviceInspections ( intBelayDeviceInspectionID )

                -- 23.
                ALTER TABLE TBelayDeviceRetiredReports ADD CONSTRAINT TBelayDeviceRetiredReports_TInspectors_FK
                FOREIGN KEY ( intInspectorID ) REFERENCES TInspectors ( intInspectorID )  
            
                -- 24.
                ALTER TABLE TBelayDeviceVisualInspections ADD CONSTRAINT TBelayDeviceVisualInspections_TBelayDeviceVisPlasticSelects_FK
                FOREIGN KEY ( intBelayDeviceVisPlasticSelectID ) REFERENCES TBelayDeviceVisPlasticSelects ( intBelayDeviceVisPlasticSelectID )

                -- 25.
                ALTER TABLE TBelayDevicePhysicalInspections ADD CONSTRAINT TBelayDevicePhysicalInspections_TBelayDevicePhysPlasticSelects_FK
                FOREIGN KEY ( intBelayDevicePhysPlasticSelectID ) REFERENCES TBelayDevicePhysPlasticSelects ( intBelayDevicePhysPlasticSelectID )

                -- 26.
                ALTER TABLE TBelayDeviceFunctionInspections ADD CONSTRAINT TBelayDeviceFunctionInspections_TBelayDeviceFunctPlasticSelects_FK
                FOREIGN KEY ( intBelayDeviceFunctPlastSelectID ) REFERENCES TBelayDeviceFunctPlasticSelects ( intBelayDeviceFunctPlastSelectID )                                                                                        
                """     

            # Execute the SQL statements
            Database.dbExeStatement(self, sqlForeignKeys)    
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating Foreign Keys: {e}")            
                        
    def insertUsrLog_Location_Data(self):
        """ 
        Function Name: insertUsrLog_Location_Data
        Function Purpose: Insert the User Login, Locations data into the tables
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the table
            sqlInsert = [
            """
            -- Add Data To Table: TStates
            INSERT INTO TStates (intStateID, strStateName)
            VALUES   (1, 'Indiana')
                    ,(2, 'Ohio')
                    ,(3, 'Kentucky')
            """,
            """
            -- Add Data To Table: TInspectors
            INSERT INTO TInspectors (intInspectorID, strFirstName, strLastName, strEmail)
            VALUES   (1, 'Root', 'Admin', 'dteixeira.betabreak@gmail.com')
                    ,(2, 'Carlos', 'Vidaurri', 'carlos@hoosierheights.com')
                    ,(3, 'David', 'Teixeira', 'David.Teixeira@gmail.com')
            """,
            """
            -- Add Data To Table: TLogins
            INSERT INTO TLogins (intLoginID, strLoginName, strPassword)
            VALUES   (1, 'Root.Admin', '203eb73702da069687657f3a5a9898e615cd6b49c10011ce7b8c8dea79bdc0420e0247d55594160ac520863ed30b882f51937adae77bd998ed650612e2a7e48d') 
                    ,(2, 'Carlos.Vidaurri', '71487d4ebb0db4bd0a3ee477add5883de7d750f0b10acb434e16bd9c343e3e89c1d257335529b7e58b6dd7f5526837a9596ab2c8a8cc2fb95a0d3c4792a7bd6f')
                    ,(3, 'David.Teixeira', '07a3c8b8d1ff7b7e57d174b37e67723f3f5c063a47f17b0e6cf214d30f03e1c94d147df394b559d4e67c7ef95840de193bd9d7551fe21e9768c0f000a624d4b4') 
            """,
            """
            -- Add Data To Table: TUserLogins
            INSERT INTO TUserLogins (intUserLoginID, intInspectorID, intLoginID)
            VALUES   (1, 1, 1)
                    ,(2, 2, 2)
                    ,(3, 3, 3)
            """,
            """
            -- Add Data To Table: TAdminUsers
            INSERT INTO TAdminUsers (intAdminUserID, intInspectorID, intLoginID)
            VALUES   (1, 1, 1)
                    ,(2, 2, 2)
                    ,(3, 3, 3)
            """,
            """
            -- Add Data To Table: TGymLocations
            INSERT INTO TGymLocations (intGymLocationID, strGymName, strAddress, intStateID, strZip)
            VALUES   (1, 'Hoosier Heights Indy', '9850 Mayflower Park Drive', 1, '46032')
                    ,(2, 'Hoosier Heights Bloomington', '1008 S. Rogers Street', 1, '47403')
                    ,(3, 'Climb Cincy', '1708 Blue Rock Street', 2, '45223')
                    ,(4, 'Climb Nulu', '1000 E. Market Street', 3, '40206')
            """]
            
            # Execute the SQL statements
            for sql in sqlInsert:
                Database.dbExeStatement(self, sql)       
            
            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")                                 

    def Create_Base_Functions(self):
        """ 
        Function Name: Create_Base_Functions
        Function Purpose: Insert the function statements into the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
        
            # Create the function names
            sqlFunctName = ['fn_GetInspectionStatus', 
                            'fn_GetInspectionType',
                            'fn_GetMetallicInspection',
                            'fn_GetTextileInspection',
                            'fn_GetPlasticInspection',
                            'fn_GetCarabiner',
                            'fn_GetCarabVisSelection',
                            'fn_GetCarabinerVisualResult',
                            'fn_GetCarabPhysSelection',
                            'fn_GetCarabinerPhysicalResult',
                            'fn_GetCarabFunctSelection',
                            'fn_GetCarabinerFunctionResult',
                            'fn_GetDeviceHandle',
                            'fn_GetHandVisSelection',
                            'fn_GetHandleVisualResult',
                            'fn_GetHandPhysSelection',
                            'fn_GetHandlePhysicalResult',
                            'fn_GetCaseCompSelection',
                            'fn_GetCaseHousingComponent',
                            'fn_GetCaseVisSelection',
                            'fn_GetCaseHousingVisualResult',
                            'fn_GetCasePhysSelection',
                            'fn_GetCaseHousingPhysicalResult',
                            'fn_GetBreakCompSelection',
                            'fn_GetBrakeHousingComponent',
                            'fn_GetBrakeVisSelection',
                            'fn_GetBrakeHousingVisualResult',
                            'fn_GetBrakePhysSelection',
                            'fn_GetBrakeHousingPhysicalResult',
                            'fn_GetLanyardLength',
                            'fn_GetLanVisSelection',
                            'fn_GetLanyardVisualResult',
                            'fn_GetLanPhysSelection',
                            'fn_GetLanyardPhysicalResult',
                            'fn_GetRetractFunctSelection',
                            'fn_GetLanyardRetractFunctionResult',
                            'fn_GetInspectorName',
                            'fn_GetInspectorEmail',
                            'fn_GetUserLoginName',
                            'fn_GetUserPassword',
                            'fn_GetWallLocationName',
                            'fn_GetAutoBelayDeviceName',
                            'fn_GetAutoBelaySerialNum',
                            'fn_GetAutoBelayBumperNum',
                            'fn_GetAutoBelayManuDate',
                            'fn_GetAutoBelayServiceDate',
                            'fn_GetAutoBelayReserviceDate',
                            'fn_GetAutoBelayInstallDate',
                            'fn_GetAutoBelayLastInspectDate',
                            'fn_GetAutoBelayNextInspectDate',
                            'fn_GetAutoBelayDeviceInUseStatus',
                            'fn_GetAutoBelayInspectComment']

            # Create the Tables for the Function
            sqlSelection = ['strInspectionStatusDesc',
                            'strInspectionTypeDesc',
                            'strMetallicInspectionDesc',
                            'strTextileInspectionDesc',
                            'strPlasticInspectionDesc',
                            'strCarabinerType',
                            'strCarabVisMetSelect',
                            'fn_GetCarabVisSelection(intCarabVisMetalSelectID)',
                            'strCarabPhysMetSelect',
                            'fn_GetCarabPhysSelection(intCarabPhysMetalSelectID)',
                            'strCarabFunctSelect',
                            'fn_GetCarabFunctSelection(intCarabFunctSelectID)',
                            'strHandleType',
                            'strHandVisMetSelect',
                            'fn_GetHandVisSelection(intHandVisMetalSelectID)',
                            'strHandPhysMetSelect',
                            'fn_GetHandPhysSelection(intHandPhysMetalSelectID)',
                            'strCaseCompSelect',
                            'strCaseComponentDesc',
                            'strCaseVisMetSelect',
                            'fn_GetCaseVisSelection(intCaseVisMetalSelectID)',
                            'strCasePhysMetSelect',
                            'fn_GetCasePhysSelection(intCasePhysMetalSelectID)',
                            'strBrakeCompSelect',
                            'strBrakeComponentDesc',
                            'strBrakeVisMetSelect',
                            'fn_GetBrakeVisSelection(intBrakeVisMetalSelectID)',
                            'strBrakePhysMetSelect',
                            'fn_GetBrakePhysSelection(intBrakePhysMetalSelectID)',
                            'strLanyardLength',
                            'strLanVisTextSelect',
                            'fn_GetLanVisSelection(intLanVisTextSelectID)',
                            'strLanPhysTextSelect',
                            'fn_GetLanPhysSelection(intLanPhysTextSelectID)',
                            'strRetractFunctSelect',
                            'fn_GetRetractFunctSelection(intRetractFunctSelectID)',
                            "strLastName || ',' || strFirstName",
                            'strEmail',
                            'strLoginName',
                            'strPassword',
                            'strWallLocationDesc',
                            'strDeviceName',
                            'strSerialNum',
                            'strBumperNum',
                            'dtmManufactureDate',
                            'dtmServiceDate',
                            'dtmReserviceDate',                            
                            'dtmInstallationDate',
                            'dtmLastInspectionDate',
                            'dtmNextInspectionDate',
                            'blnDeviceInUse',          
                            'strComment']

            # Create the Tables for the Function
            sqlTables = ['TInspectionStatus',
                        'TInspectionTypes',
                        'TMetallicInspections',
                        'TTextileInspections',
                        'TPlasticInspections',
                        'TCarabiners',
                        'TCarabVisMetalSelects',
                        'TCarabinerVisualInspections',
                        'TCarabPhysMetalSelects',
                        'TCarabinerPhysicalInspections',
                        'TCarabFunctSelects',
                        'TCarabinerFunctionInspections',
                        'TDeviceHandles',
                        'THandleVisMetalSelects',
                        'THandleVisualInspections',
                        'THandlePhysMetalSelects',
                        'THandlePhysicalInspections',
                        'TCaseCompSelects',
                        'TCaseHousings',
                        'TCaseVisMetalSelects',
                        'TCaseHousingVisualInspections',
                        'TCasePhysMetalSelects',
                        'TCaseHousingPhysicalInspections',
                        'TBrakeCompSelects',
                        'TBrakeHousings',
                        'TBrakeVisMetalSelects',
                        'TBrakeHousingVisualInspections',
                        'TBrakePhysMetalSelects',
                        'TBrakeHousingPhysicalInspections',
                        'TLanyards',
                        'TLanVisTextSelects',
                        'TLanyardVisualInspections',
                        'TLanPhysTextSelects',
                        'TLanyardPhysicalInspections',
                        'TRetractFunctSelects',
                        'TLanyardRetractFunctionInspections',
                        'TInspectors',
                        'TInspectors',
                        'TLogins',
                        'TLogins',
                        'TWallLocations',
                        'TAutoBelays',
                        'TAutoBelays',
                        'TAutoBelays',
                        'TAutoBelays',
                        'TAutoBelays',
                        'TAutoBelays',
                        'TAutoBelays',
                        'TAutoBelays',
                        'TAutoBelays',
                        'TAutoBelays',   
                        'TAutoBelayInspections']  
            
            # Create the Tables for the Function
            sqlPriKeyID = ['intInspectionStatusID',
                        'intInspectionTypeID',
                        'intMetallicInspectionID',
                        'intTextileInspectionID',
                        'intPlasticInspectionID',
                        'intCarabinerID',
                        'intCarabVisMetalSelectID',
                        'intCarabinerVisualID',
                        'intCarabPhysMetalSelectID',
                        'intCarabinerPhysicalID',
                        'intCarabFunctSelectID',
                        'intCarabinerFunctionInspectID',
                        'intDeviceHandleID',
                        'intHandVisMetalSelectID',
                        'intHandleVisualID',
                        'intHandPhysMetalSelectID',
                        'intHandlePhysicalID',
                        'intCaseCompSelectID',
                        'intCaseHousingID',
                        'intCaseVisMetalSelectID',
                        'intCaseHousingVisualInspectionID',
                        'intCasePhysMetalSelectID',
                        'intCaseHousingPhysicalInspectionID',
                        'intBrakeCompSelectID',
                        'intBrakeHousingID',
                        'intBrakeVisMetalSelectID',
                        'intBrakeHousingVisualInspectionID',
                        'intBrakePhysMetalSelectID',
                        'intBrakeHousingPhysicalInspectionID',
                        'intLanyardID',
                        'intLanVisTextSelectID',
                        'intLanyardVisualInspectionID',
                        'intLanPhysTextSelectID',
                        'intLanyardPhysicalInspectionID',
                        'intRetractFunctSelectID',
                        'intLanyardRetractFunctionInspectionID',
                        'intInspectorID',
                        'intInspectorID',
                        'intLoginID',
                        'intLoginID',
                        'intWallLocationID',
                        'intAutoBelayID',
                        'intAutoBelayID',
                        'intAutoBelayID',
                        'intAutoBelayID',
                        'intAutoBelayID',
                        'intAutoBelayID',
                        'intAutoBelayID',
                        'intAutoBelayID',
                        'intAutoBelayID',
                        'intAutoBelayID',    
                        'intAutoBelayID']  
                              
            # Execute the SQL statements
            for i in range(len(sqlFunctName)):
                Queries.dbCreate_Functions(self, sqlFunctName[i], sqlSelection[i], sqlTables[i], sqlPriKeyID[i])       

        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")   

    def Create_Base_Views(self, target_db="main"):
        """ 
        Function Name: Create_Base_Views
        Function Purpose: Insert the view statements into the database
        """          
        # Create the view names
        sqlViewName = ['vInspectors', 
                        'vInspectStatus',
                        'vInspectTypes',
                        'vMetallicInspect',
                        'vPlasticInspect',
                        'vTextileInspect',
                        'vCarabinerType',
                        'vCarabinerFunct',
                        'vDeviceHandleComp',
                        'vCaseHouseComp',
                        'vBrakeHouseComp',
                        'vRetractFunctions',
                        'vWallLocations',
                        'vLogins',
                        'vAutoBelays',
                        'vAutoBelayDates',
                        'vAutoBelayLastNextInspectDates',
                        'vAutoBelayInspectors',
                        'vAutoBelayWallLocations',
                        'vABWallLocationInspectDates',
                        'vABServiceDates',
                        'vAutoBelayInspectResults',
                        'vAutoBelayFailedResults',
                        'vAutoBelayMonitorResults',
                        'vAutoBelayFailNowResults',
                        'vAutoBelayReserviceReports',
                        'vRopes',
                        'vRopeDates',
                        'vRopeLastNextInspectDates',
                        'vRopeInspectors',
                        'vRopeWallLocations',
                        'vRopeWallLocationInspectDates',
                        'vRopeInspectResults',
                        'vRopeFailedResults',
                        'vRopeMonitorResults',
                        'vRopeFailNowResults',
                        'vConnectors',
                        'vConnectorDates',
                        'vConnectorLastNextInspectDates',
                        'vConnectorInspectors',
                        'vConnectorWallLocations',
                        'vConnectorWallLocationInspectDates',
                        'vConnectorInspectResults',
                        'vConnectorFailedResults',
                        'vConnectorMonitorResults',
                        'vConnectorFailNowResults',
                        'vBelayDevices',
                        'vBelayDeviceDates',
                        'vBelayDeviceLastNextInspectDates',
                        'vBelayDeviceInspectors',
                        'vBelayDeviceWallLocations',
                        'vBelayDeviceWallLocationInspectDates',
                        'vBelayDeviceInspectResults',
                        'vBelayDeviceFailedResults',
                        'vBelayDeviceMonitorResults',
                        'vBelayDeviceFailNowResults',
                        'vABInspectID_OutForReservice',
                        ]
                        
        # Create the sqlStatement for the view
        sqlStatement = [
                        # vInspectors
                        "SELECT strLastName || ',' || strFirstName                   AS InspectorName                FROM TInspectors",
                        # vInspectStatus
                        'SELECT strInspectionStatusDesc                              AS InspectionStatusDesc         FROM TInspectionStatus',
                        # vInspectTypes
                        'SELECT strInspectionTypeDesc                                AS InspectionTypeDesc           FROM TInspectionTypes',
                        # vMetallicInspect
                        'SELECT strMetallicInspectionDesc                            AS MetallicInspectionDesc       FROM TMetallicInspections',
                        # vPlasticInspect
                        'SELECT strPlasticInspectionDesc                             AS PlasticInspectionDesc        FROM TPlasticInspections',
                        # vTextileInspect
                        'SELECT strTextileInspectionDesc                             AS TextileInspectionDesc        FROM TTextileInspections',
                        # vCarabinerType
                        'SELECT strCarabinerType                                     AS CarabinerType                FROM TCarabiners',
                        # vCarabinerFunct
                        'SELECT strCarabinerFunctionDesc                             AS CarabinerFunctionDesc        FROM TCarabinerFunctions',
                        # vDeviceHandleComp
                        'SELECT strHandleType                                        AS HandleType                   FROM TDeviceHandles',
                        # vCaseHouseComp
                        'SELECT strCaseComponentDesc                                 AS CaseComponentDesc            FROM TCaseHousings',
                        # vBrakeHouseComp
                        'SELECT strBrakeComponentDesc                                AS BrakeComponentDesc           FROM TBrakeHousings',
                        # vRetractFunctions
                        'SELECT strRetractFunctionDesc                               AS RetractFunctionDescResult    FROM TLanyardRetractFunctionInspections',
                        # vWallLocations
                        'SELECT strWallLocationDesc                                  AS WallLocationDesc             FROM TWallLocations',
                        # vLogins
                        'SELECT strLoginName                                         AS LoginName                    FROM TLogins',
                        #vAutoBelays
                        """SELECT                                                             
                                strDeviceName                                        AS DeviceName
                                ,strSerialNum                                        AS AutoBelaySerialNumber
                                ,strBumperNum                                        AS BumperNum
                                ,dtmManufactureDate                                  AS ManufactureDate
                                ,dtmServiceDate                                      AS ServiceDate
                                ,dtmReserviceDate                                    AS ReserviceDate
                                ,dtmInstallationDate                                 AS InstallationDate
                                ,dtmLastInspectionDate                               AS LastInspectionDate
                                ,dtmNextInspectionDate                               AS NextInspectionDate
                                ,blnDeviceInUse                                      AS DeviceInUse
                            FROM
                                TAutoBelays;""",
                        #vAutoBelayDates
                        """SELECT 
                            -- Create View:  Show all AutoBelay Dates where serial number equals the selected primary key from user
                                strSerialNum                                         AS AutoBelaySerialNumber
                                ,strBumperNum                                        AS BumperNum
                                ,dtmManufactureDate                                  AS ManufactureDate
                                ,dtmServiceDate                                      AS ServiceDate
                                ,dtmReserviceDate                                    AS ReserviceDate
                                ,dtmInstallationDate                                 AS InstallationDate
                                ,dtmLastInspectionDate                               AS LastInspectionDate
                                ,dtmNextInspectionDate                               AS NextInspectionDate       
                            FROM TAutoBelays;""",
                        # vAutoBelayLastNextInspectDates
                        """SELECT
                        -- Create View: Show all AutoBelay Dates of the serial number and last/next inspection dates
                                strSerialNum                                         AS AutoBelaySerialNumber
                                ,strBumperNum                                        AS BumperNum
                                ,dtmLastInspectionDate                               AS LastInspectionDate
                                ,dtmNextInspectionDate                               AS NextInspectionDate           
                            FROM TAutoBelays;""",                            
                        # vAutoBelayInspectors
                        """SELECT
                        -- Create View: Show all AutoBelay Inspectors by date of last inspection and the upcoming inspection
                                TI.strLastName || ',' || TI.strFirstName             AS InspectorName
                                ,TAB.strSerialNum                                    AS AutoBelaySerialNumber
                                ,TAB.strBumperNum                                    AS BumperNum
                                ,TAB.dtmLastInspectionDate                           AS LastInspectionDate
                                ,TAB.dtmNextInspectionDate                           AS NextInspectionDate 
                            FROM
                                TInspectors AS TI
                                LEFT JOIN TAutoBelays AS TAB ON TAB.intAutoBelayID = TABI.intAutoBelayID
                                LEFT JOIN TAutoBelayInspectors AS TABI ON TI.intInspectorID = TABI.intInspectorID
                            WHERE 
                                TI.intInspectorID IN (SELECT TABI.intInspectorID FROM TAutoBelayInspectors AS TABI)
                                AND TAB.intAutoBelayID IN (SELECT TABI.intAutoBelayID FROM TAutoBelayInspectors AS TABI);""",
                        # vAutoBelayWallLocations
                        """SELECT
                        -- Create View: Show all Current AutoBelay Wall Location.
                                TAB.strSerialNum                                     AS AutoBelaySerialNumber
                                ,TAB.strBumperNum                                    AS BumperNum
                                ,TWL.strWallLocationDesc                             AS WallLocationName
                            FROM
                                TWallLocations AS TWL
                                INNER JOIN TAutoBelayWallLocations AS TABWL ON TWL.intWallLocationID = TABWL.intWallLocationID
                                INNER JOIN TAutoBelays AS TAB ON TAB.intAutoBelayID = TABWL.intAutoBelayID
                            WHERE
                                TAB.blnDeviceInUse = 'Yes');""",                                    
                        # vABWallLocationInspectDates
                        """SELECT
                        -- Create View: Show all Current AutoBelay Wall Location. Include last and next inspection date
                                TWL.strWallLocationDesc                              AS WallLocationName
                                ,TAB.strSerialNum                                    AS AutoBelaySerialNumber
                                ,TAB.strBumperNum                                    AS BumperNum
                                ,TAB.dtmServiceDate                                  AS ServiceDate
                                ,TAB.dtmReserviceDate                                AS ReServiceDate
                                ,TAB.dtmLastInspectionDate                           AS LastInspectionDate
                                ,TAB.dtmNextInspectionDate                           AS NextInspectionDate
                            FROM
                                TWallLocations AS TWL
                                LEFT JOIN TAutoBelayWallLocations AS TABWL ON TWL.intWallLocationID = TABWL.intWallLocationID
                                LEFT JOIN TAutoBelays AS TAB ON TAB.intAutoBelayID = TABWL.intAutoBelayID
                            WHERE
                                TAB.intAutoBelayID IN (SELECT TABWL.intAutoBelayID FROM TAutoBelayWallLocations AS TABWL);""",
                        # vABServiceDates
                        """SELECT
                        -- Create View: Show all Current AutoBelay service and reservice dates
                                TAB.strSerialNum                                    AS AutoBelaySerialNumber
                                ,TAB.strBumperNum                                   AS BumperNum
                                ,TAB.dtmServiceDate                                 AS ServiceDate
                                ,TAB.dtmReserviceDate                               AS ReServiceDate
                            FROM
                                TAutoBelays AS TAB;""",                                                                                                        
                        # vAutoBelayInspectResults
                        """SELECT
                        -- Create View: Show all AutoBelay component inspection status. Include inspector name and 
                        -- last/next inspection date 
                                TAB.strSerialNum                                    AS AutoBelaySerialNumber
                                ,TAB.strBumperNum                                   AS BumperNum
                                ,TWL.strWallLocationDesc                            AS WallLocationName
                                ,TINS.strInspectionStatusDesc                       AS OverallStatus
                                ,TIN.strLastName || ',' || TIN.strFirstName         AS InspectorName
                                ,TAB.dtmLastInspectionDate                          AS LastInspectionDate
                                ,TAB.dtmNextInspectionDate                          AS NextInspectionDate
                                ,TAB.dtmServiceDate                                 AS ServiceDate
                                ,TAB.dtmReserviceDate                               AS ReServiceDate
                            FROM 
                                TAutoBelays AS TAB 
                                LEFT JOIN TAutoBelayInspections AS TABIN ON TAB.intAutoBelayID = TABIN.intAutoBelayID
                                LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TABIN.intInspectionStatusID
                                LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TABIN.intInspectorID
                                LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TABIN.intWallLocationID
                            WHERE
                                TAB.intAutoBelayID IN (SELECT TABIN.intAutoBelayID FROM TAutoBelayInspections AS TABIN);""",
                        # vAutoBelayFailedResults
                        """SELECT
                        -- Create View: Show all failed AutoBelays
                                TAB.strSerialNum                                    AS AutoBelaySerialNumber
                                ,TAB.strBumperNum                                   AS BumperNum
                                ,TWL.strWallLocationDesc                            AS WallLocationName
                                ,TINS.strInspectionStatusDesc                       AS OverallStatus
                                ,TIN.strLastName || ',' || TIN.strFirstName         AS InspectorName
                                ,TAB.dtmLastInspectionDate                          AS LastInspectionDate
                                ,TAB.dtmNextInspectionDate                          AS NextInspectionDate
                                ,TAB.dtmServiceDate                                 AS ServiceDate
                                ,TAB.dtmReserviceDate                               AS ReServiceDate
                            FROM 
                                TAutoBelays AS TAB 
                                LEFT JOIN TAutoBelayInspections AS TABIN ON TAB.intAutoBelayID = TABIN.intAutoBelayID
                                LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TABIN.intInspectionStatusID
                                LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TABIN.intInspectorID
                                LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TABIN.intWallLocationID
                            WHERE
                                TAB.intAutoBelayID IN (SELECT TABIN.intAutoBelayID FROM TAutoBelayInspections AS TABIN
                                                        WHERE TABIN.intInspectionStatusID = 3);""",                                    
                        # vAutoBelayMonitorResults
                        """SELECT
                        -- Create View: Show all monitor AutoBelays
                                TAB.strSerialNum                                    AS AutoBelaySerialNumber
                                ,TAB.strBumperNum                                   AS BumperNum
                                ,TWL.strWallLocationDesc                            AS WallLocationName
                                ,TINS.strInspectionStatusDesc                       AS OverallStatus
                                ,TIN.strLastName || ',' || TIN.strFirstName         AS InspectorName
                                ,TAB.dtmLastInspectionDate                          AS LastInspectionDate
                                ,TAB.dtmNextInspectionDate                          AS NextInspectionDate
                                ,TAB.dtmServiceDate                                 AS ServiceDate
                                ,TAB.dtmReserviceDate                               AS ReServiceDate
                            FROM 
                                TAutoBelays AS TAB 
                                LEFT JOIN TAutoBelayInspections AS TABIN ON TAB.intAutoBelayID = TABIN.intAutoBelayID
                                LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TABIN.intInspectionStatusID
                                LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TABIN.intInspectorID
                                LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TABIN.intWallLocationID
                            WHERE
                                TAB.intAutoBelayID IN (SELECT TABIN.intAutoBelayID FROM TAutoBelayInspections AS TABIN
                                                        WHERE TABIN.intInspectionStatusID = 2);""", 
                        # vAutoBelayFailNowResults
                        """SELECT
                        -- Create View: Show all failed now AutoBelays
                                TAB.strSerialNum                                     AS AutoBelaySerialNumber
                                ,TAB.strBumperNum                                    AS BumperNum
                                ,TWL.strWallLocationDesc                             AS WallLocationName
                                ,TINS.strInspectionStatusDesc                        AS OverallStatus
                                ,TIN.strLastName || ',' || TIN.strFirstName          AS InspectorName
                                ,TAB.dtmLastInspectionDate                           AS LastInspectionDate
                                ,TAB.dtmNextInspectionDate                           AS NextInspectionDate
                                ,TAB.dtmServiceDate                                  AS ServiceDate
                                ,TAB.dtmReserviceDate                                AS ReServiceDate
                            FROM 
                                TAutoBelays AS TAB 
                                LEFT JOIN TAutoBelayInspections AS TABIN ON TAB.intAutoBelayID = TABIN.intAutoBelayID
                                LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TABIN.intInspectionStatusID
                                LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TABIN.intInspectorID
                                LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TABIN.intWallLocationID
                            WHERE
                                TABIN.intInspectionStatusID = 3 AND date(TAB.dtmLastInspectionDate) = date('now');""",
                        # vAutoBelayReserviceReports                  
                        """SELECT
                                TAB.strSerialNum                                     AS AutoBelaySerialNumber
                                ,TAB.strBumperNum                                    AS BumperNum
                                ,TAB.dtmLastInspectionDate                           AS LastInspectionDate
                                ,TAB.dtmServiceDate                                  AS ServiceDate
                                ,TAB.dtmReserviceDate                                AS ReServiceDate
                            FROM
                                TAutoBelays AS TAB
                            WHERE
                                TAB.blnDeviceInUse = 'Out For Reservice';""", 
                        #vRopes
                        """SELECT                                                                                                 
                                strSerialNum                                         AS RopeSerialNumber
                                ,strBumperNum                                        AS BumperNum
                                ,strRopeLength                                       AS RopeLength
                                ,strDiameter                                         AS Diameter
                                ,strElasticity                                       AS Elasticity
                                ,strManufactureName                                  AS ManufacturerName
                                ,dtmManufactureDate                                  AS ManufactureDate                                                                        
                                ,dtmInstallationDate                                 AS InstallationDate
                                ,dtmLastInspectionDate                               AS LastInspectionDate
                                ,dtmNextInspectionDate                               AS NextInspectionDate
                                ,strEquipInUse                                       AS RopeInUse
                            FROM
                                TRopes;""",
                        #vRopeDates
                        """SELECT 
                            -- Create View:  Show all Rope Dates where serial number equals the selected primary key from user
                                strSerialNum                                         AS RopeSerialNumber
                                ,strBumperNum                                        AS BumperNum
                                ,dtmManufactureDate                                  AS ManufactureDate                                                                        
                                ,dtmInstallationDate                                 AS InstallationDate
                                ,dtmLastInspectionDate                               AS LastInspectionDate
                                ,dtmNextInspectionDate                               AS NextInspectionDate       
                            FROM 
                                TRopes;""",
                        # vRopeLastNextInspectDates
                        """SELECT
                        -- Create View: Show all Rope Dates of the serial number and last/next inspection dates
                                strSerialNum                                         AS RopeSerialNumber
                                ,strBumperNum                                        AS BumperNum
                                ,dtmLastInspectionDate                               AS LastInspectionDate
                                ,dtmNextInspectionDate                               AS NextInspectionDate           
                            FROM TRopes;""",                            
                        # vRopeInspectors
                        """SELECT
                        -- Create View: Show all Rope Inspectors by date of last inspection and the upcoming inspection
                                TI.strLastName || ',' || TI.strFirstName             AS InspectorName
                                ,TR.strSerialNum                                     AS RopeSerialNumber
                                ,TR.strBumperNum                                     AS BumperNum
                                ,TR.dtmLastInspectionDate                            AS LastInspectionDate
                                ,TR.dtmNextInspectionDate                            AS NextInspectionDate 
                            FROM
                                TInspectors AS TI
                                LEFT JOIN TRopes AS TR ON TR.intRopeID = TRI.intRopeID
                                LEFT JOIN TRopeInspectors AS TRI ON TI.intInspectorID = TRI.intInspectorID
                            WHERE 
                                TI.intInspectorID IN (SELECT TRI.intInspectorID FROM TRopeInspectors AS TRI)
                                AND TR.intRopeID IN (SELECT TRI.intRopeID FROM TRopeInspectors AS TRI);""",
                        # vRopeWallLocations
                        """SELECT
                        -- Create View: Show all Current Rope Wall Location.
                                TR.strSerialNum                                      AS RopeSerialNumber
                                ,TR.strBumperNum                                     AS BumperNum
                                ,TWL.strWallLocationDesc                             AS WallLocationName
                            FROM
                                TWallLocations AS TWL
                                INNER JOIN TRopeWallLocations AS TRWL ON TWL.intWallLocationID = TRWL.intWallLocationID
                                INNER JOIN TRopes AS TR ON TR.intRopeID = TRWL.intRopeID
                            WHERE
                                TR.intRopeID IN (SELECT TR.intRopeID FROM TRopes AS TR WHERE TR.strEquipInUse = 'Yes');""",                                    
                        # vRopeWallLocationInspectDates
                        """SELECT
                        -- Create View: Show all Current Rope Wall Location. Include last and next inspection date
                                TWL.strWallLocationDesc                              AS WallLocationName
                                ,TR.strSerialNum                                     AS RopeSerialNumber
                                ,TR.strBumperNum                                     AS BumperNum
                                ,TR.dtmLastInspectionDate                            AS LastInspectionDate
                                ,TR.dtmNextInspectionDate                            AS NextInspectionDate
                            FROM
                                TWallLocations AS TWL
                                LEFT JOIN TRopeWallLocations AS TRWL ON TWL.intWallLocationID = TRWL.intWallLocationID
                                LEFT JOIN TRopes AS TR ON TR.intRopeID = TRWL.intRopeID
                            WHERE
                                TR.intRopeID IN (SELECT TRWL.intRopeID FROM TRopeWallLocations AS TRWL);""",                                                                                                    
                        # vRopeInspectResults
                        """SELECT
                        -- Create View: Show all Rope component inspection status. Include inspector name and 
                        -- last/next inspection date 
                                TR.strSerialNum                                     AS RopeSerialNumber
                                ,TR.strBumperNum                                    AS BumperNum
                                ,TWL.strWallLocationDesc                            AS WallLocationName
                                ,TINS.strInspectionStatusDesc                       AS OverallStatus
                                ,TIN.strLastName || ',' || TIN.strFirstName         AS InspectorName
                                ,TR.dtmLastInspectionDate                           AS LastInspectionDate
                                ,TR.dtmNextInspectionDate                           AS NextInspectionDate
                            FROM 
                                TRopes AS TR 
                                LEFT JOIN TRopeInspections AS TRIN ON TR.intRopeID = TRIN.intRopeID
                                LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TRIN.intInspectionStatusID
                                LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TRIN.intInspectorID
                                LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TRIN.intWallLocationID
                            WHERE
                                TR.intRopeID IN (SELECT TRIN.intRopeID FROM TRopeInspections AS TRIN);""",
                        # vRopeFailedResults
                        """SELECT
                        -- Create View: Show all failed Ropes
                                TR.strSerialNum                                     AS RopeSerialNumber
                                ,TR.strBumperNum                                    AS BumperNum
                                ,TWL.strWallLocationDesc                            AS WallLocationName
                                ,TINS.strInspectionStatusDesc                       AS OverallStatus
                                ,TIN.strLastName || ',' || TIN.strFirstName         AS InspectorName
                                ,TR.dtmLastInspectionDate                           AS LastInspectionDate
                                ,TR.dtmNextInspectionDate                           AS NextInspectionDate
                            FROM 
                                TRopes AS TR 
                                LEFT JOIN TRopeInspections AS TRIN ON TR.intRopeID = TRIN.intRopeID
                                LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TRIN.intInspectionStatusID
                                LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TRIN.intInspectorID
                                LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TRIN.intWallLocationID
                            WHERE
                                TR.intRopeID IN (SELECT TRIN.intRopeID FROM TRopeInspections AS TRIN
                                                        WHERE TRIN.intInspectionStatusID = 3);""",                                    
                        # vRopeMonitorResults
                        """SELECT
                        -- Create View: Show all monitor Ropes
                                TR.strSerialNum                                     AS RopeSerialNumber
                                ,TR.strBumperNum                                    AS BumperNum
                                ,TWL.strWallLocationDesc                            AS WallLocationName
                                ,TINS.strInspectionStatusDesc                       AS OverallStatus
                                ,TIN.strLastName || ',' || TIN.strFirstName         AS InspectorName
                                ,TR.dtmLastInspectionDate                           AS LastInspectionDate
                                ,TR.dtmNextInspectionDate                           AS NextInspectionDate
                            FROM 
                                TRopes AS TR 
                                LEFT JOIN TRopeInspections AS TRIN ON TR.intRopeID = TRIN.intRopeID
                                LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TRIN.intInspectionStatusID
                                LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TRIN.intInspectorID
                                LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TRIN.intWallLocationID
                            WHERE
                                TR.intRopeID IN (SELECT TRIN.intRopeID FROM TRopeInspections AS TRIN
                                                        WHERE TRIN.intInspectionStatusID = 2);""", 
                        # vRopeFailNowResults
                        """SELECT
                        -- Create View: Show all failed now Ropes
                                TR.strSerialNum                                         AS RopeSerialNumber
                                ,TR.strBumperNum                                        AS BumperNum
                                ,TWL.strWallLocationDesc                                AS WallLocationName
                                ,TINS.strInspectionStatusDesc                           AS OverallStatus
                                ,TIN.strLastName || ',' || TIN.strFirstName             AS InspectorName
                                ,TR.dtmLastInspectionDate                               AS LastInspectionDate
                                ,TR.dtmNextInspectionDate                               AS NextInspectionDate
                            FROM 
                                TRopes AS TR 
                                LEFT JOIN TRopeInspections AS TRIN ON TR.intRopeID = TRIN.intRopeID
                                LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TRIN.intInspectionStatusID
                                LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TRIN.intInspectorID
                                LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TRIN.intWallLocationID
                            WHERE
                                RIN.intInspectionStatusID = 3 AND date(TR.dtmLastInspectionDate) = date('now');""",
                        #vConnectors
                        """SELECT                                                                                                 
                                strSerialNum                                         AS ConnectorSerialNumber
                                ,strBumperNum                                        AS BumperNum
                                ,strManufactureName                                  AS ManufacturerName
                                ,dtmManufactureDate                                  AS ManufactureDate                                                                        
                                ,dtmInstallationDate                                 AS InstallationDate
                                ,dtmLastInspectionDate                               AS LastInspectionDate
                                ,dtmNextInspectionDate                               AS NextInspectionDate
                                ,strDeviceType                                       AS DeviceType
                                ,strEquipInUse                                       AS ConnectorInUse
                            FROM
                                TConnectors;""",
                        #vConnectorDates
                        """SELECT 
                            -- Create View:  Show all Connector Dates where serial number equals the selected primary key from user
                                strSerialNum                                         AS ConnectorSerialNumber
                                ,strBumperNum                                        AS BumperNum
                                ,dtmManufactureDate                                  AS ManufactureDate                                                                        
                                ,dtmInstallationDate                                 AS InstallationDate
                                ,dtmLastInspectionDate                               AS LastInspectionDate
                                ,dtmNextInspectionDate                               AS NextInspectionDate       
                            FROM TConnectors;""",
                        # vConnectorLastNextInspectDates
                        """SELECT
                        -- Create View: Show all Connector Dates of the serial number and last/next inspection dates
                                strSerialNum                                         AS ConnectorSerialNumber
                                ,strBumperNum                                        AS BumperNum
                                ,dtmLastInspectionDate                               AS LastInspectionDate
                                ,dtmNextInspectionDate                               AS NextInspectionDate           
                            FROM TConnectors;""",                            
                        # vConnectorInspectors
                        """SELECT
                        -- Create View: Show all Connector Inspectors by date of last inspection and the upcoming inspection
                                TI.strLastName || ',' || TI.strFirstName              AS InspectorName
                                ,TCN.strSerialNum                                     AS ConnectorSerialNumber
                                ,TCN.strBumperNum                                     AS BumperNum
                                ,TCN.dtmLastInspectionDate                            AS LastInspectionDate
                                ,TCN.dtmNextInspectionDate                            AS NextInspectionDate 
                            FROM
                                TInspectors AS TI
                                LEFT JOIN TConnectors AS TCN ON TCN.intConnectorID = TCNI.intConnectorID
                                LEFT JOIN TConnectorInspectors AS TCNI ON TI.intInspectorID = TCNI.intInspectorID
                            WHERE 
                                TI.intInspectorID IN (SELECT TCNI.intInspectorID FROM TConnectorInspectors AS TCNI)
                                AND TCN.intConnectorID IN (SELECT TCNI.intConnectorID FROM TConnectorInspectors AS TCNI);""",
                        # vConnectorWallLocations
                        """SELECT
                        -- Create View: Show all Current Connector Wall Location.
                                TCN.strSerialNum                                      AS ConnectorSerialNumber
                                ,TCN.strBumperNum                                     AS BumperNum
                                ,TWL.strWallLocationDesc                              AS WallLocationName
                            FROM
                                TWallLocations AS TWL
                                INNER JOIN TConnectorWallLocations AS TCNWL ON TWL.intWallLocationID = TCNWL.intWallLocationID
                                INNER JOIN TConnectors AS TCN ON TCN.intConnectorID = TCNWL.intConnectorID
                            WHERE
                                TCN.intConnectorID IN (SELECT TCN.intConnectorID FROM TConnectors AS TCN WHERE TCN.strEquipInUse = 'Yes');""",                                    
                        # vConnectorWallLocationInspectDates
                        """SELECT
                        -- Create View: Show all Current Connector Wall Location. Include last and next inspection date
                                TWL.strWallLocationDesc                               AS WallLocationName
                                ,TCN.strSerialNum                                     AS ConnectorSerialNumber
                                ,TCN.strBumperNum                                     AS BumperNum
                                ,TCN.dtmLastInspectionDate                            AS LastInspectionDate
                                ,TCN.dtmNextInspectionDate                            AS NextInspectionDate
                            FROM
                                TWallLocations AS TWL
                                LEFT JOIN TConnectorWallLocations AS TCNWL ON TWL.intWallLocationID = TCNWL.intWallLocationID
                                LEFT JOIN TConnectors AS TCN ON TCN.intConnectorID = TCNWL.intConnectorID
                            WHERE
                                TCN.intConnectorID IN (SELECT TCNWL.intConnectorID FROM TConnectorWallLocations AS TCNWL);""",                                                                                                    
                        # vConnectorInspectResults
                        """SELECT
                        -- Create View: Show all Connector component inspection status. Include inspector name and 
                        -- last/next inspection date 
                                TCN.strSerialNum                                     AS ConnectorSerialNumber
                                ,TCN.strBumperNum                                    AS BumperNum
                                ,TWL.strWallLocationDesc                             AS WallLocationName
                                ,TINS.strInspectionStatusDesc                        AS OverallStatus
                                ,TIN.strLastName || ',' || TIN.strFirstName          AS InspectorName
                                ,TCN.dtmLastInspectionDate                           AS LastInspectionDate
                                ,TCN.dtmNextInspectionDate                           AS NextInspectionDate
                            FROM 
                                TConnectors AS TCN 
                                LEFT JOIN TConnectorInspections AS TCNIN ON TCN.intConnectorID = TCNIN.intConnectorID
                                LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TCNIN.intInspectionStatusID
                                LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TCNIN.intInspectorID
                                LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TCNIN.intWallLocationID
                            WHERE
                                TCN.intConnectorID IN (SELECT TCNIN.intConnectorID FROM TConnectorInspections AS TCNIN);""",
                        # vConnectorFailedResults
                        """SELECT
                        -- Create View: Show all failed Connectors
                                TCN.strSerialNum                                     AS ConnectorSerialNumber
                                ,TCN.strBumperNum                                    AS BumperNum
                                ,TWL.strWallLocationDesc                             AS WallLocationName
                                ,TINS.strInspectionStatusDesc                        AS OverallStatus
                                ,TIN.strLastName || ',' || TIN.strFirstName          AS InspectorName
                                ,TCN.dtmLastInspectionDate                           AS LastInspectionDate
                                ,TCN.dtmNextInspectionDate                           AS NextInspectionDate
                            FROM 
                                TConnectors AS TCN 
                                LEFT JOIN TConnectorInspections AS TCNIN ON TCN.intConnectorID = TCNIN.intConnectorID
                                LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TCNIN.intInspectionStatusID
                                LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TCNIN.intInspectorID
                                LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TCNIN.intWallLocationID
                            WHERE
                                TCN.intConnectorID IN (SELECT TCNIN.intConnectorID FROM TConnectorInspections AS TCNIN
                                                        WHERE TCNIN.intInspectionStatusID = 3);""",                                    
                        # vConnectorMonitorResults
                        """SELECT
                        -- Create View: Show all monitor Connectors
                                TCN.strSerialNum                                     AS ConnectorSerialNumber
                                ,TCN.strBumperNum                                    AS BumperNum
                                ,TWL.strWallLocationDesc                             AS WallLocationName
                                ,TINS.strInspectionStatusDesc                        AS OverallStatus
                                ,TIN.strLastName || ',' || TIN.strFirstName          AS InspectorName
                                ,TCN.dtmLastInspectionDate                           AS LastInspectionDate
                                ,TCN.dtmNextInspectionDate                           AS NextInspectionDate
                            FROM 
                                TConnectors AS TCN 
                                LEFT JOIN TConnectorInspections AS TCNIN ON TCN.intConnectorID = TCNIN.intConnectorID
                                LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TCNIN.intInspectionStatusID
                                LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TCNIN.intInspectorID
                                LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TCNIN.intWallLocationID
                            WHERE
                                TCN.intConnectorID IN (SELECT TCNIN.intConnectorID FROM TConnectorInspections AS TCNIN
                                                        WHERE TCNIN.intInspectionStatusID = 2);""", 
                        # vConnectorFailNowResults
                        """SELECT
                        -- Create View: Show all failed now Connectors
                                TCN.strSerialNum                                         AS ConnectorSerialNumber
                                ,TCN.strBumperNum                                        AS BumperNum
                                ,TWL.strWallLocationDesc                                 AS WallLocationName
                                ,TINS.strInspectionStatusDesc                            AS OverallStatus
                                ,TIN.strLastName || ',' || TIN.strFirstName              AS InspectorName
                                ,TCN.dtmLastInspectionDate                               AS LastInspectionDate
                                ,TCN.dtmNextInspectionDate                               AS NextInspectionDate
                            FROM 
                                TConnectors AS TCN 
                                LEFT JOIN TConnectorInspections AS TCNIN ON TCN.intConnectorID = TCNIN.intConnectorID
                                LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TCNIN.intInspectionStatusID
                                LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TCNIN.intInspectorID
                                LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TCNIN.intWallLocationID
                            WHERE
                                TCNIN.intInspectionStatusID = 3 AND date(TCN.dtmLastInspectionDate) = date('now');""",
                        #vBelayDevices
                        """SELECT
                                strBelayDeviceName                                   AS BelayDeviceName
                                ,strSerialNum                                        AS BelayDeviceSerialNumber
                                ,strBumperNum                                        AS BumperNum
                                ,strManufactureName                                  AS ManufacturerName
                                ,dtmManufactureDate                                  AS ManufactureDate                                                                        
                                ,dtmInstallationDate                                 AS InstallationDate
                                ,dtmLastInspectionDate                               AS LastInspectionDate
                                ,dtmNextInspectionDate                               AS NextInspectionDate
                                ,strDeviceType                                       AS DeviceType
                                ,strEquipInUse                                       AS BelayDeviceInUse
                            FROM
                                TBelayDevices;""",
                        #vBelayDeviceDates
                        """SELECT 
                            -- Create View:  Show all BelayDevice Dates where serial number equals the selected primary key from user
                                strSerialNum                                         AS BelayDeviceSerialNumber
                                ,strBumperNum                                        AS BumperNum
                                ,dtmManufactureDate                                  AS ManufactureDate                                                                        
                                ,dtmInstallationDate                                 AS InstallationDate
                                ,dtmLastInspectionDate                               AS LastInspectionDate
                                ,dtmNextInspectionDate                               AS NextInspectionDate       
                            FROM TBelayDevices;""",
                        # vBelayDeviceLastNextInspectDates
                        """SELECT
                        -- Create View: Show all BelayDevice Dates of the serial number and last/next inspection dates
                                strSerialNum                                         AS BelayDeviceSerialNumber
                                ,strBumperNum                                        AS BumperNum
                                ,dtmLastInspectionDate                               AS LastInspectionDate
                                ,dtmNextInspectionDate                               AS NextInspectionDate           
                            FROM TBelayDevices;""",                            
                        # vBelayDeviceInspectors
                        """SELECT
                        -- Create View: Show all BelayDevice Inspectors by date of last inspection and the upcoming inspection
                                TI.strLastName || ',' || TI.strFirstName              AS InspectorName
                                ,TBD.strSerialNum                                     AS BelayDeviceSerialNumber
                                ,TBD.strBumperNum                                     AS BumperNum
                                ,TBD.dtmLastInspectionDate                            AS LastInspectionDate
                                ,TBD.dtmNextInspectionDate                            AS NextInspectionDate 
                            FROM
                                TInspectors AS TI
                                LEFT JOIN TBelayDevices AS TBD ON TBD.intBelayDeviceID = TBDI.intBelayDeviceID
                                LEFT JOIN TBelayDeviceInspectors AS TBDI ON TI.intInspectorID = TBDI.intInspectorID
                            WHERE 
                                TI.intInspectorID IN (SELECT TBDI.intInspectorID FROM TBelayDeviceInspectors AS TBDI)
                                AND TBD.intBelayDeviceID IN (SELECT TBDI.intBelayDeviceID FROM TBelayDeviceInspectors AS TBDI);""",
                        # vBelayDeviceWallLocations
                        """SELECT
                        -- Create View: Show all Current BelayDevice Wall Location.
                                TBD.strSerialNum                                      AS BelayDeviceSerialNumber
                                ,TBD.strBumperNum                                     AS BumperNum
                                ,TWL.strWallLocationDesc                              AS WallLocationName
                            FROM
                                TWallLocations AS TWL
                                INNER JOIN TBelayDeviceWallLocations AS TBDWL ON TWL.intWallLocationID = TBDWL.intWallLocationID
                                INNER JOIN TBelayDevices AS TBD ON TBD.intBelayDeviceID = TBDWL.intBelayDeviceID
                            WHERE
                                TBD.intBelayDeviceID IN (SELECT TBD.intBelayDeviceID FROM TBelayDevices AS TBD WHERE TBD.strEquipInUse = 'Yes');""",                                    
                        # vBelayDeviceWallLocationInspectDates
                        """SELECT
                        -- Create View: Show all Current BelayDevice Wall Location. Include last and next inspection date
                                TWL.strWallLocationDesc                               AS WallLocationName
                                ,TBD.strSerialNum                                     AS BelayDeviceSerialNumber
                                ,TBD.strBumperNum                                     AS BumperNum
                                ,TBD.dtmLastInspectionDate                            AS LastInspectionDate
                                ,TBD.dtmNextInspectionDate                            AS NextInspectionDate
                            FROM
                                TWallLocations AS TWL
                                LEFT JOIN TBelayDeviceWallLocations AS TBDWL ON TWL.intWallLocationID = TBDWL.intWallLocationID
                                LEFT JOIN TBelayDevices AS TBD ON TBD.intBelayDeviceID = TBDWL.intBelayDeviceID
                            WHERE
                                TBD.intBelayDeviceID IN (SELECT TBDWL.intBelayDeviceID FROM TBelayDeviceWallLocations AS TBDWL);""",                                                                                                    
                        # vBelayDeviceInspectResults
                        """SELECT
                        -- Create View: Show all BelayDevice component inspection status. Include inspector name and 
                        -- last/next inspection date 
                                TBD.strSerialNum                                     AS BelayDeviceSerialNumber
                                ,TBD.strBumperNum                                    AS BumperNum
                                ,TWL.strWallLocationDesc                             AS WallLocationName
                                ,TINS.strInspectionStatusDesc                        AS OverallStatus
                                ,TIN.strLastName || ',' || TIN.strFirstName          AS InspectorName
                                ,TBD.dtmLastInspectionDate                           AS LastInspectionDate
                                ,TBD.dtmNextInspectionDate                           AS NextInspectionDate
                            FROM 
                                TBelayDevices AS TBD 
                                LEFT JOIN TBelayDeviceInspections AS TBDIN ON TBD.intBelayDeviceID = TBDIN.intBelayDeviceID
                                LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TBDIN.intInspectionStatusID
                                LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TBDIN.intInspectorID
                                LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TBDIN.intWallLocationID
                            WHERE
                                TBD.intBelayDeviceID IN (SELECT TBDIN.intBelayDeviceID FROM TBelayDeviceInspections AS TBDIN);""",
                        # vBelayDeviceFailedResults
                        """SELECT
                        -- Create View: Show all failed BelayDevices
                                TBD.strSerialNum                                     AS BelayDeviceSerialNumber
                                ,TBD.strBumperNum                                    AS BumperNum
                                ,TWL.strWallLocationDesc                             AS WallLocationName
                                ,TINS.strInspectionStatusDesc                        AS OverallStatus
                                ,TIN.strLastName || ',' || TIN.strFirstName          AS InspectorName
                                ,TBD.dtmLastInspectionDate                           AS LastInspectionDate
                                ,TBD.dtmNextInspectionDate                           AS NextInspectionDate
                            FROM 
                                TBelayDevices AS TBD 
                                LEFT JOIN TBelayDeviceInspections AS TBDIN ON TBD.intBelayDeviceID = TBDIN.intBelayDeviceID
                                LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TBDIN.intInspectionStatusID
                                LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TBDIN.intInspectorID
                                LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TBDIN.intWallLocationID
                            WHERE
                                TBD.intBelayDeviceID IN (SELECT TBDIN.intBelayDeviceID FROM TBelayDeviceInspections AS TBDIN
                                                        WHERE TBDIN.intInspectionStatusID = 3);""",                                    
                        # vBelayDeviceMonitorResults
                        """SELECT
                        -- Create View: Show all monitor BelayDevices
                                TBD.strSerialNum                                     AS BelayDeviceSerialNumber
                                ,TBD.strBumperNum                                    AS BumperNum
                                ,TWL.strWallLocationDesc                             AS WallLocationName
                                ,TINS.strInspectionStatusDesc                        AS OverallStatus
                                ,TIN.strLastName || ',' || TIN.strFirstName          AS InspectorName
                                ,TBD.dtmLastInspectionDate                           AS LastInspectionDate
                                ,TBD.dtmNextInspectionDate                           AS NextInspectionDate
                            FROM 
                                TBelayDevices AS TBD 
                                LEFT JOIN TBelayDeviceInspections AS TBDIN ON TBD.intBelayDeviceID = TBDIN.intBelayDeviceID
                                LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TBDIN.intInspectionStatusID
                                LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TBDIN.intInspectorID
                                LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TBDIN.intWallLocationID
                            WHERE
                                TBD.intBelayDeviceID IN (SELECT TBDIN.intBelayDeviceID FROM TBelayDeviceInspections AS TBDIN
                                                        WHERE TBDIN.intInspectionStatusID = 2);""", 
                        # vBelayDeviceFailNowResults
                        """SELECT
                        -- Create View: Show all failed now BelayDevices
                                TBD.strSerialNum                                         AS BelayDeviceSerialNumber
                                ,TBD.strBumperNum                                        AS BumperNum
                                ,TWL.strWallLocationDesc                                 AS WallLocationName
                                ,TINS.strInspectionStatusDesc                            AS OverallStatus
                                ,TIN.strLastName || ',' || TIN.strFirstName              AS InspectorName
                                ,TBD.dtmLastInspectionDate                               AS LastInspectionDate
                                ,TBD.dtmNextInspectionDate                               AS NextInspectionDate
                            FROM 
                                TBelayDevices AS TBD 
                                LEFT JOIN TBelayDeviceInspections AS TBDIN ON TBD.intBelayDeviceID = TBDIN.intBelayDeviceID
                                LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TBDIN.intInspectionStatusID
                                LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TBDIN.intInspectorID
                                LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TBDIN.intWallLocationID
                            WHERE
                                TBDIN.intInspectionStatusID = 3 AND date(TBD.dtmLastInspectionDate) = date('now');""",
                        # vABInspectID_OutForReservice
                        """ SELECT
                        -- Create View: Show all 'AB InspectionID's Out For Reservice' AutoBelays
                                MAX(TABIN.intAutoBelayInspectionID)                 AS InspectionID
                                ,TAB.strSerialNum                                   AS AutoBelaySerialNumber
                                ,TAB.strBumperNum                                   AS BumperNum
                                ,TINS.strInspectionStatusDesc                       AS OverallStatus
                                ,TAB.dtmLastInspectionDate                          AS LastInspectionDate
                                ,TAB.dtmServiceDate                                 AS ServiceDate
                                ,TAB.dtmReserviceDate                               AS ReServiceDate
                            FROM 
                                TAutoBelays AS TAB 
                                LEFT JOIN TAutoBelayInspections AS TABIN ON TAB.intAutoBelayID = TABIN.intAutoBelayID
                                LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TABIN.intInspectionStatusID
                                LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TABIN.intInspectorID
                                LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TABIN.intWallLocationID
                            WHERE
                                TAB.blnDeviceInUse = 'Out For Reservice' AND TAB.intAutoBelayID = TABIN.intAutoBelayID
                            GROUP BY TAB.intAutoBelayID;""",                                    
                        ]

        # First check if connected to the database outside the loop (optional based on your use case)
        if not self.conn:
            raise Exception("Database is not connected.")

        # Execute the SQL statements
        for view_name, sql in zip(sqlViewName, sqlStatement):
            full_view_name = f"{target_db + '.' if target_db != 'main' else ''}{view_name}"
            try:
                Queries.dbCreateViews(self, full_view_name, sql)
            except sqlite3.Error as e:
                print(f"Error creating view '{view_name}': {e}")
            except Exception as e:
                print(f"An unexpected error occurred: {e}")

                                    
    # ------------------------------------------------------------------------------------------------
    # Load the Database 
    # ------------------------------------------------------------------------------------------------
    def dbLoadDrops(self):
        """ 
        Function Name: dbLoadDrops
        Function Purpose: Loads all the drop tables, procedures, views, functions, and audit tables for the database
        """      
        # Drop the primary tables
        Database.dbDropTables(self)
        
        # Drop the stored procedures 
        # Database.dbDropProcedures(self)
        
        # Drop the audit tables
        Database.dbDrop_Z_Tables(self)
        
        # Drop the functions
        # Database.dbDropFunctions(self) 
                
        # Drop the views
        Database.dbDropViews(self)         

    def dbLoadTables(self):
        """ 
        Function Name: dbLoadTables
        Function Purpose: Loads all the tables and audit tables for the database
        """     
        # ----------------------------------------------------------------
        # Generic Types
        # ----------------------------------------------------------------           
        # Load the Inspection Status tables
        Database.createInspectionStatusTable(self)
        
        # Load the Inspection Types tables
        Database.createInspectionTypesTables(self)        
        
        # Load the Metallic Component Inspection tables
        Database.createMetallicInspectionsTable(self)      
        
        # Load the Textile Component Inspection tables
        Database.createTextileInspectionsTable(self)      

        # Load the Plastic Component Inspection tables
        Database.createPlasticInspectionsTable(self)         

        # ----------------------------------------------------------------
        # Carabiner
        # ----------------------------------------------------------------    
        # Load the Carabiner Types tables
        Database.createCarabinerTables(self)         
                               
        # Load the Carabiner Visual Metal Selection tables
        Database.createCarabVisMetSelTables(self)                                        

        # Load the Carabiner Visual Inspection tables
        Database.createCarabVisInspectTables(self)                                        

        # Load the Carabiner Physical Metal Selection tables
        Database.createCarabPhysMetSelTables(self)    

        # Load the Carabiner Physical Inspection tables
        Database.createCarabPhysInspectTables(self)                  

        # Load the Carabiner Functional Types tables
        Database.createCarabinerFunctionsTables(self)            

        # Load the Carabiner Functional Selection tables
        Database.createCarabFunctSelTables(self)         

        # Load the Carabiner Functional Inspection tables
        Database.createCarabFunctInspectTables(self) 
        
        # Load the Standard Carabiner Inspection tables
        Database.createStandCarabInspectTables(self)                     

        # ----------------------------------------------------------------
        # Device Handle
        # ----------------------------------------------------------------    
        # Load the Device Handle Types tables
        Database.createDeviceHandleTables(self)         
                               
        # Load the Device Handle Visual Metal Selection tables
        Database.createHandVisMetSelTables(self)                                        

        # Load the Device Handle Visual Inspection tables
        Database.createHandVisInspectTables(self)                                        

        # Load the Device Handle Physical Metal Selection tables
        Database.createHandPhysMetSelTables(self)    

        # Load the Device Handle Physical Inspection tables
        Database.createHandPhysInspectTables(self)                  

        # Load the Standard Device Handle Inspection tables
        Database.createStandHandInspectTables(self)                       

        # ----------------------------------------------------------------
        # Case Housing
        # ----------------------------------------------------------------    
        # Load the Case Housing Types tables
        Database.createCaseTables(self)         

        # Load the Case Housing Component Selection tables
        Database.createCaseCompSelTables(self)    
                                       
        # Load the Case Housing Visual Metal Selection tables
        Database.createCaseVisMetSelTables(self)                                        

        # Load the Case Housing Visual Inspection tables
        Database.createCaseVisInspectTables(self)                                        

        # Load the Case Housing Physical Metal Selection tables
        Database.createCasePhysMetSelTables(self)    

        # Load the Case Housing Physical Inspection tables
        Database.createCasePhysInspectTables(self)                  

        # Load the Standard Case Housing Inspection tables
        Database.createStandCaseInspectTables(self)           

        # ----------------------------------------------------------------
        # Brake Housing
        # ----------------------------------------------------------------    
        # Load the Brake Housing Types tables
        Database.createBrakeTables(self)         

        # Load the Brake Housing Component Selection tables
        Database.createBrakeCompSelTables(self)    
                                       
        # Load the Brake Housing Visual Metal Selection tables
        Database.createBrakeVisMetSelTables(self)                                        

        # Load the Brake Housing Visual Inspection tables
        Database.createBrakeVisInspectTables(self)                                        

        # Load the Brake Housing Physical Metal Selection tables
        Database.createBrakePhysMetSelTables(self)    

        # Load the Brake Housing Physical Inspection tables
        Database.createBrakePhysInspectTables(self)                  

        # Load the Standard Brake Housing Inspection tables
        Database.createStandBrakeInspectTables(self)         

        # ----------------------------------------------------------------
        # Lanyard/Retractor
        # ----------------------------------------------------------------    
        # Load the Lanyard/Retractor Types tables
        Database.createLanyardTables(self)         
                               
        # Load the Lanyard/Retractor Visual Metal Selection tables
        Database.createLanVisTextSelTables(self)                                        

        # Load the Lanyard/Retractor Visual Inspection tables
        Database.createLanVisInspectTables(self)                                        

        # Load the Lanyard/Retractor Physical Metal Selection tables
        Database.createLanPhysTextSelTables(self)    

        # Load the Lanyard/Retractor Physical Inspection tables
        Database.createLanPhysInspectTables(self)                  

        # Load the Lanyard/Retractor Functional Types tables
        Database.createRetractFunctionsTables(self)            

        # Load the Lanyard/Retractor Functional Selection tables
        Database.createRetractFunctSelTables(self)         

        # Load the Lanyard/Retractor Functional Inspection tables
        Database.createRetractFunctInspectTables(self) 
        
        # Load the Standard Lanyard/Retractor Inspection tables
        Database.createStandLanyInspectTables(self)             

        # ----------------------------------------------------------------
        # Load State, Inspector, Logins, Locations, and Auto Belay data
        # ----------------------------------------------------------------    
        # Load the State tables
        Database.createStateTables(self)         

        # Load the Inspector tables
        Database.createInspectorTables(self)                                        

        # Load the Login tables
        Database.createLoginTables(self)                                        

        # Load the User Login tables
        Database.createUserLoginTables(self)    

        # Load the Admin User tables
        Database.createAdminUserTables(self)                  

        # Load the Gym Location tables
        Database.createGymLocationTables(self)            

        # Load the Auto Belay tables
        Database.createAutoBelayTables(self)         

        # Load the Wall Location tables
        Database.createWallLocationTables(self) 

        # Load the Location tables
        Database.createLocationTables(self) 
        
        # Load the AutoBelay Wall Locations
        Database.createAutoBelay_WallLocationTables(self)
        
        # Load the Auto Belay Inspector tables
        Database.createAutoBelay_InspectorTables(self)         

        # Load the Standard Inspection tables
        Database.createStandardInspectionTables(self)  

        # Load the Auto Belay Inspection tables
        Database.createAutoBelay_InspectionTables(self)    
        
        # Load the Auto Belay Reservice Reports
        Database.createAutoBelay_ReserviceReports_Tables(self)
        Database.dbLoad_AutoBelay_Reservice_Reports   

        # Load the Auto Belay Insepction ID' Out For Reservice 
        Database.dbLoad_ABInspectID_OutForReservice_View              

        # ----------------------------------------------------------------
        # Custome Rope Systems Tables
        # ----------------------------------------------------------------    
        # Load the Custom Rope Sytems tables
        Database.createCustomRopeSystemsTables(self) 
        
        # ----------------------------------------------------------------
        # Ropes Tables
        # ----------------------------------------------------------------    
        # Load the Rope Types tables
        Database.createRopeTables(self)         
        
        # Load the Rope Visual Textile Selection tables
        Database.createRopeVisTextSelTables(self)                                        

        # Load the Rope Visual Inspection tables
        Database.createRopeVisInspectTables(self)                                        

        # Load the Rope Physical Textile Selection tables
        Database.createRopePhysTextSelTables(self)    

        # Load the Rope Physical Inspection tables
        Database.createRopePhysInspectTables(self)                  

        # Load the Standard Rope Inspection tables
        Database.createStandRopeInspectTables(self)
        
        # Load the Rope Inspection tables
        Database.createRope_InspectionTables(self)
        
        # Load the Rope Wall Inspection tables
        Database.createRope_WallLocationTables(self)
        
        # Load the Rope Inspector Inspection tables
        Database.createRope_InspectorTables(self)

        # Load the Rope Retired Reports Inspection tables
        Database.createRope_RetiredReports_Tables(self)             
        
        # ----------------------------------------------------------------
        # Connectors Tables
        # ----------------------------------------------------------------    
        # Load the Connector Types tables
        Database.createConnectorTables(self)         
        
        # Load the Connector Visual Metal Selection tables
        Database.createConnectorVisMetalSelTables(self)                                        

        # Load the Connector Visual Inspection tables
        Database.createConnectorVisInspectTables(self)                                        

        # Load the Connector Physical Metal Selection tables
        Database.createConnectorPhysMetalSelTables(self)    

        # Load the Connector Physical Inspection tables
        Database.createConnectorPhysInspectTables(self)                  

        # Load the Connector Functional Metal Selection tables
        Database.createConnectorFunctionsTables(self)    

        # Load the Connector Functional Inspection tables
        Database.createConnectorFunctSelTables(self) 
        
        # Load the Connector Functional Inspection tables
        Database.createConnectorFunctInspectTables(self)          
        
        # Load the Standard Connector Inspection tables
        Database.createStandConnectorInspectTables(self)
        
        # Load the Connector Inspection tables
        Database.createConnector_InspectionTables(self)
        
        # Load the Connector Wall Inspection tables
        Database.createConnector_WallLocationTables(self)
        
        # Load the Connector Inspector Inspection tables
        Database.createConnector_InspectorTables(self)

        # Load the Connector Retired Reports tables
        Database.createConnector_RetiredReports_Tables(self)        
        
        # ----------------------------------------------------------------
        # BelayDevices Tables
        # ----------------------------------------------------------------    
        # Load the BelayDevice Types tables
        Database.createBelayDevicesTables(self)         
        
        # Load the BelayDevice Visual Metal Selection tables
        Database.createBelayDevicesVisMetalSelTables(self)                                        

        # Load the BelayDevice Visual Inspection tables
        Database.createBelayDevicesVisInspectTables(self)                                        

        # Load the BelayDevice Physical Metal Selection tables
        Database.createBelayDevicesPhysMetalSelTables(self)    

        # Load the BelayDevice Physical Inspection tables
        Database.createBelayDevicesPhysInspectTables(self)                  

        # Load the BelayDevice Functional Metal Selection tables
        Database.createBelayDevicesFunctionsTables(self)    

        # Load the BelayDevice Functional Inspection tables
        Database.createBelayDevicesFunctSelTables(self) 
        
        # Load the BelayDevice Functional Inspection tables
        Database.createBelayDevicesFunctInspectTables(self)          
        
        # Load the Standard BelayDevice Inspection tables
        Database.createStandBelayDevicesInspectTables(self)
        
        # Load the BelayDevice Inspection tables
        Database.createBelayDevices_InspectionTables(self)
        
        # Load the BelayDevice Wall Inspection tables
        Database.createBelayDevices_WallLocationTables(self)
        
        # Load the BelayDevice Inspector Inspection tables
        Database.createBelayDevices_InspectorTables(self)
        
        # Load the BelayDevice Retired Reports tables
        Database.createBelayDevice_RetiredReports_Tables(self)
        
        # Load the BelayDevice Visual Plastic Selection tables
        Database.createBelayDevicesVisPlasticSelTables(self)
        
        # Load the BelayDevice Physical Plastic Selection tables
        Database.createBelayDevicesPhysPlasticSelTables(self)
        
        # Load the BelayDevice Functional Plastic Selection tables
        Database.createBelayDevicesFunctPlasticSelTables(self)
                                        
        # ----------------------------------------------------------------
        # Load Test Data
        # ----------------------------------------------------------------    
        # Insert the Inspection Status/Types, Metal, Textile, Plastic, and Carabiner data into the tables
        Database.insertStatus_Types_MetTextPlast_Carab_Data(self)          
        
        # Insert the Handle data into the tables
        Database.insertHandData(self)                

        # Insert the Case data into the tables
        Database.insertCaseData(self)                          

        # Insert the Brake data into the tables
        Database.insertBrakeData(self)             

        # Insert the Retractor data into the tables
        Database.insertRetractData(self)                         

        # Insert the User Login, Locations data into the tables
        Database.insertUsrLog_Location_Data(self)            

    def dbLoadForeignKeys(self):
        """ 
        Function Name: dbLoadForeignKeys
        Function Purpose: Loads all the foreign key constraints
        """     
        # ----------------------------------------------------------------
        # Load the constraints
        # ----------------------------------------------------------------   
        # Insert the user login, locations, autobelays foreign keys       
        Database.createUsrLog_Location_AutoBelay_ForeignKeys(self),
        
        # Insert the user Ropes foreign keys
        Database.createRopes_ForeignKeys(self),
        
        # Insert the user Connectors foreign keys
        Database.createConnector_ForeignKeys(self),
        
        # Insert the user Belay Device foreign keys
        Database.createBelayDevices_ForeignKeys(self),

    def dbLoadViews(self):
        """ 
        Function Name: dbLoadViews
        Function Purpose: Loads all the views for the database
        """     
        # ----------------------------------------------------------------
        # Load the Views
        # ----------------------------------------------------------------          
        self.Create_Base_Views()
                
    def dbLoadFunctions(self):
        """ 
        Function Name: dbLoadFunctions
        Function Purpose: Loads all the functions for the database
        """     
        # ----------------------------------------------------------------
        # Load the Functions
        # ----------------------------------------------------------------          
        Database.Create_Base_Functions(self)

    def dbLoad_New_ABFail_View(self):
        """ 
        Function Name: dbLoad_New_ABFail_View
        Function Purpose: Loads the newly added view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vAutoBelayFailedResults'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                                TAB.strSerialNum                                    AS AutoBelaySerialNumber
                                ,TAB.strBumperNum                                   AS BumperNum
                                ,TWL.strWallLocationDesc                            AS WallLocationName
                                ,TINS.strInspectionStatusDesc                       AS OverallStatus
                                ,TIN.strLastName || ',' || TIN.strFirstName         AS InspectorName
                                ,TAB.dtmLastInspectionDate                          AS LastInspectionDate
                                ,TAB.dtmNextInspectionDate                          AS NextInspectionDate
                                ,TAB.dtmServiceDate                                 AS ServiceDate
                                ,TAB.dtmReserviceDate                               AS ReServiceDate
                            FROM 
                                TAutoBelays AS TAB 
                                LEFT JOIN TAutoBelayInspections AS TABIN ON TAB.intAutoBelayID = TABIN.intAutoBelayID
                                LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TABIN.intInspectionStatusID
                                LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TABIN.intInspectorID
                                LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TABIN.intWallLocationID
                            WHERE
                                TABIN.intInspectionStatusID = 3;"""                    
        
                        
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")       
            
    def dbLoad_New_ABMonitor_View(self):
        """ 
        Function Name: dbLoad_New_ABMonitor_View
        Function Purpose: Loads the newly added view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vAutoBelayMonitorResults'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                                TAB.strSerialNum                                    AS AutoBelaySerialNumber
                                ,TAB.strBumperNum                                   AS BumperNum
                                ,TWL.strWallLocationDesc                            AS WallLocationName
                                ,TINS.strInspectionStatusDesc                       AS OverallStatus
                                ,TIN.strLastName || ',' || TIN.strFirstName         AS InspectorName
                                ,TAB.dtmLastInspectionDate                          AS LastInspectionDate
                                ,TAB.dtmNextInspectionDate                          AS NextInspectionDate
                                ,TAB.dtmServiceDate                                 AS ServiceDate
                                ,TAB.dtmReserviceDate                               AS ReServiceDate
                            FROM 
                                TAutoBelays AS TAB 
                                LEFT JOIN TAutoBelayInspections AS TABIN ON TAB.intAutoBelayID = TABIN.intAutoBelayID
                                LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TABIN.intInspectionStatusID
                                LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TABIN.intInspectorID
                                LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TABIN.intWallLocationID
                            WHERE
                                TABIN.intInspectionStatusID = 2;"""                      
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")   

    def dbLoad_Now_ABFail_View(self):
        """ 
        Function Name: dbLoad_Now_ABFail_View
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vAutoBelayFailNowResults'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                                TAB.strSerialNum                                    AS AutoBelaySerialNumber
                                ,TAB.strBumperNum                                   AS BumperNum
                                ,TWL.strWallLocationDesc                            AS WallLocationName
                                ,TINS.strInspectionStatusDesc                       AS OverallStatus
                                ,TIN.strLastName || ',' || TIN.strFirstName         AS InspectorName
                                ,TAB.dtmLastInspectionDate                          AS LastInspectionDate
                                ,TAB.dtmNextInspectionDate                          AS NextInspectionDate
                                ,TAB.dtmServiceDate                                 AS ServiceDate
                                ,TAB.dtmReserviceDate                               AS ReServiceDate
                            FROM 
                                TAutoBelays AS TAB 
                                LEFT JOIN TAutoBelayInspections AS TABIN ON TAB.intAutoBelayID = TABIN.intAutoBelayID
                                LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TABIN.intInspectionStatusID
                                LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TABIN.intInspectorID
                                LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TABIN.intWallLocationID
                            WHERE
                                TABIN.intInspectionStatusID = 3 AND date(TAB.dtmLastInspectionDate) = date('now');"""                      
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")  

    def dbLoad_Now_ABMonitor_View(self):
        """ 
        Function Name: dbLoad_Now_ABMonitor_View
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vAutoBelayMonitorNowResults'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                                TAB.strSerialNum                                    AS AutoBelaySerialNumber
                                ,TAB.strBumperNum                                   AS BumperNum
                                ,TWL.strWallLocationDesc                            AS WallLocationName
                                ,TINS.strInspectionStatusDesc                       AS OverallStatus
                                ,TIN.strLastName || ',' || TIN.strFirstName         AS InspectorName
                                ,TAB.dtmLastInspectionDate                          AS LastInspectionDate
                                ,TAB.dtmNextInspectionDate                          AS NextInspectionDate
                                ,TAB.dtmServiceDate                                 AS ServiceDate
                                ,TAB.dtmReserviceDate                               AS ReServiceDate
                            FROM 
                                TAutoBelays AS TAB 
                                LEFT JOIN TAutoBelayInspections AS TABIN ON TAB.intAutoBelayID = TABIN.intAutoBelayID
                                LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TABIN.intInspectionStatusID
                                LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TABIN.intInspectorID
                                LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TABIN.intWallLocationID
                            WHERE
                                TABIN.intInspectionStatusID = 2 AND date(TAB.dtmLastInspectionDate) = date('now');"""                      
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}") 

    def dbLoad_AutoBelay_Reservice_Reports(self):
        """ 
        Function Name: AutoBelay_Reservice_Reports
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vAutoBelayReserviceReports'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                                    TAB.strSerialNum                 AS AutoBelaySerialNumber
                                    ,TAB.strBumperNum                AS BumperNum
                                    ,TAB.dtmLastInspectionDate       AS LastInspectionDate
                                    ,TAB.dtmServiceDate              AS ServiceDate
                                    ,TAB.dtmReserviceDate            AS ReServiceDate
                                    -- Add any other relevant columns you need
                                FROM
                                    TAutoBelays AS TAB
                                WHERE
                                    TAB.blnDeviceInUse = 'Out For Reservice';"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}") 

    def dbLoad_vRopes(self):
        """ 
        Function Name: dbLoad_vRopes
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vRopes'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT                                                                                                 
                                strSerialNum                                         AS RopeSerialNumber
                                ,strBumperNum                                        AS BumperNum
                                ,strRopeLength                                       AS RopeLength
                                ,strDiameter                                         AS Diameter
                                ,strElasticity                                       AS Elasticity
                                ,strManufactureName                                  AS ManufacturerName
                                ,dtmManufactureDate                                  AS ManufactureDate                                                                        
                                ,dtmInstallationDate                                 AS InstallationDate
                                ,dtmLastInspectionDate                               AS LastInspectionDate
                                ,dtmNextInspectionDate                               AS NextInspectionDate
                                ,strEquipInUse                                       AS RopeInUse
                            FROM
                                TRopes;"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")            

    def dbLoad_vRopeDates(self):
        """ 
        Function Name: dbLoad_vRopes
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vRopeDates'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT 
                             -- Create View:  Show all Rope Dates where serial number equals the selected primary key from user
                                    strSerialNum                                         AS RopeSerialNumber
                                    ,strBumperNum                                        AS BumperNum
                                    ,dtmManufactureDate                                  AS ManufactureDate                                                                        
                                    ,dtmInstallationDate                                 AS InstallationDate
                                    ,dtmLastInspectionDate                               AS LastInspectionDate
                                    ,dtmNextInspectionDate                               AS NextInspectionDate       
                                FROM 
                                    TRopes;"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")

    def dbLoad_vRopeLastNextInspectDates(self):
        """ 
        Function Name: dbLoad_vRopeLastNextInspectDates
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vRopeLastNextInspectDates'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                            -- Create View: Show all Rope Dates of the serial number and last/next inspection dates
                                    strSerialNum                                         AS RopeSerialNumber
                                    ,strBumperNum                                        AS BumperNum
                                    ,dtmLastInspectionDate                               AS LastInspectionDate
                                    ,dtmNextInspectionDate                               AS NextInspectionDate           
                                FROM 
                                    TRopes;"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")
            
    def dbLoad_vRopeInspectors(self):
        """ 
        Function Name: dbLoad_vRopeInspectors
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vRopeInspectors'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                            -- Create View: Show all Rope Inspectors by date of last inspection and the upcoming inspection
                                    TI.strLastName || ',' || TI.strFirstName             AS InspectorName
                                    ,TR.strSerialNum                                     AS RopeSerialNumber
                                    ,TR.strBumperNum                                     AS BumperNum
                                    ,TR.dtmLastInspectionDate                            AS LastInspectionDate
                                    ,TR.dtmNextInspectionDate                            AS NextInspectionDate 
                                FROM
                                    TInspectors AS TI
                                    LEFT JOIN TRopes AS TR ON TR.intRopeID = TRI.intRopeID
                                    LEFT JOIN TRopeInspectors AS TRI ON TI.intInspectorID = TRI.intInspectorID
                                WHERE 
                                    TI.intInspectorID IN (SELECT TRI.intInspectorID FROM TRopeInspectors AS TRI)
                                    AND TR.intRopeID IN (SELECT TRI.intRopeID FROM TRopeInspectors AS TRI);"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}") 

    def dbLoad_vRopeWallLocations(self):
        """ 
        Function Name: dbLoad_vRopeWallLocations
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vRopeWallLocations'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                            -- Create View: Show all Current Rope Wall Location.
                                    TR.strSerialNum                                      AS RopeSerialNumber
                                    ,TR.strBumperNum                                     AS BumperNum
                                    ,TWL.strWallLocationDesc                             AS WallLocationName
                                FROM
                                    TWallLocations AS TWL
                                    INNER JOIN TRopeWallLocations AS TRWL ON TWL.intWallLocationID = TRWL.intWallLocationID
                                    INNER JOIN TRopes AS TR ON TR.intRopeID = TRWL.intRopeID
                                WHERE
                                    TR.intRopeID IN (SELECT TR.intRopeID FROM TRopes AS TR WHERE TR.strEquipInUse = 'Yes');"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")

    def dbLoad_vRopeWallLocationInspectDates(self):
        """ 
        Function Name: dbLoad_vRopeWallLocationInspectDates
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vRopeWallLocationInspectDates'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                            -- Create View: Show all Current Rope Wall Location. Include last and next inspection date
                                    TWL.strWallLocationDesc                              AS WallLocationName
                                    ,TR.strSerialNum                                     AS RopeSerialNumber
                                    ,TR.strBumperNum                                     AS BumperNum
                                    ,TR.dtmLastInspectionDate                            AS LastInspectionDate
                                    ,TR.dtmNextInspectionDate                            AS NextInspectionDate
                                FROM
                                    TWallLocations AS TWL
                                    LEFT JOIN TRopeWallLocations AS TRWL ON TWL.intWallLocationID = TRWL.intWallLocationID
                                    LEFT JOIN TRopes AS TR ON TR.intRopeID = TRWL.intRopeID
                                WHERE
                                    TR.intRopeID IN (SELECT TRWL.intRopeID FROM TRopeWallLocations AS TRWL);"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")                                               

    def dbLoad_vRopeInspectResults(self):
        """ 
        Function Name: dbLoad_vRopeInspectResults
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vRopeInspectResults'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                            -- Create View: Show all Rope component inspection status. Include inspector name and 
                            -- last/next inspection date 
                                    TR.strSerialNum                                     AS RopeSerialNumber
                                    ,TR.strBumperNum                                    AS BumperNum
                                    ,TWL.strWallLocationDesc                            AS WallLocationName
                                    ,TINS.strInspectionStatusDesc                       AS OverallStatus
                                    ,TIN.strLastName || ',' || TIN.strFirstName         AS InspectorName
                                    ,TR.dtmLastInspectionDate                           AS LastInspectionDate
                                    ,TR.dtmNextInspectionDate                           AS NextInspectionDate
                                FROM 
                                    TRopes AS TR 
                                    LEFT JOIN TRopeInspections AS TRIN ON TR.intRopeID = TRIN.intRopeID
                                    LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TRIN.intInspectionStatusID
                                    LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TRIN.intInspectorID
                                    LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TRIN.intWallLocationID
                                WHERE
                                    TR.intRopeID IN (SELECT TRIN.intRopeID FROM TRopeInspections AS TRIN);"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")

    def dbLoad_vRopeFailedResults(self):
        """ 
        Function Name: dbLoad_vRopeFailedResults
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vRopeFailedResults'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                            -- Create View: Show all failed Ropes
                                    TR.strSerialNum                                     AS RopeSerialNumber
                                    ,TR.strBumperNum                                    AS BumperNum
                                    ,TWL.strWallLocationDesc                            AS WallLocationName
                                    ,TINS.strInspectionStatusDesc                       AS OverallStatus
                                    ,TIN.strLastName || ',' || TIN.strFirstName         AS InspectorName
                                    ,TR.dtmLastInspectionDate                           AS LastInspectionDate
                                    ,TR.dtmNextInspectionDate                           AS NextInspectionDate
                                FROM 
                                    TRopes AS TR 
                                    LEFT JOIN TRopeInspections AS TRIN ON TR.intRopeID = TRIN.intRopeID
                                    LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TRIN.intInspectionStatusID
                                    LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TRIN.intInspectorID
                                    LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TRIN.intWallLocationID
                                WHERE
                                    TR.intRopeID IN (SELECT TRIN.intRopeID FROM TRopeInspections AS TRIN
                                                           WHERE TRIN.intInspectionStatusID = 3);"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")

    def dbLoad_vRopeMonitorResults(self):
        """ 
        Function Name: dbLoad_vRopeMonitorResults
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vRopeMonitorResults'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                            -- Create View: Show all monitor Ropes
                                    TR.strSerialNum                                     AS RopeSerialNumber
                                    ,TR.strBumperNum                                    AS BumperNum
                                    ,TWL.strWallLocationDesc                            AS WallLocationName
                                    ,TINS.strInspectionStatusDesc                       AS OverallStatus
                                    ,TIN.strLastName || ',' || TIN.strFirstName         AS InspectorName
                                    ,TR.dtmLastInspectionDate                           AS LastInspectionDate
                                    ,TR.dtmNextInspectionDate                           AS NextInspectionDate
                                FROM 
                                    TRopes AS TR 
                                    LEFT JOIN TRopeInspections AS TRIN ON TR.intRopeID = TRIN.intRopeID
                                    LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TRIN.intInspectionStatusID
                                    LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TRIN.intInspectorID
                                    LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TRIN.intWallLocationID
                                WHERE
                                    TR.intRopeID IN (SELECT TRIN.intRopeID FROM TRopeInspections AS TRIN
                                                           WHERE TRIN.intInspectionStatusID = 2);"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")                        

    def dbLoad_vRopeFailNowResults(self):
        """ 
        Function Name: dbLoad_vRopeFailNowResults
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vRopeFailNowResults'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                            -- Create View: Show all failed now Ropes
                                    TR.strSerialNum                                         AS RopeSerialNumber
                                    ,TR.strBumperNum                                        AS BumperNum
                                    ,TWL.strWallLocationDesc                                AS WallLocationName
                                    ,TINS.strInspectionStatusDesc                           AS OverallStatus
                                    ,TIN.strLastName || ',' || TIN.strFirstName             AS InspectorName
                                    ,TR.dtmLastInspectionDate                               AS LastInspectionDate
                                    ,TR.dtmNextInspectionDate                               AS NextInspectionDate
                                FROM 
                                    TRopes AS TR 
                                    LEFT JOIN TRopeInspections AS TRIN ON TR.intRopeID = TRIN.intRopeID
                                    LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TRIN.intInspectionStatusID
                                    LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TRIN.intInspectorID
                                    LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TRIN.intWallLocationID
                                WHERE
                                    RIN.intInspectionStatusID = 3 AND date(TR.dtmLastInspectionDate) = date('now');"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")

    def dbLoad_vConnectors(self):
        """ 
        Function Name: dbLoad_vConnectors
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vConnectors'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT                                                                                                 
                                strSerialNum                                         AS ConnectorSerialNumber
                                ,strBumperNum                                        AS BumperNum
                                ,strManufactureName                                  AS ManufacturerName
                                ,dtmManufactureDate                                  AS ManufactureDate                                                                        
                                ,dtmInstallationDate                                 AS InstallationDate
                                ,dtmLastInspectionDate                               AS LastInspectionDate
                                ,dtmNextInspectionDate                               AS NextInspectionDate
                                ,strDeviceType                                       AS DeviceType
                                ,strEquipInUse                                       AS ConnectorInUse
                            FROM
                                TConnectors;"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")

    def dbLoad_vConnectorDates(self):
        """ 
        Function Name: dbLoad_vConnectorDates
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vConnectorDates'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT 
                             -- Create View:  Show all Connector Dates where serial number equals the selected primary key from user
                                    strSerialNum                                         AS ConnectorSerialNumber
                                    ,strBumperNum                                        AS BumperNum
                                    ,dtmManufactureDate                                  AS ManufactureDate                                                                        
                                    ,dtmInstallationDate                                 AS InstallationDate
                                    ,dtmLastInspectionDate                               AS LastInspectionDate
                                    ,dtmNextInspectionDate                               AS NextInspectionDate       
                                FROM 
                                    TConnectors;"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")

    def dbLoad_vConnectorLastNextInspectDates(self):
        """ 
        Function Name: dbLoad_vConnectorLastNextInspectDates
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vConnectorLastNextInspectDates'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                            -- Create View: Show all Connector Dates of the serial number and last/next inspection dates
                                    strSerialNum                                         AS ConnectorSerialNumber
                                    ,strBumperNum                                        AS BumperNum
                                    ,dtmLastInspectionDate                               AS LastInspectionDate
                                    ,dtmNextInspectionDate                               AS NextInspectionDate           
                                FROM 
                                    TConnectors;"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")

    def dbLoad_vConnectorInspectors(self):
        """ 
        Function Name: dbLoad_vConnectorInspectors
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vConnectorInspectors'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                            -- Create View: Show all Connector Inspectors by date of last inspection and the upcoming inspection
                                    TI.strLastName || ',' || TI.strFirstName              AS InspectorName
                                    ,TCN.strSerialNum                                     AS ConnectorSerialNumber
                                    ,TCN.strBumperNum                                     AS BumperNum
                                    ,TCN.dtmLastInspectionDate                            AS LastInspectionDate
                                    ,TCN.dtmNextInspectionDate                            AS NextInspectionDate 
                                FROM
                                    TInspectors AS TI
                                    LEFT JOIN TConnectors AS TCN ON TCN.intConnectorID = TCNI.intConnectorID
                                    LEFT JOIN TConnectorInspectors AS TCNI ON TI.intInspectorID = TCNI.intInspectorID
                                WHERE 
                                    TI.intInspectorID IN (SELECT TCNI.intInspectorID FROM TConnectorInspectors AS TCNI)
                                    AND TCN.intConnectorID IN (SELECT TCNI.intConnectorID FROM TConnectorInspectors AS TCNI);"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")

    def dbLoad_vConnectorWallLocations(self):
        """ 
        Function Name: dbLoad_vConnectorWallLocations
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vConnectorWallLocations'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                            -- Create View: Show all Current Connector Wall Location.
                                    TCN.strSerialNum                                      AS ConnectorSerialNumber
                                    ,TCN.strBumperNum                                     AS BumperNum
                                    ,TWL.strWallLocationDesc                              AS WallLocationName
                                FROM
                                    TWallLocations AS TWL
                                    INNER JOIN TConnectorWallLocations AS TCNWL ON TWL.intWallLocationID = TCNWL.intWallLocationID
                                    INNER JOIN TConnectors AS TCN ON TCN.intConnectorID = TCNWL.intConnectorID
                                WHERE
                                    TCN.intConnectorID IN (SELECT TCN.intConnectorID FROM TConnectors AS TCN WHERE TCN.strEquipInUse = 'Yes');"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")

    def dbLoad_vConnectorWallLocationInspectDates(self):
        """ 
        Function Name: dbLoad_vConnectorWallLocationInspectDates
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vConnectorWallLocationInspectDates'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                            -- Create View: Show all Current Connector Wall Location. Include last and next inspection date
                                    TWL.strWallLocationDesc                               AS WallLocationName
                                    ,TCN.strSerialNum                                     AS ConnectorSerialNumber
                                    ,TCN.strBumperNum                                     AS BumperNum
                                    ,TCN.dtmLastInspectionDate                            AS LastInspectionDate
                                    ,TCN.dtmNextInspectionDate                            AS NextInspectionDate
                                FROM
                                    TWallLocations AS TWL
                                    LEFT JOIN TConnectorWallLocations AS TCNWL ON TWL.intWallLocationID = TCNWL.intWallLocationID
                                    LEFT JOIN TConnectors AS TCN ON TCN.intConnectorID = TCNWL.intConnectorID
                                WHERE
                                    TCN.intConnectorID IN (SELECT TCNWL.intConnectorID FROM TConnectorWallLocations AS TCNWL);"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")

    def dbLoad_vConnectorInspectResults(self):
        """ 
        Function Name: dbLoad_vConnectorInspectResults
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vConnectorInspectResults'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                            -- Create View: Show all Connector component inspection status. Include inspector name and 
                            -- last/next inspection date 
                                    TCN.strSerialNum                                     AS ConnectorSerialNumber
                                    ,TCN.strBumperNum                                    AS BumperNum
                                    ,TWL.strWallLocationDesc                             AS WallLocationName
                                    ,TINS.strInspectionStatusDesc                        AS OverallStatus
                                    ,TIN.strLastName || ',' || TIN.strFirstName          AS InspectorName
                                    ,TCN.dtmLastInspectionDate                           AS LastInspectionDate
                                    ,TCN.dtmNextInspectionDate                           AS NextInspectionDate
                                FROM 
                                    TConnectors AS TCN 
                                    LEFT JOIN TConnectorInspections AS TCNIN ON TCN.intConnectorID = TCNIN.intConnectorID
                                    LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TCNIN.intInspectionStatusID
                                    LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TCNIN.intInspectorID
                                    LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TCNIN.intWallLocationID
                                WHERE
                                    TCN.intConnectorID IN (SELECT TCNIN.intConnectorID FROM TConnectorInspections AS TCNIN);"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")

    def dbLoad_vConnectorFailedResults(self):
        """ 
        Function Name: dbLoad_vConnectorFailedResults
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vConnectorFailedResults'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                            -- Create View: Show all failed Connectors
                                    TCN.strSerialNum                                     AS ConnectorSerialNumber
                                    ,TCN.strBumperNum                                    AS BumperNum
                                    ,TWL.strWallLocationDesc                             AS WallLocationName
                                    ,TINS.strInspectionStatusDesc                        AS OverallStatus
                                    ,TIN.strLastName || ',' || TIN.strFirstName          AS InspectorName
                                    ,TCN.dtmLastInspectionDate                           AS LastInspectionDate
                                    ,TCN.dtmNextInspectionDate                           AS NextInspectionDate
                                FROM 
                                    TConnectors AS TCN 
                                    LEFT JOIN TConnectorInspections AS TCNIN ON TCN.intConnectorID = TCNIN.intConnectorID
                                    LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TCNIN.intInspectionStatusID
                                    LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TCNIN.intInspectorID
                                    LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TCNIN.intWallLocationID
                                WHERE
                                    TCN.intConnectorID IN (SELECT TCNIN.intConnectorID FROM TConnectorInspections AS TCNIN
                                                           WHERE TCNIN.intInspectionStatusID = 3);"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")

    def dbLoad_vConnectorMonitorResults(self):
        """ 
        Function Name: dbLoad_vConnectorMonitorResults
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vConnectorMonitorResults'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                            -- Create View: Show all monitor Connectors
                                    TCN.strSerialNum                                     AS ConnectorSerialNumber
                                    ,TCN.strBumperNum                                    AS BumperNum
                                    ,TWL.strWallLocationDesc                             AS WallLocationName
                                    ,TINS.strInspectionStatusDesc                        AS OverallStatus
                                    ,TIN.strLastName || ',' || TIN.strFirstName          AS InspectorName
                                    ,TCN.dtmLastInspectionDate                           AS LastInspectionDate
                                    ,TCN.dtmNextInspectionDate                           AS NextInspectionDate
                                FROM 
                                    TConnectors AS TCN 
                                    LEFT JOIN TConnectorInspections AS TCNIN ON TCN.intConnectorID = TCNIN.intConnectorID
                                    LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TCNIN.intInspectionStatusID
                                    LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TCNIN.intInspectorID
                                    LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TCNIN.intWallLocationID
                                WHERE
                                    TCN.intConnectorID IN (SELECT TCNIN.intConnectorID FROM TConnectorInspections AS TCNIN
                                                           WHERE TCNIN.intInspectionStatusID = 2);"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}") 

    def dbLoad_vConnectorFailNowResults(self):
        """ 
        Function Name: dbLoad_vConnectorMonitorResults
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vConnectorFailNowResults'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                            -- Create View: Show all failed now Connectors
                                    TCN.strSerialNum                                         AS ConnectorSerialNumber
                                    ,TCN.strBumperNum                                        AS BumperNum
                                    ,TWL.strWallLocationDesc                                 AS WallLocationName
                                    ,TINS.strInspectionStatusDesc                            AS OverallStatus
                                    ,TIN.strLastName || ',' || TIN.strFirstName              AS InspectorName
                                    ,TCN.dtmLastInspectionDate                               AS LastInspectionDate
                                    ,TCN.dtmNextInspectionDate                               AS NextInspectionDate
                                FROM 
                                    TConnectors AS TCN 
                                    LEFT JOIN TConnectorInspections AS TCNIN ON TCN.intConnectorID = TCNIN.intConnectorID
                                    LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TCNIN.intInspectionStatusID
                                    LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TCNIN.intInspectorID
                                    LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TCNIN.intWallLocationID
                                WHERE
                                    TCNIN.intInspectionStatusID = 3 AND date(TCN.dtmLastInspectionDate) = date('now');"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")

    def dbLoad_vBelayDevices(self):
        """ 
        Function Name: dbLoad_vConnectorMonitorResults
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vBelayDevices'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                                    strBelayDeviceName                                   AS BelayDeviceName
                                    ,strSerialNum                                        AS BelayDeviceSerialNumber
                                    ,strBumperNum                                        AS BumperNum
                                    ,strManufactureName                                  AS ManufacturerName
                                    ,dtmManufactureDate                                  AS ManufactureDate                                                                        
                                    ,dtmInstallationDate                                 AS InstallationDate
                                    ,dtmLastInspectionDate                               AS LastInspectionDate
                                    ,dtmNextInspectionDate                               AS NextInspectionDate
                                    ,strDeviceType                                       AS DeviceType
                                    ,strEquipInUse                                       AS BelayDeviceInUse
                                FROM
                                    TBelayDevices;"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")

    def dbLoad_vBelayDeviceDates(self):
        """ 
        Function Name: dbLoad_vBelayDeviceDates
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vBelayDeviceDates'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT 
                             -- Create View:  Show all BelayDevice Dates where serial number equals the selected primary key from user
                                    strSerialNum                                         AS BelayDeviceSerialNumber
                                    ,strBumperNum                                        AS BumperNum
                                    ,dtmManufactureDate                                  AS ManufactureDate                                                                        
                                    ,dtmInstallationDate                                 AS InstallationDate
                                    ,dtmLastInspectionDate                               AS LastInspectionDate
                                    ,dtmNextInspectionDate                               AS NextInspectionDate       
                                FROM 
                                    TBelayDevices;"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")
            
    def dbLoad_vBelayDeviceLastNextInspectDates(self):
        """ 
        Function Name: dbLoad_vBelayDeviceLastNextInspectDates
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vBelayDeviceLastNextInspectDates'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                            -- Create View: Show all BelayDevice Dates of the serial number and last/next inspection dates
                                    strSerialNum                                         AS BelayDeviceSerialNumber
                                    ,strBumperNum                                        AS BumperNum
                                    ,dtmLastInspectionDate                               AS LastInspectionDate
                                    ,dtmNextInspectionDate                               AS NextInspectionDate           
                                FROM 
                                    TBelayDevices;"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")     
            
    def dbLoad_vBelayDeviceInspectors(self):
        """ 
        Function Name: dbLoad_vBelayDeviceInspectors
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vBelayDeviceInspectors'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                            -- Create View: Show all BelayDevice Inspectors by date of last inspection and the upcoming inspection
                                    TI.strLastName || ',' || TI.strFirstName              AS InspectorName
                                    ,TBD.strSerialNum                                     AS BelayDeviceSerialNumber
                                    ,TBD.strBumperNum                                     AS BumperNum
                                    ,TBD.dtmLastInspectionDate                            AS LastInspectionDate
                                    ,TBD.dtmNextInspectionDate                            AS NextInspectionDate 
                                FROM
                                    TInspectors AS TI
                                    LEFT JOIN TBelayDevices AS TBD ON TBD.intBelayDeviceID = TBDI.intBelayDeviceID
                                    LEFT JOIN TBelayDeviceInspectors AS TBDI ON TI.intInspectorID = TBDI.intInspectorID
                                WHERE 
                                    TI.intInspectorID IN (SELECT TBDI.intInspectorID FROM TBelayDeviceInspectors AS TBDI)
                                    AND TBD.intBelayDeviceID IN (SELECT TBDI.intBelayDeviceID FROM TBelayDeviceInspectors AS TBDI);"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")

    def dbLoad_vBelayDeviceWallLocations(self):
        """ 
        Function Name: dbLoad_vBelayDeviceWallLocations
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vBelayDeviceWallLocations'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                            -- Create View: Show all Current BelayDevice Wall Location.
                                    TBD.strSerialNum                                      AS BelayDeviceSerialNumber
                                    ,TBD.strBumperNum                                     AS BumperNum
                                    ,TWL.strWallLocationDesc                              AS WallLocationName
                                FROM
                                    TWallLocations AS TWL
                                    INNER JOIN TBelayDeviceWallLocations AS TBDWL ON TWL.intWallLocationID = TBDWL.intWallLocationID
                                    INNER JOIN TBelayDevices AS TBD ON TBD.intBelayDeviceID = TBDWL.intBelayDeviceID
                                WHERE
                                    TBD.intBelayDeviceID IN (SELECT TBD.intBelayDeviceID FROM TBelayDevices AS TBD WHERE TBD.strEquipInUse = 'Yes');"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")
            
    def dbLoad_vBelayDeviceWallLocationInspectDates(self):
        """ 
        Function Name: dbLoad_vBelayDeviceWallLocationInspectDates
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vBelayDeviceWallLocationInspectDates'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                            -- Create View: Show all Current BelayDevice Wall Location. Include last and next inspection date
                                    TWL.strWallLocationDesc                               AS WallLocationName
                                    ,TBD.strSerialNum                                     AS BelayDeviceSerialNumber
                                    ,TBD.strBumperNum                                     AS BumperNum
                                    ,TBD.dtmLastInspectionDate                            AS LastInspectionDate
                                    ,TBD.dtmNextInspectionDate                            AS NextInspectionDate
                                FROM
                                    TWallLocations AS TWL
                                    LEFT JOIN TBelayDeviceWallLocations AS TBDWL ON TWL.intWallLocationID = TBDWL.intWallLocationID
                                    LEFT JOIN TBelayDevices AS TBD ON TBD.intBelayDeviceID = TBDWL.intBelayDeviceID
                                WHERE
                                    TBD.intBelayDeviceID IN (SELECT TBDWL.intBelayDeviceID FROM TBelayDeviceWallLocations AS TBDWL);"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")
            
    def dbLoad_vBelayDeviceInspectResults(self):
        """ 
        Function Name: dbLoad_vBelayDeviceInspectResults
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vBelayDeviceInspectResults'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                            -- Create View: Show all BelayDevice component inspection status. Include inspector name and 
                            -- last/next inspection date 
                                    TBD.strSerialNum                                     AS BelayDeviceSerialNumber
                                    ,TBD.strBumperNum                                    AS BumperNum
                                    ,TWL.strWallLocationDesc                             AS WallLocationName
                                    ,TINS.strInspectionStatusDesc                        AS OverallStatus
                                    ,TIN.strLastName || ',' || TIN.strFirstName          AS InspectorName
                                    ,TBD.dtmLastInspectionDate                           AS LastInspectionDate
                                    ,TBD.dtmNextInspectionDate                           AS NextInspectionDate
                                FROM 
                                    TBelayDevices AS TBD 
                                    LEFT JOIN TBelayDeviceInspections AS TBDIN ON TBD.intBelayDeviceID = TBDIN.intBelayDeviceID
                                    LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TBDIN.intInspectionStatusID
                                    LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TBDIN.intInspectorID
                                    LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TBDIN.intWallLocationID
                                WHERE
                                    TBD.intBelayDeviceID IN (SELECT TBDIN.intBelayDeviceID FROM TBelayDeviceInspections AS TBDIN);"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")

    def dbLoad_vBelayDeviceFailedResults(self):
        """ 
        Function Name: dbLoad_vBelayDeviceFailedResults
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vBelayDeviceFailedResults'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                            -- Create View: Show all failed BelayDevices
                                    TBD.strSerialNum                                     AS BelayDeviceSerialNumber
                                    ,TBD.strBumperNum                                    AS BumperNum
                                    ,TWL.strWallLocationDesc                             AS WallLocationName
                                    ,TINS.strInspectionStatusDesc                        AS OverallStatus
                                    ,TIN.strLastName || ',' || TIN.strFirstName          AS InspectorName
                                    ,TBD.dtmLastInspectionDate                           AS LastInspectionDate
                                    ,TBD.dtmNextInspectionDate                           AS NextInspectionDate
                                FROM 
                                    TBelayDevices AS TBD 
                                    LEFT JOIN TBelayDeviceInspections AS TBDIN ON TBD.intBelayDeviceID = TBDIN.intBelayDeviceID
                                    LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TBDIN.intInspectionStatusID
                                    LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TBDIN.intInspectorID
                                    LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TBDIN.intWallLocationID
                                WHERE
                                    TBD.intBelayDeviceID IN (SELECT TBDIN.intBelayDeviceID FROM TBelayDeviceInspections AS TBDIN
                                                           WHERE TBDIN.intInspectionStatusID = 3);"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}") 

    def dbLoad_vBelayDeviceMonitorResults(self):
        """ 
        Function Name: dbLoad_vBelayDeviceMonitorResults
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vBelayDeviceMonitorResults'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                            -- Create View: Show all monitor BelayDevices
                                    TBD.strSerialNum                                     AS BelayDeviceSerialNumber
                                    ,TBD.strBumperNum                                    AS BumperNum
                                    ,TWL.strWallLocationDesc                             AS WallLocationName
                                    ,TINS.strInspectionStatusDesc                        AS OverallStatus
                                    ,TIN.strLastName || ',' || TIN.strFirstName          AS InspectorName
                                    ,TBD.dtmLastInspectionDate                           AS LastInspectionDate
                                    ,TBD.dtmNextInspectionDate                           AS NextInspectionDate
                                FROM 
                                    TBelayDevices AS TBD 
                                    LEFT JOIN TBelayDeviceInspections AS TBDIN ON TBD.intBelayDeviceID = TBDIN.intBelayDeviceID
                                    LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TBDIN.intInspectionStatusID
                                    LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TBDIN.intInspectorID
                                    LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TBDIN.intWallLocationID
                                WHERE
                                    TBD.intBelayDeviceID IN (SELECT TBDIN.intBelayDeviceID FROM TBelayDeviceInspections AS TBDIN
                                                           WHERE TBDIN.intInspectionStatusID = 2);"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")

    def dbLoad_vBelayDeviceFailNowResults(self):
        """ 
        Function Name: dbLoad_vBelayDeviceFailNowResults
        Function Purpose: Loads the view for the database if the view does not exists in the db
        """     
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Create the view names
            sqlViewName = 'vBelayDeviceFailNowResults'
            
            # Create the sqlStatement for the view
            sqlStatement = """SELECT
                            -- Create View: Show all failed now BelayDevices
                                    TBD.strSerialNum                                         AS BelayDeviceSerialNumber
                                    ,TBD.strBumperNum                                        AS BumperNum
                                    ,TWL.strWallLocationDesc                                 AS WallLocationName
                                    ,TINS.strInspectionStatusDesc                            AS OverallStatus
                                    ,TIN.strLastName || ',' || TIN.strFirstName              AS InspectorName
                                    ,TBD.dtmLastInspectionDate                               AS LastInspectionDate
                                    ,TBD.dtmNextInspectionDate                               AS NextInspectionDate
                                FROM 
                                    TBelayDevices AS TBD 
                                    LEFT JOIN TBelayDeviceInspections AS TBDIN ON TBD.intBelayDeviceID = TBDIN.intBelayDeviceID
                                    LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TBDIN.intInspectionStatusID
                                    LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TBDIN.intInspectorID
                                    LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TBDIN.intWallLocationID
                                WHERE
                                    TBDIN.intInspectionStatusID = 3 AND date(TBD.dtmLastInspectionDate) = date('now');"""      
                                                
            # Execute the SQL statements
            Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
            
            # Commit the changes 
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Error Inserting Data: {e}")                                                                                                                                                                                                                                            

    def dbLoad_ABInspectID_OutForReservice_View(self):
            """ 
            Function Name: dbLoad_ABInspectID_OutForReservice_View
            Function Purpose: Loads the newly added view for the database if the view does not exists in the db
            """     
            try:
                # First check if connected to the database   
                if not self.conn:
                    raise Exception("Database is not connected.")

                # Create the view names
                sqlViewName = 'vABInspectID_OutForReservice'
                
                # Create the sqlStatement for the view
                sqlStatement = """SELECT
                                -- Create View: Show all 'Out For Reservice' AutoBelays
                                        MAX(TABIN.intAutoBelayInspectionID)                 AS MaxInspectionID
                                        ,TAB.strSerialNum                                   AS AutoBelaySerialNumber
                                        ,TAB.strBumperNum                                   AS BumperNum
                                        ,TINS.strInspectionStatusDesc                       AS OverallStatus
                                        ,TAB.dtmLastInspectionDate                          AS LastInspectionDate
                                        ,TAB.dtmServiceDate                                 AS ServiceDate
                                        ,TAB.dtmReserviceDate                               AS ReServiceDate
                                        ,TABR.dtmReportDate                                 AS ReportDate
                                    FROM 
                                        TAutoBelayReserivceReports AS TABR 
                                        JOIN TAutoBelayInspections AS TABIN ON TABR.intAutoBelayInspectionID = TABIN.intAutoBelayInspectionID
                                        JOIN TAutoBelays AS TAB ON TAB.intAutoBelayID = TABIN.intAutoBelayID
                                        LEFT JOIN TInspectionStatus AS TINS ON TINS.intInspectionStatusID = TABIN.intInspectionStatusID
                                        LEFT JOIN TInspectors AS TIN ON TIN.intInspectorID = TABIN.intInspectorID
                                        LEFT JOIN TWallLocations AS TWL ON TWL.intWallLocationID = TABIN.intWallLocationID
                                    WHERE
                                        TAB.blnDeviceInUse = 'Out For Reservice' AND TAB.intAutoBelayID = TABIN.intAutoBelayID
                                    GROUP BY TAB.intAutoBelayID;"""                     
            
                            
                # Execute the SQL statements
                Queries.Create_New_Views(Queries, sqlViewName, sqlStatement)       
                
                # Commit the changes 
                self.conn.commit()
            except sqlite3.Error as e:
                print(f"Error Inserting Data: {e}") 
                                                            
class Queries(Database):
    """
    Class Name: SQLQueries
    Class Description: This class is the main NOSQL database for the entire program
    """    
    def Get_MaxPrimaryKeys(self, strTable, intPrimaryKeyCol):
        """
        Function Name: Get_MaxPrimaryKeys
        Function Description: This function is used to get the max value
        """
        # Declare Primary sql statement and pass in the table and primary key column
        sqlQuery = f"SELECT MAX({intPrimaryKeyCol}) FROM {strTable}"
        sqlMaxQuery = f"SELECT MAX({intPrimaryKeyCol}) + 1 FROM {strTable}"
        sqlCoalQuery = f"SELECT COALESCE(MAX({intPrimaryKeyCol}), 1) FROM {strTable}"
                
        try:
            # # Connect to the database
            # Database.dbConnect(self)
            
            # Execute the SQL query
            cursor = self.conn.cursor()
            cursor.execute(sqlQuery)
            
            # Fetch the result
            result = cursor.fetchone()
            maxPrimaryKey = result[0]
            
            # Check if the table is empty 
            if maxPrimaryKey is None:
                cursor.execute(sqlCoalQuery)
                
                # Fetch the result
                result = cursor.fetchone()
                maxPrimaryKey = result[0]
                
            else:
                cursor.execute(sqlMaxQuery)
                                
                # Fetch the result
                result = cursor.fetchone()
                maxPrimaryKey = result[0]  
 
            # Close the connection 
            cursor.close()

            # Return the result
            return maxPrimaryKey
            
        except sqlite3.Error as e:
            print(f"Error executing SQL statement: {e}")   

    def Insert_Values(self, params):
        """
        Function Name: Insert_Values
        Function Description: This function is used to insert values into non-Standard Tables.
        """
        strTable = params[0]
        aTableCol = params[1]
        aTableValues = params[2]
                
        try:
            # Declare Primary sql statement and pass in the table and primary key column
            sqlQuery = f"INSERT INTO {strTable} {aTableCol} VALUES {aTableValues}"
            
            # Execute the SQL query
            cursor = self.conn.cursor()
            cursor.execute(sqlQuery)

            # Close the connection 
            cursor.close()
                        
        except sqlite3.Error as e:
            print(f"Error executing SQL statement: {e}")   
            self.conn.rollback()

    def Update_Values(self, params):
        """
        Function Name: Update_Values
        Function Description: This function is used to update values into non-Standard Tables.
        """
        strTable = params[0]
        aTableCol = params[1]
        aTableValues = params[2]
        intPrimID = params[3]
        intKeyID = params[4]
        
        try:
            # Create the SET part of the SQL query dynamically
            set_part = ', '.join(f"{col} = ?" for col in aTableCol)
            sqlQuery = f"UPDATE {strTable} SET {set_part} WHERE {intPrimID} = ?"

            # Prepare values for the parameterized query
            values_to_update = list(aTableValues) 
            values_to_update.append(intKeyID)

            # Execute the SQL query
            cursor = self.conn.cursor()
            cursor.execute(sqlQuery, values_to_update)
            cursor.close()

        except sqlite3.Error as e:
            print(f"Error executing SQL statement: {e}")
            self.conn.rollback()       
            
    def dbExeUSP_AddValues(self, params):
        """
        Function Name: dbExeUSP_AddValues
        Function Description: This function is used to insert values into non-Standard Tables.
        """
        try:
            if self.conn:            
                # Begin the transaction
                cursor = self.conn.cursor()
                cursor.execute("BEGIN TRANSACTION")
                
                # Insert the new primary key into the database
                Queries.Insert_Values(self, params)
                
                # Commit the transaction
                cursor.execute("COMMIT")
                self.conn.commit()
                
                # Close the connection 
                cursor.close()

        except sqlite3.Error as e:
            print(f"Error executing SQL statement: {e}")   
            self.conn.rollback()

    def dbExeUSP_UpdateValues(self, params):
        """
        Function Name: dbExeUSP_UpdateValues
        Function Description: This function is used to update values into non-Standard Tables.
        """
        try:
            if self.conn:            
                # Begin the transaction
                cursor = self.conn.cursor()
                cursor.execute("BEGIN TRANSACTION")
                
                # Update the specific primary key and columns
                Queries.Update_Values(self, params)

                # Commit the transaction
                cursor.execute("COMMIT")
                self.conn.commit()
                
                # Close the connection 
                cursor.close()                
                
        except sqlite3.Error as e:
            print(f"Error executing SQL statement: {e}")   
            self.conn.rollback()

    def Remove_Attribute_Query(self, strTable, intPrimID, intKeyID):
        """ 
        Function Name: Remove_Attribute_Query
        Function Description: This function updates the database by removing objects inside the db table
        """    
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
                                    
            # Create the SQL delete query
            sqlQuery = f"DELETE FROM {strTable} WHERE {intPrimID} = {intKeyID}"

            # Execute the SQL query   
            Database.dbExeStatement(self, sqlQuery)
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error executing SQL statement: {e}")
            self.conn.rollback()   
                        
    def dbCreateViews(self, full_view_name, sql):
        """
        Function Name: dbCreateViews
        Function Description: This function is used to create views inside the sqlite3 database
        """
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")

            # Split the full_view_name to extract the target_db and viewName
            parts = full_view_name.split('.')
            if len(parts) == 2:
                target_db, viewName = parts
            else:
                # If the view name does not include a period, assume it's in the main database
                target_db = "main"
                viewName = full_view_name

            # Construct the SQL command
            sql_command = f"CREATE VIEW IF NOT EXISTS {viewName} AS {sql}"
            
            # Determine which execution method to use
            if target_db == "main":
                # Execute the SQL command
                self.dbExeStatement(sql_command)

                # Commit the changes
                self.conn.commit()
            else:
                # Direct execution for other databases
                cursor = self.conn.cursor()
                try:                
                    # Execute the SQL command
                    cursor.executescript(sql_command)

                    # Commit the changes
                    self.conn.commit()
                    # print(f"SQL executed successfully for {target_db} database:", sql)
                except sqlite3.Error as e:
                    print(f"Error executing SQL statement: {e}")
                finally:
                    cursor.close()
                    
        except sqlite3.Error as e:
            print(f"Error creating SQLite view: {e}")   

    def Create_New_Views(self, viewName, sql):
        """
        Function Name: dbCreateViews
        Function Description: This function is used to create views inside the sqlite3 database
        """
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
                        
            # Declare Local Variables
            sql_command = f"CREATE VIEW IF NOT EXISTS {viewName} AS {sql}"
            
            # Create the view
            Database.dbExeStatement(Database, sql_command)

            # Commit the changes and close the connection
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating SQLite view: {e}") 
            
    def dbCreate_Functions(self, fnName, selParams, strTable, intPriKeyID):
        """ 
        Function Name: fn_GetInspectionStatus
        Function Purpose: Creates the GetInspectionStatus function for the database
        """          
        try:
            # First check if connected to the database   
            if not self.conn:
                raise Exception("Database is not connected.")
            
            # Define the SQL statement for the function
            sql = f'SELECT {selParams} FROM {strTable} WHERE {intPriKeyID} = ?'
            
            # Get the result of the SQL statement 
            sqlResult = Database.dbExeQuery(self, sql, (intPriKeyID,))

            # Register the function with SQLite
            self.conn.create_function(fnName, 1, sqlResult)

            # Commit the changes 
            self.conn.commit()

        except sqlite3.Error as e:
            print(f"Error creating function: {e}")  

    def Get_StatusID(self, params):
        """
        Function Name: Get_StatusID
        Function Purpose: Execute the given SQL statement and return the statusID
        """
        # Set the sql statement
        sql = f"""SELECT intInspectionStatusID FROM {params[0]} WHERE {params[1]} = {params[2]}"""
                
        try:
            # Create the cursor object
            cursor = self.conn.cursor()
            cursor.execute(sql)
            
            # Fetch the result
            result = cursor.fetchone()  
            intStatusID = result[0]
            
            # Close the cursor
            cursor.close()

            # Extract and return the inspection status if found, or None otherwise
            if intStatusID:
                return intStatusID
            else:
                return None     
            
        except sqlite3.Error as e:
            print(f"Error executing SQL statement: {e}") 

    def Get_Specific_DB_Value(self, value, table):
        """
        Function Name: Get_Specific_DB_Value
        Function Purpose: Execute the given SQL statement and return the desired DB value
        """
        # Set the sql statement
        sql = f"""SELECT {value} FROM {table}"""
                
        try:
            # Create the cursor object
            cursor = self.conn.cursor()
            cursor.execute(sql)
            
            # Fetch the result
            result = cursor.fetchone()  
            value = result[0]
            
            # Close the cursor
            cursor.close()

            # Extract and return the inspection status if found, or None otherwise
            if value:
                return value
            else:
                return None     
            
        except sqlite3.Error as e:
            print(f"Error executing SQL statement: {e}")
            
    def Get_DB_Value(self, value, params):
        """
        Function Name: Get_DB_Value
        Function Purpose: Execute the given SQL statement and return the desired DB value
        """
        # Set the sql statement
        sql = f"""SELECT {value} FROM {params[0]} WHERE {params[1]} = {params[2]}"""
                
        try:
            # Create the cursor object
            cursor = self.conn.cursor()
            cursor.execute(sql)
            
            # Fetch the result
            result = cursor.fetchone()  
            value = result[0]
            
            # Close the cursor
            cursor.close()

            # Extract and return the inspection status if found, or None otherwise
            if value:
                return value
            else:
                return None     
            
        except sqlite3.Error as e:
            print(f"Error executing SQL statement: {e}")

    def Get_All_DB_Values_OnePrimKey(self, params):
        """
        Function Name: Get_All_DB_Values_OnePrimKey
        Function Purpose: Execute the given SQL statement and return the desired DB values
        """
        # Set the sql statement
        sql = f"""SELECT * FROM {params[0]} WHERE {params[1]} = {params[2]}"""

        try:
            # Create the cursor object
            cursor = self.conn.cursor()
            cursor.execute(sql)
            
            # Fetch the result
            result = cursor.fetchall()  
            value = result[0]
            
            # Close the cursor
            cursor.close()

            # Extract and return the inspection status if found, or None otherwise
            if value:
                return value
            else:
                return None     
            
        except sqlite3.Error as e:
            print(f"Error executing SQL statement: {e}")

    def Get_All_DB_Values_Not_Connected(self, strTable):
        """
        Function Name: Get_All_DB_Values_Not_Connected
        Function Purpose: Execute the given SQL statement and return the desired DB values
        """
        # Set the sql statement
        sql = f"""SELECT * FROM {strTable}"""

        try:
            # Create the cursor object
            self.conn = sqlite3.connect(Database.db_path)
            cursor = self.conn.cursor()
            cursor.execute(sql)
            
            # Fetch the result
            result = cursor.fetchall()  
            
            # Close the cursor
            cursor.close()
            
            return result

        except sqlite3.Error as e:
            print(f"Error executing SQL statement: {e}")
            return None

    def Get_All_DB_Values(self, strTable):
        """
        Function Name: Get_All_DB_Values
        Function Purpose: Execute the given SQL statement and return the desired DB values
        """
        # Set the sql statement
        sql = f"""SELECT * FROM {strTable}"""

        try:
            # Create the cursor object
            self.conn = sqlite3.connect(Database.db_path)
            cursor = self.conn.cursor()
            cursor.execute(sql)
            
            # Fetch the result
            result = cursor.fetchall()  
            
            # Close the cursor
            cursor.close()
            
            return result

        except sqlite3.Error as e:
            print(f"Error executing SQL statement: {e}")
            return None        
            
    def Get_Duplicate_Data(self, params):
        """
        Function Name: Get_Duplicate_Data
        Function Purpose: Execute the given SQL statement and check the table for any duplicate data entry and
        use the data ID and associated data attributes as the result
        """
        # Set the SQL statement to find duplicate data entries
        if len(params) > 2:
            sql = f"SELECT {params[1]}, {params[2]} FROM {params[0]}"
        else:
            sql = f"SELECT {params[1]} FROM {params[0]}"

        try:
            # Create the cursor object
            cursor = self.conn.cursor()
            cursor.execute(sql)

            # Fetch the result
            result = cursor.fetchall()

            # Close the cursor
            cursor.close()

            # Return the duplicate data and associated primary keys
            return result

        except sqlite3.Error as e:
            print(f"Error executing SQL statement: {e}")
                
    def Open_XL_File(self): 
        """ 
        Function Name: Open_XL_File
        Function Purpose: This button executes when the user wants to open the parent dir and open XL files. Pops up
        the file dialog and displays contents within the parent dir.
        """    
        # Create the file list available to the user
        fList = [("Excel files","*.xlsx")]
        
        # Mask the file types available for input
        file = filedialog.askopenfilename(title="Select File",
                                        initialdir="RecordsDir", 
                                        filetypes=fList) 
        # Create a new variable for file so python doesn't get confused
        fName = file

        # Open the file in read mode    
        with open(fName, 'r') as f:
            while True:
                try:
                    # Open and load the file
                    if sys.platform.startswith('darwin'):
                        subprocess.call(('open', fName))
                        f.close()
                        break
                    elif sys.platform.startswith('win32'):
                        subprocess.call(('open', fName))
                        f.close()
                        break
                    elif sys.platform.startswith('linux'):
                        subprocess.call(('open', fName))
                        f.close()
                        break                    
                except EOFError:
                    f.close()
                    break

    def Select_XL_File(self): 
        """ 
        Function Name: Select_XL_File
        Function Purpose: This button executes when the user wants to select the parent dir and send XL files via email. 
        Pops up the file dialog and displays contents within the parent dir. Return the file name to send
        """    
        # Get the parent directory of the current working directory
        parentDir = os.path.abspath(os.path.join(os.getcwd(), os.pardir))
        
        # Open a directory dialog with the parent directory as the initial directory
        dirPath = filedialog.askdirectory(initialdir=parentDir)
        
        # Get the string of the file
        fString = str(dirPath)
        
        # Get the file selected
        dirList = fString.split(os.path.sep)  
        lastFile = dirList[-1]
        
        # Set the file name selected by the user
        self.strFileName = lastFile

    def SaveAs_XL_File(self, sqlQuery): 
        """ 
        Function Name: SaveAs_XL_File
        Function Purpose: This button executes when the user wants to open the parent dir and write/save XL files. Pops up
        the file dialog and displays contents within the dir to name and save under the parent dir.
        """    
        # Get todays date
        dtmToday = datetime.date(datetime.now())
        dtmToday = datetime.strftime(dtmToday, '%m/%d/%Y')
        dtmToday = str(dtmToday)
            
        # Create the file list available to the user
        fList = [("Excel files","*.xlsx")]
        
        # Mask the file types available for input
        file = filedialog.asksaveasfilename(title = "Save File", 
                                            initialdir='RecordsDir',
                                            initialfile='_'+dtmToday,
                                            filetypes = fList,
                                            defaultextension=(".xlsx"),
                                            confirmoverwrite=True) 
        # Create a new file variable name so that python doesn't get confused
        fNewFile = str(file)
        
        # Check if the file exists
        if fNewFile.exists() is True: 

            # Open the file in write mode 
            while True:
                try:
                    # Open and load the file
                    if sys.platform.startswith('darwin'):
                        # Create the excel file by passing in the sql query from the db for UNIX
                        Queries.Create_XL_Write(sqlQuery, fNewFile)
                        break
                        
                    elif sys.platform.startswith('win32'):
                        # Create the excel file by passing in the sql query from the db for Win
                        Queries.Create_XL_Write(sqlQuery, fNewFile)
                        break
                    
                    elif sys.platform.startswith('linux'):
                        # Create the excel file by passing in the sql query from the db for Linux
                        Queries.Create_XL_Write(sqlQuery, fNewFile)
                        break
                                
                except EOFError:
                    break
   
#######################################################################################################
# Backup Database Class
#######################################################################################################
class BackupDb():
    """
    Class Name: BackupDb
    Class Description: This class is the main Backup database for the entire program
    """    
    def __init__(self, dbBackup_path=None):
        """ 
        Function Name: __init__
        Function Purpose: Instantiate the class objects and attributes 
        """           
        self.dbBackup_path = dbBackup_path    

    # Create a string of the backup database name
    def __str__(self):
        return str(self.dbBackup_path)
        
    # Property decorator object get function to access private dbBackup_path
    @property
    def dbBackup_path(self):
        return self._dbBackup_path

    # setter method 
    @dbBackup_path.setter 
    def dbBackup_path(self, dbBackup_path): 
        # Return true if specified object is of str type
        if not isinstance(dbBackup_path, str): 
            raise TypeError('Backup Database Path must be a string') 
        # Check if the value is empty, otherwise check if the value is alpha
        if dbBackup_path.isspace(): 
            raise ValueError('Backup Database Path cannot be empty') 
        # Set the attribute to the value if true
        elif dbBackup_path.isascii():
            self._dbBackup_path = dbBackup_path

    def Find_Oldest_BackupFile(self):
        """ 
        Function Name: Find_Oldest_BackupFile
        Function Purpose: Find the oldest .zip file matching the pattern 
                        Database.db_YYYYMMDD_HHMMSS.zip in the specified backup directory 
                        and create a copy of the unzipped database.
        """         
        backup_dir = os.path.join(BackupDb.dbBackup_path, "BackupDir")

        if not os.path.exists(backup_dir):
            print(f"'BackupDir' not found at: {backup_dir}")
            return None

        # First, check the contents of BackupDir for .zip files
        zip_files = glob.glob(f"{backup_dir}/Database.db_????????_??????.zip")

        # If no .zip files were found in BackupDir, check its subdirectories
        if not zip_files:
            zip_files = glob.glob(f"{backup_dir}/**/Database.db_????????_??????.zip", recursive=True)

        # If still no .zip files are found, return None
        if not zip_files:
            return None

        oldest_zip = max(zip_files, key=os.path.getctime)
        
        # Assuming the destination directory is the same as backup_dir, adjust as needed
        decryption_password = Database.db_password
        strOldestBackupFile = Database.Unzip_Decrypt(oldest_zip, backup_dir, decryption_password)

        # Constructing the path to the unzipped database file based on the zip file's name
        unzipped_db_filename = os.path.basename(strOldestBackupFile)
        unzipped_db_path = os.path.join(backup_dir, unzipped_db_filename)

        # Copy the unzipped database to a new file
        db = Database(db_name=unzipped_db_filename, db_path=unzipped_db_path, db_password=Database.db_password)

        return db.db_name, db.db_path
        

#######################################################################################################
# Records Class
#######################################################################################################
class Records():
    """
    Class Name: Records
    Class Description: This class is the main Records for the entire program
    """    
    def __init__(self, strRecords_path=None):
        """ 
        Function Name: __init__
        Function Purpose: Instantiate the class objects and attributes 
        """           
        self.strRecords_path = strRecords_path    

    # Create a string of the records name
    def __str__(self):
        return str(self.strRecords_path)
        
    # Property decorator object get function to access private strRecords_path
    @property
    def strRecords_path(self):
        return self._strRecords_path

    # setter method 
    @strRecords_path.setter 
    def strRecords_path(self, strRecords_path): 
        # Return true if specified object is of str type
        if not isinstance(strRecords_path, str): 
            raise TypeError('Records Path must be a string') 
        # Check if the value is empty, otherwise check if the value is alpha
        if strRecords_path.isspace(): 
            raise ValueError('Records Path cannot be empty') 
        # Set the attribute to the value if true
        elif strRecords_path.isascii():
            self._strRecords_path = strRecords_path                                                

    def Open_Records_Dir(): 
        """ 
        Function Name: Open_Records_Dir
        Function Purpose: This button executes when the user wants to open the parent dir. Pops up
        the file dialog and displays contents within the parent dir.
        """    
        # Get the directory path
        dirPath = Records.strRecords_path
            
        # Open the directory using the default file explorer for the respective OS
        if sys.platform.startswith('darwin'):
            # MacOS (using Finder)
            subprocess.call(['open', dirPath])
        elif sys.platform.startswith('linux'):
            # For Linux using xdg-open (generic)
            subprocess.call(['xdg-open', dirPath])
        elif sys.platform.startswith('win32'):
            # Windows (using Explorer)
            subprocess.call(['start', dirPath], shell=True)